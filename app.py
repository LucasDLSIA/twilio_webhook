import os
import io
import hmac
import hashlib
import logging
import re
import time

import json
import threading
from datetime import datetime
import datetime as _dt

from typing import Optional, Dict, List
from urllib.parse import urlencode, quote
from functools import wraps

import pandas as pd
from flask import Flask, request, redirect, Response, jsonify, session, g, has_app_context
from werkzeug.middleware.proxy_fix import ProxyFix

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.http import MediaIoBaseUpload
 

from twilio.rest import Client
from twilio.request_validator import RequestValidator


# =========================
# Config
# =========================
ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN", "").strip()
EMPRESAS_FILE_ID = os.environ.get("EMPRESAS_FILE_ID", "").strip()

# --- Seguridad (ver CAMBIOS.md) ---
SECRET_KEY = os.environ.get("SECRET_KEY", "").strip()
# Secreto para firmar las URLs de /media/pdf y /media/terms. Si no se define, usa SECRET_KEY.
MEDIA_SECRET = os.environ.get("MEDIA_SECRET", "").strip() or SECRET_KEY
# Vigencia (segundos) de los links de media firmados. Default: 24 hs.
MEDIA_TOKEN_TTL = int(os.environ.get("MEDIA_TOKEN_TTL", "86400"))
# Validación de firma de webhooks de Twilio (X-Twilio-Signature). Poner "0" solo para debug local.
TWILIO_VALIDATE_WEBHOOKS = os.environ.get("TWILIO_VALIDATE_WEBHOOKS", "1").strip() != "0"

# Google Service Account
GOOGLE_SA_JSON = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
GOOGLE_SA_FILE = (
    os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE", "").strip()
    or os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "").strip()
    or ("/etc/secrets/Service_account.json" if os.path.exists("/etc/secrets/Service_account.json") else "")
)

# Twilio
TWILIO_ACCOUNT_SID = os.environ.get("TWILIO_ACCOUNT_SID", "").strip()
TWILIO_AUTH_TOKEN = os.environ.get("TWILIO_AUTH_TOKEN", "").strip()
TWILIO_WHATSAPP_FROM = os.environ.get("TWILIO_WHATSAPP_FROM", "").strip()
TWILIO_MESSAGING_SERVICE_SID = os.environ.get("TWILIO_MESSAGING_SERVICE_SID", "").strip()

# Plantillas (Content Templates)
TWILIO_TEMPLATE_SID = os.environ.get("TWILIO_TEMPLATE_SID", "").strip()            # template con botón VIEW_NOW
TWILIO_SIGN_TEMPLATE_SID = os.environ.get("TWILIO_SIGN_TEMPLATE_SID", "").strip()  # (opcional) template con botones SIGN_OK / SIGN_OBS

# Casilla que recibe el aviso de facturación cuando termina un envío INITIAL.
BILLING_NOTIFY_EMAIL = os.environ.get("BILLING_NOTIFY_EMAIL", "").strip()

# Cache
_EMP_CACHE = {"ts": 0.0, "rows": []}
_ENV_CACHE: Dict[str, Dict] = {}  # tenant_slug -> {"ts":..., "rows":[...] }
CACHE_TTL = int(os.environ.get("CACHE_TTL", "120"))


PUBLIC_BASE_URL = os.environ.get("PUBLIC_BASE_URL", "").strip()  # ej: https://miapp.com
# =========================



# Flask
# =========================
app = Flask(__name__)

if not SECRET_KEY:
    # Fail-fast: sin SECRET_KEY las sesiones del portal quedan rotas/inseguras en silencio.
    raise RuntimeError("Falta SECRET_KEY en ENV: configurala antes de iniciar la app.")
app.secret_key = SECRET_KEY

# La app corre detrás de un proxy (Render/Heroku/etc.): respetar X-Forwarded-Proto/Host
# para que request.url y request.host_url reflejen la URL pública real (https).
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

# --- Logging ---
# Nivel configurable con LOG_LEVEL (DEBUG/INFO/WARNING/ERROR). Default: INFO.
# Los errores dentro de except ahora loguean el traceback completo (log.exception).
_log_level = os.environ.get("LOG_LEVEL", "INFO").strip().upper()
logging.basicConfig(
    level=getattr(logging, _log_level, logging.INFO),
    format="%(asctime)s %(levelname)s %(message)s",
)
log = logging.getLogger("app")
# La librería de Twilio loguea cada request/response a su API a nivel INFO
# (bloques "-- BEGIN Twilio API Request --"): puro ruido, lo subimos a WARNING.
logging.getLogger("twilio").setLevel(logging.WARNING)

# --- Cookies de sesión (admin y portal) ---
# SameSite=Lax bloquea CSRF en los POST autenticados por sesión: el navegador
# no manda la cookie en POSTs originados en otros sitios.
# COOKIE_SECURE=0 solo para desarrollo local sin HTTPS.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("COOKIE_SECURE", "1").strip() != "0",
)

# --- Sentry: alertas de errores por mail (opcional) ---
# Se activa solo si SENTRY_DSN está definida en el entorno. Sin la variable,
# este bloque no hace nada y la app funciona exactamente igual que siempre.
# Solo errores (sin tracing de performance: cuida la cuota gratis) y sin PII
# por defecto; Sentry además enmascara campos tipo token/password al recibir.
# Los log.error / log.exception del código llegan como eventos con traceback;
# los log.info quedan como "migas" de contexto del evento.
SENTRY_DSN = (os.environ.get("SENTRY_DSN") or "").strip()
if SENTRY_DSN:
    import sentry_sdk
    sentry_sdk.init(
        dsn=SENTRY_DSN,
        send_default_pii=False,
        traces_sample_rate=0.0,
        environment=os.environ.get("SENTRY_ENVIRONMENT", "production"),
    )
    log.info("Sentry activo (alertas de errores encendidas)")



# =========================
# Helpers
# =========================
def esc(s: str) -> str:
    return (
        (str(s or ""))
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )

def slugify(name: str) -> str:
    name = (name or "").strip().lower()
    name = re.sub(r"[^a-z0-9]+", "-", name)
    name = re.sub(r"-{2,}", "-", name).strip("-")
    return name or "empresa"

def norm_digits(s: str) -> str:
    return re.sub(r"\D", "", str(s or ""))

def norm_cuil(s: str) -> str:
    return re.sub(r"\D+", "", (s or "")).strip()


def strip_pdf(name: str) -> str:
    s = str(name or "").strip()
    if s.lower().endswith(".pdf"):
        s = s[:-4]
    return s.strip()

def norm_whatsapp(s: str) -> str:
    d = norm_digits(s)
    if not d:
        return ""

    # ya viene correcto
    if d.startswith("549"):
        return "whatsapp:+" + d

    # viene con 54 pero sin 9
    if d.startswith("54"):
        return "whatsapp:+549" + d[2:]

    # número local
    return "whatsapp:+549" + d

def parse_period_folder(name: str) -> Optional[str]:
    """
    Acepta: '01-2026', '01/2026'
    Devuelve siempre 'mm/aaaa' o None si no reconoce.
    """
    n = (name or "").strip()
    m = re.search(r"(\d{2})[-/](\d{4})", n)
    if m:
        mm, yyyy = m.group(1), m.group(2)
        if 1 <= int(mm) <= 12:
            return f"{mm}/{yyyy}"
    return None

def period_to_folder_name(period: str) -> str:
    """
    Recibe 'mm/aaaa' o 'mm-aaaa' y devuelve 'mm-aaaa' (nombre de carpeta en Drive).
    """
    p = (period or "").strip().replace("-", "/")
    mm, yyyy = p.split("/")
    return f"{mm}-{yyyy}"

# =========================
# Auth admin (token por query/header)
# =========================
def make_media_token(kind: str, tenant: str = "", cuil: str = "", period: str = "", ttl: int | None = None) -> str:
    """Token firmado (HMAC) con vencimiento para los endpoints /media/*.
    Reemplaza el uso de ADMIN_TOKEN en URLs que viajan a Twilio."""
    exp = int(time.time()) + int(ttl if ttl is not None else MEDIA_TOKEN_TTL)
    msg = f"{kind}|{tenant}|{cuil}|{period}|{exp}".encode("utf-8")
    sig = hmac.new(MEDIA_SECRET.encode("utf-8"), msg, hashlib.sha256).hexdigest()
    return f"{exp}.{sig}"

def check_media_token(token: str, kind: str, tenant: str = "", cuil: str = "", period: str = "") -> bool:
    """Valida un token generado por make_media_token (firma + vencimiento + alcance)."""
    try:
        exp_s, sig = (token or "").split(".", 1)
        exp = int(exp_s)
    except (ValueError, AttributeError):
        return False
    if exp < int(time.time()):
        return False
    msg = f"{kind}|{tenant}|{cuil}|{period}|{exp}".encode("utf-8")
    expected = hmac.new(MEDIA_SECRET.encode("utf-8"), msg, hashlib.sha256).hexdigest()
    return hmac.compare_digest(sig, expected)

# =========================
# Autenticación del admin (sesión + token para automatizaciones)
# =========================
# Vigencia de la sesión del admin (segundos). Default: 8 horas.
ADMIN_SESSION_TTL = int(os.environ.get("ADMIN_SESSION_TTL", str(8 * 3600)))

def admin_session_active() -> bool:
    """True si hay una sesión de admin válida y no vencida."""
    if not session.get("admin_authed"):
        return False
    since = session.get("admin_authed_at", 0)
    try:
        since = int(since)
    except (TypeError, ValueError):
        return False
    return (int(time.time()) - since) < ADMIN_SESSION_TTL

def _admin_token_match(tok: str) -> bool:
    if not ADMIN_TOKEN:
        return False
    tok = (tok or "").strip()
    return bool(tok) and hmac.compare_digest(tok, ADMIN_TOKEN)

def admin_token_valid() -> bool:
    """Token por CUALQUIER vía (query, form o header).
    Reservado a los endpoints de automatización que lo necesitan por URL:
    queue_tick (cron y self-call de fondo), reset_reenvios,
    send_template_preview y media_certificado. El resto del panel NO
    acepta token por URL: humanos por login, máquinas por header."""
    return _admin_token_match(request.args.get("token") or request.form.get("token")
                              or request.headers.get("X-Admin-Token", ""))

def admin_token_valid_header() -> bool:
    """Token solo por header X-Admin-Token (no queda en access logs ni historial)."""
    return _admin_token_match(request.headers.get("X-Admin-Token", ""))

def admin_authenticated() -> bool:
    """Humanos: sesión (login). Máquinas: header X-Admin-Token.
    El token por URL ya NO abre el panel (ver allowlist en admin_token_valid)."""
    return admin_session_active() or admin_token_valid_header()

def admin_ok() -> bool:
    # Mantengo el nombre por compatibilidad: ahora acepta sesión o token.
    # (Antes devolvía True si ADMIN_TOKEN estaba vacío, dejando el admin abierto.)
    return admin_authenticated()

def _admin_next_path() -> str:
    """Destino post-login: path actual con su query pero SIN el token,
    para que un ?token= viejo no siga paseando por la URL del login."""
    qs = [(k, v) for k, v in request.args.items(multi=True) if k != "token"]
    return request.path + ("?" + urlencode(qs) if qs else "")

def require_admin():
    if admin_authenticated():
        return None
    # Si parece un navegador (acepta HTML) y no es POST, lo mandamos al login
    # preservando el destino. Para clientes no-navegador / POST, 401 directo.
    accepts_html = "text/html" in (request.headers.get("Accept") or "")
    if request.method == "GET" and accepts_html:
        return redirect("/admin/login?next=" + quote(_admin_next_path(), safe=""))
    return Response("Unauthorized (admin login requerido)", status=401)

def _get_admin_token_from_request() -> str:
    return (request.args.get("token") or request.form.get("token") or "").strip()

def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        # Acepta sesión de admin (humano logueado) o ADMIN_TOKEN (cron/background/self-call).
        if admin_authenticated():
            return fn(*args, **kwargs)
        accepts_html = "text/html" in (request.headers.get("Accept") or "")
        if request.method == "GET" and accepts_html:
            return redirect("/admin/login?next=" + quote(_admin_next_path(), safe=""))
        return Response("Unauthorized", status=401)
    return wrapper

# =========================
# Google Drive
# =========================
def drive_service():
    scopes = ["https://www.googleapis.com/auth/drive"]
    if GOOGLE_SA_JSON:
        info = json.loads(GOOGLE_SA_JSON) if isinstance(GOOGLE_SA_JSON, str) else GOOGLE_SA_JSON
        creds = service_account.Credentials.from_service_account_info(info, scopes=scopes)
    elif GOOGLE_SA_FILE:
        creds = service_account.Credentials.from_service_account_file(GOOGLE_SA_FILE, scopes=scopes)
    else:
        raise RuntimeError("Falta GOOGLE_SERVICE_ACCOUNT_JSON o GOOGLE_SERVICE_ACCOUNT_FILE/GOOGLE_APPLICATION_CREDENTIALS en ENV")
    return build("drive", "v3", credentials=creds, cache_discovery=False)

def download_excel_df(file_id: str) -> pd.DataFrame:
    service = drive_service()
    req = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.seek(0)
    df = pd.read_excel(fh)
    df.columns = [str(c).strip() for c in df.columns]
    return df

# =========================
# Tenants (Empresas.xlsx)
# =========================
_TENANTS_CACHE = {"ts": 0, "items": []}
_TENANTS_TTL = 60  # segundos

def load_tenants(force: bool = False) -> list[dict]:
    """
    Lee el Excel maestro (EMPRESAS_FILE_ID) y devuelve tenants normalizados.
    Soporta headers: Empresa, Envios_File_ID, Drive_Root_ID (case-insensitive)
    y también: slug, display_name, envios_file_id, recibos_root_id, drive_root_id, root_id.
    """
    now = time.time()
    if (not force) and _TENANTS_CACHE["items"] and (now - _TENANTS_CACHE["ts"] < _TENANTS_TTL):
        return _TENANTS_CACHE["items"]

    df = download_excel_df(EMPRESAS_FILE_ID)
    if df is None or df.empty:
        _TENANTS_CACHE.update({"ts": now, "items": []})
        return []

    df.columns = [str(c).strip().lower() for c in df.columns]

    def pick(*names):
        for n in names:
            if n in df.columns:
                return n
        return None

    c_slug = pick("slug", "empresa", "tenant")
    c_name = pick("display_name", "nombre", "name", "empresa")
    c_env  = pick("envios_file_id", "envios_file_id ", "envios", "envios_id", "envios_file", "enviosfileid")
    c_root = pick("drive_root_id", "recibos_root_id", "root_id", "drive_root_folder_id", "carpeta_root_id", "drive_root")

    # MUY IMPORTANTE: tu Excel real trae Envios_File_ID y Drive_Root_ID
    # al bajar a lower quedan: envios_file_id y drive_root_id → con esto lo levanta.

    items = []
    for _, r in df.iterrows():
        raw_slug = str(r.get(c_slug, "")).strip()
        if not raw_slug:
            continue

        slug = slugify(raw_slug)  # o tu normalizador de slug
        display_name = str(r.get(c_name, raw_slug)).strip() if c_name else raw_slug

        envios_file_id = str(r.get(c_env, "")).strip() if c_env else ""
        drive_root_id  = str(r.get(c_root, "")).strip() if c_root else ""

        # Normalizamos claves: SIEMPRE devolvemos ambos nombres
        items.append({
            "slug": slug,
            "display_name": display_name,
            "envios_file_id": envios_file_id,
            "drive_root_id": drive_root_id,
            "recibos_root_id": drive_root_id,   # compatibilidad con funciones viejas
        })

    _TENANTS_CACHE.update({"ts": now, "items": items})
    return items

def get_tenant(slug: str) -> dict | None:
    slug = (slug or "").strip().lower()
    if not slug:
        return None

    for t in load_tenants():
        if (t.get("slug") or "").strip().lower() == slug:
            # fallback por si falta alguno
            if not t.get("recibos_root_id") and t.get("drive_root_id"):
                t["recibos_root_id"] = t["drive_root_id"]
            if not t.get("drive_root_id") and t.get("recibos_root_id"):
                t["drive_root_id"] = t["recibos_root_id"]
            return t
    return None

# =========================
# Envios por tenant
# =========================
def load_envios_rows(tenant_slug: str, force: bool = False) -> List[dict]:
    t = get_tenant(tenant_slug)
    if not t:
        return []
    now = time.time()
    cached = _ENV_CACHE.get(tenant_slug)
    if (not force) and cached and (now - cached["ts"] < CACHE_TTL):
        return cached["rows"]

    df = download_excel_df(t["envios_file_id"])
    df.columns = [str(c).strip() for c in df.columns]
    rows = df.fillna("").to_dict(orient="records")

    _ENV_CACHE[tenant_slug] = {"ts": now, "rows": rows}
    return rows

def find_person_by_cuil(envios_rows: List[dict], cuil: str) -> Optional[dict]:
    """
    Tu formato real por empresa:
      nombre | telefono | archivo | DNI
    donde archivo = '20-xxxxxxxx-x.pdf'
    """
    target = norm_cuil(cuil)
    if not target:
        return None

    for r in envios_rows:
        archivo = strip_pdf(r.get("archivo") or r.get("Archivo") or "")
        if norm_cuil(archivo) == target:
            nombre = str(r.get("nombre") or r.get("Nombre") or "").strip()
            tel = str(r.get("telefono") or r.get("teléfono") or r.get("Telefono") or "").strip()
            dni = str(r.get("dni") or r.get("DNI") or "").strip()
            return {
                "cuil": archivo,
                "nombre": nombre,
                "telefono_raw": tel,
                "to_whatsapp": norm_whatsapp(tel),
                "dni": dni,
            }
    return None

# =========================
def _norm_period_variants(period: str) -> list[str]:
    """
    Devuelve posibles nombres de carpeta según el período.
    Ej: "01/2026" => ["01-2026", "01_2026", "01 2026", "01/2026", "012026", "2026-01"]
    """
    p = (period or "").strip()
    if not p:
        return []
    p2 = p.replace("-", "/").replace("_", "/").replace(" ", "/")
    parts = p2.split("/")
    if len(parts) == 2:
        mm = parts[0].zfill(2)
        yyyy = parts[1]
    else:
        mm = p[:2].zfill(2)
        yyyy = p[-4:]
    return list(dict.fromkeys([
        f"{mm}/{yyyy}",
        f"{mm}-{yyyy}",
        f"{mm}_{yyyy}",
        f"{mm} {yyyy}",
        f"{mm}{yyyy}",
        f"{yyyy}-{mm}",
        f"{yyyy}{mm}",
    ]))


def _drive_list_children(service, parent_id: str, mime_type: str | None = None, page_size: int = 200) -> list[dict]:
    """
    Lista hijos directos de una carpeta.
    """
    q = f"'{parent_id}' in parents and trashed=false"
    if mime_type:
        q += f" and mimeType='{mime_type}'"
    res = service.files().list(
        q=q,
        fields="files(id,name,mimeType)",
        pageSize=page_size
    ).execute()
    return res.get("files", [])


def _drive_find_child_by_exact_name(service, parent_id: str, name: str, mime_type: str | None = None) -> str | None:
    """
    Busca un hijo por nombre exacto dentro de una carpeta.
    """
    # OJO: name necesita escapar comillas simples si las hubiera (raro en PDFs)
    safe_name = (name or "").replace("'", "\\'")
    q = f"'{parent_id}' in parents and trashed=false and name='{safe_name}'"
    if mime_type:
        q += f" and mimeType='{mime_type}'"
    res = service.files().list(
        q=q,
        fields="files(id,name,mimeType)",
        pageSize=1
    ).execute()
    files = res.get("files", [])
    return files[0]["id"] if files else None

def debug_list_pdfs_in_folder(folder_id: str):
    service = drive_service()
    q = f"'{folder_id}' in parents and trashed=false and mimeType='application/pdf'"
    res = service.files().list(
        q=q,
        fields="files(id,name)",
        pageSize=50
    ).execute()

    files = res.get("files", [])
    log.info("%s %s", "📂 DEBUG PDFs en carpeta", folder_id)
    if not files:
        log.info("   (no hay PDFs)")
    for f in files:
        log.info("%s %s %s %s", "   -", f["name"], "| id:", f["id"])


# Drive: PDF
# =========================
def format_cuil_with_dashes(cuil: str) -> str:
    d = norm_digits(cuil)
    if len(d) != 11:
        return cuil.strip()
    return f"{d[0:2]}-{d[2:10]}-{d[10:11]}"


def find_pdf_file_id_for_cuil_period(tenant: str, cuil: str, period: str, *, quiet: bool = False) -> str | None:
    t = get_tenant(tenant)
    if not t:
        if not quiet: log.error("%s %s", "❌ tenant inválido:", tenant)
        return None

    root_id = (t.get("recibos_root_id") or t.get("drive_root_id") or "").strip()
    if not root_id:
        if not quiet: log.error("%s %s", "❌ tenant sin recibos_root_id:", tenant)
        return None

    cuil_digits = norm_digits(strip_pdf(cuil).strip())
    if len(cuil_digits) != 11:
        if not quiet: log.error("%s %s", "❌ CUIL inválido:", cuil)
        return None

    cuil_dash = format_cuil_with_dashes(cuil_digits)
    filename_exact = f"{cuil_dash}.pdf"
    period_folder_name = normalize_period_for_drive((period or "").strip())

    service = drive_service()

    q_folder = (
        f"'{root_id}' in parents and trashed=false "
        f"and mimeType='application/vnd.google-apps.folder' "
        f"and name='{period_folder_name}'"
    )
    res = service.files().list(q=q_folder, fields="files(id,name)", pageSize=5).execute()
    folders = res.get("files", [])
    if not folders:
        if not quiet:
            log.error(f"❌ No encontré carpeta período '{period_folder_name}' en root {root_id}")
        return None

    period_id = folders[0]["id"]

    q_pdf_exact = f"'{period_id}' in parents and trashed=false and name='{filename_exact}'"
    res2 = service.files().list(q=q_pdf_exact, fields="files(id,name)", pageSize=5).execute()
    files = res2.get("files", [])
    if files:
        return files[0]["id"]

    q_pdf_contains = (
        f"'{period_id}' in parents and trashed=false "
        f"and mimeType='application/pdf' "
        f"and name contains '{cuil_dash}'"
    )
    res3 = service.files().list(q=q_pdf_contains, fields="files(id,name)", pageSize=10).execute()
    files2 = res3.get("files", [])
    if files2:
        return files2[0]["id"]

    if not quiet:
        log.error(f"❌ No encontré {filename_exact} dentro de carpeta período {period_folder_name} ({period_id})")
    return None


def list_recibos_for_period(tenant: str, period_label: str) -> list[dict]:
    """
    Lista los recibos (PDFs) de un período leyendo la carpeta de Drive del tenant.
    period_label esperado: "MM/YYYY". Devuelve [{cuil, nombre, file_id}] ordenado por nombre.
    """
    t = get_tenant(tenant)
    if not t:
        return []
    root_id = (t.get("recibos_root_id") or t.get("drive_root_id") or "").strip()
    if not root_id:
        return []

    folder_name = label_to_period_folder(period_label)  # "MM/YYYY" -> "MM-YYYY"
    if not folder_name:
        return []

    service = drive_service()
    period_id = _drive_find_child_folder_id(service, root_id, folder_name)
    if not period_id:
        return []

    files = _drive_list_children(
        service, parent_id=period_id, mime_type="application/pdf", page_size=1000
    )

    # Mapa CUIL(11 dígitos) -> nombre desde la planilla de envíos (una sola lectura, ya cacheada)
    nombre_por_cuil = {}
    for r in load_envios_rows(tenant):
        archivo = strip_pdf(r.get("archivo") or r.get("Archivo") or "")
        cd = norm_digits(archivo)
        if len(cd) == 11:
            nombre_por_cuil[cd] = str(r.get("nombre") or r.get("Nombre") or "").strip()

    recibos = []
    for f in files:
        cuil_digits = norm_digits(strip_pdf((f.get("name") or "").strip()))
        if len(cuil_digits) != 11:
            continue
        recibos.append({
            "cuil": format_cuil_with_dashes(cuil_digits),
            "nombre": nombre_por_cuil.get(cuil_digits, ""),
            "file_id": f.get("id"),
        })

    recibos.sort(key=lambda x: (x["nombre"] or "zzz").lower())
    return recibos

from typing import List, Optional
import re

def list_periods_for_cuil2(tenant_slug: str, cuil: str) -> List[str]:
    t = get_tenant(tenant_slug)
    if not t:
        return []

    root_id = (t.get("recibos_root_id") or t.get("drive_root_id") or "").strip()
    if not root_id:
        return []

    # normalizar cuil a 11 dígitos y con guiones
    cuil_digits = norm_digits(strip_pdf(cuil).strip())
    if len(cuil_digits) != 11:
        return []

    cuil_dash = format_cuil_with_dashes(cuil_digits)
    filename_exact = f"{cuil_dash}.pdf"

    service = drive_service()

    folders = service.files().list(
        q=f"'{root_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id,name)",
        pageSize=1000,
    ).execute().get("files", [])

    periods = []
    for f in folders:
        label = parse_period_folder(f.get("name", ""))
        if not label:
            continue

        # 1) exacto
        q_exact = (
            f"'{f['id']}' in parents and trashed=false "
            f"and name='{filename_exact}'"
        )
        res = service.files().list(q=q_exact, fields="files(id)", pageSize=1).execute().get("files", [])
        if res:
            periods.append(label)
            continue

        # 2) fallback contains (por si cambia el prefijo/sufijo)
        q_contains = (
            f"'{f['id']}' in parents and trashed=false "
            f"and mimeType='application/pdf' "
            f"and name contains '{cuil_dash}'"
        )
        res2 = service.files().list(q=q_contains, fields="files(id)", pageSize=1).execute().get("files", [])
        if res2:
            periods.append(label)

    def key(p: str):
        mm, yyyy = p.split("/")
        return int(yyyy) * 100 + int(mm)

    return sorted(set(periods), key=key, reverse=True)

def list_periods_for_cuil(tenant_slug: str, cuil: str) -> List[str]:
    t = get_tenant(tenant_slug)
    if not t:
        return []

    root_id = t["drive_root_id"]
    filename = f"{cuil}.pdf"
    service = drive_service()

    folders = service.files().list(
        q=f"'{root_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false",
        fields="files(id,name)",
        pageSize=1000,
    ).execute().get("files", [])

    periods = []
    for f in folders:
        label = parse_period_folder(f.get("name", ""))
        if not label:
            continue
        res = service.files().list(
            q=f"'{f['id']}' in parents and name='{filename}' and mimeType='application/pdf' and trashed=false",
            fields="files(id)",
            pageSize=1,
        ).execute().get("files", [])
        if res:
            periods.append(label)

    def key(p: str):
        mm, yyyy = p.split("/")
        return int(yyyy) * 100 + int(mm)

    return sorted(set(periods), key=key, reverse=True)


def ensure_sqlite_columns(table: str, columns: dict[str, str]) -> None:
    """
    columns: {"colname": "INTEGER", "error_message": "TEXT", ...}
    Agrega columnas faltantes via ALTER TABLE.
    """
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT column_name AS name
        FROM information_schema.columns
        WHERE table_name = %s
    """, (table,))
    existing = {row['name'] for row in cur.fetchall()}  # row['name'] = name

    for col, col_type in columns.items():
        if col not in existing:
            try:
                cur.execute(f"ALTER TABLE {table} ADD COLUMN {col} {col_type};")
            except Exception:
                pass

    conn.commit()
    conn.close()

# =========================
# Media endpoint (Twilio descarga PDF desde acá)
# =========================
# =========================
@app.get("/media/pdf")
def media_pdf():
    token = request.args.get("token", "").strip()
    tenant = (request.args.get("tenant") or "").strip().lower()
    cuil = (request.args.get("cuil") or "").strip()
    period = (request.args.get("period") or "").strip()

    if not (tenant and cuil and period):
        return Response("Faltan parámetros tenant/cuil/period", status=400)

    # Token firmado y acotado a este tenant/cuil/period (ya no se usa ADMIN_TOKEN acá).
    if not check_media_token(token, "pdf", tenant=tenant, cuil=cuil, period=period):
        return Response("Unauthorized", status=401)

    file_id = find_pdf_file_id_for_cuil_period(tenant, cuil, period, quiet=True)
    if not file_id:
        return Response("PDF no encontrado", status=404)

    try:
        service = drive_service()
        
        # ✅ OPTIMIZACIÓN 1: Obtener metadata primero (más rápido)
        file_metadata = service.files().get(fileId=file_id, fields='size,name').execute()
        file_size = int(file_metadata.get('size', 0))
        file_name = file_metadata.get('name', f"{strip_pdf(cuil)}.pdf")
        
        # ✅ OPTIMIZACIÓN 2: Descargar con chunks más grandes (más rápido)
        req = service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, req, chunksize=5*1024*1024)  # 5MB chunks
        
        done = False
        while not done:
            status, done = downloader.next_chunk()
        
        fh.seek(0)
        data = fh.read()
        
        # ✅ OPTIMIZACIÓN 3: Headers correctos para mejor cache
        resp = Response(data, mimetype="application/pdf")
        resp.headers["Content-Disposition"] = f'inline; filename="{file_name}"'
        resp.headers["Content-Length"] = str(len(data))
        resp.headers["Cache-Control"] = "public, max-age=300"  # 5 min cache
        resp.headers["Accept-Ranges"] = "bytes"
        
        return resp
        
    except Exception as e:
        log.exception(f"❌ Error descargando PDF: {e}")
        return Response("Error descargando PDF de Drive", status=500)
    

# Twilio senders
# =========================
def _twilio_signature_ok() -> bool:
    """Valida la firma X-Twilio-Signature de los webhooks entrantes.
    Evita que terceros simulen mensajes/estados. Desactivable con TWILIO_VALIDATE_WEBHOOKS=0."""
    if not TWILIO_VALIDATE_WEBHOOKS:
        return True
    if not TWILIO_AUTH_TOKEN:
        log.warning("⚠️ TWILIO_AUTH_TOKEN no configurado: webhook rechazado (TWILIO_VALIDATE_WEBHOOKS=0 lo desactiva).")
        return False
    sig = request.headers.get("X-Twilio-Signature", "")
    if not sig:
        return False
    validator = RequestValidator(TWILIO_AUTH_TOKEN)
    params = request.form.to_dict()
    if validator.validate(request.url, params, sig):
        return True
    # Fallback por si el proxy no reporta bien el esquema (http vs https).
    if request.url.startswith("http://"):
        return validator.validate("https://" + request.url[len("http://"):], params, sig)
    return False


def _twilio_client() -> Client:
    if not (TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN):
        raise RuntimeError("Faltan TWILIO_ACCOUNT_SID / TWILIO_AUTH_TOKEN en ENV")
    return Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

STATUS_CALLBACK_URL = os.environ.get("STATUS_CALLBACK_URL", "").strip()

def send_whatsapp_pdf(to_whatsapp: str, media_url: str, body: str, status_callback: str | None = None) -> str:
    if not (TWILIO_WHATSAPP_FROM or TWILIO_MESSAGING_SERVICE_SID):
        raise RuntimeError("Falta TWILIO_WHATSAPP_FROM o TWILIO_MESSAGING_SERVICE_SID en ENV")

    client = _twilio_client()

    payload = {
        "to": to_whatsapp,
        "body": body or " ",
        "media_url": [media_url],
    }

    if status_callback:
        payload["status_callback"] = status_callback

    if TWILIO_MESSAGING_SERVICE_SID:
        payload["messaging_service_sid"] = TWILIO_MESSAGING_SERVICE_SID
    else:
        payload["from_"] = TWILIO_WHATSAPP_FROM

    msg = client.messages.create(**payload)
    return msg.sid

# ========================================
# Sistema de emails
# ========================================

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

SMTP_HOST = os.environ.get("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")
SMTP_FROM = os.environ.get("SMTP_FROM", SMTP_USER)


def send_email(to_email: str, subject: str, html_body: str) -> bool:
    """
    Envía un email usando Gmail SMTP.
    Returns True si se envió correctamente, False si hubo error.
    """
    if not SMTP_USER or not SMTP_PASSWORD:
        log.error("ERROR: SMTP credentials not configured")
        return False
    
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = SMTP_FROM
        msg['To'] = to_email
        
        html_part = MIMEText(html_body, 'html')
        msg.attach(html_part)
        
        server = smtplib.SMTP(SMTP_HOST, SMTP_PORT)
        server.starttls()
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.sendmail(SMTP_FROM, to_email, msg.as_string())
        server.quit()
        
        log.info(f"Email sent to {to_email}")
        return True
        
    except Exception as e:
        log.exception(f"Error sending email to {to_email}: {e}")
        return False


def send_welcome_email(email: str, username: str, temp_password: str, portal_url: str) -> bool:
    """
    Envía email de bienvenida con contraseña temporal.
    """
    subject = "Acceso al Portal de Recibos"
    
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <h2 style="color: #5aa7ff;">Bienvenido al Portal de Recibos</h2>
        
        <p>Hola,</p>
        
        <p>Tu acceso al portal está listo. Podés ingresar con los siguientes datos:</p>
        
        <div style="background: #f5f5f5; padding: 15px; border-radius: 8px; margin: 20px 0;">
            <p style="margin: 5px 0;"><strong>URL:</strong> <a href="{portal_url}">{portal_url}</a></p>
            <p style="margin: 5px 0;"><strong>Usuario:</strong> {username}</p>
            <p style="margin: 5px 0;"><strong>Contraseña temporal:</strong> <code style="background: #fff; padding: 2px 6px; border-radius: 3px;">{temp_password}</code></p>
        </div>
        
        <p style="color: #d97706; font-weight: bold;">⚠️ Por seguridad, al ingresar por primera vez se te pedirá que cambies tu contraseña.</p>
        
        <p>Si tenés algún problema para acceder, contactanos.</p>
        
        <p>Saludos,<br>
        Equipo RecibosApp</p>
        
        <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;">
        <p style="color: #999; font-size: 12px;">Este es un email automático, por favor no respondas a este mensaje.</p>
    </body>
    </html>
    """
    
    return send_email(email, subject, html_body)


def send_password_reset_email(email: str, username: str, reset_url: str) -> bool:
    """
    Envía email con link para resetear contraseña.
    """
    subject = "Recuperar contraseña - Portal de Recibos"
    
    html_body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
        <h2 style="color: #5aa7ff;">Recuperar contraseña</h2>
        
        <p>Hola {username},</p>
        
        <p>Recibimos una solicitud para resetear tu contraseña. Hacé click en el siguiente botón para crear una nueva:</p>
        
        <div style="text-align: center; margin: 30px 0;">
            <a href="{reset_url}" style="background: #5aa7ff; color: white; padding: 12px 30px; text-decoration: none; border-radius: 6px; display: inline-block; font-weight: bold;">Resetear contraseña</a>
        </div>
        
        <p>O copiá y pegá este link en tu navegador:</p>
        <p style="background: #f5f5f5; padding: 10px; border-radius: 4px; word-break: break-all;">{reset_url}</p>
        
        <p style="color: #d97706;">⚠️ Este link expira en 1 hora.</p>
        
        <p style="color: #999;">Si no solicitaste este cambio, podés ignorar este email. Tu contraseña permanecerá sin cambios.</p>
        
        <p>Saludos,<br>
        Equipo RecibosApp</p>
        
        <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;">
        <p style="color: #999; font-size: 12px;">Este es un email automático, por favor no respondas a este mensaje.</p>
    </body>
    </html>
    """
    
    return send_email(email, subject, html_body)

# ========================================

def get_tenant_user_emails(tenant: str) -> list[str]:
    """Emails de los usuarios activos del portal de una empresa."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT email FROM client_users
        WHERE tenant = %s AND active = TRUE AND email IS NOT NULL AND email <> ''
    """, (tenant,))
    rows = cur.fetchall()
    conn.close()
    emails = []
    for r in rows:
        e = (dict(r).get('email') or '').strip()
        if e:
            emails.append(e)
    return emails


def send_certificado_notification(tenant: str, nombre: str, cuil: str, tipo: str = "medico"):
    """Avisa por email a los usuarios del portal que llegó un certificado nuevo."""
    try:
        emails = get_tenant_user_emails(tenant)
        if not emails:
            return

        t = get_tenant(tenant)
        empresa = (t.get('display_name', tenant) if t else tenant) or tenant

        if tipo == "familia":
            tipo_label = "certificado de asignaciones familiares"
            portal_path = "/portal/certificados-familia"
        else:
            tipo_label = "certificado médico"
            portal_path = "/portal/certificados"

        base_url = PUBLIC_BASE_URL
        portal_url = base_url + portal_path
        fecha = ts_str(int(time.time()))

        subject = f"Nuevo {tipo_label} recibido - {empresa}"
        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <h2 style="color: #5aa7ff;">Nuevo {tipo_label}</h2>
            <p>Se recibió un nuevo {tipo_label} en el portal de <strong>{empresa}</strong>.</p>
            <div style="background: #f5f5f5; padding: 15px; border-radius: 8px; margin: 20px 0;">
                <p style="margin: 5px 0;"><strong>Empleado:</strong> {nombre or 'Sin nombre'}</p>
                <p style="margin: 5px 0;"><strong>CUIL:</strong> {cuil or '-'}</p>
                <p style="margin: 5px 0;"><strong>Fecha:</strong> {fecha}</p>
            </div>
            <p>
                <a href="{portal_url}" style="display:inline-block; background:#2E3B8E; color:#fff; padding:10px 18px; border-radius:8px; text-decoration:none;">Ver en el portal</a>
            </p>
            <hr style="border: none; border-top: 1px solid #ddd; margin: 30px 0;">
            <p style="color: #999; font-size: 12px;">Este es un email automático, por favor no respondas a este mensaje.</p>
        </body>
        </html>
        """

        for email in emails:
            try:
                send_email(email, subject, html_body)
            except Exception as e:
                log.warning(f"No se pudo notificar a {email}: {e}")
    except Exception as e:
        log.exception(f"Error en send_certificado_notification: {e}")

# Gestión de usuarios del portal
# ========================================

import secrets
import string
from werkzeug.security import generate_password_hash, check_password_hash


def generate_temp_password(length=12):
    """
    Genera una contraseña temporal aleatoria.
    """
    chars = string.ascii_letters + string.digits
    return ''.join(secrets.choice(chars) for _ in range(length))


def create_client_user(tenant: str, username: str, email: str, full_name: str, created_by: str) -> dict:
    """
    Crea un nuevo usuario del portal.
    Genera contraseña temporal y envía email.
    El email es la credencial de acceso.
    Returns: {'ok': True/False, 'message': str, 'temp_password': str}
    """
    tenant = tenant.strip().lower()
    email = email.strip().lower()
    # El usuario de acceso es el email
    username = (username.strip().lower() or email)
    
    if not tenant or not email:
        return {'ok': False, 'message': 'Faltan datos requeridos (empresa y email)'}
    
    # Verificar que el email no esté ya registrado (es la credencial de acceso)
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id FROM client_users WHERE LOWER(email) = %s", (email,))
    if cur.fetchone():
        conn.close()
        return {'ok': False, 'message': 'Ya existe un usuario con ese email'}
    
    # Generar contraseña temporal
    temp_password = generate_temp_password()
    password_hash = generate_password_hash(temp_password)
    
    # Crear usuario
    now = int(time.time())
    cur.execute("""
        INSERT INTO client_users
        (tenant, username, password_hash, email, full_name, role, active, must_change_password, created_at, created_by)
        VALUES (%s, %s, %s, %s, %s, 'admin', TRUE, TRUE, %s, %s)
        RETURNING id
    """, (tenant, username, password_hash, email, full_name, now, created_by))

    user_id = cur.fetchone()['id']
    conn.commit()
    conn.close()
    
    # Enviar email
    portal_url = f"{PUBLIC_BASE_URL}/portal/login"
    email_sent = send_welcome_email(email, username, temp_password, portal_url)
    
    if not email_sent:
        return {
            'ok': True, 
            'message': 'Usuario creado pero el email no se pudo enviar. Enviá manualmente la contraseña.',
            'temp_password': temp_password
        }
    
    return {
        'ok': True,
        'message': 'Usuario creado y email enviado correctamente',
        'temp_password': temp_password
    }

def get_all_client_users():
    """
    Obtiene todos los usuarios del portal.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, tenant, username, email, full_name, role, active, 
               must_change_password, created_at, last_login
        FROM client_users
        ORDER BY tenant, username
    """)
    rows = cur.fetchall()
    conn.close()
    
    return [dict(r) for r in rows]


def toggle_client_user_active(user_id: int):
    """
    Activa/desactiva un usuario.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE client_users SET active = NOT active WHERE id = %s", (user_id,))
    conn.commit()
    conn.close()


def delete_client_user(user_id: int):
    """
    Elimina un usuario del portal.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM client_users WHERE id = %s", (user_id,))
    conn.commit()
    conn.close()


# ========================================
# Autenticación del portal
# ========================================

def authenticate_portal_user(login: str, password: str) -> dict:
    """
    Autentica un usuario del portal por email (o username, retrocompat).
    Returns: {'ok': True/False, 'user': dict, 'message': str}
    """
    login = login.strip().lower()
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT id, tenant, username, password_hash, email, full_name, 
               role, active, must_change_password
        FROM client_users
        WHERE LOWER(email) = %s OR username = %s
    """, (login, login))
    
    user = cur.fetchone()
    conn.close()
    
    if not user:
        return {'ok': False, 'message': 'Usuario o contraseña incorrectos'}
    
    if not user['active']:
        return {'ok': False, 'message': 'Usuario desactivado. Contactá al administrador.'}
    
    if not check_password_hash(user['password_hash'], password):
        return {'ok': False, 'message': 'Usuario o contraseña incorrectos'}
    
    # Actualizar último login
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE client_users SET last_login = %s WHERE id = %s", 
                (int(time.time()), user['id']))
    conn.commit()
    conn.close()
    
    return {
        'ok': True,
        'user': dict(user),
        'message': 'Login exitoso'
    }

def get_portal_user_by_id(user_id: int):
    """
    Obtiene un usuario del portal por ID.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, tenant, username, password_hash, email, full_name, role, active, must_change_password
        FROM client_users
        WHERE id = %s
    """, (user_id,))
    user = cur.fetchone()
    conn.close()
    return dict(user) if user else None


def change_portal_password(user_id: int, new_password: str):
    """
    Cambia la contraseña de un usuario del portal.
    """
    password_hash = generate_password_hash(new_password)
    
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE client_users 
        SET password_hash = %s, must_change_password = FALSE
        WHERE id = %s
    """, (password_hash, user_id))
    conn.commit()
    conn.close()


def log_portal_action(user_id: int, tenant: str, action: str, details: str = "", ip_address: str = ""):
    """
    Registra una acción en el audit log del portal.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO client_audit_log (user_id, tenant, action, details, ip_address, created_at)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (user_id, tenant, action, details, ip_address, int(time.time())))
    conn.commit()
    conn.close()


def get_user_audit_log(user_id: int, limit: int = 100) -> list[dict]:
    """Historial de actividad de un usuario del portal (más reciente primero)."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT action, details, ip_address, created_at
        FROM client_audit_log
        WHERE user_id = %s
        ORDER BY created_at DESC
        LIMIT %s
    """, (user_id, limit))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _accion_label(action: str) -> str:
    """Traduce el código de acción a un texto legible."""
    return {
        'login': '🔓 Inicio de sesión',
        'logout': '🔒 Cierre de sesión',
        'change_password': '🔑 Cambio de contraseña',
        'password_reset': '🔑 Restablecimiento de contraseña',
        'delete_certificado': '🗑 Borró un certificado',
    }.get(action or '', action or '—')

def require_portal_login():
    """
    Middleware: verifica que el usuario esté logueado.
    Returns None si está logueado, o un redirect si no lo está.
    """
    user_id = session.get('portal_user_id')
    if not user_id:
        return redirect('/portal/login')
    
    user = get_portal_user_by_id(user_id)
    if not user or not user['active']:
        session.clear()
        return redirect('/portal/login?msg=session_expired')
    
    return None

@app.route("/portal/login", methods=["GET", "POST"])
def portal_login():
    """
    Login del portal de clientes.
    """
    if request.method == "POST":
        login = request.form.get("email", "").strip() or request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        
        result = authenticate_portal_user(login, password)
        
        if result['ok']:
            user = result['user']
            session['portal_user_id'] = user['id']
            session['portal_tenant'] = user['tenant']
            session.permanent = True  # 7 días
            
            # Log
            ip = request.headers.get('X-Forwarded-For', request.remote_addr)
            log_portal_action(user['id'], user['tenant'], 'login', '', ip)
            
            # Si debe cambiar contraseña, redirigir a cambio
            if user['must_change_password']:
                return redirect('/portal/change_password?first_time=1')
            
            return redirect('/portal')
        else:
            return redirect(f"/portal/login?error={result['message']}")
    
    # GET: mostrar formulario
    error = request.args.get("error", "")
    msg = request.args.get("msg", "")
    
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Portal - Login</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    /* Solo estilos específicos de esta página si es necesario */
    body {
      display: flex;
      align-items: center;
      justify-content: center;
      min-height: 100vh;
      padding: 20px;
    }
    .forgot {
      text-align: center;
      margin-top: 15px;
      font-size: 13px;
    }
    .forgot a {
      color: var(--accent);
      text-decoration: none;
      font-weight: 600;
    }
    .forgot a:hover {
      text-decoration: underline;
    }
  </style>
</head>
<body class="login-page">
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
  
  <div class="login-card">
    <h1>Portal de Recibos</h1>
    <div class="subtitle">Acceso para clientes</div>
""")
    
    if error:
        html.append(f"<div class='alert alert-error'>❌ {esc(error)}</div>")

    if msg == "session_expired":
        html.append("<div class='alert alert-info'>⚠️ Tu sesión expiró. Ingresá nuevamente.</div>")
    elif msg == "logout":
        html.append("<div class='alert alert-info'>✅ Sesión cerrada correctamente.</div>")
    elif msg == "password_reset":
        html.append("<div class='alert alert-info'>✅ Contraseña cambiada correctamente. Ya podés ingresar.</div>")

    html.append("""
    <form method="post">
      <label>Email</label>
      <input type="email" name="email" required autofocus placeholder="tu@email.com">
      
      <label>Contraseña</label>
      <input type="password" name="password" required placeholder="••••••••">
      
      <button type="submit" class="btn primary">Ingresar</button>
    </form>
    
    <div class="forgot">
      <a href="/portal/forgot">¿Olvidaste tu contraseña?</a>
    </div>
  </div>
  <script>
    if ('serviceWorker' in navigator) {
      navigator.serviceWorker.register('/static/sw.js')
        .then(() => console.log('Service Worker registrado'))
        .catch(err => console.log('Service Worker error:', err));
    }
  </script>
</body>
</html>
""")
    
    return Response("".join(html), mimetype="text/html")


@app.route("/portal")
def portal_dashboard():
    """
    Dashboard principal del portal de clientes.
    """
    # Verificar login
    auth = require_portal_login()
    if auth:
        return auth
    
    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = user['tenant']
    
    # Obtener períodos disponibles
    period_folders = list_tenant_period_folders(tenant)
    period_labels = []
    for p in period_folders:
        lbl = period_folder_to_label(p)
        if lbl:
            period_labels.append(lbl)
    
    # Período seleccionado
    selected_period = request.args.get("period", "")
    if not selected_period and period_labels:
        selected_period = period_labels[0]
    
    msg = request.args.get("msg", "")
    
    # Obtener info del tenant
    t = get_tenant(tenant)
    empresa_nombre = t.get('display_name', tenant) if t else tenant
    
    # KPIs del período actual
    kpis = None
    chart_data = None
    
    if selected_period:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # KPIs actuales
        cur.execute("""
            SELECT COUNT(DISTINCT cuil) AS total
            FROM message_status
            WHERE tenant = %s AND period = %s AND kind = 'template'
        """, (tenant, selected_period))
        enviados = cur.fetchone()['total'] or 0

        cur.execute("""
            SELECT COUNT(DISTINCT cuil) AS total
            FROM sent_pdfs
            WHERE tenant = %s AND period = %s
        """, (tenant, selected_period))
        vistos = cur.fetchone()['total'] or 0

        cur.execute("""
            SELECT COUNT(DISTINCT cuil) AS total
            FROM recibo_estado
            WHERE tenant = %s AND period = %s AND estado IN ('FIRMADO', 'OBSERVADO')
        """, (tenant, selected_period))
        firmados = cur.fetchone()['total'] or 0

        # Tiempo promedio de firma
        cur.execute("""
            SELECT AVG(re.updated_at - ms.created_at) / 86400.0 AS avg_days
            FROM recibo_estado re
            JOIN message_status ms ON ms.tenant = re.tenant
                AND ms.cuil = re.cuil
                AND ms.period = re.period
                AND ms.kind = 'template'
            WHERE re.tenant = %s AND re.period = %s
                AND re.estado IN ('FIRMADO', 'OBSERVADO')
        """, (tenant, selected_period))
        avg_days = cur.fetchone()['avg_days']
        avg_days = round(avg_days, 1) if avg_days else 0
        
        pendientes = enviados - firmados
        pct_vistos = int((vistos / enviados * 100)) if enviados > 0 else 0
        pct_firmados = int((firmados / enviados * 100)) if enviados > 0 else 0
        
        kpis = {
            'enviados': enviados,
            'vistos': vistos,
            'firmados': firmados,
            'pendientes': pendientes,
            'pct_vistos': pct_vistos,
            'pct_firmados': pct_firmados,
            'avg_days': avg_days
        }
        
        # Datos para gráficos: últimos 6 períodos
        last_6_periods = period_labels[:6]
        
        periods_data = []
        for p in last_6_periods:
            cur.execute("""
                SELECT COUNT(DISTINCT cuil) AS total
                FROM message_status
                WHERE tenant = %s AND period = %s AND kind = 'template'
            """, (tenant, p))
            env = cur.fetchone()['total'] or 0

            cur.execute("""
                SELECT COUNT(DISTINCT cuil) AS total
                FROM sent_pdfs
                WHERE tenant = %s AND period = %s
            """, (tenant, p))
            vis = cur.fetchone()['total'] or 0

            cur.execute("""
                SELECT COUNT(DISTINCT cuil) AS total
                FROM recibo_estado
                WHERE tenant = %s AND period = %s AND estado IN ('FIRMADO', 'OBSERVADO')
            """, (tenant, p))
            fir = cur.fetchone()['total'] or 0
            
            periods_data.append({
                'period': p,
                'enviados': env,
                'vistos': vis,
                'firmados': fir,
                'pct_firmados': int((fir / env * 100)) if env > 0 else 0
            })
        
        conn.close()
        
        # Invertir para que el más reciente esté a la derecha
        periods_data.reverse()
        
        chart_data = {
            'labels': [d['period'] for d in periods_data],
            'enviados': [d['enviados'] for d in periods_data],
            'vistos': [d['vistos'] for d in periods_data],
            'firmados': [d['firmados'] for d in periods_data],
            'pct_firmados': [d['pct_firmados'] for d in periods_data]
        }
    
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Portal - Dashboard</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
  <style>
    .actions {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
      gap: 16px;
      margin-top: 20px;
    }
    .action-card {
      background: var(--card);
      border: 1px solid var(--line);
      border-radius: var(--radius);
      padding: 24px;
      text-align: center;
      text-decoration: none;
      color: var(--text);
      transition: all 0.3s ease;
    }
    .action-card:hover {
      background: var(--card-hover);
      border-color: var(--accent);
      transform: translateY(-4px);
      box-shadow: 0 8px 24px rgba(244, 196, 48, 0.15);
    }
    .action-icon {
      font-size: 48px;
      margin-bottom: 12px;
    }
    .action-title {
      font-size: 18px;
      font-weight: 700;
      margin-bottom: 8px;
    }
    .action-desc {
      font-size: 14px;
      color: var(--text-muted);
    }
    .header-actions {
      display: flex;
      gap: 12px;
      flex-wrap: wrap;
    }
    .charts-grid {
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
      gap: 20px;
      margin-top: 20px;
    }
    .chart-container {
      position: relative;
      height: 300px;
    }
    @media (max-width: 768px) {
      .charts-grid {
        grid-template-columns: 1fr;
      }
    }
    select{
      background:#0f1b33; color:#eaf0ff; border:1px solid rgba(255,255,255,.2);
      padding:10px 12px; border-radius:10px; font-size:15px; color-scheme:dark;
    }
    select option{ background-color:#0f1b33; color:#eaf0ff; }
  </style>
</head>
<body>
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
  
  <div class="container">
    <div class="header">
      <div style="display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:16px">
        <div>
          <h1>👋 Hola, """ + esc(user.get('full_name') or user['username']) + """</h1>
          <div class="subtitle">🏢 """ + esc(empresa_nombre) + """</div>
        </div>
        <div class="header-actions">
          <a href="/portal/change_password" class="btn">🔐 Cambiar contraseña</a>
          <a href="/portal/logout" class="btn">🚪 Salir</a>
        </div>
      </div>
    </div>
""")
    
    if msg == "password_changed":
        html.append("<div class='alert alert-success'>✅ Contraseña cambiada correctamente</div>")
    
    # Selector de período
    html.append("<div class='card'>")
    html.append("<h2>📅 Seleccionar período</h2>")
    html.append("<form method='get'>")
    html.append("<select name='period' onchange='this.form.submit()'>")
    if not period_labels:
        html.append("<option>No hay períodos disponibles</option>")
    else:
        for lbl in period_labels:
            sel = "selected" if lbl == selected_period else ""
            html.append(f"<option value='{esc(lbl)}' {sel}>{esc(lbl)}</option>")
    html.append("</select>")
    html.append("</form>")
    html.append("</div>")
    
    # KPIs
    if kpis:
        html.append("<div class='card'>")
        html.append(f"<h2>📊 Resumen - {esc(selected_period)}</h2>")
        html.append("<div class='stat-grid'>")
        
        html.append("<div class='stat'>")
        html.append(f"<div class='stat-value'>{kpis['enviados']}</div>")
        html.append("<div class='stat-label'>📤 Enviados</div>")
        html.append("</div>")
        
        html.append("<div class='stat'>")
        html.append(f"<div class='stat-value'>{kpis['pct_vistos']}%</div>")
        html.append("<div class='stat-label'>👁️ Tasa de apertura</div>")
        html.append("</div>")
        
        html.append("<div class='stat'>")
        html.append(f"<div class='stat-value'>{kpis['pct_firmados']}%</div>")
        html.append("<div class='stat-label'>✅ Tasa de firma</div>")
        html.append("</div>")
        
        html.append("<div class='stat'>")
        html.append(f"<div class='stat-value'>{kpis['avg_days']}</div>")
        html.append("<div class='stat-label'>📅 Días prom. de firma</div>")
        html.append("</div>")
        
        html.append("</div>")
        html.append("</div>")
        
        # GRÁFICOS
        if chart_data and len(chart_data['labels']) > 1:
            html.append("<div class='card'>")
            html.append("<h2>📈 Tendencias</h2>")
            html.append("<div class='charts-grid'>")
            
            # Gráfico 1: Evolución de firmas
            html.append("<div>")
            html.append("<h3 style='font-size:16px; margin-bottom:16px'>Evolución últimos 6 meses</h3>")
            html.append("<div class='chart-container'><canvas id='evolutionChart'></canvas></div>")
            html.append("</div>")
            
            # Gráfico 2: Tasa de firma
            html.append("<div>")
            html.append("<h3 style='font-size:16px; margin-bottom:16px'>Tasa de firma por mes</h3>")
            html.append("<div class='chart-container'><canvas id='rateChart'></canvas></div>")
            html.append("</div>")
            
            html.append("</div>")
            html.append("</div>")
        
        # Acciones rápidas
        html.append("<div class='card'>")
        html.append("<h2>🎯 Acciones rápidas</h2>")
        html.append("<div class='actions'>")
        
        html.append(f"<a href='/portal/search?period={esc(selected_period)}' class='action-card'>")
        html.append("<div class='action-icon'>🔍</div>")
        html.append("<div class='action-title'>Buscar empleado</div>")
        html.append("<div class='action-desc'>Por nombre, CUIL o DNI</div>")
        html.append("</a>")

        html.append(f"<a href='/portal/pendientes?period={esc(selected_period)}' class='action-card'>")
        html.append("<div class='action-icon'>⚠️</div>")
        html.append("<div class='action-title'>Ver pendientes</div>")
        html.append(f"<div class='action-desc'>{kpis['pendientes']} sin firmar</div>")
        html.append("</a>")

        html.append(f"<a href='/portal/reports?period={esc(selected_period)}' class='action-card'>")
        html.append("<div class='action-icon'>📊</div>")
        html.append("<div class='action-title'>Reportes</div>")
        html.append("<div class='action-desc'>Descargar PDF y Excel</div>")
        html.append("</a>")
        html.append("<a href='/portal/calendario' class='action-card'>")
        html.append("<div class='action-icon'>📅</div>")
        html.append("<div class='action-title'>Calendario</div>")
        html.append("<div class='action-desc'>Ver todos los períodos</div>")
        html.append("</a>")
        html.append("<a href='/portal/certificados' class='action-card'>")
        html.append("<div class='action-icon'>🩺</div>")
        html.append("<div class='action-title'>Certificados médicos</div>")
        html.append("<div class='action-desc'>Ver los certificados recibidos</div>")
        html.append("</a>")
        html.append("<a href='/portal/certificados-familia' class='action-card'>")
        html.append("<div class='action-icon'>🧑\u200d🧑\u200d🧒</div>")
        html.append("<div class='action-title'>Certificados de familia</div>")
        html.append("<div class='action-desc'>Asignaciones familiares recibidas</div>")
        html.append(f"<a href='/portal/recibos?period={esc(selected_period)}' class='action-card'>")
        html.append("<div class='action-icon'>🧾</div>")
        html.append("<div class='action-title'>Recibos</div>")
        html.append("<div class='action-desc'>Ver y descargar los del período</div>")
        html.append("</a>")
        html.append("</div>")
        html.append("</div>")
    
    html.append("</div>")
    
    # JavaScript para los gráficos
    if chart_data and len(chart_data['labels']) > 1:
        import json
        html.append("<script>")
        html.append(f"const chartData = {json.dumps(chart_data)};")
        html.append("""
// Configuración de colores
const colors = {
  primary: '#2E3B8E',
  accent: '#F4C430',
  success: '#10b981',
  text: '#c7d0e8'
};

// Gráfico de evolución
const ctx1 = document.getElementById('evolutionChart').getContext('2d');
new Chart(ctx1, {
  type: 'line',
  data: {
    labels: chartData.labels,
    datasets: [
      {
        label: 'Firmados',
        data: chartData.firmados,
        borderColor: colors.success,
        backgroundColor: 'rgba(16, 185, 129, 0.1)',
        tension: 0.4,
        fill: true
      },
      {
        label: 'Vistos',
        data: chartData.vistos,
        borderColor: colors.accent,
        backgroundColor: 'rgba(244, 196, 48, 0.1)',
        tension: 0.4,
        fill: true
      },
      {
        label: 'Enviados',
        data: chartData.enviados,
        borderColor: colors.primary,
        backgroundColor: 'rgba(46, 59, 142, 0.1)',
        tension: 0.4,
        fill: true
      }
    ]
  },
  options: {
    responsive: true,
    maintainAspectRatio: false,
    plugins: {
      legend: {
        labels: { color: colors.text }
      }
    },
    scales: {
      y: {
        beginAtZero: true,
        ticks: { color: colors.text },
        grid: { color: 'rgba(255, 255, 255, 0.1)' }
      },
      x: {
        ticks: { color: colors.text },
        grid: { color: 'rgba(255, 255, 255, 0.1)' }
      }
    }
  }
});

// Gráfico de tasa de firma
const ctx2 = document.getElementById('rateChart').getContext('2d');
new Chart(ctx2, {
  type: 'bar',
  data: {
    labels: chartData.labels,
    datasets: [{
      label: '% Firmados',
      data: chartData.pct_firmados,
      backgroundColor: colors.accent,
      borderColor: colors.accent,
      borderWidth: 1
    }]
  },
  options: {
    responsive: true,
    maintainAspectRatio: false,
    plugins: {
      legend: {
        labels: { color: colors.text }
      }
    },
    scales: {
      y: {
        beginAtZero: true,
        max: 100,
        ticks: { 
          color: colors.text,
          callback: function(value) { return value + '%'; }
        },
        grid: { color: 'rgba(255, 255, 255, 0.1)' }
      },
      x: {
        ticks: { color: colors.text },
        grid: { color: 'rgba(255, 255, 255, 0.1)' }
      }
    }
  }
});
        """)
        html.append("</script>")
    
    html.append("</body></html>")
    
    return Response("".join(html), mimetype="text/html")

@app.route("/portal/calendario")
def portal_calendario():
    """
    Vista de calendario con todos los períodos.
    """
    # Verificar login
    auth = require_portal_login()
    if auth:
        return auth
    
    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = user['tenant']
    
    # Obtener info del tenant
    t = get_tenant(tenant)
    empresa_nombre = t.get('display_name', tenant) if t else tenant
    
    # Obtener todos los períodos
    period_folders = list_tenant_period_folders(tenant)
    period_labels = []
    for p in period_folders:
        lbl = period_folder_to_label(p)
        if lbl:
            period_labels.append(lbl)
    
    # Calcular stats por período
    conn = get_db_connection()
    cur = conn.cursor()
    
    periods_stats = []
    for period in period_labels:
        cur.execute("""
            SELECT COUNT(DISTINCT cuil) AS total
            FROM message_status
            WHERE tenant = %s AND period = %s AND kind = 'template'
        """, (tenant, period))
        enviados = cur.fetchone()['total'] or 0

        cur.execute("""
            SELECT COUNT(DISTINCT cuil) AS total
            FROM recibo_estado
            WHERE tenant = %s AND period = %s AND estado IN ('FIRMADO', 'OBSERVADO')
        """, (tenant, period))
        firmados = cur.fetchone()['total'] or 0
        
        pct = int((firmados / enviados * 100)) if enviados > 0 else 0
        
        # Determinar color
        if enviados == 0:
            status = 'empty'
            color = '#3a4258'
            emoji = '⚪'
        elif pct >= 80:
            status = 'good'
            color = '#10b981'
            emoji = '🟢'
        elif pct >= 50:
            status = 'warning'
            color = '#f59e0b'
            emoji = '🟡'
        else:
            status = 'bad'
            color = '#ef4444'
            emoji = '🔴'
        
        # Parsear mes y año
        try:
            parts = period.split('/')
            mes = int(parts[0])
            anio = int(parts[1])
        except:
            mes = 0
            anio = 0
        
        periods_stats.append({
            'period': period,
            'mes': mes,
            'anio': anio,
            'enviados': enviados,
            'firmados': firmados,
            'pct': pct,
            'status': status,
            'color': color,
            'emoji': emoji
        })
    
    conn.close()
    
    # Agrupar por año
    years = {}
    for p in periods_stats:
        if p['anio'] > 0:
            if p['anio'] not in years:
                years[p['anio']] = []
            years[p['anio']].append(p)
    
    # Ordenar años descendente
    sorted_years = sorted(years.keys(), reverse=True)
    
    # Nombres de meses
    month_names = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 
                   'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Calendario</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    .calendar-year {
      margin-bottom: 40px;
    }
    .year-title {
      font-size: 24px;
      font-weight: 700;
      margin-bottom: 20px;
      color: var(--accent);
    }
    .months-grid {
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(150px, 1fr));
      gap: 16px;
    }
    .month-card {
      background: var(--card);
      border: 2px solid var(--line);
      border-radius: var(--radius);
      padding: 20px;
      text-align: center;
      cursor: pointer;
      transition: all 0.3s ease;
      text-decoration: none;
      color: var(--text);
    }
    .month-card:hover {
      transform: translateY(-4px);
      box-shadow: 0 8px 24px rgba(0, 0, 0, 0.3);
    }
    .month-emoji {
      font-size: 32px;
      margin-bottom: 8px;
    }
    .month-name {
      font-size: 18px;
      font-weight: 700;
      margin-bottom: 12px;
    }
    .month-stats {
      font-size: 13px;
      color: var(--text-muted);
      line-height: 1.6;
    }
    .month-pct {
      font-size: 24px;
      font-weight: 800;
      margin: 8px 0;
    }
    .legend {
      display: flex;
      gap: 24px;
      flex-wrap: wrap;
      margin-top: 20px;
      padding: 16px;
      background: rgba(0, 0, 0, 0.2);
      border-radius: var(--radius-sm);
    }
    .legend-item {
      display: flex;
      align-items: center;
      gap: 8px;
      font-size: 13px;
    }
    @media (max-width: 768px) {
      .months-grid {
        grid-template-columns: repeat(auto-fill, minmax(120px, 1fr));
      }
    }
  </style>
</head>
<body>
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
  
  <div class="container">
    <div class="header">
      <a href="/portal" class="btn">← Volver al dashboard</a>
    </div>
    
    <div class="card">
      <h2>📅 Calendario de períodos</h2>
      <div class="muted" style="margin-bottom:16px">
        🏢 """ + esc(empresa_nombre) + """
      </div>
      
      <div class="legend">
        <div class="legend-item">
          <span style="font-size:20px">🟢</span>
          <span>Excelente (>80%)</span>
        </div>
        <div class="legend-item">
          <span style="font-size:20px">🟡</span>
          <span>Bien (50-80%)</span>
        </div>
        <div class="legend-item">
          <span style="font-size:20px">🔴</span>
          <span>Bajo (<50%)</span>
        </div>
        <div class="legend-item">
          <span style="font-size:20px">⚪</span>
          <span>Sin envíos</span>
        </div>
      </div>
    </div>
""")
    
    # Mostrar años
    for year in sorted_years:
        html.append("<div class='card calendar-year'>")
        html.append(f"<div class='year-title'>{year}</div>")
        html.append("<div class='months-grid'>")
        
        # Ordenar meses de este año
        year_months = sorted(years[year], key=lambda x: x['mes'])
        
        for p in year_months:
            month_name = month_names[p['mes']] if p['mes'] < len(month_names) else p['period']
            
            html.append(f"<a href='/portal?period={esc(p['period'])}' class='month-card' style='border-color:{p['color']}'>")
            html.append(f"<div class='month-emoji'>{p['emoji']}</div>")
            html.append(f"<div class='month-name'>{month_name}</div>")
            
            if p['enviados'] > 0:
                html.append(f"<div class='month-pct' style='color:{p['color']}'>{p['pct']}%</div>")
                html.append("<div class='month-stats'>")
                html.append(f"{p['firmados']} / {p['enviados']}<br>firmados")
                html.append("</div>")
            else:
                html.append("<div class='month-stats'>Sin envíos</div>")
            
            html.append("</a>")
        
        html.append("</div>")
        html.append("</div>")
    
    if not sorted_years:
        html.append("<div class='card'>")
        html.append("<div style='text-align:center; padding:60px 20px; color:var(--text-muted)'>")
        html.append("📅 No hay períodos disponibles todavía")
        html.append("</div>")
        html.append("</div>")
    
    html.append("</div>")
    html.append("</body></html>")
    
    return Response("".join(html), mimetype="text/html")

@app.route("/portal/certificados")
def portal_certificados():
    """Certificados médicos recibidos (portal)."""
    return _portal_certificados_render("medico")


@app.route("/portal/certificados-familia")
def portal_certificados_familia():
    """Certificados de asignaciones familiares recibidos (portal)."""
    return _portal_certificados_render("familia")


def _portal_certificados_render(tipo: str = "medico"):
    """
    Render de la grilla de certificados de un tipo, para el portal de clientes.
    Cada cliente ve SOLO los de su empresa (sale de la sesión).
    """
    # Textos según el tipo
    if tipo == "familia":
        titulo = "🧑\u200d🧑\u200d🧒 Certificados de familia"
        titulo_tab = "Certificados de familia"
    else:
        titulo = "🩺 Certificados médicos"
        titulo_tab = "Certificados"

    # Verificar login
    auth = require_portal_login()
    if auth:
        return auth

    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = user['tenant']

    t = get_tenant(tenant)
    empresa_nombre = t.get('display_name', tenant) if t else tenant

    certificados = get_certificados(tenant, tipo=tipo)
    msg = request.args.get("msg", "")

    # Botón "Descargar todos" (solo si hay certificados)
    dl_btn = ""
    if certificados:
        dl_btn = f"<a href='/portal/certificados-zip?tipo={esc(tipo)}' class='btn primary'>⬇️ Descargar todos</a>"

    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Portal - """ + titulo_tab + """</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    .table-wrap { overflow:auto; max-height:520px; border:1px solid var(--line); border-radius:var(--radius); margin-top:16px; }
    table { width:100%; border-collapse:separate; border-spacing:0; }
    th, td { padding:12px 14px; border-bottom:1px solid var(--line); font-size:14px; text-align:left; }
    thead th { position:sticky; top:0; background:var(--card); z-index:1; color:var(--text-muted); font-size:13px; }
    tr:hover td { background:var(--card-hover); }
    .btn-sm { padding:6px 12px; font-size:13px; border-radius:8px; }
    th.sortable { cursor:pointer; user-select:none; }
    th.sortable:hover { color:var(--text); }
    th .arrow { font-size:11px; opacity:.7; }
    .btn-del { padding:6px 12px; font-size:13px; border-radius:8px; background:#b91c1c; color:#fff; border:none; cursor:pointer; }
    .btn-del:hover { background:#991b1b; }
  </style>
</head>
<body>
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
 
  <div class="container">
    <div class="header">
      <div style="display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:16px">
        <div>
          <h1>""" + titulo + """</h1>
          <div class="subtitle">🏢 """ + esc(empresa_nombre) + """</div>
        </div>
        <div class="header-actions" style="display:flex; gap:10px; flex-wrap:wrap">
          """ + dl_btn + """
          <a href="/portal" class="btn">← Volver al inicio</a>
        </div>
      </div>
    </div>
 
    <div class="card">
""")

    # Mensajes
    if msg == "deleted":
        html.append("<div class='alert alert-info'>✅ Certificado borrado.</div>")
    elif msg == "forbidden":
        html.append("<div class='alert alert-error'>❌ No tenés permiso sobre ese certificado.</div>")
    elif msg == "not_found":
        html.append("<div class='alert alert-error'>❌ No se encontró el certificado.</div>")

    html.append(f"<h2>Recibidos: {len(certificados)}</h2>")
 
    if certificados:
        html.append("<div class='table-wrap'>")
        html.append("<table id='cert-table'>")
        html.append(
            "<thead><tr>"
            "<th class='sortable' data-key='fecha'>Fecha <span class='arrow'></span></th>"
            "<th class='sortable' data-key='nombre'>Nombre <span class='arrow'></span></th>"
            "<th>CUIL</th><th>WhatsApp</th><th>Archivo</th><th>Acciones</th>"
            "</tr></thead>"
        )
        html.append("<tbody>")
        for c in certificados:
            ver_url = f"/portal/certificado?id={c['id']}"
            # Si el nombre quedó vacío al guardar, lo buscamos por CUIL
            nombre_cert = c.get('nombre','') or get_nombre_for_cuil(tenant, c.get('cuil',''))
            ts = int(c.get('created_at') or 0)
            html.append("<tr>")
            html.append(f"<td data-ts='{ts}'>{esc(ts_str(c.get('created_at')))}</td>")
            html.append(f"<td>{esc(nombre_cert)}</td>")
            html.append(f"<td>{esc(c.get('cuil',''))}</td>")
            html.append(f"<td>{esc(c.get('to_whatsapp','') or '')}</td>")
            html.append(f"<td><a class='btn primary btn-sm' href='{ver_url}' target='_blank'>Ver / Descargar</a></td>")
            html.append(
                "<td>"
                "<form method='post' action='/portal/certificado/delete' style='margin:0' "
                "onsubmit=\"return confirm('¿Borrar este certificado? Se quita del portal y no se puede recuperar.');\">"
                f"<input type='hidden' name='id' value='{c['id']}'>"
                f"<input type='hidden' name='tipo' value='{esc(tipo)}'>"
                "<button type='submit' class='btn-del'>🗑 Borrar</button>"
                "</form>"
                "</td>"
            )
            html.append("</tr>")
        html.append("</tbody></table></div>")
    else:
        html.append("<p class='subtitle'>Todavía no se recibieron certificados.</p>")
 
    html.append("""
    </div>
  </div>
  <script>
  (function(){
    var table = document.getElementById('cert-table');
    if(!table) return;
    var tbody = table.querySelector('tbody');
    var dir = {};
    table.querySelectorAll('th.sortable').forEach(function(th){
      th.addEventListener('click', function(){
        var key = th.getAttribute('data-key');
        var idx = Array.prototype.indexOf.call(th.parentNode.children, th);
        dir[key] = !dir[key];
        var asc = dir[key];
        var rows = Array.prototype.slice.call(tbody.querySelectorAll('tr'));
        rows.sort(function(a, b){
          var ca = a.children[idx], cb = b.children[idx], va, vb;
          if(key === 'fecha'){
            va = parseInt(ca.getAttribute('data-ts') || '0', 10);
            vb = parseInt(cb.getAttribute('data-ts') || '0', 10);
          } else {
            va = (ca.textContent || '').trim().toLowerCase();
            vb = (cb.textContent || '').trim().toLowerCase();
          }
          if(va < vb) return asc ? -1 : 1;
          if(va > vb) return asc ? 1 : -1;
          return 0;
        });
        rows.forEach(function(r){ tbody.appendChild(r); });
        table.querySelectorAll('th.sortable .arrow').forEach(function(a){ a.textContent = ''; });
        var arrow = th.querySelector('.arrow');
        if(arrow) arrow.textContent = asc ? '▲' : '▼';
      });
    });
  })();
  </script>
</body>
</html>
""")
 
    return Response("".join(html), mimetype="text/html") 
 
@app.route("/portal/recibos")
def portal_recibos():
    """Recibos de sueldo de un período (portal). Cada cliente ve solo los de su empresa."""
    auth = require_portal_login()
    if auth:
        return auth

    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = (user['tenant'] or "").strip().lower()

    t = get_tenant(tenant)
    empresa_nombre = t.get('display_name', tenant) if t else tenant

    # Períodos disponibles (etiquetas "MM/YYYY")
    period_labels = [period_folder_to_label(p) for p in list_tenant_period_folders(tenant)]
    period_labels = [p for p in period_labels if p]

    selected_period = request.args.get("period", "").strip()
    if selected_period:
        selected_period = normalize_period_label(selected_period)
    if not selected_period and period_labels:
        selected_period = period_labels[0]

    recibos = list_recibos_for_period(tenant, selected_period) if selected_period else []

    # Selector de período
    sel = ["<form method='get' style='margin:0'>",
           "<select name='period' onchange='this.form.submit()' class='btn' "
           "style='padding:8px 12px; min-width:140px'>"]
    for p in period_labels:
        selflag = " selected" if p == selected_period else ""
        sel.append(f"<option value='{esc(p)}'{selflag}>{esc(p)}</option>")
    sel.append("</select></form>")
    sel_html = "".join(sel) if period_labels else ""
    dl_btn = ""
    if recibos:
        from urllib.parse import quote
        dl_btn = (f"<a href='/portal/recibos-zip?period={quote(selected_period)}' "
                  f"class='btn primary'>⬇️ Descargar todos</a>")
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Portal - Recibos</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    .table-wrap { overflow:auto; max-height:560px; border:1px solid var(--line); border-radius:var(--radius); margin-top:16px; }
    table { width:100%; border-collapse:separate; border-spacing:0; }
    th, td { padding:12px 14px; border-bottom:1px solid var(--line); font-size:14px; text-align:left; }
    thead th { position:sticky; top:0; background:var(--card); z-index:1; color:var(--text-muted); font-size:13px; }
    tr:hover td { background:var(--card-hover); }
    .btn-sm { padding:6px 12px; font-size:13px; border-radius:8px; }
    th.sortable { cursor:pointer; user-select:none; }
    th.sortable:hover { color:var(--text); }
    th .arrow { font-size:11px; opacity:.7; }
    select{ color-scheme:dark; }
    select option{ background-color:#0f1b33; color:#eaf0ff; }
  </style>
</head>
<body>
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>

  <div class="container">
    <div class="header">
      <div style="display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:16px">
        <div>
          <h1>🧾 Recibos</h1>
          <div class="subtitle">🏢 """ + esc(empresa_nombre) + """</div>
        </div>
        <div class="header-actions" style="display:flex; gap:10px; flex-wrap:wrap; align-items:center">
          """ + dl_btn + """
          """ + sel_html + """
          <a href="/portal" class="btn">← Volver al inicio</a>
        </div>
      </div>
    </div>

    <div class="card">
""")

    if not period_labels:
        html.append("<p class='subtitle'>Todavía no hay períodos cargados.</p>")
    else:
        html.append(f"<h2>Período {esc(selected_period)} · {len(recibos)} recibos</h2>")
        if recibos:
            html.append("<div class='table-wrap'>")
            html.append("<table id='rec-table'>")
            html.append(
                "<thead><tr>"
                "<th class='sortable' data-key='nombre'>Nombre <span class='arrow'></span></th>"
                "<th class='sortable' data-key='cuil'>CUIL <span class='arrow'></span></th>"
                "<th>Recibo</th>"
                "</tr></thead><tbody>"
            )
            for r in recibos:
                qs = urlencode({
                    "tenant": tenant,
                    "cuil": r["cuil"],
                    "period": selected_period,
                    "token": make_media_token("pdf", tenant=tenant, cuil=r["cuil"], period=selected_period),
                })
                ver_url = f"/media/pdf?{qs}"
                html.append("<tr>")
                html.append(f"<td>{esc(r['nombre'] or '—')}</td>")
                html.append(f"<td>{esc(r['cuil'])}</td>")
                html.append(f"<td><a class='btn primary btn-sm' href='{ver_url}' target='_blank' rel='noopener'>Ver / Descargar</a></td>")
                html.append("</tr>")
            html.append("</tbody></table></div>")
        else:
            html.append("<p class='subtitle'>No hay recibos en este período.</p>")

    html.append("""
    </div>
  </div>
  <script>
  (function(){
    var table = document.getElementById('rec-table');
    if(!table) return;
    var tbody = table.querySelector('tbody');
    var dir = {};
    table.querySelectorAll('th.sortable').forEach(function(th){
      th.addEventListener('click', function(){
        var key = th.getAttribute('data-key');
        var idx = Array.prototype.indexOf.call(th.parentNode.children, th);
        dir[key] = !dir[key];
        var asc = dir[key];
        var rows = Array.prototype.slice.call(tbody.querySelectorAll('tr'));
        rows.sort(function(a, b){
          var va = (a.children[idx].textContent || '').trim().toLowerCase();
          var vb = (b.children[idx].textContent || '').trim().toLowerCase();
          if(va < vb) return asc ? -1 : 1;
          if(va > vb) return asc ? 1 : -1;
          return 0;
        });
        rows.forEach(function(r){ tbody.appendChild(r); });
        table.querySelectorAll('th.sortable .arrow').forEach(function(a){ a.textContent = ''; });
        var arrow = th.querySelector('.arrow');
        if(arrow) arrow.textContent = asc ? '▲' : '▼';
      });
    });
  })();
  </script>
</body>
</html>
""")

    return Response("".join(html), mimetype="text/html")

@app.get("/portal/recibos-zip")
def portal_recibos_zip():
    """Descarga todos los recibos de un período en un ZIP (solo del tenant del usuario)."""
    import zipfile
    auth = require_portal_login()
    if auth:
        return auth

    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = (user['tenant'] or "").strip().lower()

    selected_period = normalize_period_label((request.args.get("period") or "").strip())
    if not selected_period:
        return Response("Falta el período", status=400)

    recibos = list_recibos_for_period(tenant, selected_period)
    if not recibos:
        return Response("No hay recibos para descargar", status=404)

    service = drive_service()
    mem = io.BytesIO()
    usados = set()

    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in recibos:
            file_id = (r.get("file_id") or "").strip()
            if not file_id:
                continue

            # Nombre legible dentro del zip: nombre_cuil.pdf
            base = f"{r.get('nombre') or 'sin_nombre'}_{r.get('cuil', '')}".strip('_')
            base = re.sub(r'[^\w\-. ]', '_', base)
            arcname = f"{base}.pdf"

            # Evitar nombres repetidos dentro del zip
            final, n = arcname, 1
            while final.lower() in usados:
                final = f"{base}_{n}.pdf"
                n += 1
            usados.add(final.lower())

            try:
                req = service.files().get_media(fileId=file_id, supportsAllDrives=True)
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, req, chunksize=5 * 1024 * 1024)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
                fh.seek(0)
                zf.writestr(final, fh.read())
            except Exception as e:
                log.exception(f"Error agregando recibo {r.get('cuil')} al zip: {e}")
                continue

    mem.seek(0)
    data = mem.read()

    empresa = (t['display_name'] if (t := get_tenant(tenant)) else tenant) or tenant
    empresa_slug = re.sub(r'[^\w\-]', '_', empresa)
    period_slug = selected_period.replace('/', '-')
    zip_name = f"recibos_{period_slug}_{empresa_slug}.zip"

    resp = Response(data, mimetype="application/zip")
    resp.headers["Content-Disposition"] = f'attachment; filename="{zip_name}"'
    resp.headers["Content-Length"] = str(len(data))
    return resp

@app.get("/portal/certificados-zip")
def portal_certificados_zip():
    """Descarga todos los certificados de un tipo en un ZIP (solo del tenant del usuario)."""
    import zipfile
    auth = require_portal_login()
    if auth:
        return auth

    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = user['tenant']

    tipo = (request.args.get("tipo") or "medico").strip().lower()
    if tipo not in ("medico", "familia"):
        tipo = "medico"

    certificados = get_certificados(tenant, tipo=tipo)
    if not certificados:
        return Response("No hay certificados para descargar", status=404)

    service = drive_service()
    mem = io.BytesIO()
    usados = set()

    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for c in certificados:
            file_id = (c.get("file_id") or "").strip()
            if not file_id:
                continue

            # Nombre original (para conservar la extensión)
            try:
                meta = service.files().get(fileId=file_id, fields='name', supportsAllDrives=True).execute()
                drive_name = meta.get('name') or c.get('file_name') or f"certificado_{c['id']}"
            except Exception as e:
                log.warning(f"No se pudo leer meta del cert {c.get('id')}: {e}")
                drive_name = c.get('file_name') or f"certificado_{c['id']}"

            ext = ('.' + drive_name.rsplit('.', 1)[1]) if '.' in drive_name else ''

            # Nombre legible dentro del zip: fecha_nombre_cuil.ext
            nombre_cert = c.get('nombre', '') or get_nombre_for_cuil(tenant, c.get('cuil', ''))
            fecha = ts_str(c.get('created_at')).replace('/', '-').replace(':', '.').replace(' ', '_')
            base = f"{fecha}_{nombre_cert or 'sin_nombre'}_{c.get('cuil', '')}".strip('_')
            base = re.sub(r'[^\w\-. ]', '_', base)
            arcname = f"{base}{ext}"

            # Evitar nombres repetidos dentro del zip
            final, n = arcname, 1
            while final.lower() in usados:
                final = f"{base}_{n}{ext}"
                n += 1
            usados.add(final.lower())

            try:
                req = service.files().get_media(fileId=file_id, supportsAllDrives=True)
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, req, chunksize=5 * 1024 * 1024)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
                fh.seek(0)
                zf.writestr(final, fh.read())
            except Exception as e:
                log.exception(f"Error agregando cert {c.get('id')} al zip: {e}")
                continue

    mem.seek(0)
    data = mem.read()

    empresa = (t['display_name'] if (t := get_tenant(tenant)) else tenant) or tenant
    empresa_slug = re.sub(r'[^\w\-]', '_', empresa)
    tipo_label = "familia" if tipo == "familia" else "medicos"
    zip_name = f"certificados_{tipo_label}_{empresa_slug}.zip"

    resp = Response(data, mimetype="application/zip")
    resp.headers["Content-Disposition"] = f'attachment; filename="{zip_name}"'
    resp.headers["Content-Length"] = str(len(data))
    return resp

@app.post("/portal/certificado/delete")
def portal_certificado_delete():
    """Borra un certificado del portal (solo si pertenece al tenant del usuario)."""
    auth = require_portal_login()
    if auth:
        return auth

    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = user['tenant']

    try:
        cert_id = int(request.form.get("id") or "0")
    except ValueError:
        cert_id = 0

    tipo = (request.form.get("tipo") or "medico").strip().lower()
    if tipo not in ("medico", "familia"):
        tipo = "medico"
    dest = "/portal/certificados-familia" if tipo == "familia" else "/portal/certificados"

    if not cert_id:
        return redirect(f"{dest}?msg=not_found")

    # Buscar y validar que sea del tenant del usuario
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT id, tenant, file_id FROM certificados WHERE id = %s", (cert_id,))
    row = cur.fetchone()

    if not row:
        conn.close()
        return redirect(f"{dest}?msg=not_found")

    cert = dict(row)
    if (cert.get("tenant") or "") != tenant:
        conn.close()
        return redirect(f"{dest}?msg=forbidden")

    # Borrar el registro de la BD
    cur.execute("DELETE FROM certificados WHERE id = %s", (cert_id,))
    conn.commit()
    conn.close()

    # Mandar el archivo de Drive a la papelera (recuperable)
    file_id = (cert.get("file_id") or "").strip()
    if file_id:
        try:
            service = drive_service()
            service.files().update(fileId=file_id, body={'trashed': True}, supportsAllDrives=True).execute()
        except Exception as e:
            log.warning(f"No se pudo enviar a papelera el archivo {file_id}: {e}")

    # Auditoría
    try:
        ip = request.headers.get('X-Forwarded-For', request.remote_addr)
        log_portal_action(user_id, tenant, 'delete_certificado', f"id={cert_id} tipo={tipo}", ip)
    except Exception:
        pass

    return redirect(f"{dest}?msg=deleted")


def _accion_label(action: str) -> str:
    """Traduce el código de acción a un texto legible."""
    return {
        'login': '🔓 Inicio de sesión',
        'logout': '🔒 Cierre de sesión',
        'change_password': '🔑 Cambio de contraseña',
        'password_reset': '🔑 Restablecimiento de contraseña',
        'delete_certificado': '🗑 Borró un certificado',
    }.get(action or '', action or '—')


def get_admin_audit_log(tenant: str = "", action: str = "", limit: int = 300) -> list[dict]:
    """Historial de actividad de TODOS los usuarios del portal (vista admin)."""
    conn = get_db_connection()
    cur = conn.cursor()
    sql = """
        SELECT a.created_at, a.action, a.details, a.ip_address, a.tenant, a.user_id,
               u.username, u.email, u.full_name
        FROM client_audit_log a
        LEFT JOIN client_users u ON u.id = a.user_id
        WHERE 1=1
    """
    params = []
    if tenant:
        sql += " AND a.tenant = %s"
        params.append(tenant)
    if action:
        sql += " AND a.action = %s"
        params.append(action)
    sql += " ORDER BY a.created_at DESC LIMIT %s"
    params.append(limit)
    cur.execute(sql, tuple(params))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]


# ╔══════════════════════════════════════════════════════════════════════╗
# ║ PIEZA B — DESCARGA /portal/certificado                                 ║
# ║ Sirve el archivo desde Drive. Protegido por login del portal.          ║
# ║ SEGURIDAD: valida que el certificado sea del tenant del usuario.       ║
# ╚══════════════════════════════════════════════════════════════════════╝
 
@app.get("/portal/certificado")
def portal_certificado():
    """Sirve un certificado desde Drive (solo si pertenece al tenant del usuario)."""
    auth = require_portal_login()
    if auth:
        return auth
 
    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = user['tenant']
 
    try:
        cert_id = int(request.args.get("id") or "0")
    except ValueError:
        cert_id = 0
    if not cert_id:
        return Response("Falta id", status=400)
 
    # Buscar el certificado y VALIDAR que sea del tenant del usuario
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, tenant, file_id, file_name
        FROM certificados
        WHERE id = %s
    """, (cert_id,))
    row = cur.fetchone()
    conn.close()
 
    if not row:
        return Response("No encontrado", status=404)
 
    cert = dict(row)
    # 🔒 Seguridad: que el cliente no pueda ver certificados de otra empresa
    if (cert.get("tenant") or "") != tenant:
        return Response("Unauthorized", status=403)
 
    file_id = (cert.get("file_id") or "").strip()
    if not file_id:
        return Response("Sin archivo", status=404)
 
    try:
        service = drive_service()
        meta = service.files().get(fileId=file_id, fields='name,mimeType', supportsAllDrives=True).execute()
        file_name = meta.get('name', cert.get('file_name', 'certificado'))
        mime = meta.get('mimeType', 'application/octet-stream')
 
        req = service.files().get_media(fileId=file_id, supportsAllDrives=True)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, req, chunksize=5*1024*1024)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        data = fh.read()
 
        resp = Response(data, mimetype=mime)
        resp.headers["Content-Disposition"] = f'inline; filename="{file_name}"'
        resp.headers["Content-Length"] = str(len(data))
        resp.headers["Cache-Control"] = "private, max-age=60"
        return resp
    except Exception as e:
        log.exception(f"❌ Error descargando certificado (portal): {e}")
        return Response("Error descargando certificado", status=500)
 
@app.route("/portal/search")
def portal_search():
    """
    Buscar empleado en el portal.
    """
    # Verificar login
    auth = require_portal_login()
    if auth:
        return auth
    
    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = user['tenant']
    
    period = request.args.get("period", "")
    query = request.args.get("q", "").strip()
    
    # Obtener info del tenant
    t = get_tenant(tenant)
    empresa_nombre = t.get('display_name', tenant) if t else tenant
    
    results = []
    
    if query and len(query) >= 2:
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Primero obtener los CUILs que tienen template enviado en este período
        cur.execute("""
            SELECT DISTINCT cuil FROM message_status
            WHERE tenant = %s AND period = %s AND kind = 'template'
        """, (tenant, period))
        cuils_enviados = [r['cuil'] for r in cur.fetchall()]
        
        # Cargar envíos
        envios = load_envios_rows(tenant)
        
        q_lower = query.lower()
        
        # Buscar solo en los que tienen envío
        for cuil_enviado in cuils_enviados:
            # Buscar datos del empleado
            person = find_person_by_cuil(envios, cuil_enviado)
            if not person:
                continue
            
            nombre = person.get('nombre', '').lower()
            cuil = norm_cuil(person.get('cuil', ''))
            dni = cuil.replace('-', '')[-8:] if cuil else ''
            whatsapp = person.get('whatsapp', '')
            
            # Match por nombre, CUIL o DNI
            if q_lower in nombre or q_lower in cuil or q_lower in dni:
                # Ver si tiene template enviado
                cur.execute("""
                    SELECT created_at FROM message_status
                    WHERE tenant = %s AND cuil = %s AND period = %s AND kind = 'template'
                    ORDER BY created_at DESC LIMIT 1
                """, (tenant, cuil, period))
                template_row = cur.fetchone()
                
                # Ver si vio el recibo (pidió PDF)
                cur.execute("""
                    SELECT created_at FROM sent_pdfs
                    WHERE tenant = %s AND cuil = %s AND period = %s
                    ORDER BY created_at DESC LIMIT 1
                """, (tenant, cuil, period))
                pdf_row = cur.fetchone()
                
                # Ver si firmó
                cur.execute("""
                    SELECT estado FROM recibo_estado
                    WHERE tenant = %s AND cuil = %s AND period = %s
                    LIMIT 1
                """, (tenant, cuil, period))
                estado_row = cur.fetchone()
                
                # Determinar estado
                if estado_row and estado_row['estado'] in ('FIRMADO', 'OBSERVADO'):
                    status = 'firmado'
                    status_emoji = '✅'
                    status_text = 'Firmado' if estado_row['estado'] == 'FIRMADO' else 'Observado'
                elif pdf_row:
                    status = 'visto'
                    status_emoji = '👁️'
                    status_text = 'Visto, no firmado'
                    days_ago = int((time.time() - pdf_row['created_at']) / 86400) if pdf_row['created_at'] else 0
                    if days_ago > 0:
                        status_text += f' (hace {days_ago}d)'
                elif template_row:
                    status = 'enviado'
                    status_emoji = '⚠️'
                    status_text = 'No visto'
                    days_ago = int((time.time() - template_row['created_at']) / 86400) if template_row['created_at'] else 0
                    if days_ago > 0:
                        status_text += f' (hace {days_ago}d)'
                else:
                    # No debería pasar porque filtramos por template, pero por si acaso
                    continue
                
                results.append({
                    'nombre': person.get('nombre', ''),
                    'cuil': cuil,
                    'dni': dni,
                    'whatsapp': whatsapp,
                    'status': status,
                    'status_emoji': status_emoji,
                    'status_text': status_text
                })
        
        conn.close()
        
        # Limitar a 20 resultados
        results = results[:20]
    
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Buscar empleado</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    .search-box {
      display: flex;
      gap: 12px;
      margin-bottom: 20px;
    }
    .search-box input {
      flex: 1;
    }
    .search-box button {
      padding: 12px 32px;
      white-space: nowrap;
    }
    .result {
      background: rgba(0, 0, 0, 0.2);
      border: 1px solid var(--line);
      border-radius: var(--radius-sm);
      padding: 16px;
      margin-bottom: 12px;
      transition: all 0.2s ease;
    }
    .result:hover {
      background: rgba(0, 0, 0, 0.3);
      border-color: var(--accent);
    }
    .result-header {
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-bottom: 8px;
      flex-wrap: wrap;
      gap: 8px;
    }
    .result-name {
      font-size: 16px;
      font-weight: 600;
    }
    .result-details {
      font-size: 13px;
      color: var(--text-muted);
    }
    .no-results {
      text-align: center;
      padding: 60px 20px;
      color: var(--text-muted);
    }
  </style>
</head>
<body>
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
  
  <div class="container">
    <div class="header">
      <a href="/portal" class="btn">← Volver al dashboard</a>
    </div>
    
    <div class="card">
      <h2>🔍 Buscar empleado</h2>
      <div class="muted" style="margin-bottom:16px">
        🏢 """ + esc(empresa_nombre) + """ · 📅 """ + esc(period) + """
      </div>
      
      <form method="get" class="search-box">
        <input type="hidden" name="period" value='""" + esc(period) + """'>
        <input type="text" name="q" placeholder="Nombre, CUIL o DNI..." 
               value='""" + esc(query) + """' autofocus>
        <button type="submit" class="btn primary">Buscar</button>
      </form>
""")
    
    if query and len(query) < 2:
        html.append("<div class='no-results'>⚠️ Ingresá al menos 2 caracteres para buscar</div>")
    elif query and not results:
        html.append("<div class='no-results'>No se encontraron resultados para: <strong>" + esc(query) + "</strong></div>")
    elif results:
        html.append(f"<div class='muted' style='margin-bottom:16px'>✨ {len(results)} resultado(s) encontrado(s)</div>")
        
        for r in results:
            html.append("<div class='result'>")
            html.append("<div class='result-header'>")
            html.append(f"<div class='result-name'>{r['status_emoji']} {esc(r['nombre'])}</div>")
            html.append(f"<span class='badge badge-{r['status'] if r['status'] != 'sin_enviar' else 'error'}'>{esc(r['status_text'])}</span>")
            html.append("</div>")
            html.append("<div class='result-details'>")
            html.append(f"CUIL: {esc(r['cuil'])} · DNI: {esc(r['dni'])} · WhatsApp: {esc(r['whatsapp'])}")
            html.append("</div>")
            html.append(f"<div style='margin-top:12px'><a href='/portal/historial/{esc(r['cuil'])}' class='btn' style='font-size:13px; padding:8px 16px'>📜 Ver historial</a></div>")
            html.append("</div>")
    
    html.append("</div>")
    html.append("</div>")
    html.append("</body></html>")
    
    return Response("".join(html), mimetype="text/html")

@app.route("/portal/historial/<cuil>")
def portal_historial(cuil):
    """
    Ver historial completo de un empleado.
    """
    # Verificar login
    auth = require_portal_login()
    if auth:
        return auth
    
    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = user['tenant']
    
    cuil = norm_cuil(cuil)
    
    # Obtener info del tenant
    t = get_tenant(tenant)
    empresa_nombre = t.get('display_name', tenant) if t else tenant
    
    # Buscar datos del empleado
    envios = load_envios_rows(tenant)
    person = find_person_by_cuil(envios, cuil)
    
    if not person:
        return redirect('/portal/search')
    
    nombre = person.get('nombre', '')
    whatsapp = person.get('whatsapp', '')
    dni = cuil.replace('-', '')[-8:] if cuil else ''
    
    # Obtener historial de todos los períodos
    conn = get_db_connection()
    cur = conn.cursor()
    
    # Obtener todos los períodos donde tuvo envío
    cur.execute("""
        SELECT DISTINCT period FROM message_status
        WHERE tenant = %s AND cuil = %s AND kind = 'template'
        ORDER BY period DESC
    """, (tenant, cuil))
    periods = [r['period'] for r in cur.fetchall()]
    
    historial = []
    for period in periods:
        # Template enviado
        cur.execute("""
            SELECT created_at FROM message_status
            WHERE tenant = %s AND cuil = %s AND period = %s AND kind = 'template'
            ORDER BY created_at DESC LIMIT 1
        """, (tenant, cuil, period))
        template_row = cur.fetchone()
        
        # PDF enviado
        cur.execute("""
            SELECT created_at FROM sent_pdfs
            WHERE tenant = %s AND cuil = %s AND period = %s
            ORDER BY created_at DESC LIMIT 1
        """, (tenant, cuil, period))
        pdf_row = cur.fetchone()
        
        # Firmado
        cur.execute("""
            SELECT estado, updated_at FROM recibo_estado
            WHERE tenant = %s AND cuil = %s AND period = %s
            LIMIT 1
        """, (tenant, cuil, period))
        estado_row = cur.fetchone()
        
        # Calcular tiempos
        enviado_ts = template_row['created_at'] if template_row else None
        visto_ts = pdf_row['created_at'] if pdf_row else None
        firmado_ts = estado_row['updated_at'] if estado_row else None

        tiempo_ver = None
        tiempo_firmar = None

        if enviado_ts and visto_ts:
            tiempo_ver = int((visto_ts - enviado_ts) / 3600)  # horas

        if visto_ts and firmado_ts:
            tiempo_firmar = int((firmado_ts - visto_ts) / 3600)  # horas
        elif enviado_ts and firmado_ts:
            tiempo_firmar = int((firmado_ts - enviado_ts) / 3600)  # horas

        # Determinar estado
        if estado_row and estado_row['estado'] in ('FIRMADO', 'OBSERVADO'):
            status = 'firmado'
            status_emoji = '✅'
            status_text = 'Firmado' if estado_row['estado'] == 'FIRMADO' else 'Observado'
            status_class = 'success'
        elif pdf_row:
            status = 'visto'
            status_emoji = '👁️'
            status_text = 'Visto, no firmado'
            status_class = 'warning'
        else:
            status = 'enviado'
            status_emoji = '⚠️'
            status_text = 'No visto'
            status_class = 'error'
        
        historial.append({
            'period': period,
            'enviado': ts_str(enviado_ts) if enviado_ts else '-',
            'visto': ts_str(visto_ts) if visto_ts else '-',
            'firmado': ts_str(firmado_ts) if firmado_ts else '-',
            'tiempo_ver': tiempo_ver,
            'tiempo_firmar': tiempo_firmar,
            'status': status,
            'status_emoji': status_emoji,
            'status_text': status_text,
            'status_class': status_class
        })
    
    conn.close()
    
    # Estadísticas generales
    total_envios = len(historial)
    firmados = len([h for h in historial if h['status'] == 'firmado'])
    vistos = len([h for h in historial if h['status'] in ('visto', 'firmado')])
    
    tiempos_firmar = [h['tiempo_firmar'] for h in historial if h['tiempo_firmar']]
    promedio_horas = int(sum(tiempos_firmar) / len(tiempos_firmar)) if tiempos_firmar else 0
    
    stats = {
        'total': total_envios,
        'firmados': firmados,
        'vistos': vistos,
        'pct_firmados': int((firmados / total_envios * 100)) if total_envios > 0 else 0,
        'promedio_horas': promedio_horas
    }
    
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Historial empleado</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    .employee-header {
      background: rgba(0, 0, 0, 0.3);
      border: 1px solid var(--line);
      border-radius: var(--radius);
      padding: 24px;
      margin-bottom: 20px;
    }
    .employee-name {
      font-size: 24px;
      font-weight: 700;
      margin-bottom: 8px;
    }
    .employee-details {
      color: var(--text-muted);
      font-size: 14px;
    }
    .timeline {
      position: relative;
      padding-left: 40px;
    }
    .timeline::before {
      content: '';
      position: absolute;
      left: 15px;
      top: 0;
      bottom: 0;
      width: 2px;
      background: var(--line);
    }
    .timeline-item {
      position: relative;
      padding-bottom: 32px;
    }
    .timeline-dot {
      position: absolute;
      left: -29px;
      top: 6px;
      width: 12px;
      height: 12px;
      border-radius: 50%;
      border: 2px solid var(--line);
      background: var(--card);
    }
    .timeline-dot.success {
      border-color: var(--success);
      background: var(--success);
    }
    .timeline-dot.warning {
      border-color: var(--warning);
      background: var(--warning);
    }
    .timeline-dot.error {
      border-color: var(--error);
      background: var(--error);
    }
    .timeline-content {
      background: rgba(0, 0, 0, 0.2);
      border: 1px solid var(--line);
      border-radius: var(--radius-sm);
      padding: 16px;
    }
    .timeline-period {
      font-size: 18px;
      font-weight: 700;
      margin-bottom: 12px;
      display: flex;
      justify-content: space-between;
      align-items: center;
    }
    .timeline-details {
      font-size: 13px;
      color: var(--text-muted);
      line-height: 1.8;
    }
    .time-badge {
      display: inline-block;
      padding: 4px 10px;
      border-radius: 12px;
      font-size: 12px;
      font-weight: 600;
      background: rgba(244, 196, 48, 0.15);
      color: var(--accent);
      margin-left: 8px;
    }
  </style>
</head>
<body>
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
  
  <div class="container">
    <div class="header">
      <a href="/portal/search" class="btn">← Volver a búsqueda</a>
    </div>
    
    <div class="employee-header">
      <div class="employee-name">👤 """ + esc(nombre) + """</div>
      <div class="employee-details">
        CUIL: """ + esc(cuil) + """ · DNI: """ + esc(dni) + """ · WhatsApp: """ + esc(whatsapp) + """
      </div>
    </div>
    
    <div class="card">
      <h2>📊 Estadísticas generales</h2>
      <div class="stat-grid">
        <div class="stat">
          <div class="stat-value">""" + str(stats['total']) + """</div>
          <div class="stat-label">Total envíos</div>
        </div>
        <div class="stat">
          <div class="stat-value">""" + str(stats['pct_firmados']) + """%</div>
          <div class="stat-label">Tasa de firma</div>
        </div>
        <div class="stat">
          <div class="stat-value">""" + str(stats['firmados']) + """</div>
          <div class="stat-label">Firmados</div>
        </div>
        <div class="stat">
          <div class="stat-value">""" + str(stats['promedio_horas']) + """h</div>
          <div class="stat-label">Promedio de respuesta</div>
        </div>
      </div>
    </div>
    
    <div class="card">
      <h2>📜 Historial completo</h2>
      <div class="timeline" style="margin-top:24px">
""")
    
    for h in historial:
        html.append("<div class='timeline-item'>")
        html.append(f"<div class='timeline-dot {h['status_class']}'></div>")
        html.append("<div class='timeline-content'>")
        html.append("<div class='timeline-period'>")
        html.append(f"<span>{h['status_emoji']} {esc(h['period'])}</span>")
        html.append(f"<span class='badge badge-{h['status_class']}'>{esc(h['status_text'])}</span>")
        html.append("</div>")
        
        html.append("<div class='timeline-details'>")
        html.append(f"📤 Enviado: {esc(h['enviado'])}")
        
        if h['tiempo_ver']:
            html.append(f"<br>👁️ Visto: {esc(h['visto'])}")
            html.append(f"<span class='time-badge'>{h['tiempo_ver']}h después</span>")
        
        if h['tiempo_firmar']:
            html.append(f"<br>✅ Firmado: {esc(h['firmado'])}")
            html.append(f"<span class='time-badge'>{h['tiempo_firmar']}h después</span>")
        
        html.append("</div>")
        html.append("</div>")
        html.append("</div>")
    
    if not historial:
        html.append("<div style='text-align:center; padding:60px 20px; color:var(--text-muted)'>")
        html.append("📭 No hay historial de envíos para este empleado")
        html.append("</div>")
    
    html.append("</div>")
    html.append("</div>")
    html.append("</div>")
    html.append("</body></html>")
    
    return Response("".join(html), mimetype="text/html")


@app.get("/portal/report.pdf")
def portal_report_pdf():
    """
    Reporte PDF para el portal de clientes.
    """
    # Verificar login
    auth = require_portal_login()
    if auth:
        return auth
    
    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = user['tenant']
    
    period = (request.args.get("period") or "").strip()
    
    # Generar PDF
    buf = generate_pdf_report_v2(tenant, period_filter=period)
    
    filename = f"reporte_{tenant}_{(norm_period_label(period).replace('/','-') if period else 'todos')}.pdf"
    
    from flask import send_file
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )

@app.route("/portal/reports")
def portal_reports():
    """
    Reportes y exportación de datos.
    """
    # Verificar login
    auth = require_portal_login()
    if auth:
        return auth
    
    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = user['tenant']
    
    period = request.args.get("period", "")
    
    # Obtener info del tenant
    t = get_tenant(tenant)
    empresa_nombre = t.get('display_name', tenant) if t else tenant
    
    # Si se pidió exportar
    action = request.args.get("action", "")
    if action in ("export_all", "export_pending_sigs", "export_pending_views") and period:
        # Generar Excel
        import io
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        if action == "export_all":
            ws.title = "Todos"
            ws.append(["Nombre", "CUIL", "WhatsApp", "Estado", "Enviado", "Visto", "Firmado"])
            
            # Obtener solo los CUILs que tienen template enviado
            cur.execute("""
                SELECT DISTINCT cuil FROM message_status
                WHERE tenant = %s AND period = %s AND kind = 'template'
                ORDER BY cuil
            """, (tenant, period))
            cuils = [r['cuil'] for r in cur.fetchall()]
            
        elif action == "export_pending_views":
            ws.title = "No vieron"
            ws.append(["Nombre", "CUIL", "WhatsApp", "Enviado hace", "Días sin ver"])
            
            pending = get_pending_views_over_7days(tenant, period)
            cuils = [p['cuil'] for p in pending]
            
        elif action == "export_pending_sigs":
            ws.title = "No firmaron"
            ws.append(["Nombre", "CUIL", "WhatsApp", "PDF enviado", "Días sin firmar"])
            
            pending = get_pending_signatures_over_7days(tenant, period)
            cuils = [p['cuil'] for p in pending]
        
        # Cargar envios
        envios = load_envios_rows(tenant)
        
        for cuil in cuils:
            # Buscar datos del empleado
            person = find_person_by_cuil(envios, cuil)
            nombre = person.get('nombre', '') if person else ''
            whatsapp = ''
            if person:
                # La función devuelve 'to_whatsapp' ya normalizado
                whatsapp = person.get('to_whatsapp', '')
                # Fallback a telefono_raw si to_whatsapp está vacío
                if not whatsapp:
                    whatsapp = person.get('telefono_raw', '')
            if action == "export_all":
                # Ver estado completo
                cur.execute("""
                    SELECT created_at FROM message_status
                    WHERE tenant = %s AND cuil = %s AND period = %s AND kind = 'template'
                    LIMIT 1
                """, (tenant, cuil, period))
                template_row = cur.fetchone()
                
                cur.execute("""
                    SELECT created_at FROM sent_pdfs
                    WHERE tenant = %s AND cuil = %s AND period = %s
                    LIMIT 1
                """, (tenant, cuil, period))
                pdf_row = cur.fetchone()
                
                cur.execute("""
                    SELECT estado, updated_at FROM recibo_estado
                    WHERE tenant = %s AND cuil = %s AND period = %s
                    LIMIT 1
                """, (tenant, cuil, period))
                estado_row = cur.fetchone()
                
                if estado_row and estado_row['estado'] in ('FIRMADO', 'OBSERVADO'):
                    status = estado_row['estado']
                    firmado_fecha = ts_str(estado_row['updated_at']) if estado_row['updated_at'] else ""
                elif pdf_row:
                    status = "VISTO"
                    firmado_fecha = ""
                elif template_row:
                    status = "ENVIADO"
                    firmado_fecha = ""
                else:
                    status = "NO_ENVIADO"
                    firmado_fecha = ""

                enviado_fecha = ts_str(template_row['created_at']) if template_row else ""
                visto_fecha = ts_str(pdf_row['created_at']) if pdf_row else ""
                
                ws.append([nombre, cuil, whatsapp, status, enviado_fecha, visto_fecha, firmado_fecha])
            
            elif action == "export_pending_views":
                # Solo los que no vieron
                cur.execute("""
                    SELECT created_at FROM message_status
                    WHERE tenant = %s AND cuil = %s AND period = %s AND kind = 'template'
                    LIMIT 1
                """, (tenant, cuil, period))
                template_row = cur.fetchone()
                
                enviado_fecha = ts_str(template_row['created_at']) if template_row else ""
                days_ago = int((time.time() - template_row['created_at']) / 86400) if template_row else 0

                ws.append([nombre, cuil, whatsapp, enviado_fecha, days_ago])

            elif action == "export_pending_sigs":
                # Solo los que no firmaron
                cur.execute("""
                    SELECT created_at FROM sent_pdfs
                    WHERE tenant = %s AND cuil = %s AND period = %s
                    LIMIT 1
                """, (tenant, cuil, period))
                pdf_row = cur.fetchone()

                pdf_fecha = ts_str(pdf_row['created_at']) if pdf_row else ""
                days_ago = int((time.time() - pdf_row['created_at']) / 86400) if pdf_row else 0
                
                ws.append([nombre, cuil, whatsapp, pdf_fecha, days_ago])
        
        conn.close()
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        filename_map = {
            'export_all': f'reporte_completo_{tenant}_{period.replace("/", "-")}.xlsx',
            'export_pending_views': f'pendientes_no_vieron_{tenant}_{period.replace("/", "-")}.xlsx',
            'export_pending_sigs': f'pendientes_no_firmaron_{tenant}_{period.replace("/", "-")}.xlsx'
        }
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename_map.get(action, 'reporte.xlsx')
        )
    
    # KPIs
    stats = None
    if period:
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            SELECT COUNT(DISTINCT cuil) AS total FROM message_status
            WHERE tenant = %s AND period = %s AND kind = 'template'
        """, (tenant, period))
        enviados = cur.fetchone()['total'] or 0

        cur.execute("""
            SELECT COUNT(DISTINCT cuil) AS total FROM sent_pdfs
            WHERE tenant = %s AND period = %s
        """, (tenant, period))
        vistos = cur.fetchone()['total'] or 0

        cur.execute("""
            SELECT COUNT(DISTINCT cuil) AS total FROM recibo_estado
            WHERE tenant = %s AND period = %s AND estado IN ('FIRMADO', 'OBSERVADO')
        """, (tenant, period))
        firmados = cur.fetchone()['total'] or 0

        # Tiempo promedio de firma
        cur.execute("""
            SELECT AVG(re.updated_at - ms.created_at) / 86400.0 AS avg_days
            FROM recibo_estado re
            JOIN message_status ms ON ms.tenant = re.tenant
                AND ms.cuil = re.cuil
                AND ms.period = re.period
                AND ms.kind = 'template'
            WHERE re.tenant = %s AND re.period = %s
                AND re.estado IN ('FIRMADO', 'OBSERVADO')
        """, (tenant, period))
        avg_days = cur.fetchone()['avg_days']
        avg_days = round(avg_days, 1) if avg_days else 0
        
        conn.close()
        
        pct_vistos = int((vistos / enviados * 100)) if enviados > 0 else 0
        pct_firmados = int((firmados / enviados * 100)) if enviados > 0 else 0
        
        stats = {
            'enviados': enviados,
            'vistos': vistos,
            'firmados': firmados,
            'pct_vistos': pct_vistos,
            'pct_firmados': pct_firmados,
            'avg_days': avg_days
        }
    
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Reportes</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    .download-section {
      display: grid;
      gap: 16px;
      margin-top: 20px;
    }
    .download-card {
      background: rgba(255, 255, 255, 0.02);
      border: 1px solid var(--line);
      border-radius: var(--radius);
      padding: 20px;
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 16px;
      transition: all 0.3s ease;
    }
    .download-card:hover {
      background: rgba(255, 255, 255, 0.05);
      border-color: var(--accent);
    }
    .download-info h3 {
      font-size: 16px;
      margin-bottom: 6px;
    }
    .download-info p {
      font-size: 13px;
      color: var(--text-muted);
      margin: 0;
    }
  </style>
</head>
<body>
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
  
  <div class="container">
    <div class="header">
      <a href="/portal" class="btn">← Volver al dashboard</a>
    </div>
    
    <div class="card">
      <h2>📊 Reportes</h2>
      <div class="muted" style="margin-bottom:16px">
        🏢 """ + esc(empresa_nombre) + """ · 📅 """ + esc(period) + """
      </div>
""")
    
    if stats:
        html.append("<h3 style='margin:24px 0 16px 0'>📈 Estadísticas del período</h3>")
        html.append("<div class='stat-grid'>")
        
        html.append("<div class='stat'>")
        html.append(f"<div class='stat-value'>{stats['enviados']}</div>")
        html.append("<div class='stat-label'>📤 Enviados</div>")
        html.append("</div>")
        
        html.append("<div class='stat'>")
        html.append(f"<div class='stat-value'>{stats['pct_vistos']}%</div>")
        html.append("<div class='stat-label'>👁️ Tasa de apertura</div>")
        html.append("</div>")
        
        html.append("<div class='stat'>")
        html.append(f"<div class='stat-value'>{stats['pct_firmados']}%</div>")
        html.append("<div class='stat-label'>✅ Tasa de firma</div>")
        html.append("</div>")
        
        html.append("<div class='stat'>")
        html.append(f"<div class='stat-value'>{stats['avg_days']}</div>")
        html.append("<div class='stat-label'>📅 Días prom. de firma</div>")
        html.append("</div>")
        
        html.append("</div>")
        
        # Descargas
        html.append("<h3 style='margin:32px 0 16px 0'>📥 Descargas</h3>")
        html.append("<div class='download-section'>")
        
        html.append("<div class='download-card'>")
        html.append("<div class='download-info'>")
        html.append("<h3>📑 Reporte PDF completo</h3>")
        html.append("<p>PDF con KPIs, gráficos y tabla detallada</p>")
        html.append("</div>")
        html.append(f"<a href='/portal/report.pdf?period={esc(period)}' class='btn primary'>Descargar PDF</a>")
        html.append("</div>")
        
        html.append("<div class='download-card'>")
        html.append("<div class='download-info'>")
        html.append("<h3>📄 Excel - Reporte completo</h3>")
        html.append("<p>Todos los empleados con su estado</p>")
        html.append("</div>")
        html.append(f"<a href='/portal/reports?period={esc(period)}&action=export_all' class='btn primary'>Descargar Excel</a>")
        html.append("</div>")

        html.append("<div class='download-card'>")
        html.append("<div class='download-info'>")
        html.append("<h3>📝 Constancias de firma (ZIP)</h3>")
        html.append("<p>Un PDF por empleado firmado/observado, con código de verificación</p>")
        html.append("</div>")
        html.append(f"<a href='/portal/constancias-firma.zip?period={esc(period)}' class='btn primary'>Descargar ZIP</a>")
        html.append("</div>")
        
        html.append("<div class='download-card'>")
        html.append("<div class='download-info'>")
        html.append("<h3>🔴 Excel - Solo no vieron</h3>")
        html.append("<p>Empleados que no pidieron el PDF (>7 días)</p>")
        html.append("</div>")
        html.append(f"<a href='/portal/reports?period={esc(period)}&action=export_pending_views' class='btn primary'>Descargar Excel</a>")
        html.append("</div>")
        
        html.append("<div class='download-card'>")
        html.append("<div class='download-info'>")
        html.append("<h3>🟡 Excel - Solo no firmaron</h3>")
        html.append("<p>Empleados que no firmaron (>7 días)</p>")
        html.append("</div>")
        html.append(f"<a href='/portal/reports?period={esc(period)}&action=export_pending_sigs' class='btn primary'>Descargar Excel</a>")
        html.append("</div>")
        
        html.append("</div>")
    
    html.append("</div>")
    html.append("</div>")
    html.append("</body></html>")
    
    return Response("".join(html), mimetype="text/html")


@app.route("/portal/pendientes")
def portal_pendientes():
    """
    Ver pendientes en el portal de clientes.
    """
    # Verificar login
    auth = require_portal_login()
    if auth:
        return auth
    
    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    tenant = user['tenant']
    
    period = request.args.get("period", "")
    
    # Obtener info del tenant
    t = get_tenant(tenant)
    empresa_nombre = t.get('display_name', tenant) if t else tenant
    
    # Obtener pendientes
    pending_views = []
    pending_sigs = []
    
    if period:
        pending_views = get_pending_views_over_7days(tenant, period)
        pending_sigs = get_pending_signatures_over_7days(tenant, period)
    
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Pendientes</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    .section-header {
      display: flex;
      justify-content: space-between;
      align-items: center;
      margin-bottom: 16px;
      padding-bottom: 12px;
      border-bottom: 1px solid var(--line);
      flex-wrap: wrap;
      gap: 12px;
    }
    .empty {
      text-align: center;
      padding: 60px 20px;
      color: var(--text-muted);
    }
  </style>
</head>
<body>
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
  
  <div class="container">
    <div class="header">
      <a href="/portal" class="btn">← Volver al dashboard</a>
    </div>
    
    <div class="card">
      <h2>⚠️ Pendientes</h2>
      <div class="muted" style="margin-bottom:16px">
        🏢 """ + esc(empresa_nombre) + """ · 📅 """ + esc(period) + """
      </div>
    </div>
""")
    
    # No vieron (>7 días)
    html.append("<div class='card'>")
    html.append("<div class='section-header'>")
    html.append("<div>")
    html.append("<h2>🔴 No vieron el recibo</h2>")
    html.append("<div class='muted'>Template enviado hace más de 7 días, nunca pidieron el PDF</div>")
    html.append("</div>")
    html.append(f"<span class='badge badge-error'>{len(pending_views)}</span>")
    html.append("</div>")
    
    if pending_views:
        html.append("<table>")
        html.append("<thead><tr>")
        html.append("<th>Empleado</th><th>CUIL</th><th>WhatsApp</th><th>Hace</th>")
        html.append("</tr></thead><tbody>")
        
        for p in pending_views[:20]:  # Limitar a 20
            html.append("<tr>")
            html.append(f"<td>{esc(p.get('nombre', ''))}</td>")
            html.append(f"<td>{esc(p.get('cuil', ''))}</td>")
            html.append(f"<td>{esc(p.get('whatsapp', ''))}</td>")
            html.append(f"<td>{p.get('days_ago', 0)} días</td>")
            html.append("</tr>")
        
        html.append("</tbody></table>")
        
        if len(pending_views) > 20:
            html.append(f"<div class='muted' style='margin-top:12px; text-align:center'>... y {len(pending_views) - 20} más</div>")
    else:
        html.append("<div class='empty'>✅ No hay pendientes en esta categoría</div>")
    
    html.append("</div>")
    
    # No firmaron (>7 días)
    html.append("<div class='card'>")
    html.append("<div class='section-header'>")
    html.append("<div>")
    html.append("<h2>🟡 No firmaron</h2>")
    html.append("<div class='muted'>PDF recibido hace más de 7 días, nunca firmaron</div>")
    html.append("</div>")
    html.append(f"<span class='badge badge-warning'>{len(pending_sigs)}</span>")
    html.append("</div>")
    
    if pending_sigs:
        html.append("<table>")
        html.append("<thead><tr>")
        html.append("<th>Empleado</th><th>CUIL</th><th>WhatsApp</th><th>Hace</th>")
        html.append("</tr></thead><tbody>")
        
        for p in pending_sigs[:20]:  # Limitar a 20
            html.append("<tr>")
            html.append(f"<td>{esc(p.get('nombre', ''))}</td>")
            html.append(f"<td>{esc(p.get('cuil', ''))}</td>")
            html.append(f"<td>{esc(p.get('whatsapp', ''))}</td>")
            html.append(f"<td>{p.get('days_ago', 0)} días</td>")
            html.append("</tr>")
        
        html.append("</tbody></table>")
        
        if len(pending_sigs) > 20:
            html.append(f"<div class='muted' style='margin-top:12px; text-align:center'>... y {len(pending_sigs) - 20} más</div>")
    else:
        html.append("<div class='empty'>✅ No hay pendientes en esta categoría</div>")
    
    html.append("</div>")
    
    html.append("</div>")
    html.append("</body></html>")
    
    return Response("".join(html), mimetype="text/html")

@app.route("/portal/logout")
def portal_logout():
    """
    Logout del portal.
    """
    user_id = session.get('portal_user_id')
    tenant = session.get('portal_tenant')
    
    if user_id:
        log_portal_action(user_id, tenant or '', 'logout', '', '')
    
    session.clear()
    return redirect('/portal/login?msg=logout')

@app.route("/portal/change_password", methods=["GET", "POST"])
def portal_change_password():
    """
    Cambio de contraseña (obligatorio en primer login).
    """
    # Verificar login
    auth = require_portal_login()
    if auth:
        return auth
    
    user_id = session.get('portal_user_id')
    user = get_portal_user_by_id(user_id)
    
    first_time = request.args.get("first_time") == "1"
    
    if request.method == "POST":
        current_password = request.form.get("current_password", "").strip()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()
        
        # Validaciones
        if not check_password_hash(user['password_hash'], current_password):
            return redirect("/portal/change_password?error=Contraseña actual incorrecta" + ("&first_time=1" if first_time else ""))
        
        if len(new_password) < 8:
            return redirect("/portal/change_password?error=La nueva contraseña debe tener al menos 8 caracteres" + ("&first_time=1" if first_time else ""))
        
        if new_password != confirm_password:
            return redirect("/portal/change_password?error=Las contraseñas no coinciden" + ("&first_time=1" if first_time else ""))
        
        # Cambiar contraseña
        change_portal_password(user_id, new_password)
        
        # Log
        log_portal_action(user_id, user['tenant'], 'change_password', '', '')
        
        return redirect("/portal?msg=password_changed")
    
    # GET: mostrar formulario
    error = request.args.get("error", "")
    
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Cambiar contraseña</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    body {
      display: flex;
      align-items: center;
      justify-content: center;
      min-height: 100vh;
      padding: 20px;
    }
    .hint {
      font-size: 12px;
      color: var(--text-muted);
      margin-top: 6px;
    }
    .back-link {
      text-align: center;
      margin-top: 16px;
      font-size: 13px;
    }
    .back-link a {
      color: var(--accent);
      text-decoration: none;
      font-weight: 600;
    }
    .back-link a:hover {
      text-decoration: underline;
    }
  </style>
</head>
<body class="login-page">
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
  
  <div class="login-card">
    <h1>🔐 Cambiar contraseña</h1>
""")
    
    if first_time:
        html.append("<div class='subtitle'>Hola, " + esc(user['full_name'] or user['username']) + "</div>")
        html.append("""
        <div class="alert alert-warning" style="margin-top:16px">
          ⚠️ <strong>Primer ingreso</strong><br>
          Por seguridad, debés cambiar tu contraseña temporal.
        </div>
        """)
    else:
        html.append("<div class='subtitle'>Usuario: " + esc(user['username']) + "</div>")
    
    if error:
        html.append(f"<div class='alert alert-error' style='margin-top:16px'>❌ {esc(error)}</div>")
    
    html.append("""
    <form method="post" style="margin-top:20px">
      <label>Contraseña actual</label>
      <input type="password" name="current_password" required autofocus>
      
      <label style="margin-top:16px">Nueva contraseña</label>
      <input type="password" name="new_password" required minlength="8">
      <div class="hint">Mínimo 8 caracteres</div>
      
      <label style="margin-top:16px">Confirmar nueva contraseña</label>
      <input type="password" name="confirm_password" required minlength="8">
      
      <button type="submit" class="btn primary" style="margin-top:24px; width:100%">Cambiar contraseña</button>
    </form>
""")
    
    if not first_time:
        html.append("<div class='back-link'><a href='/portal'>← Volver al portal</a></div>")
    
    html.append("</div></body></html>")
    
    return Response("".join(html), mimetype="text/html")

# ========================================
# Reset de contraseña
# ========================================

def create_password_reset_token(user_id: int) -> str:
    """
    Crea un token para resetear contraseña.
    Expira en 1 hora.
    """
    import uuid
    
    token = str(uuid.uuid4())
    expires_at = int(time.time()) + 3600  # 1 hora
    
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO password_reset_tokens (user_id, token, expires_at, created_at)
        VALUES (%s, %s, %s, %s)
    """, (user_id, token, expires_at, int(time.time())))
    conn.commit()
    conn.close()
    
    return token


def get_password_reset_token(token: str):
    """
    Valida un token de reset.
    Returns: user_id si es válido, None si no.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    
    cur.execute("""
        SELECT user_id, expires_at, used
        FROM password_reset_tokens
        WHERE token = %s
    """, (token,))
    
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return None
    
    if row['used']:
        return None
    
    if row['expires_at'] < int(time.time()):
        return None
    
    return row['user_id']


def mark_reset_token_used(token: str):
    """
    Marca un token como usado.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE password_reset_tokens SET used = TRUE WHERE token = %s", (token,))
    conn.commit()
    conn.close()


def get_client_user_by_email(email: str):
    """
    Busca un usuario del portal por email.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, tenant, username, email, full_name, active
        FROM client_users
        WHERE LOWER(email) = LOWER(%s)
    """, (email.strip(),))
    user = cur.fetchone()
    conn.close()
    return dict(user) if user else None

@app.route("/portal/forgot", methods=["GET", "POST"])
def portal_forgot_password():
    """
    Solicitar reset de contraseña.
    """
    if request.method == "POST":
        email = request.form.get("email", "").strip()
        
        user = get_client_user_by_email(email)
        
        if user and user['active']:
            # Crear token
            token = create_password_reset_token(user['id'])
            
            # Enviar email
            reset_url = PUBLIC_BASE_URL + f"/portal/reset/{token}"
            send_password_reset_email(user['email'], user['username'], reset_url)
        
        # Siempre mostrar el mismo mensaje (seguridad)
        return redirect("/portal/forgot?msg=sent")
    
    # GET: mostrar formulario
    msg = request.args.get("msg", "")
    
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Recuperar contraseña</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    body {
      display: flex;
      align-items: center;
      justify-content: center;
      min-height: 100vh;
      padding: 20px;
    }
    .back-link {
      text-align: center;
      margin-top: 16px;
      font-size: 13px;
    }
    .back-link a {
      color: var(--accent);
      text-decoration: none;
      font-weight: 600;
    }
    .back-link a:hover {
      text-decoration: underline;
    }
  </style>
</head>
<body class="login-page">
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
  
  <div class="login-card">
    <h1>🔑 Recuperar contraseña</h1>
    <div class="subtitle">Te enviaremos un link para resetear tu contraseña</div>
""")
    
    if msg == "sent":
        html.append("""
        <div class="alert alert-success" style="margin-top:20px">
          ✅ <strong>Email enviado</strong><br>
          Si el email existe en nuestro sistema, recibirás un link para resetear tu contraseña.
          <br><br>
          Revisá tu casilla (incluso spam) y seguí las instrucciones.
        </div>
        <div class="back-link">
          <a href="/portal/login">← Volver al login</a>
        </div>
        """)
    else:
        html.append("""
        <form method="post" style="margin-top:20px">
          <label>Email</label>
          <input type="email" name="email" required autofocus placeholder="tu@email.com">
          
          <button type="submit" class="btn primary" style="margin-top:24px; width:100%">Enviar link de recuperación</button>
        </form>
        
        <div class="back-link">
          <a href="/portal/login">← Volver al login</a>
        </div>
        """)
    
    html.append("</div></body></html>")
    
    return Response("".join(html), mimetype="text/html")


@app.route("/portal/reset/<token>", methods=["GET", "POST"])
def portal_reset_password(token):
    """
    Resetear contraseña con token.
    """
    user_id = get_password_reset_token(token)
    
    if not user_id:
        html = """<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Link inválido</title>
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    body {
      display: flex;
      align-items: center;
      justify-content: center;
      min-height: 100vh;
      padding: 20px;
    }
    .links {
      text-align: center;
      margin-top: 16px;
      font-size: 14px;
    }
    .links a {
      color: var(--accent);
      text-decoration: none;
      font-weight: 600;
    }
    .links a:hover {
      text-decoration: underline;
    }
  </style>
</head>
<body class="login-page">
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
  
  <div class="login-card">
    <div class="alert alert-error">
      ❌ <strong>Link inválido o expirado</strong><br><br>
      Este link ya fue usado o expiró (válido por 1 hora).
    </div>
    <div class="links">
      <a href="/portal/forgot">Solicitar nuevo link</a> · <a href="/portal/login">Ir al login</a>
    </div>
  </div>
</body>
</html>"""
        return Response(html, mimetype="text/html")
    
    user = get_portal_user_by_id(user_id)
    
    if request.method == "POST":
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()
        
        if len(new_password) < 8:
            return redirect(f"/portal/reset/{token}?error=La contraseña debe tener al menos 8 caracteres")
        
        if new_password != confirm_password:
            return redirect(f"/portal/reset/{token}?error=Las contraseñas no coinciden")
        
        # Cambiar contraseña
        change_portal_password(user_id, new_password)
        
        # Marcar token como usado
        mark_reset_token_used(token)
        
        # Log
        log_portal_action(user_id, user['tenant'], 'password_reset', '', '')
        
        return redirect("/portal/login?msg=password_reset")
    
    # GET: mostrar formulario
    error = request.args.get("error", "")
    
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Nueva contraseña</title>
  <link rel="manifest" href="/static/manifest.json">
  <meta name="theme-color" content="#2E3B8E">
  <meta name="apple-mobile-web-app-capable" content="yes">
  <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
  <meta name="apple-mobile-web-app-title" content="Recibos">
  <link rel="apple-touch-icon" href="/static/icon-192.png">
  <link rel="stylesheet" href="/static/portal-theme.css">
  <style>
    body {
      display: flex;
      align-items: center;
      justify-content: center;
      min-height: 100vh;
      padding: 20px;
    }
    .hint {
      font-size: 12px;
      color: var(--text-muted);
      margin-top: 6px;
    }
  </style>
</head>
<body class="login-page">
  <div class="top-logo">
    <img src="/static/icon-192.png" alt="SIA Sueldos">
    <span class="top-logo-text">SIA</span>
  </div>
  
  <div class="login-card">
    <h1>🔐 Nueva contraseña</h1>
    <div class="subtitle">Usuario: """ + esc(user['username']) + """</div>
""")
    
    if error:
        html.append(f"<div class='alert alert-error' style='margin-top:16px'>❌ {esc(error)}</div>")
    
    html.append("""
    <form method="post" style="margin-top:20px">
      <label>Nueva contraseña</label>
      <input type="password" name="new_password" required autofocus minlength="8">
      <div class="hint">Mínimo 8 caracteres</div>
      
      <label style="margin-top:16px">Confirmar contraseña</label>
      <input type="password" name="confirm_password" required minlength="8">
      
      <button type="submit" class="btn primary" style="margin-top:24px; width:100%">Cambiar contraseña</button>
    </form>
  </div>
</body>
</html>
""")
    
    return Response("".join(html), mimetype="text/html")

@app.route("/static/<path:filename>")
def serve_static(filename):
    """
    Servir archivos estáticos (manifest, iconos, etc).
    """
    import os
    from flask import send_from_directory
    
    static_dir = os.path.join(os.path.dirname(__file__), 'static')
    return send_from_directory(static_dir, filename)

def send_whatsapp_template(
    to_whatsapp: str,
    content_vars: Optional[dict] = None,
    template_sid: Optional[str] = None,
    status_callback: Optional[str] = None,
) -> str:
    """
    Envío de plantilla aprobada (WhatsApp) usando Content Templates.
    template_sid:
      - por defecto usa TWILIO_TEMPLATE_SID (la de VIEW_NOW)
      - podés pasar TWILIO_SIGN_TEMPLATE_SID para la de firma/observa
    """
    tpl = (template_sid or TWILIO_TEMPLATE_SID or "").strip()
    if not tpl:
        raise RuntimeError("Falta TWILIO_TEMPLATE_SID (ContentSid) en ENV")
    if not (TWILIO_WHATSAPP_FROM or TWILIO_MESSAGING_SERVICE_SID):
        raise RuntimeError("Falta TWILIO_WHATSAPP_FROM o TWILIO_MESSAGING_SERVICE_SID en ENV")

    client = _twilio_client()

    payload = {
        "to": to_whatsapp,
        "content_sid": tpl,
    }
    if TWILIO_MESSAGING_SERVICE_SID:
        payload["messaging_service_sid"] = TWILIO_MESSAGING_SERVICE_SID
    else:
        payload["from_"] = TWILIO_WHATSAPP_FROM

    if content_vars:
        payload["content_variables"] = json.dumps(content_vars)

    if status_callback:
        payload["status_callback"] = status_callback

    msg = client.messages.create(**payload)
    return msg.sid
# =========================

def _set_status_on_table(table: str, sid: str, status: str, error_code=None, error_message=None):
    now = int(time.time())
    conn = get_db_connection()
    cur = conn.cursor()

    if status == "delivered":
        cur.execute(f"UPDATE {table} SET delivered_at=%s WHERE message_sid=%s;", (now, sid))
    elif status == "read":
        cur.execute(f"UPDATE {table} SET read_at=%s WHERE message_sid=%s;", (now, sid))
    elif status in ("failed", "undelivered"):
        cur.execute(
            f"UPDATE {table} SET failed_at=%s, error_code=%s, error_message=%s WHERE message_sid=%s;",
            (now, str(error_code or ""), str(error_message or ""), sid),
        )

    conn.commit()
    conn.close()

def is_pdf_sid(message_sid: str) -> bool:
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM sent_pdfs WHERE message_sid = %s LIMIT 1", (message_sid,))
    ok = cur.fetchone() is not None
    conn.close()
    return ok


@app.post("/twilio/status")
def twilio_status():
    if not _twilio_signature_ok():
        log.warning("⚠️ /twilio/status: firma Twilio inválida, request rechazado")
        return Response("Forbidden", status=403)

    sid = (request.form.get("MessageSid") or "").strip()
    status = (request.form.get("MessageStatus") or "").strip().lower()
    error_code = (request.form.get("ErrorCode") or "").strip()
    error_message = (request.form.get("ErrorMessage") or "").strip()
    now = int(time.time())

    if not sid:
        return Response("OK", status=200)

    conn = get_db_connection()
    cur = conn.cursor()

    # 1) intentar update (si existe fila)
    cur.execute("""
        UPDATE message_status
        SET last_status = %s, last_status_at = %s,
            error_code = CASE WHEN %s != '' THEN %s ELSE error_code END,
            error_message = CASE WHEN %s != '' THEN %s ELSE error_message END
        WHERE message_sid = %s
    """, (status, now, error_code, error_code, error_message, error_message, sid))

    # 2) ✅ si no existía, crearla desde sent_pdfs (caso PDF/media)
    if cur.rowcount == 0:
        cur.execute("""
            SELECT tenant, cuil, period, to_whatsapp
            FROM sent_pdfs
            WHERE message_sid = %s
            LIMIT 1
        """, (sid,))
        r = cur.fetchone()

        if r:
            tenant = r['tenant']
            cuil = r['cuil']
            period = r['period']
            to_whatsapp = r['to_whatsapp']
            # creamos fila mínima en message_status para que el callback pueda registrar delivered/read
            cur.execute("""
                INSERT INTO message_status
                (message_sid, to_whatsapp, tenant, cuil, period, nombre, kind, created_at, last_status, last_status_at)
                VALUES (%s, %s, %s, %s, %s, '', 'media', %s, %s, %s)
                ON CONFLICT (message_sid) DO NOTHING
            """, (sid, to_whatsapp, tenant, cuil, period, now, status, now))

            # volver a aplicar update (por si insertó recién)
            cur.execute("""
                UPDATE message_status
                SET last_status = %s, last_status_at = %s,
                    error_code = CASE WHEN %s != '' THEN %s ELSE error_code END,
                    error_message = CASE WHEN %s != '' THEN %s ELSE error_message END
                WHERE message_sid = %s
            """, (status, now, error_code, error_code, error_message, error_message, sid))

    # 3) timestamps por estado
    # 3) timestamps por estado
    if status == "delivered":
        cur.execute("""
            UPDATE message_status
            SET delivered_at = COALESCE(delivered_at, %s)
            WHERE message_sid = %s
        """, (now, sid))
        
        # ✅ TAMBIÉN actualizar sent_pdfs
        cur.execute("""
            UPDATE sent_pdfs
            SET delivered_at = COALESCE(delivered_at, %s),
                status = 'delivered'
            WHERE message_sid = %s
        """, (now, sid))

    # 3) timestamps por estado
    if status == "delivered":
        cur.execute("""
            UPDATE message_status
            SET delivered_at = COALESCE(delivered_at, %s)
            WHERE message_sid = %s
        """, (now, sid))
        
        # ✅ TAMBIÉN actualizar sent_pdfs
        cur.execute("""
            UPDATE sent_pdfs
            SET delivered_at = COALESCE(delivered_at, %s),
                status = 'delivered'
            WHERE message_sid = %s
        """, (now, sid))

    if status == "failed":
        cur.execute("""
            UPDATE message_status
            SET failed_at = COALESCE(failed_at, %s),
                error_code = COALESCE(NULLIF(%s, ''), error_code),
                error_message = COALESCE(NULLIF(%s, ''), error_message)
            WHERE message_sid = %s
        """, (now, error_code, error_message, sid))

        # ✅ TAMBIÉN actualizar sent_pdfs
        cur.execute("""
            UPDATE sent_pdfs
            SET failed_at = COALESCE(failed_at, %s),
                error_code = COALESCE(NULLIF(%s, ''), error_code),
                error_message = COALESCE(NULLIF(%s, ''), error_message),
                status = 'failed'
            WHERE message_sid = %s
        """, (now, error_code, error_message, sid))

    # 4) SIGN después de delivered (solo INITIAL)
        # 4) SIGN después de delivered (solo INITIAL) + estado A_FINALIZAR
    if status == "delivered":
        cur.execute("""
            SELECT tenant, cuil, period, to_whatsapp, sign_sent_at, COALESCE(origin, 'INITIAL') AS origin
            FROM sent_pdfs
            WHERE message_sid = %s
            LIMIT 1
        """, (sid,))
        row = cur.fetchone()

        if row:
            tenant = row['tenant']
            cuil = row['cuil']
            period = row['period']
            to_whatsapp = row['to_whatsapp']
            sign_sent_at = row['sign_sent_at']
            origin = row['origin']
            tenant = (tenant or "").strip().lower()
            cuil = norm_cuil(cuil)
            period = norm_period_label(period)


            # ✅ Marcar A_FINALIZAR al entregar (pero NO pisar si ya está FIRMADO/OBSERVADO)
            cur.execute("""
                INSERT INTO recibo_estado (tenant, cuil, period, estado, updated_at)
                VALUES (%s, %s, %s, 'A_FINALIZAR', %s)
                ON CONFLICT(tenant, cuil, period) DO UPDATE SET
                  estado='A_FINALIZAR',
                  updated_at=excluded.updated_at
                WHERE recibo_estado.estado NOT IN ('FIRMADO','OBSERVADO');
            """, (tenant, cuil, period, now))

            # 🔒 Si ya está cerrado, no mandes firma
            cur.execute("""
                SELECT estado FROM recibo_estado
                WHERE tenant=%s AND cuil=%s AND period=%s
                LIMIT 1
            """, (tenant, cuil, period))
            _est_row = cur.fetchone()
            est = _est_row['estado'] if _est_row else None

            # Si es reenvío, NO firmar nunca
            if origin != "INITIAL":
                log.info("%s %s %s %s", "SKIP SIGN AFTER PDF (origin=", origin, "):", sid)

            elif est in ("FIRMADO", "OBSERVADO"):
                log.info("%s %s %s %s %s", "SKIP SIGN (already closed):", tenant, cuil, period, est)

            else:
                if not sign_sent_at and TWILIO_SIGN_TEMPLATE_SID:
                    try:
                        sid_sign = send_whatsapp_template(
                            to_whatsapp,
                            content_vars={"1": period},
                            template_sid=TWILIO_SIGN_TEMPLATE_SID
                        )

                        cur.execute("""
                            INSERT INTO message_status
                            (message_sid, to_whatsapp, tenant, cuil, period, kind, created_at, last_status, last_status_at)
                            VALUES (%s, %s, %s, %s, %s, 'sign', %s, 'sent', %s)
                            ON CONFLICT (message_sid) DO NOTHING
                        """, (sid_sign, to_whatsapp, tenant, cuil, period, now, now))

                        cur.execute("UPDATE sent_pdfs SET sign_sent_at = %s WHERE message_sid = %s", (now, sid))
                        log.info("%s %s", "SENT SIGN AFTER PDF DELIVERED:", sid_sign)
                    except Exception as e:
                        log.warning("%s %s", "WARN: could not send SIGN:", e)


    conn.commit()
    conn.close()
    return Response("OK", status=200)


def is_template_sid(message_sid: str) -> bool:
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM sent_templates WHERE message_sid=%s LIMIT 1;", (message_sid,))
    row = cur.fetchone()
    conn.close()
    return bool(row)

def get_sent_pdf_by_sid(message_sid: str) -> dict | None:
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT tenant,cuil,period,to_whatsapp,sign_sent_at FROM sent_pdfs WHERE message_sid=%s LIMIT 1;", (message_sid,))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None

def mark_sign_sent(pdf_sid: str):
    now = int(time.time())
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE sent_pdfs SET sign_sent_at=%s WHERE message_sid=%s;", (now, pdf_sid))
    conn.commit()
    conn.close()


import psycopg2
from psycopg2.extras import RealDictCursor

# =========================
# DB: pending view + estado firma
# =========================
DB_PATH = os.environ.get("DB_PATH", "/data/app.db")

def get_db_connection():
    DATABASE_URL = os.environ.get("DATABASE_URL")
    if not DATABASE_URL:
        raise RuntimeError("Falta DATABASE_URL en ENV (la app requiere PostgreSQL)")
    conn = psycopg2.connect(DATABASE_URL)
    conn.cursor_factory = RealDictCursor
    # Red de seguridad contra fugas: registramos la conexión para cerrarla al final
    # del request si algún handler la dejó abierta (p. ej. por una excepción).
    if has_app_context():
        if not hasattr(g, "_db_conns"):
            g._db_conns = []
        g._db_conns.append(conn)
    return conn


@app.teardown_appcontext
def _close_leaked_db_connections(exc):
    conns = getattr(g, "_db_conns", None) or []
    for conn in conns:
        try:
            if getattr(conn, "closed", 1) == 0:
                try:
                    conn.rollback()
                except Exception:
                    pass
                conn.close()
        except Exception:
            pass
    g._db_conns = []


def get_latest_context_for_whatsapp(to_whatsapp: str) -> dict | None:
    """
    Devuelve el último contexto conocido (tenant, cuil, period) para ese WhatsApp,
    mirando message_status (template/pdf).
    """
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT tenant, cuil, period
        FROM message_status
        WHERE to_whatsapp = %s
          AND COALESCE(tenant,'') != ''
          AND COALESCE(cuil,'') != ''
        ORDER BY COALESCE(created_at,0) DESC, id DESC
        LIMIT 1
    """, (to_whatsapp,))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None


def resolve_best_period_with_pdf(tenant: str, cuil: str, *, max_months_back: int = 36) -> str | None:
    """
    Devuelve el período MÁS CERCANO a hoy con PDF:
      - prueba mes actual
      - si no, mes anterior, y así hacia atrás hasta encontrar.
    """
    import datetime as _dt

    def add_months(dt: _dt.datetime, delta_months: int) -> _dt.datetime:
        # delta_months negativo para ir hacia atrás
        y = dt.year + (dt.month - 1 + delta_months) // 12
        m = (dt.month - 1 + delta_months) % 12 + 1
        return _dt.datetime(y, m, 1)

    now = _dt.datetime.now()

    for back in range(0, max_months_back + 1):
        d = add_months(now, -back)
        period = f"{d.month:02d}/{d.year:04d}"

        # quiet=True para que no te tire ❌ por cada mes que no exista
        try:
            fid = find_pdf_file_id_for_cuil_period(tenant, cuil, period, quiet=True)
        except Exception:
            fid = None

        if fid:
            return period

    return None

def get_receipt_request_count(tenant: str, cuil: str, period: str, to_whatsapp: str) -> int:
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT request_count
        FROM receipt_requests
        WHERE tenant=%s AND cuil=%s AND period=%s AND to_whatsapp=%s
        LIMIT 1
    """, (tenant, cuil, period, to_whatsapp))
    row = cur.fetchone()
    conn.close()
    return int(row['request_count']) if row and row['request_count'] is not None else 0


def inc_receipt_request_count(tenant: str, cuil: str, period: str, to_whatsapp: str) -> int:
    now = int(time.time())
    conn = get_db_connection()
    cur = conn.cursor()

    # upsert compatible
    cur.execute("""
        SELECT id, request_count, first_requested_at
        FROM receipt_requests
        WHERE tenant=%s AND cuil=%s AND period=%s AND to_whatsapp=%s
        LIMIT 1
    """, (tenant, cuil, period, to_whatsapp))
    row = cur.fetchone()

    if row:
        rid = row['id']
        cnt = row['request_count']
        first_ts = row['first_requested_at']
        cnt = int(cnt or 0) + 1
        cur.execute("""
            UPDATE receipt_requests
            SET request_count=%s, last_requested_at=%s
            WHERE id=%s
        """, (cnt, now, rid))
    else:
        cnt = 1
        cur.execute("""
            INSERT INTO receipt_requests
            (tenant,cuil,period,to_whatsapp,request_count,first_requested_at,last_requested_at)
            VALUES (%s,%s,%s,%s,%s,%s,%s)
        """, (tenant, cuil, period, to_whatsapp, cnt, now, now))

    conn.commit()
    conn.close()
    return cnt



def get_origin_by_message_sid(message_sid: str) -> str | None:
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT origin FROM receipt_request_events
            WHERE message_sid = %s
            ORDER BY id DESC LIMIT 1
        """, (message_sid,))
        row = cur.fetchone()
    finally:
        conn.close()
    return row['origin'] if row and row['origin'] else None


def _log_receipt_request_event(
    tenant: str,
    cuil: str,
    period: str,
    to_whatsapp: str,
    source: str,
    result: str,
    message_sid: str | None = None,
    origin: str | None = None,
):
    ts = int(time.time())
    origin = origin or source

    conn = get_db_connection()
    try:
        cur = conn.cursor()

        # 2) Ver columnas actuales
        cur.execute("""
            SELECT column_name AS name
            FROM information_schema.columns
            WHERE table_name = 'receipt_request_events'
        """)
        cols = {r['name'] for r in cur.fetchall()}

        def _add_col(colname: str, coltype: str):
            nonlocal cols
            if colname not in cols:
                cur.execute(f"ALTER TABLE receipt_request_events ADD COLUMN {colname} {coltype}")
                cols.add(colname)

        # 3) Asegurar columnas usadas (compatibilidad)
        _add_col("tenant", "TEXT")
        _add_col("cuil", "TEXT")
        _add_col("period", "TEXT")
        _add_col("source", "TEXT")
        _add_col("result", "TEXT")
        _add_col("message_sid", "TEXT")
        _add_col("created_at", "INTEGER")
        _add_col("requested_at", "INTEGER")   # ✅ clave para tu NOT NULL viejo
        _add_col("origin", "TEXT")

        # compatibilidad con nombres viejos/nuevos
        if "to_whatsapp" not in cols and "whatsapp" not in cols:
            _add_col("to_whatsapp", "TEXT")
        else:
            _add_col("to_whatsapp", "TEXT")
            _add_col("whatsapp", "TEXT")

        # 4) Preparar data
        data = {
            "tenant": tenant,
            "cuil": cuil,
            "period": period,
            "source": source,
            "result": result,
            "message_sid": message_sid,
            "created_at": ts,
            "requested_at": ts,   # ✅ siempre seteamos ambos si existen
            "origin": origin,
        }

        # guardar whatsapp en la columna que exista
        if "to_whatsapp" in cols:
            data["to_whatsapp"] = to_whatsapp
        if "whatsapp" in cols:
            data["whatsapp"] = to_whatsapp

        insert_cols = [k for k in data.keys() if k in cols]
        placeholders = ",".join(["%s"] * len(insert_cols))
        sql = f"INSERT INTO receipt_request_events ({','.join(insert_cols)}) VALUES ({placeholders})"

        cur.execute(sql, tuple(data[k] for k in insert_cols))
        conn.commit()

    except Exception as e:
        log.warning("%s %s", "WARN: _log_receipt_request_event failed:", e)
        try:
            conn.rollback()
        except Exception:
            pass
    finally:
        conn.close()

# =========================
# DB Initialization / Migration
def get_receipt_event_origin_by_sid(message_sid: str) -> str | None:
    if not message_sid:
        return None
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        # origin puede ser NULL en filas viejas
        cur.execute("""
            SELECT origin, source
            FROM receipt_request_events
            WHERE message_sid = %s
            ORDER BY id DESC
            LIMIT 1
        """, (message_sid,))
        row = cur.fetchone()
        if not row:
            return None
        return row['origin'] or row['source']  # origin si existe, si no source
    finally:
        conn.close()



import time, hashlib

def _try_alter(cur, sql: str):
    try:
        cur.execute(sql)
    except Exception:
        pass

def init_db():
    conn = get_db_connection()
    cur = conn.cursor()
    
    DATABASE_URL = os.environ.get("DATABASE_URL")
    
    if not DATABASE_URL:
        # SQLite only
        cur.execute("PRAGMA journal_mode=WAL;")
        cur.execute("PRAGMA foreign_keys=ON;")

    if DATABASE_URL:
        # PostgreSQL - tablas ya creadas, solo agregar columnas faltantes
        _try_alter(cur, "ALTER TABLE pending_views ADD COLUMN IF NOT EXISTS step TEXT;")
        _try_alter(cur, "ALTER TABLE pending_views ADD COLUMN IF NOT EXISTS dni_attempts INTEGER;")
        _try_alter(cur, "ALTER TABLE pending_views ADD COLUMN IF NOT EXISTS origin TEXT;")
        _try_alter(cur, "ALTER TABLE pending_views ADD COLUMN IF NOT EXISTS period_offset INTEGER DEFAULT 0;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS to_whatsapp TEXT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS tenant TEXT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS cuil TEXT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS period TEXT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS nombre TEXT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS kind TEXT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS created_at BIGINT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS last_status TEXT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS last_status_at BIGINT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS delivered_at BIGINT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS read_at BIGINT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS failed_at BIGINT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS error_code TEXT;")
        _try_alter(cur, "ALTER TABLE message_status ADD COLUMN IF NOT EXISTS error_message TEXT;")
        _try_alter(cur, "ALTER TABLE sent_pdfs ADD COLUMN IF NOT EXISTS sign_sent_at BIGINT;")
        _try_alter(cur, "ALTER TABLE sent_pdfs ADD COLUMN IF NOT EXISTS origin TEXT;")
        _try_alter(cur, "ALTER TABLE sent_pdfs ADD COLUMN IF NOT EXISTS delivered_at BIGINT;")
        _try_alter(cur, "ALTER TABLE sent_pdfs ADD COLUMN IF NOT EXISTS read_at BIGINT;")
        _try_alter(cur, "ALTER TABLE sent_pdfs ADD COLUMN IF NOT EXISTS failed_at BIGINT;")
        _try_alter(cur, "ALTER TABLE sent_pdfs ADD COLUMN IF NOT EXISTS error_code TEXT;")
        _try_alter(cur, "ALTER TABLE sent_pdfs ADD COLUMN IF NOT EXISTS error_message TEXT;")
        _try_alter(cur, "ALTER TABLE sent_pdfs ADD COLUMN IF NOT EXISTS status TEXT;")
        _try_alter(cur, "ALTER TABLE recibo_estado ADD COLUMN IF NOT EXISTS to_whatsapp TEXT;")
        _try_alter(cur, "ALTER TABLE recibo_estado ADD COLUMN IF NOT EXISTS observaciones TEXT;")
        _try_alter(cur, "ALTER TABLE recibo_estado ADD COLUMN IF NOT EXISTS created_at BIGINT;")
        _try_alter(cur, "ALTER TABLE receipt_request_events ADD COLUMN IF NOT EXISTS requested_at BIGINT;")
        _try_alter(cur, "ALTER TABLE receipt_request_events ADD COLUMN IF NOT EXISTS created_at BIGINT;")
        _try_alter(cur, "ALTER TABLE receipt_request_events ADD COLUMN IF NOT EXISTS whatsapp TEXT;")
        _try_alter(cur, "ALTER TABLE verifications ADD COLUMN IF NOT EXISTS nombre TEXT;")
        _try_alter(cur, "ALTER TABLE verifications ADD COLUMN IF NOT EXISTS dni_hash TEXT;")
        _try_alter(cur, "ALTER TABLE verifications ADD COLUMN IF NOT EXISTS dni_last4 TEXT;")

        # ✅ NUEVO: certificados médicos recibidos (PostgreSQL)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS certificados (
                id SERIAL PRIMARY KEY,
                tenant TEXT NOT NULL,
                cuil TEXT NOT NULL,
                nombre TEXT,
                to_whatsapp TEXT,
                file_id TEXT NOT NULL,
                file_name TEXT,
                mime_type TEXT,
                created_at BIGINT NOT NULL
            )
        ''')
        _try_alter(cur, "ALTER TABLE certificados ADD COLUMN tipo TEXT DEFAULT 'medico';")
        _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_cert_tenant ON certificados(tenant, created_at);")
        _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_cert_tenant_tipo ON certificados(tenant, tipo, created_at);")

        conn.commit()
        conn.close()
        return

    # SQLite only - crear tablas
    cur.execute("""
    CREATE TABLE IF NOT EXISTS pending_views (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        to_whatsapp TEXT NOT NULL,
        tenant TEXT NOT NULL,
        cuil TEXT NOT NULL,
        period TEXT NOT NULL,
        created_at INTEGER NOT NULL,
        step TEXT DEFAULT 'READY',
        dni_attempts INTEGER DEFAULT 0,
        UNIQUE(to_whatsapp, tenant, cuil, period)
    );
    """)

    _try_alter(cur, "ALTER TABLE pending_views ADD COLUMN step TEXT;")
    _try_alter(cur, "ALTER TABLE pending_views ADD COLUMN dni_attempts INTEGER;")
    _try_alter(cur, "ALTER TABLE pending_views ADD COLUMN origin TEXT;")
    _try_alter(cur, "ALTER TABLE pending_views ADD COLUMN period_offset INTEGER DEFAULT 0;")

    _try_alter(cur, """
    DELETE FROM pending_views
    WHERE id NOT IN (
    SELECT pv.id
    FROM pending_views pv
    JOIN (
        SELECT to_whatsapp, MAX(created_at) AS max_created
        FROM pending_views
        GROUP BY to_whatsapp
    ) x
    ON x.to_whatsapp = pv.to_whatsapp AND x.max_created = pv.created_at
    );
    """)

    _try_alter(cur, """
    CREATE UNIQUE INDEX IF NOT EXISTS ux_pending_views_to_whatsapp
    ON pending_views(to_whatsapp);
    """)

    cur.execute("""
      CREATE TABLE IF NOT EXISTS recibo_estado (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tenant TEXT NOT NULL,
        cuil TEXT NOT NULL,
        period TEXT NOT NULL,
        estado TEXT NOT NULL,
        updated_at INTEGER NOT NULL,
        UNIQUE(tenant, cuil, period)
      );
    """)

    cur.execute("""
      CREATE TABLE IF NOT EXISTS message_status (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        message_sid TEXT UNIQUE NOT NULL,
        to_whatsapp TEXT,
        tenant TEXT,
        cuil TEXT,
        period TEXT,
        nombre TEXT,
        kind TEXT,
        created_at INTEGER,
        last_status TEXT,
        last_status_at INTEGER,
        delivered_at INTEGER,
        read_at INTEGER,
        failed_at INTEGER,
        error_code TEXT,
        error_message TEXT
      );
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS template_send_queue (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tenant TEXT NOT NULL,
        period TEXT NOT NULL,
        to_whatsapp TEXT NOT NULL,
        cuil TEXT NOT NULL,
        nombre TEXT,
        require_pdf BOOLEAN DEFAULT TRUE,
        status TEXT DEFAULT 'PENDING',
        error TEXT,
        created_at INTEGER NOT NULL,
        updated_at INTEGER,
        sent_sid TEXT,
        sent_at INTEGER,
        UNIQUE(tenant, period, to_whatsapp, cuil)
    );
    """)

    cur.execute("""
      CREATE TABLE IF NOT EXISTS sent_pdfs (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tenant TEXT NOT NULL,
        cuil TEXT NOT NULL,
        period TEXT NOT NULL,
        to_whatsapp TEXT NOT NULL,
        message_sid TEXT NOT NULL UNIQUE,
        created_at INTEGER NOT NULL,
        sign_sent_at INTEGER,
        origin TEXT
      );
    """)

    cur.execute("""
      CREATE TABLE IF NOT EXISTS verifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tenant TEXT NOT NULL,
        cuil TEXT NOT NULL,
        to_whatsapp TEXT NOT NULL,
        nombre TEXT,
        dni_hash TEXT,
        dni_last4 TEXT,
        verified_at INTEGER NOT NULL,
        updated_at INTEGER NOT NULL,
        UNIQUE(tenant, cuil, to_whatsapp)
      );
    """)

    cur.execute("""
      CREATE TABLE IF NOT EXISTS receipt_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tenant TEXT NOT NULL,
        cuil TEXT NOT NULL,
        period TEXT NOT NULL,
        to_whatsapp TEXT NOT NULL,
        request_count INTEGER NOT NULL DEFAULT 0,
        first_requested_at INTEGER,
        last_requested_at INTEGER,
        UNIQUE(tenant, cuil, period, to_whatsapp)
      );
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS receipt_request_events (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        tenant TEXT,
        cuil TEXT,
        period TEXT,
        to_whatsapp TEXT,
        source TEXT,
        result TEXT,
        message_sid TEXT,
        created_at INTEGER,
        origin TEXT
    )
    """)

    cur.execute("""
      CREATE TABLE IF NOT EXISTS terms_accepted (
        whatsapp TEXT PRIMARY KEY,
        accepted_at INTEGER NOT NULL,
        ip_address TEXT,
        user_agent TEXT
      );
    """)

    cur.execute("""
      CREATE TABLE IF NOT EXISTS pending_terms (
        whatsapp TEXT PRIMARY KEY,
        tenant TEXT,
        cuil TEXT,
        period TEXT,
        origin TEXT DEFAULT 'INITIAL',
        created_at INTEGER
      );
    """)

    cur.execute("""
      CREATE TABLE IF NOT EXISTS inbound_dedup (
        message_sid TEXT PRIMARY KEY,
        created_at INTEGER
      );
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS multi_tenant_selection (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            whatsapp TEXT NOT NULL UNIQUE,
            tenants_json TEXT NOT NULL,
            created_at INTEGER NOT NULL,
            expires_at INTEGER NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS client_users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tenant TEXT NOT NULL,
            username TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            full_name TEXT,
            email TEXT,
            role TEXT DEFAULT 'admin',
            active BOOLEAN DEFAULT TRUE,
            must_change_password BOOLEAN DEFAULT TRUE,
            created_at INTEGER NOT NULL,
            last_login INTEGER,
            created_by TEXT,
            UNIQUE(tenant, username)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS password_reset_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token TEXT NOT NULL,
            expires_at INTEGER NOT NULL,
            used BOOLEAN DEFAULT FALSE,
            created_at INTEGER NOT NULL,
            UNIQUE(token),
            FOREIGN KEY(user_id) REFERENCES client_users(id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS client_audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            tenant TEXT,
            action TEXT,
            details TEXT,
            ip_address TEXT,
            created_at INTEGER NOT NULL
        )
    """)

    # ✅ NUEVO: certificados médicos recibidos (SQLite)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS certificados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tenant TEXT NOT NULL,
            cuil TEXT NOT NULL,
            nombre TEXT,
            to_whatsapp TEXT,
            file_id TEXT NOT NULL,
            file_name TEXT,
            mime_type TEXT,
            created_at INTEGER NOT NULL
        )
    """)
    # Candado para avisar 1 sola vez por tanda de envío INITIAL (facturación)
    cur.execute('''
        CREATE TABLE IF NOT EXISTS template_batch_notice (
            tenant TEXT NOT NULL,
            period TEXT NOT NULL,
            enviados INTEGER,
            notified_at BIGINT NOT NULL,
            PRIMARY KEY (tenant, period)
        )
    ''')

    _try_alter(cur, "ALTER TABLE certificados ADD COLUMN tipo TEXT DEFAULT 'medico';")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_cert_tenant ON certificados(tenant, created_at);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_cert_tenant_tipo ON certificados(tenant, tipo, created_at);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_pending_to_created ON pending_views(to_whatsapp, created_at);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_estado_key ON recibo_estado(tenant, cuil, period);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_msg_key ON message_status(tenant, cuil, period, kind);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_msg_sid ON message_status(message_sid);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_sentpdfs_sid ON sent_pdfs(message_sid);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_verif_tenant_cuil ON verifications(tenant, cuil);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_verif_tenant_wa ON verifications(tenant, to_whatsapp);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_rr_key ON receipt_requests(tenant, cuil, period, to_whatsapp);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_rre_key ON receipt_request_events(tenant, cuil, period, to_whatsapp, created_at);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_multi_tenant_expires ON multi_tenant_selection(expires_at);")
    _try_alter(cur, "CREATE INDEX IF NOT EXISTS idx_ts_queue_pending ON template_send_queue(status, tenant, period, created_at);")

    cur.execute("""
      CREATE TABLE IF NOT EXISTS tenant_branding (
        tenant TEXT PRIMARY KEY,
        accent TEXT DEFAULT '',
        logo BYTEA,
        logo_mime TEXT DEFAULT '',
        updated_at INTEGER
      );
    """)

    conn.commit()
    conn.close()

    
init_db()

def inbound_seen(message_sid: str) -> bool:
    if not message_sid:
        return False
    now = int(time.time())
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("INSERT INTO inbound_dedup(message_sid, created_at) VALUES (%s, %s) ON CONFLICT (message_sid) DO NOTHING", (message_sid, now))
    conn.commit()
    
    inserted = (cur.rowcount == 1)
    conn.close()
    return (not inserted)  # True si ya existía







def save_pdf_sid(tenant: str, cuil: str, period: str, to_whatsapp: str, message_sid: str, origin: str = "INITIAL"):
    now = int(time.time())
    conn = get_db_connection()
    cur = conn.cursor()

    # 1) histórico de PDFs
    cur.execute("""
        INSERT INTO sent_pdfs
        (tenant, cuil, period, to_whatsapp, message_sid, created_at, sign_sent_at, origin)
        VALUES (%s, %s, %s, %s, %s, COALESCE((SELECT created_at FROM sent_pdfs WHERE message_sid = %s), %s),
                COALESCE((SELECT sign_sent_at FROM sent_pdfs WHERE message_sid = %s), NULL),
                %s)
        ON CONFLICT (message_sid) DO UPDATE SET
            tenant = EXCLUDED.tenant,
            cuil = EXCLUDED.cuil,
            period = EXCLUDED.period,
            to_whatsapp = EXCLUDED.to_whatsapp,
            created_at = EXCLUDED.created_at,
            sign_sent_at = EXCLUDED.sign_sent_at,
            origin = EXCLUDED.origin
    """, (tenant, cuil, period, to_whatsapp, message_sid, message_sid, now, message_sid, origin))

    # 2) ✅ tracking para reportes + callbacks
    cur.execute("""
        INSERT INTO message_status
        (message_sid, to_whatsapp, tenant, cuil, period, nombre, kind, created_at, last_status, last_status_at)
        VALUES (%s, %s, %s, %s, %s, '', 'media', %s, 'sent', %s)
        ON CONFLICT (message_sid) DO NOTHING
    """, (message_sid, to_whatsapp, tenant, cuil, period, now, now))

    conn.commit()
    conn.close()

def _digits(s: str) -> str:
    return "".join(ch for ch in (s or "") if ch.isdigit())

def cuil_to_dni(cuil: str) -> str | None:
    """
    Para personas físicas AR: CUIL = XX + DNI(8) + X
    Ej: 20-28169249-3 -> DNI 28169249
    """
    d = _digits(cuil)
    if len(d) != 11:
        return None
    return d[2:10]  # 8 dígitos

import hashlib

def _hash_dni(dni: str, tenant: str = "") -> tuple[str, str]:
    """Hash de DNI + últimos 4 dígitos. Delega en _dni_hash() (salt por tenant)
    para que toda la tabla verifications use un único esquema de hash."""
    dni_digits = "".join(ch for ch in (dni or "") if ch.isdigit())
    last4 = dni_digits[-4:] if len(dni_digits) >= 4 else dni_digits
    h = _dni_hash(dni_digits, tenant) if dni_digits else ""
    return h, last4

def is_verified(tenant: str, cuil: str, to_whatsapp: str) -> bool:
    """
    Wrapper legacy. Mantener por compatibilidad.
    Usa is_verified_contact como única fuente de verdad.
    """
    return is_verified_contact(tenant, cuil, to_whatsapp)

def upsert_verification(tenant: str, cuil: str, to_whatsapp: str, dni: str | None = None):
    now = int(time.time())
    dni_hash, dni_last4 = _hash_dni(dni or "", tenant)
    conn = get_db_connection()
    cur = conn.cursor()

    # Upsert “manual” compatible sin depender de UNIQUE en DB vieja
    cur.execute("""
      SELECT id FROM verifications
      WHERE tenant=%s AND cuil=%s AND to_whatsapp=%s
      LIMIT 1
    """, (tenant, cuil, to_whatsapp))
    row = cur.fetchone()

    if row:
        cur.execute("""
          UPDATE verifications
          SET dni_hash=%s, dni_last4=%s, updated_at=%s
          WHERE id=%s
        """, (dni_hash or None, dni_last4 or None, now, row['id']))
    else:
        cur.execute("""
          INSERT INTO verifications (tenant, cuil, to_whatsapp, dni_hash, dni_last4, verified_at, updated_at)
          VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (tenant, cuil, to_whatsapp, dni_hash or None, dni_last4 or None, now, now))

    conn.commit()
    conn.close()

def delete_verification(tenant: str, cuil: str, to_whatsapp: str):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
      DELETE FROM verifications
      WHERE tenant=%s AND cuil=%s AND to_whatsapp=%s
    """, (tenant, cuil, to_whatsapp))
    conn.commit()
    conn.close()

def get_verifications_rows(tenant: str):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT
            cuil,
            to_whatsapp,
            COALESCE(nombre,'') AS nombre,
            COALESCE(dni_last4,'') AS dni_last4,
            verified_at,
            updated_at
        FROM verifications
        WHERE tenant=%s
        ORDER BY updated_at DESC, verified_at DESC
        LIMIT 1000
    """, (tenant,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def norm_cuil_digits(x: str) -> str:
    return "".join(ch for ch in (x or "") if ch.isdigit())


def is_verified_contact(tenant: str, cuil: str, to_whatsapp: str) -> bool:
    cuil_d = norm_cuil_digits(cuil)
    w = normalize_whatsapp(to_whatsapp)

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
      SELECT 1
      FROM verifications
      WHERE tenant = %s
        AND (
          -- match por cuil normalizado (saca guiones, espacios, etc.)
          replace(replace(cuil, '-', ''), ' ', '') = %s
        )
        AND (
          -- match por whatsapp normalizado
          to_whatsapp = %s
          OR replace(replace(replace(to_whatsapp,'whatsapp:',''),'+',''),' ','') =
             replace(replace(replace(%s,'whatsapp:',''),'+',''),' ','')
        )
      LIMIT 1
    """, (tenant, cuil_d, w, w))
    ok = cur.fetchone() is not None
    conn.close()
    return ok


import hashlib

def _dni_hash(dni: str, tenant: str) -> str:
    # Hash canónico de DNI (salt por tenant). _hash_dni() delega acá:
    # antes había dos esquemas distintos escribiendo en verifications.dni_hash.
    raw = f"{tenant}|{dni}".encode("utf-8")
    return hashlib.sha256(raw).hexdigest()


def set_verified_contact(tenant: str, cuil: str, to_whatsapp: str, dni: str, nombre: str = ""):
    now = int(time.time())
    dni = "".join(ch for ch in (dni or "") if ch.isdigit())
    dni_h = _dni_hash(dni, tenant)
    dni_last4 = dni[-4:] if len(dni) >= 4 else None

    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO verifications (tenant, cuil, to_whatsapp, nombre, dni_hash, dni_last4, verified_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT(tenant, cuil, to_whatsapp)
        DO UPDATE SET
            nombre = CASE
                        WHEN excluded.nombre IS NOT NULL AND excluded.nombre <> '' THEN excluded.nombre
                        ELSE verifications.nombre
                     END,
            dni_hash=excluded.dni_hash,
            dni_last4=excluded.dni_last4,
            verified_at=excluded.verified_at,
            updated_at=excluded.updated_at
    """, (tenant, cuil, to_whatsapp, (nombre or "").strip(), dni_h, dni_last4, now, now))

    conn.commit()
    conn.close()

@app.get("/admin/reset_reenvios")
def admin_reset_reenvios():
    token = (request.args.get("token") or "").strip()
    tenant = (request.args.get("tenant") or "").strip().lower()
    cuil = (request.args.get("cuil") or "").strip()
    whatsapp = (request.args.get("whatsapp") or "").strip()

    if not token or token != ADMIN_TOKEN:
        return Response("Unauthorized", status=401)
    if not tenant or not cuil:
        return Response("Missing tenant/cuil", status=400)

    conn = get_db_connection()
    cur = conn.cursor()

    def _safe(sql, params=()):
        try:
            cur.execute(sql, params)
            return cur.rowcount
        except Exception as e:
            log.warning("%s %s", "WARN reset_reenvios:", e)
            return 0

    deleted = {}

    # Borra SOLO eventos de reenvío (RESEND_LAST)
    # soporta schema viejo/nuevo: origin o source
    if whatsapp:
        deleted["receipt_request_events"] = _safe("""
            DELETE FROM receipt_request_events
            WHERE tenant=%s AND cuil=%s AND (whatsapp=%s OR to_whatsapp=%s)
              AND (
                origin='RESEND_LAST' OR source='RESEND_LAST'
              )
        """, (tenant, cuil, whatsapp, whatsapp))
    else:
        deleted["receipt_request_events"] = _safe("""
            DELETE FROM receipt_request_events
            WHERE tenant=%s AND cuil=%s
              AND (
                origin='RESEND_LAST' OR source='RESEND_LAST'
              )
        """, (tenant, cuil))

    conn.commit()
    conn.close()

    return jsonify({
        "ok": True,
        "mode": "safe_reset_reenvios",
        "tenant": tenant,
        "cuil": cuil,
        "whatsapp": whatsapp or None,
        "deleted": deleted
    })

@app.get("/admin/send_template_preview")
def admin_send_template_preview():
    # Endpoint de acción por URL (dispara envíos): solo ADMIN_TOKEN, sin sesión.
    # Mismo criterio que /admin/reset_reenvios — inmune a CSRF por link.
    if not admin_token_valid():
        return Response("Unauthorized (requiere token=ADMIN_TOKEN)", status=401)

    tenant = (request.args.get("tenant") or "").strip().lower()
    period_label = (request.args.get("period") or "").strip()  # "01/2026"
    limit = int((request.args.get("limit") or "0").strip() or 0)
    require_pdf = (request.args.get("require_pdf") or "true").strip().lower() in ("1", "true", "yes", "on")

    if not tenant:
        return jsonify({"ok": False, "error": "Falta tenant"}), 400

    t = get_tenant(tenant)
    if not t:
        return jsonify({"ok": False, "error": "Tenant inválido"}), 400

    envios_rows = load_envios_rows(tenant, force=False) or []

    def pick(r, keys):
        for k in keys:
            v = r.get(k)
            if v is not None and str(v).strip():
                return str(v).strip()
        return ""

    # --- 1) Si require_pdf: obtener set de CUILs que tienen PDF en ese período ---
    pdf_cuils = None
    period_folder_id = ""

    if require_pdf:
        period_folder_id = get_tenant_period_folder_id(tenant, period_label)  # root -> "MM-YYYY"
        if not period_folder_id:
            return jsonify({
                "ok": True,
                "tenant": tenant,
                "period": period_label,
                "limit": limit,
                "total_match": 0,
                "showing": 0,
                "recipients": [],
                "note": f"No existe carpeta de período para {period_label} (o get_tenant_period_folder_id no la encontró)."
            })

        service = drive_service()

        # IMPORTANTE: esto debe listar archivos del folder (no solo subfolders)
        children = _drive_list_children(service, parent_id=period_folder_id, mime_type=None, page_size=1000) or []

        pdf_cuils = set()
        pdf_names_sample = []

        for c in children:
            name = (c.get("name") or "").strip()
            if not name:
                continue
            if name.lower().endswith(".pdf"):
                base = name[:-4].strip()
                nc = norm_cuil(base)
                if nc:
                    pdf_cuils.add(nc)
                    if len(pdf_names_sample) < 5:
                        pdf_names_sample.append(name)

        # Si no encontró ningún pdf, te devolvemos note con diagnóstico
        if not pdf_cuils:
            return jsonify({
                "ok": True,
                "tenant": tenant,
                "period": period_label,
                "limit": limit,
                "total_match": 0,
                "showing": 0,
                "recipients": [],
                "note": "Encontré la carpeta del período pero no vi PDFs adentro (o _drive_list_children no está listando archivos)."
            })

    # --- 2) Construir recipients desde Excel y filtrar por PDF ---
    recipients = []
    excel_cuil_sample = []

    for r in envios_rows:
        cuil_raw = pick(r, ["cuil", "CUIL"])
        cuil = norm_cuil(cuil_raw)
        nombre = pick(r, ["nombre", "Nombre", "NOMBRE", "name"])
        whatsapp = pick(r, ["to_whatsapp", "whatsapp", "telefono", "tel", "phone"])

        if len(excel_cuil_sample) < 5 and cuil_raw:
            excel_cuil_sample.append(cuil_raw)

        if require_pdf:
            if not cuil:
                continue
            if cuil not in pdf_cuils:
                continue

        recipients.append({"nombre": nombre, "whatsapp": whatsapp, "cuil": cuil})

    total = len(recipients)
    if limit and limit > 0:
        recipients = recipients[:limit]

    resp = {
        "ok": True,
        "tenant": tenant,
        "period": period_label,
        "limit": limit,
        "total_match": total,
        "showing": len(recipients),
        "recipients": recipients,
    }

    # Debug suave (te ayuda si vuelve a quedar en 0)
    if require_pdf and total == 0:
        resp["note"] = "0 match. Revisar formato CUIL. Muestras:"
        resp["excel_cuil_sample"] = excel_cuil_sample
        resp["pdf_cuil_sample"] = list(sorted(list(pdf_cuils)))[:5]

    return jsonify(resp)


def precache_pdfs_for_period(tenant, period):
    """
    Carga todos los PDFs de un período en un dict.
    Returns: {cuil: file_id, ...}
    """
    try:
        # ✅ USAR LA FUNCIÓN CORRECTA
        folder_id = get_tenant_period_folder_id(tenant, period)
        if not folder_id:
            log.warning(f"⚠️ No se encontró carpeta para {tenant}/{period}")
            return {}
        
        service = drive_service()
        results = service.files().list(
            q=f"'{folder_id}' in parents and mimeType='application/pdf'",
            fields="files(id, name)",
            pageSize=1000
        ).execute()
        
        files = results.get('files', [])
        
        cache = {}
        for f in files:
            filename = f['name']
            cuil = strip_pdf(filename)  # "27-12345678-9.pdf" → "27123456789"
            cuil = norm_cuil(cuil)
            if cuil:
                cache[cuil] = f['id']
        
        log.info(f"✅ Pre-cached {len(cache)} PDFs for {tenant}/{period}")
        return cache
        
    except Exception as e:
        log.exception(f"❌ Error pre-caching PDFs: {e}")
        return {}

def set_pending_step(pending_id: int, step: str):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE pending_views SET step=%s WHERE id=%s", (step, pending_id))
    conn.commit()
    conn.close()

def inc_pending_dni_attempts(pending_id: int) -> int:
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("UPDATE pending_views SET dni_attempts = COALESCE(dni_attempts,0) + 1 WHERE id=%s", (pending_id,))
    cur.execute("SELECT dni_attempts FROM pending_views WHERE id=%s", (pending_id,))
    _row = cur.fetchone()
    n = _row['dni_attempts'] if _row and _row['dni_attempts'] is not None else 0
    conn.commit()
    conn.close()
    return int(n)

def get_nombre_for_cuil(tenant: str, cuil: str) -> str:
    try:
        df = get_envios_df_for_tenant(tenant)
        if df is None or df.empty:
            return ""
        df.columns = [str(c).strip().lower() for c in df.columns]

        # columnas posibles
        c_nombre = None
        for n in ("nombre", "name", "empleado", "persona"):
            if n in df.columns:
                c_nombre = n
                break

        c_arch = None
        for n in ("archivo", "cuil", "archivo_norm"):
            if n in df.columns:
                c_arch = n
                break

        if not c_nombre or not c_arch:
            return ""

        def norm(x):
            s = str(x or "").strip().replace(".pdf","")
            try:
                s = strip_pdf(s)
            except Exception:
                pass
            return _digits(s)   # 👈 deja SOLO números: "20-44143190-3" → "20441431903"

        target = norm(cuil)
        df["_cuil_norm"] = df[c_arch].apply(norm)

        row = df[df["_cuil_norm"] == target]
        if row.empty:
            return ""
        return str(row.iloc[0][c_nombre] or "").strip()
    except Exception:
        return ""

def _db_fetchall_dict(sql, params=()):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def get_recibo_estado_rows(tenant: str, period_label: str = ""):
    if period_label:
        return _db_fetchall_dict(
            "SELECT tenant,cuil,period,estado,updated_at FROM recibo_estado WHERE tenant=%s AND period=%s",
            (tenant, period_label)
        )
    return _db_fetchall_dict(
        "SELECT tenant,cuil,period,estado,updated_at FROM recibo_estado WHERE tenant=%s",
        (tenant,)
    )

def get_verifications_rows_for_report(tenant: str):
    return _db_fetchall_dict(
        "SELECT tenant,cuil,to_whatsapp,nombre,verified_at FROM verifications WHERE tenant=%s",
        (tenant,)
    )

def get_message_status_rows(tenant: str, period_label: str = ""):
    if period_label:
        return _db_fetchall_dict(
            """
            SELECT tenant,cuil,period,to_whatsapp,nombre,kind,created_at,last_status,last_status_at,
                   delivered_at,read_at,failed_at,error_code,error_message
            FROM message_status
            WHERE tenant=%s AND period=%s
            """,
            (tenant, period_label)
        )
    return _db_fetchall_dict(
        """
        SELECT tenant,cuil,period,to_whatsapp,nombre,kind,created_at,last_status,last_status_at,
               delivered_at,read_at,failed_at,error_code,error_message
        FROM message_status
        WHERE tenant=%s
        """,
        (tenant,)
    )

from io import BytesIO
import time
import re
from flask import Response, request, send_file
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def period_to_label(p: str) -> str:
    p = (p or "").strip()
    if not p:
        return ""
    # admite 01/2026 o 01-2026
    p = p.replace("-", "/")
    # normaliza 1/2026 -> 01/2026
    m = re.match(r"^(\d{1,2})/(\d{4})$", p)
    if not m:
        return p
    mm = int(m.group(1))
    yyyy = m.group(2)
    return f"{mm:02d}/{yyyy}"


def _autosize_ws(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)


from io import BytesIO
import time
import pandas as pd
from flask import request, Response, send_file

def norm_period_label(p: str) -> str:
    """
    Normaliza a 'MM/AAAA'
    Acepta: 'MM/AAAA', 'MM-AAAA', 'AAAA-MM', 'AAAA/MM'
    """
    p = (p or "").strip()
    if not p:
        return ""
    p = p.replace("\\", "/").replace("-", "/")
    parts = [x for x in p.split("/") if x.strip()]
    if len(parts) != 2:
        return ""

    a, b = parts[0].zfill(2), parts[1]
    # si viene AAAA/MM
    if len(a) == 4:
        yyyy = a
        mm = b.zfill(2)
    else:
        mm = a.zfill(2)
        yyyy = b

    if not (mm.isdigit() and yyyy.isdigit()):
        return ""
    if not (1 <= int(mm) <= 12):
        return ""
    if len(yyyy) != 4:
        return ""

    return f"{mm}/{yyyy}"

from io import BytesIO

import time
import os
from reportlab.platypus import Image

def _load_icon_flowable():
    import os
    from reportlab.platypus import Image
    from reportlab.lib.units import cm

    # __file__ = /opt/render/project/src/app.py  (en Render)
    base_dir = os.path.dirname(os.path.abspath(__file__))

    candidates = [
        os.path.join(base_dir, "static", "icon_dashboard.png"),           # /src/static/...
        os.path.join(base_dir, "..", "static", "icon_dashboard.png"),     # /static/... (si tu static quedó 1 nivel arriba)
        os.path.join(os.getcwd(), "static", "icon_dashboard.png"),
        "static/icon_dashboard.png",
        r"C:\Users\lucasdl\Desktop\TWILIO\twilio_webhook\icon_dashboard.png",
    ]

    for p in candidates:
        p = os.path.abspath(p)
        try:
            if os.path.exists(p):
                return Image(p, width=1.8*cm, height=1.8*cm)
        except Exception:
            pass

    return None



def generate_pdf_report_v2(tenant: str, period_filter: str = ""):
    """
    Página 1: KPIs + 1 donut (Estados por jerarquía)
    Página 2+: Tabla con TODAS las personas (sin recortes) y sin duplicar "FIRMADO"
    Jerarquía: FIRMADO > OBSERVADO > PEND. RESPUESTA > PEND. ENVÍO
    Latest-wins por timestamps.
    """
    from io import BytesIO
    import time
    

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak

    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    tenant = (tenant or "").strip().lower()
    period_filter = norm_period_label(period_filter)

    def _safe_int(x):
        try:
            return int(x or 0)
        except Exception:
            return 0

    # ========= DB =========
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        SELECT
            to_whatsapp, tenant, cuil, period, nombre, kind,
            created_at, delivered_at, read_at, failed_at,
            last_status, last_status_at
        FROM message_status
        WHERE tenant = %s
    """, (tenant,))
    msg_rows = cur.fetchall()

    cur.execute("""
        SELECT tenant, cuil, period, estado, updated_at
        FROM recibo_estado
        WHERE tenant = %s
    """, (tenant,))
    estado_rows = cur.fetchall()

    cur.execute("""
        SELECT tenant,cuil,period,to_whatsapp,request_count,last_requested_at
        FROM receipt_requests
        WHERE tenant = %s
    """, (tenant,))
    rr_rows = cur.fetchall()

    conn.close()

    # ========= maps (más reciente por key) =========
    estado_map = {}
    for r in estado_rows:
        c = norm_cuil(r["cuil"])
        p = norm_period_label(r["period"] or "")
        ts = _safe_int(r["updated_at"])
        if not (c and p):
            continue
        key = (c, p)
        prev = estado_map.get(key)
        if (prev is None) or (ts >= _safe_int(prev.get("ts"))):
            estado_map[key] = {"estado": (r["estado"] or "").strip(), "ts": ts}

    rr_map = {}
    for r in rr_rows:
        c = norm_cuil(r["cuil"])
        p = norm_period_label(r["period"] or "")
        w = (r["to_whatsapp"] or "").strip()
        ts = _safe_int(r["last_requested_at"])
        if not (c and p and w):
            continue
        key = (w, c, p)
        prev = rr_map.get(key)
        if (prev is None) or (ts >= _safe_int(prev.get("last"))):
            rr_map[key] = {"count": int(r["request_count"] or 0), "last": ts}

    # ========= agregación por (whatsapp, periodo) con latest-wins =========
    agg = {}

    def _ensure(k, base):
        if k not in agg:
            agg[k] = {
                "Periodo": base.get("Periodo", ""),
                "Nombre": base.get("Nombre", ""),
                "CUIL": base.get("CUIL", ""),
                "WhatsApp": base.get("WhatsApp", ""),

                "Respuesta": "",          # se usa solo para clasificar, no se imprime
                "Pedidos": 0,
                "Ultimo_pedido": "",
                "_last_status": "",
                "_last_status_at": 0,
                "_sent_ts": 0,
                "_delivered_ts": 0,
                "_read_ts": 0,
                "_failed_ts": 0,
                "_estado_ts": 0,
                "_pedido_ts": 0,
                "_last_ts": 0,
            }

    for row in msg_rows:
        wpp = (row["to_whatsapp"] or "").strip()
        if not wpp:
            continue

        cuil = norm_cuil(row["cuil"])
        nombre = (row["nombre"] or "").strip()
        per = norm_period_label((row["period"] or "").strip())
        if period_filter and per != period_filter:
            continue
        if not per:
            continue

        k = (wpp, per)
        _ensure(k, {"Periodo": per, "Nombre": nombre, "CUIL": cuil, "WhatsApp": wpp})
        rec = agg[k]

        last_status = (row["last_status"] or "").strip().upper()
        last_status_at = _safe_int(row["last_status_at"])

        # guardar el last_status más reciente
        if last_status and last_status_at >= rec.get("_last_status_at", 0):
            rec["_last_status"] = last_status
            rec["_last_status_at"] = last_status_at
            rec["_last_ts"] = max(rec["_last_ts"], last_status_at)

        if nombre and not rec["Nombre"]:
            rec["Nombre"] = nombre
        if cuil and not rec["CUIL"]:
            rec["CUIL"] = cuil

        kind = (row["kind"] or "").strip().lower()
        if kind not in ("pdf", "media"):
            continue

        sent_ts = _safe_int(row["created_at"])
        deliv_ts = _safe_int(row["delivered_at"])
        read_ts = _safe_int(row["read_at"])
        fail_ts = _safe_int(row["failed_at"])

        if sent_ts and sent_ts >= rec["_sent_ts"]:
            rec["_sent_ts"] = sent_ts
        if deliv_ts and deliv_ts >= rec["_delivered_ts"]:
            rec["_delivered_ts"] = deliv_ts
        if read_ts and read_ts >= rec["_read_ts"]:
            rec["_read_ts"] = read_ts
        if fail_ts and fail_ts >= rec["_failed_ts"]:
            rec["_failed_ts"] = fail_ts

        rec["_last_ts"] = max(rec["_last_ts"], sent_ts, deliv_ts, read_ts, fail_ts)

    # mezclar estado + pedidos (latest-wins)
    for rec in agg.values():
        cuil = rec.get("CUIL", "")
        per = rec.get("Periodo", "")

        st = estado_map.get((cuil, per))
        if st:
            ts = _safe_int(st.get("ts"))
            if ts >= rec["_estado_ts"]:
                rec["_estado_ts"] = ts
                rec["Respuesta"] = st.get("estado", "") or ""
            rec["_last_ts"] = max(rec["_last_ts"], ts)

        rr = rr_map.get((rec["WhatsApp"], cuil, per))
        if rr:
            ts = _safe_int(rr.get("last"))
            if ts >= rec["_pedido_ts"]:
                rec["_pedido_ts"] = ts
                rec["Pedidos"] = int(rr.get("count") or 0)
                rec["Ultimo_pedido"] = ts_to_str(ts) if ts else ""
            rec["_last_ts"] = max(rec["_last_ts"], ts)

    rows = list(agg.values())

    # ========= Estado ÚNICO con jerarquía =========
    def classify_status(r: dict) -> str:
        resp = (r.get("Respuesta", "") or "").strip().upper()

        # "Se envió" = cualquier evidencia de PDF en status:
        any_sent = (
            _safe_int(r.get("_sent_ts")) > 0
            or _safe_int(r.get("_delivered_ts")) > 0
            or _safe_int(r.get("_read_ts")) > 0
        )

        if resp == "FIRMADO":
            return "FIRMADO"
        if resp == "OBSERVADO":
            return "OBSERVADO"

        if any_sent:
            return "PEND. RESPUESTA"

        return "PEND. ENVÍO"

    for r in rows:
        r["_status"] = classify_status(r)

    def count_status(s):
        return sum(1 for r in rows if r.get("_status") == s)

    c_firm = count_status("FIRMADO")
    c_obs = count_status("OBSERVADO")
    c_presp = count_status("PEND. RESPUESTA")
    c_penv = count_status("PEND. ENVÍO")
    c_plec = count_status("PEND. LECTURA")
    c_ok = count_status("OK")

    # ========= Donut estados =========
    def donut_png(labels, values, title):
        from io import BytesIO
        import matplotlib.pyplot as plt

        # Colores por etiqueta
        color_map = {
            "Firmados": "#22c55e",
            "Observados": "#f59e0b",
            "Pend. respuesta": "#3b82f6",
            "Pend. envío": "#ef4444",
            "Pend. lectura": "#a855f7",
            "OK": "#64748b",
            "Sin datos": "#cbd5e1",
        }

        # Normalizar a enteros >= 0
        pairs = []
        for l, v in zip(labels, values):
            try:
                iv = int(v or 0)
            except Exception:
                iv = 0
            if iv < 0:
                iv = 0
            pairs.append((l, iv))

        total = sum(v for _, v in pairs)

        # Si todo es 0 -> donut "Sin datos" (evita NaN)
        if total == 0:
            pairs = [("Sin datos", 1)]
            total = 0

        # (opcional) sacar ceros cuando hay datos reales
        if pairs and pairs[0][0] != "Sin datos":
            pairs = [(l, v) for (l, v) in pairs if v > 0]
            # por si filtrando quedó vacío
            if not pairs:
                pairs = [("Sin datos", 1)]
                total = 0

        labels2 = [l for l, _ in pairs]
        values2 = [v for _, v in pairs]
        colors_list = [color_map.get(l, "#94a3b8") for l in labels2]

        # FIG CUADRADA (no chata)
        fig = plt.figure(figsize=(4.8, 4.8), dpi=220)
        ax = fig.add_subplot(111)

        wedges, _ = ax.pie(
            values2,
            startangle=90,
            colors=colors_list,
            wedgeprops=dict(width=0.42, edgecolor="white", linewidth=2),
        )
        ax.set_aspect("equal")
        ax.set_title(title, fontsize=13, pad=10)

        # Centro: total (si sin datos, mostramos 0)
        ax.text(0, 0.06, f"{total}", ha="center", va="center", fontsize=22, fontweight="bold")
        ax.text(0, -0.12, "total", ha="center", va="center", fontsize=10, color="#6b7280")

        # Leyenda ABAJO (no panorámico)
        ax.legend(
            wedges,
            [f"{l} • {v if l!='Sin datos' else 0}" for l, v in pairs],
            loc="upper center",
            bbox_to_anchor=(0.5, -0.08),
            ncol=2,
            fontsize=9,
            frameon=False,
            handlelength=1.2,
            columnspacing=1.4,
        )

        fig.tight_layout()
        buf = BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight", transparent=True)
        plt.close(fig)
        buf.seek(0)
        return buf

    labels_est = ["Firmados", "Observados", "Pend. respuesta", "Pend. envío"]
    values_est = [c_firm, c_obs, c_presp, c_penv]
    if c_plec:
        labels_est.append("Pend. lectura")
        values_est.append(c_plec)
    if c_ok:
        labels_est.append("OK")
        values_est.append(c_ok)

    donut_estados = donut_png(labels_est, values_est, "Estados (jerarquía)")

    # ========= Orden de tabla =========
    order_rank = {
        "FIRMADO": 1,
        "OBSERVADO": 2,
        "PEND. RESPUESTA": 3,
        "PEND. ENVÍO": 4,
        "PEND. LECTURA": 5,
        "OK": 6,
    }
    rows_sorted = sorted(
        rows,
        key=lambda r: (order_rank.get(r.get("_status"), 99), -int(r.get("_last_ts") or 0))
    )

    # ========= PDF =========
    out = BytesIO()
    page_w, page_h = landscape(A4)

    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4),
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        topMargin=1.0 * cm,
        bottomMargin=1.0 * cm,
        title=f"Recibos - {tenant}",
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleCool", fontSize=18, leading=20, spaceAfter=2))
    styles.add(ParagraphStyle(name="SubCool", fontSize=10, leading=12, textColor=colors.HexColor("#6b7280")))
    styles.add(ParagraphStyle(name="HSection", fontSize=12, leading=14, spaceBefore=8, spaceAfter=6))
    styles.add(ParagraphStyle(name="Cell7", fontSize=7, leading=8))
    styles.add(ParagraphStyle(name="Cell7B", fontSize=7, leading=8, fontName="Helvetica-Bold"))

    period_label = period_filter if period_filter else "TODOS"
    gen_ts = time.strftime("%Y-%m-%d %H:%M")

    story = []
    # --- Página 1: solo gráficos / KPIs ---
    icon = _load_icon_flowable()

    left = [
        Paragraph(f"📌 Control de Recibos • <b>{tenant}</b>", styles["TitleCool"]),
        Paragraph(f"Período: <b>{period_label}</b> • Generado: {gen_ts}", styles["SubCool"]),
    ]

    # si no hay icono, ponemos un Spacer para que no explote el layout
    right = icon if icon else Spacer(1, 1.8*cm)

    header = Table([[left, right]], colWidths=[24.7*cm, 2.6*cm])
    header.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING", (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
    ]))
    story.append(header)
    story.append(Spacer(1, 0.25*cm))


    def kpi_box(items):
        data = []
        for lab, val in items:
            data.append([
                Paragraph(f"<font color='#6b7280'>{lab}</font>", styles["BodyText"]),
                Paragraph(f"<b><font size='16'>{val}</font></b>", styles["BodyText"])
            ])
        t = Table(data, colWidths=[4.2*cm, 1.8*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#f6f7fb")),
            ("BOX", (0,0), (-1,-1), 0.6, colors.HexColor("#e5e7eb")),
            ("INNERGRID", (0,0), (-1,-1), 0.4, colors.HexColor("#e5e7eb")),
            ("LEFTPADDING", (0,0), (-1,-1), 10),
            ("RIGHTPADDING", (0,0), (-1,-1), 10),
            ("TOPPADDING", (0,0), (-1,-1), 7),
            ("BOTTOMPADDING", (0,0), (-1,-1), 7),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]))
        return t

    kpis = kpi_box([
        ("Firmados", c_firm),
        ("Observados", c_obs),
        ("Pend. respuesta", c_presp),
        ("Pend. envío", c_penv),
        ("Pend. lectura", c_plec),
        ("OK", c_ok),
        ("Total", len(rows)),
    ])

    top_grid = Table(
        [[kpis, Image(donut_estados, width=18.0*cm, height=10.2*cm)]],
        colWidths=[7.0*cm, 20.3*cm]
    )
    top_grid.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING", (0,0), (-1,-1), 0),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0),
    ]))
    story.append(top_grid)

    # --- salto: tabla arranca en hoja 2 ---
    story.append(PageBreak())

    # --- Página 2+: tabla ---
    story.append(Paragraph("📋 Detalle (todas las personas)", styles["HSection"]))

    def clip(s, n):
        s = str(s or "")
        return s if len(s) <= n else s[:n-1] + "…"

    def fmt_ts(ts):
        ts = _safe_int(ts)
        return ts_to_str(ts) if ts else ""

    # ✅ Sin columna "Respuesta" para que no duplique FIRMADO/OBSERVADO
    headers = ["Estado", "Período", "Nombre", "CUIL", "WhatsApp", "Enviado", "Leído", "Pedidos", "Últ. pedido", "Últ. act."]
    table_data = [[Paragraph(h, styles["Cell7B"]) for h in headers]]

    for r in rows_sorted:
        table_data.append([
            Paragraph(r.get("_status",""), styles["Cell7"]),
            Paragraph(r.get("Periodo","") or "", styles["Cell7"]),
            Paragraph(clip(r.get("Nombre","") or "—", 34), styles["Cell7"]),
            Paragraph(clip(r.get("CUIL","") or "—", 16), styles["Cell7"]),
            Paragraph(clip(r.get("WhatsApp","") or "—", 18), styles["Cell7"]),
            Paragraph(fmt_ts(r.get("_sent_ts")), styles["Cell7"]),
            Paragraph(fmt_ts(r.get("_read_ts")), styles["Cell7"]),
            Paragraph(str(int(r.get("Pedidos") or 0)), styles["Cell7"]),
            Paragraph(r.get("Ultimo_pedido","") or "", styles["Cell7"]),
            Paragraph(fmt_ts(r.get("_last_ts")), styles["Cell7"]),
        ])

    # Anchos que entran en A4 apaisado con tus márgenes (~27.3 cm útiles)
    col_widths = [
        2.6*cm,  # Estado
        2.2*cm,  # Período
        6.6*cm,  # Nombre
        3.0*cm,  # CUIL
        3.2*cm,  # WhatsApp
        2.6*cm,  # Enviado
        2.6*cm,  # Leído
        1.5*cm,  # Pedidos
        3.0*cm,  # Últ. pedido
        2.6*cm,  # Últ. act.
    ]

    people_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    people_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#e5e7eb")),
        ("BOX", (0,0), (-1,-1), 0.8, colors.HexColor("#e5e7eb")),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (7,1), (7,-1), "CENTER"),  # Pedidos
    ]))

    for i in range(1, len(table_data)):
        if i % 2 == 0:
            people_table.setStyle(TableStyle([("BACKGROUND", (0,i), (-1,i), colors.HexColor("#f9fafb"))]))

    story.append(people_table)

    def on_page(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#9ca3af"))
        canvas.drawRightString(page_w - 1.2*cm, 0.8*cm, f"Página {doc_obj.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    out.seek(0)
    return out


@app.get("/admin/report_recibos.pdf")
@admin_required
def admin_report_recibos_pdf():
    token = _get_admin_token_from_request()
    tenant = (request.args.get("tenant") or "").strip().lower()
    period = (request.args.get("period") or "").strip()

    if not tenant:
        return Response("Falta tenant", status=400)

    buf = generate_pdf_report_v2(tenant, period_filter=period)

    filename = f"reporte_recibos_{tenant}_{(norm_period_label(period).replace('/','-') if period else 'todos')}.pdf"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/pdf"
    )


@app.get("/admin/report_recibos.xlsx")
@admin_required
def admin_report_recibos_xlsx():
    token = _get_admin_token_from_request()
    tenant = (request.args.get("tenant") or "").strip().lower()
    period = (request.args.get("period") or "").strip()  # MM/AAAA desde el panel (o vacío)

    if not tenant:
        return Response("Falta tenant", status=400)

    buf = generate_excel_report_v2(tenant, period_filter=period)

    filename = f"reporte_recibos_{tenant}_{(norm_period_label(period).replace('/','-') if period else 'todos')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def list_verifications(tenant: str, q: str = ""):
    conn = get_db_connection()
    cur = conn.cursor()

    q = (q or "").strip().lower()
    params = [tenant]

    sql = """
      SELECT id, tenant, cuil, to_whatsapp, nombre, dni_last4, verified_at, updated_at
      FROM verifications
      WHERE tenant = %s
    """

    if q:
        sql += """
          AND (
            lower(cuil) LIKE %s
            OR lower(to_whatsapp) LIKE %s
            OR lower(COALESCE(nombre,'')) LIKE %s
            OR lower(COALESCE(dni_last4,'')) LIKE %s
          )
        """
        like = f"%{q}%"
        params += [like, like, like, like]

    sql += " ORDER BY updated_at DESC, verified_at DESC LIMIT 200"

    cur.execute(sql, params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows



def get_pdf_by_sid(sid):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM sent_pdfs WHERE message_sid = %s", (sid,))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None


from functools import wraps
from flask import request, Response

# (_get_admin_token_from_request y admin_required están definidos arriba,
#  junto a los demás helpers de auth: los decoradores se evalúan al importar
#  y acá quedaban DESPUÉS de su primer uso.)



def add_pending_view(to_whatsapp: str, tenant: str, cuil: str, period: str, origin: str = "INITIAL"):
    now = int(time.time())
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
      INSERT INTO pending_views
        (to_whatsapp, tenant, cuil, period, created_at, step, dni_attempts, origin)
      VALUES (%s, %s, %s, %s, %s, 'READY', 0, %s)
      ON CONFLICT(to_whatsapp) DO UPDATE SET
        tenant=excluded.tenant,
        cuil=excluded.cuil,
        period=excluded.period,
        created_at=excluded.created_at,
        step='READY',
        dni_attempts=0,
        origin=excluded.origin
    """, (to_whatsapp, tenant, cuil, period, now, origin))

    conn.commit()
    conn.close()



def get_latest_pending_view(to_whatsapp: str):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
      SELECT
        id,
        to_whatsapp,
        tenant,
        cuil,
        period,
        created_at,
        COALESCE(step, 'READY') AS step,
        COALESCE(dni_attempts, 0) AS dni_attempts,
        COALESCE(origin, 'INITIAL') AS origin,
        COALESCE(period_offset, 0) AS period_offset
      FROM pending_views
      WHERE to_whatsapp=%s
      ORDER BY created_at DESC
      LIMIT 1
    """, (to_whatsapp,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return None

    try:
        return dict(row)
    except Exception:
        # fallback si fetchone devuelve tuple
        keys = ["id","to_whatsapp","tenant","cuil","period","created_at","step","dni_attempts","origin","period_offset"]
        return dict(zip(keys, row))
    
def consume_pending_view(pending_id: int):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM pending_views WHERE id=%s;", (pending_id,))
    conn.commit()
    conn.close()

def set_recibo_estado(tenant: str, cuil: str, period: str, estado: str):
    tenant = (tenant or "").strip().lower()
    cuil = norm_cuil(cuil)
    period = norm_period_label(period)
    estado = (estado or "").strip().upper()

    conn = get_db_connection()
    cur = conn.cursor()
    now = int(time.time())

    cur.execute("""
        SELECT estado FROM recibo_estado
        WHERE tenant=%s AND cuil=%s AND period=%s
    """, (tenant, cuil, period))
    row = cur.fetchone()
    if row and (row['estado'] or "").strip().upper() == "FIRMADO":
        conn.close()
        return

    cur.execute("""
      INSERT INTO recibo_estado (tenant, cuil, period, estado, updated_at)
      VALUES (%s, %s, %s, %s, %s)
      ON CONFLICT(tenant, cuil, period) DO UPDATE SET
        estado=excluded.estado,
        updated_at=excluded.updated_at;
    """, (tenant, cuil, period, estado, now))
    conn.commit()
    conn.close()


def generate_signature_certificate_pdf(tenant: str, cuil: str, period: str) -> bytes | None:
    """
    Genera una constancia de firma de recibo en PDF (bytes).
    Devuelve None si el recibo todavía NO está FIRMADO ni OBSERVADO.
    """
    from io import BytesIO
    import datetime as _dt
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    tenant = (tenant or "").strip().lower()
    cuil = norm_cuil(cuil)
    period = norm_period_label(period)

    try:
        from zoneinfo import ZoneInfo
        _AR_TZ = ZoneInfo("America/Argentina/Buenos_Aires")
    except Exception:
        _AR_TZ = _dt.timezone(_dt.timedelta(hours=-3))

    def _fmt_ar(ts):
        if not ts:
            return "—"
        # ts es epoch UTC (int(time.time())); lo mostramos en hora de Argentina.
        return _dt.datetime.fromtimestamp(int(ts), _AR_TZ).strftime("%d/%m/%Y %H:%M:%S") + " hs (ART)"

    # --- 1) Estado de firma (obligatorio) ---
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT estado, updated_at
        FROM recibo_estado
        WHERE tenant=%s AND cuil=%s AND period=%s
    """, (tenant, cuil, period))
    est_row = cur.fetchone()

    estado = (est_row["estado"].strip().upper() if est_row and est_row["estado"] else "")
    if estado not in ("FIRMADO", "OBSERVADO"):
        conn.close()
        log.info("Constancia: recibo no firmado (%s/%s/%s estado=%s)",
                 tenant, cuil, period, estado or "NINGUNO")
        return None
    firma_ts = est_row["updated_at"]

    # --- 2) Identidad verificada (la más reciente) ---
    # --- 2) Identidad verificada (la más reciente). Normalizamos el CUIL en SQL
    #         porque verifications puede guardarlo con guiones. ---
    cur.execute("""
        SELECT nombre, dni_last4, dni_hash, to_whatsapp, verified_at
        FROM verifications
        WHERE tenant=%s
          AND regexp_replace(COALESCE(cuil,''), '\\D', '', 'g') = %s
        ORDER BY verified_at DESC NULLS LAST
        LIMIT 1
    """, (tenant, cuil))
    ver = cur.fetchone() or {}

    # --- 3) Trazabilidad del envío del PDF (entregado / leído) ---
    # --- 3) Trazabilidad del envío del PDF (entregado / leído).
    #         El PDF se registra con kind='media' (a veces 'pdf') y el CUIL/período
    #         pueden venir con guiones, así que normalizamos ambos lados en SQL. ---
    cur.execute("""
        SELECT delivered_at, read_at, to_whatsapp
        FROM message_status
        WHERE tenant=%s
          AND regexp_replace(COALESCE(cuil,''), '\\D', '', 'g') = %s
          AND replace(COALESCE(period,''), '-', '/') = %s
          AND lower(COALESCE(kind,'')) IN ('media', 'pdf')
        ORDER BY COALESCE(delivered_at, read_at, created_at, 0) DESC
        LIMIT 1
    """, (tenant, cuil, period))
    msg = cur.fetchone() or {}
    conn.close()

    empresa = (get_tenant(tenant) or {}).get("display_name") or tenant
    nombre = (ver.get("nombre") or get_nombre_for_cuil(tenant, cuil) or "—").strip()
    whatsapp = (ver.get("to_whatsapp") or msg.get("to_whatsapp") or "").replace("whatsapp:", "").strip() or "—"
    cuil_fmt = format_cuil_with_dashes(cuil)
    dni_last4 = (ver.get("dni_last4") or "").strip()
    dni_hash = (ver.get("dni_hash") or "").strip()
    estado_txt = "FIRMADO" if estado == "FIRMADO" else "OBSERVADO"
    emitido_ts = int(time.time())

    base = f"sigcert|{tenant}|{cuil}|{period}|{estado}|{firma_ts}"
    codigo = hmac.new(MEDIA_SECRET.encode("utf-8"), base.encode("utf-8"),
                      hashlib.sha256).hexdigest()[:16].upper()
    codigo_fmt = "-".join(codigo[i:i + 4] for i in range(0, len(codigo), 4))

    AZUL = colors.HexColor("#1f2766")
    ORO = colors.HexColor("#F4C430")
    GRIS = colors.HexColor("#5b6478")
    LINEA = colors.HexColor("#d7dbe6")

    ss = getSampleStyleSheet()
    st_title = ParagraphStyle("ct_title", parent=ss["Title"], fontName="Helvetica-Bold",
                              fontSize=18, textColor=AZUL, spaceAfter=2, leading=22)
    st_sub = ParagraphStyle("ct_sub", parent=ss["Normal"], fontName="Helvetica",
                            fontSize=10, textColor=GRIS, spaceAfter=2)
    st_section = ParagraphStyle("ct_section", parent=ss["Normal"], fontName="Helvetica-Bold",
                                fontSize=11, textColor=AZUL, spaceBefore=14, spaceAfter=6)
    st_body = ParagraphStyle("ct_body", parent=ss["Normal"], fontName="Helvetica",
                             fontSize=10, textColor=colors.black, leading=15)
    st_label = ParagraphStyle("ct_label", parent=ss["Normal"], fontName="Helvetica",
                              fontSize=9.5, textColor=GRIS)
    st_value = ParagraphStyle("ct_value", parent=ss["Normal"], fontName="Helvetica-Bold",
                              fontSize=10.5, textColor=colors.black)
    st_small = ParagraphStyle("ct_small", parent=ss["Normal"], fontName="Helvetica",
                              fontSize=8, textColor=GRIS, leading=11)
    st_code = ParagraphStyle("ct_code", parent=ss["Normal"], fontName="Courier-Bold",
                             fontSize=13, textColor=AZUL)

    def kv_table(rows):
        data = [[Paragraph(k, st_label), Paragraph(v, st_value)] for k, v in rows]
        t = Table(data, colWidths=[5.5 * cm, 11.0 * cm])
        t.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("LINEBELOW", (0, 0), (-1, -2), 0.4, LINEA),
        ]))
        return t

    story = []
    try:
        icon = _load_icon_flowable()
    except Exception:
        icon = None
    title_block = [Paragraph("Constancia de firma de recibo de sueldo", st_title),
                   Paragraph(f"{empresa}", st_sub)]
    if icon is not None:
        header = Table([[icon, title_block]], colWidths=[2.2 * cm, 14.3 * cm])
        header.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ]))
        story.append(header)
    else:
        story.extend(title_block)

    story.append(Spacer(1, 6))
    story.append(Table([[""]], colWidths=[16.5 * cm],
                       style=TableStyle([("LINEABOVE", (0, 0), (-1, -1), 2, ORO)])))
    story.append(Spacer(1, 10))

    verbo = "firmó de conformidad" if estado == "FIRMADO" else "registró con observaciones"
    story.append(Paragraph(
        f"Se deja constancia de que el/la empleado/a identificado/a a continuación "
        f"{verbo} el recibo de sueldo correspondiente al período <b>{period}</b>, "
        f"a través del canal oficial de WhatsApp de la empresa, "
        f"previa verificación de identidad.", st_body))

    story.append(Paragraph("Datos del recibo", st_section))
    story.append(kv_table([
        ("Empresa", empresa),
        ("Empleado/a", nombre),
        ("CUIL", cuil_fmt),
        ("Período", period),
        ("Estado", estado_txt),
    ]))

    story.append(Paragraph("Trazabilidad de la operación", st_section))
    story.append(kv_table([
        ("PDF entregado", _fmt_ar(msg.get("delivered_at"))),
        ("Fecha y hora de firma", _fmt_ar(firma_ts)),
        ("WhatsApp del firmante", whatsapp),
    ]))

    story.append(Paragraph("Verificación de identidad", st_section))
    id_rows = [("Identidad verificada", _fmt_ar(ver.get("verified_at")) if ver.get("verified_at") else "—")]
    if dni_last4:
        id_rows.append(("DNI (últimos 4 dígitos)", f"•••••{dni_last4}"))
    if dni_hash:
        id_rows.append(("Huella del DNI (SHA-256)", dni_hash[:32] + "…"))
    story.append(kv_table(id_rows))

    story.append(Spacer(1, 16))
    code_block = Table(
        [[Paragraph("Código de verificación", st_label)],
         [Paragraph(codigo_fmt, st_code)]],
        colWidths=[16.5 * cm])
    code_block.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f3f5fb")),
        ("BOX", (0, 0), (-1, -1), 0.6, LINEA),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
    ]))
    story.append(code_block)

    story.append(Spacer(1, 10))
    story.append(Paragraph(
        f"Documento generado electrónicamente el {_fmt_ar(emitido_ts)}. "
        f"El código de verificación es único e inalterable: cualquier modificación "
        f"de los datos de esta constancia invalida dicho código. "
        f"Este comprobante refleja los registros del sistema de distribución de "
        f"recibos al momento de su emisión.", st_small))

    def _on_page(canvas, doc_obj):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(GRIS)
        canvas.drawString(2 * cm, 1.2 * cm, f"Constancia {codigo_fmt}")
        canvas.drawRightString(A4[0] - 2 * cm, 1.2 * cm, "Página %d" % doc_obj.page)
        canvas.restoreState()

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=1.8 * cm, bottomMargin=1.8 * cm,
        title=f"Constancia de firma {cuil_fmt} {period}",
        author=empresa,
    )
    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    return buf.getvalue()


@app.get("/portal/constancias-firma.zip")
def portal_constancias_firma_zip():
    """Descarga en un ZIP la constancia de firma de TODOS los empleados
    firmados/observados del período (scoped al tenant del usuario logueado)."""
    import zipfile

    auth = require_portal_login()
    if auth:
        return auth

    user_id = session.get("portal_user_id")
    user = get_portal_user_by_id(user_id)
    tenant = user["tenant"]  # tenant fijado por la sesión, NO por la URL

    period = norm_period_label((request.args.get("period") or "").strip())

    filtro = (request.args.get("estado") or "todos").strip().lower()
    if filtro == "firmado":
        estados_ok = {"FIRMADO"}
    elif filtro == "observado":
        estados_ok = {"OBSERVADO"}
    else:
        estados_ok = {"FIRMADO", "OBSERVADO"}

    if not period:
        return Response("Falta el período (?period=MM/AAAA)", status=400)

    rows = get_recibo_estado_rows(tenant, period) or []
    rows = [r for r in rows if (r.get("estado") or "").strip().upper() in estados_ok]
    if not rows:
        return Response("No hay recibos firmados/observados para ese período.", status=404)

    mem = io.BytesIO()
    usados = set()
    generadas = 0
    omitidas = []

    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in rows:
            cuil = (r.get("cuil") or "").strip()
            if not cuil:
                continue
            try:
                pdf = generate_signature_certificate_pdf(tenant, cuil, period)
            except Exception as e:
                log.exception("Portal constancias ZIP: error cuil=%s: %s", cuil, e)
                pdf = None

            if not pdf:
                omitidas.append(cuil)
                continue

            nombre = get_nombre_for_cuil(tenant, cuil) or "sin_nombre"
            cuil_d = norm_cuil(cuil)
            base = re.sub(r"[^\w\-. ]", "_", f"constancia_{nombre}_{cuil_d}").strip("_")

            arcname, n = f"{base}.pdf", 1
            while arcname.lower() in usados:
                arcname = f"{base}_{n}.pdf"
                n += 1
            usados.add(arcname.lower())

            zf.writestr(arcname, pdf)
            generadas += 1

    if generadas == 0:
        return Response("No se pudo generar ninguna constancia.", status=404)

    mem.seek(0)
    data = mem.read()

    try:
        log_portal_action(
            user_id, tenant, "descarga_constancias_firma",
            details=f"period={period} estado={filtro} generadas={generadas} omitidas={len(omitidas)}",
            ip_address=request.remote_addr or "",
        )
    except Exception:
        pass

    empresa = (get_tenant(tenant) or {}).get("display_name") or tenant
    empresa_slug = re.sub(r"[^\w\-]", "_", empresa)
    zip_name = f"constancias_firma_{empresa_slug}_{period.replace('/', '-')}.zip"

    resp = Response(data, mimetype="application/zip")
    resp.headers["Content-Disposition"] = f'attachment; filename="{zip_name}"'
    resp.headers["Content-Length"] = str(len(data))
    return resp


def get_recibo_estado(tenant: str, cuil: str, period: str):
    tenant = (tenant or "").strip().lower()
    cuil = norm_cuil(cuil)
    period = norm_period_label(period)

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
      SELECT estado FROM recibo_estado
      WHERE tenant=%s AND cuil=%s AND period=%s
      LIMIT 1;
    """, (tenant, cuil, period))
    row = cur.fetchone()
    conn.close()
    return (row["estado"].strip().upper() if row and row["estado"] else None)

import requests

def pdf_exists_for_tenant_period_cuil(tenant, cuil, period):
    url = f"{os.environ.get('PUBLIC_BASE_URL','').rstrip('/')}/media/pdf"
    if not url.startswith("http"):
        # fallback: no podemos verificar
        return True
    t_q = str(tenant or "").strip().lower()
    c_q = str(cuil or "").strip()
    p_q = str(period or "").strip()
    r = requests.get(url, params={
        "tenant": t_q,
        "cuil": c_q,
        "period": p_q,
        "token": make_media_token("pdf", tenant=t_q, cuil=c_q, period=p_q)
    }, timeout=12)
    return r.status_code == 200


# =========================
# Branding por empresa (portal de clientes) — nivel 1
# =========================
# Personalización visual del portal por tenant: logo + color de acento.
# - Se configura desde /admin/branding (requiere sesión de admin).
# - Se aplica a TODAS las páginas del portal sin tocarlas una por una:
#   un hook after_request inyecta un <style> con el override de --accent
#   y reemplaza el logo/texto de la barra superior por los de la empresa.
# - El logo se guarda en Postgres (BYTEA, máx. 400 KB) y se sirve desde
#   /branding/logo con cache de 1 día; la URL lleva ?v=updated_at para
#   que el cambio se vea al instante cuando se actualiza.

BRANDING_MAX_LOGO_BYTES = 400 * 1024
BRANDING_ALLOWED_MIMES = {"image/png", "image/jpeg", "image/webp"}
BRANDING_PRESETS = [
    ("dorado",  "Dorado (SIA)", "#F4C430"),
    ("azul",    "Azul",         "#5aa7ff"),
    ("verde",   "Verde",        "#34d399"),
    ("violeta", "Violeta",      "#a78bfa"),
    ("coral",   "Coral",        "#fb7185"),
    ("celeste", "Celeste",      "#22d3ee"),
]

_BRANDING_CACHE: Dict[str, Dict] = {}  # tenant -> {"ts": float, "data": dict|None}

def _branding_hex_ok(c: str) -> bool:
    return bool(re.fullmatch(r"#[0-9a-fA-F]{6}", (c or "").strip()))

def _branding_hex_rgb(c: str):
    c = (c or "").lstrip("#")
    return int(c[0:2], 16), int(c[2:4], 16), int(c[4:6], 16)

def get_tenant_branding(tenant: str, force: bool = False) -> Optional[Dict]:
    """Config de branding del tenant (sin el binario del logo) o None si no hay."""
    t = (tenant or "").strip().lower()
    if not t:
        return None
    now = time.time()
    hit = _BRANDING_CACHE.get(t)
    if hit and not force and (now - hit["ts"] < CACHE_TTL):
        return hit["data"]
    data = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(
            "SELECT accent, logo_mime, (logo IS NOT NULL) AS has_logo, updated_at "
            "FROM tenant_branding WHERE tenant = %s",
            (t,),
        )
        row = cur.fetchone()
        conn.close()
        if row:
            data = {
                "accent": (row.get("accent") or "").strip(),
                "logo_mime": (row.get("logo_mime") or "").strip(),
                "has_logo": bool(row.get("has_logo")),
                "updated_at": int(row.get("updated_at") or 0),
            }
    except Exception as e:
        log.exception("branding: error leyendo config de %s: %s", t, e)
    _BRANDING_CACHE[t] = {"ts": now, "data": data}
    return data

def save_tenant_branding(tenant: str, accent: str, logo: bytes | None = None,
                         logo_mime: str = "", clear_logo: bool = False) -> None:
    """Upsert de branding. logo=None deja el logo actual; clear_logo=True lo borra."""
    t = (tenant or "").strip().lower()
    now = int(time.time())
    conn = get_db_connection()
    cur = conn.cursor()
    if clear_logo:
        cur.execute(
            "INSERT INTO tenant_branding (tenant, accent, logo, logo_mime, updated_at) "
            "VALUES (%s, %s, NULL, '', %s) "
            "ON CONFLICT (tenant) DO UPDATE SET accent = EXCLUDED.accent, "
            "logo = NULL, logo_mime = '', updated_at = EXCLUDED.updated_at",
            (t, accent, now),
        )
    elif logo is not None:
        cur.execute(
            "INSERT INTO tenant_branding (tenant, accent, logo, logo_mime, updated_at) "
            "VALUES (%s, %s, %s, %s, %s) "
            "ON CONFLICT (tenant) DO UPDATE SET accent = EXCLUDED.accent, "
            "logo = EXCLUDED.logo, logo_mime = EXCLUDED.logo_mime, "
            "updated_at = EXCLUDED.updated_at",
            (t, accent, psycopg2.Binary(logo), logo_mime, now),
        )
    else:
        cur.execute(
            "INSERT INTO tenant_branding (tenant, accent, updated_at) "
            "VALUES (%s, %s, %s) "
            "ON CONFLICT (tenant) DO UPDATE SET accent = EXCLUDED.accent, "
            "updated_at = EXCLUDED.updated_at",
            (t, accent, now),
        )
    conn.commit()
    conn.close()
    _BRANDING_CACHE.pop(t, None)

@app.get("/branding/logo")
def branding_logo():
    """Sirve el logo del tenant (público: un logo no es información sensible)."""
    t = (request.args.get("t") or "").strip().lower()
    if not t:
        return Response("Not found", status=404)
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("SELECT logo, logo_mime FROM tenant_branding WHERE tenant = %s", (t,))
        row = cur.fetchone()
        conn.close()
    except Exception as e:
        log.exception("branding: error sirviendo logo de %s: %s", t, e)
        return Response("Error", status=500)
    if not row or row.get("logo") is None:
        return Response("Not found", status=404)
    data = bytes(row["logo"])
    resp = Response(data, mimetype=(row.get("logo_mime") or "image/png"))
    resp.headers["Content-Length"] = str(len(data))
    # Cache fuerte: la URL cambia con ?v= cuando se actualiza el logo.
    resp.headers["Cache-Control"] = "public, max-age=86400"
    return resp

@app.after_request
def inject_portal_branding(resp):
    """Aplica el branding del tenant a todas las páginas HTML del portal."""
    try:
        if not request.path.startswith("/portal"):
            return resp
        if resp.status_code != 200 or resp.direct_passthrough:
            return resp
        if "text/html" not in (resp.headers.get("Content-Type") or ""):
            return resp
        tenant = (session.get("portal_tenant") or "").strip().lower()
        if not tenant:
            return resp
        b = get_tenant_branding(tenant)
        if not b:
            return resp

        html = resp.get_data(as_text=True)
        changed = False

        # 1) Color de acento: pisa la variable del theme + el glow dorado hardcodeado.
        accent = b.get("accent") or ""
        if _branding_hex_ok(accent) and "</head>" in html:
            r_, g_, b_ = _branding_hex_rgb(accent)
            style = (
                '<style id="tenant-branding">'
                ':root{--accent:%s;}'
                '.action-card:hover{box-shadow:0 8px 24px rgba(%d,%d,%d,.18);}'
                '</style>' % (accent, r_, g_, b_)
            )
            html = html.replace("</head>", style + "</head>", 1)
            changed = True

        # 2) Logo: reemplaza el bloque .top-logo (idéntico en todas las páginas).
        if b.get("has_logo"):
            old_img = '<img src="/static/icon-192.png" alt="SIA Sueldos">'
            if old_img in html:
                v = int(b.get("updated_at") or 0)
                new_img = ('<img src="/branding/logo?t=%s&v=%d" alt="Logo" '
                           'style="object-fit:contain">' % (quote(tenant), v))
                html = html.replace(old_img, new_img)
                t_info = get_tenant(tenant) or {}
                disp = esc(t_info.get("display_name") or "")
                if disp:
                    html = html.replace(
                        '<span class="top-logo-text">SIA</span>',
                        '<span class="top-logo-text">%s</span>' % disp,
                    )
                changed = True

        if changed:
            resp.set_data(html)
    except Exception as e:
        log.exception("branding: error inyectando branding en %s: %s", request.path, e)
    return resp

@app.route("/admin/branding", methods=["GET", "POST"])
def admin_branding():
    """Pantalla de personalización del portal por empresa (logo + color)."""
    auth = require_admin()
    if auth:
        return auth

    tenant = (request.values.get("tenant") or "").strip().lower()

    if request.method == "POST":
        t_info = get_tenant(tenant)
        if not t_info:
            return redirect("/admin/branding?msg=err_tenant")

        # Color: preset o personalizado
        preset = (request.form.get("preset") or "").strip()
        accent = ""
        if preset == "custom":
            accent = (request.form.get("accent_custom") or "").strip()
            if not _branding_hex_ok(accent):
                return redirect(f"/admin/branding?tenant={quote(tenant)}&msg=err_color")
        else:
            for key, _label, hexv in BRANDING_PRESETS:
                if key == preset:
                    accent = hexv
                    break
            if not accent:
                return redirect(f"/admin/branding?tenant={quote(tenant)}&msg=err_color")

        clear_logo = request.form.get("clear_logo") == "1"
        logo_bytes = None
        logo_mime = ""
        f = request.files.get("logo")
        if f and f.filename and not clear_logo:
            mime = (f.mimetype or "").lower()
            if mime not in BRANDING_ALLOWED_MIMES:
                return redirect(f"/admin/branding?tenant={quote(tenant)}&msg=err_tipo")
            data = f.read()
            if not data:
                return redirect(f"/admin/branding?tenant={quote(tenant)}&msg=err_tipo")
            if len(data) > BRANDING_MAX_LOGO_BYTES:
                return redirect(f"/admin/branding?tenant={quote(tenant)}&msg=err_peso")
            logo_bytes, logo_mime = data, mime

        try:
            save_tenant_branding(tenant, accent, logo=logo_bytes,
                                 logo_mime=logo_mime, clear_logo=clear_logo)
        except Exception as e:
            log.exception("branding: error guardando %s: %s", tenant, e)
            return redirect(f"/admin/branding?tenant={quote(tenant)}&msg=err_db")
        return redirect(f"/admin/branding?tenant={quote(tenant)}&msg=ok")

    # ---- GET ----
    msg = (request.args.get("msg") or "").strip()
    tenants = get_all_tenants()
    b = get_tenant_branding(tenant, force=True) if tenant else None
    cur_accent = (b or {}).get("accent") or ""

    preset_checked = ""
    if cur_accent:
        for key, _label, hexv in BRANDING_PRESETS:
            if hexv.lower() == cur_accent.lower():
                preset_checked = key
                break
        if not preset_checked:
            preset_checked = "custom"
    else:
        preset_checked = BRANDING_PRESETS[0][0]
    custom_value = cur_accent if (preset_checked == "custom" and _branding_hex_ok(cur_accent)) else "#F4C430"

    MSGS = {
        "ok": ("success", "✅ Guardado. Los usuarios del portal ya lo ven (con varios workers puede demorar hasta 2 minutos por cache)."),
        "err_tenant": ("error", "❌ Empresa inválida."),
        "err_color": ("error", "❌ Color inválido: elegí un preset o un hex tipo #AABBCC."),
        "err_tipo": ("error", "❌ Formato de logo no soportado: subí PNG, JPG o WebP."),
        "err_peso": ("error", "❌ El logo es muy pesado: máximo 400 KB."),
        "err_db": ("error", "❌ Error guardando en la base. Mirá los logs."),
    }

    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Personalización del portal</title>
  <style>
    :root{
      --bg:#0b1220; --card:#0f1b33; --muted:#9fb2d0; --text:#eaf0ff;
      --line:rgba(255,255,255,.08); --radius:14px;
    }
    *{box-sizing:border-box}
    body{
      margin:0; font-family:system-ui; color:var(--text); padding:20px;
      background: radial-gradient(1200px 700px at 20% -20%, rgba(90,167,255,.25), transparent 60%), var(--bg);
    }
    .wrap{max-width:760px;margin:0 auto}
    .card{
      border:1px solid var(--line); background:rgba(255,255,255,.03);
      border-radius:var(--radius); padding:20px; margin:16px 0;
    }
    h2{margin:0 0 6px 0} h3{margin:0 0 10px 0}
    .muted{color:var(--muted);font-size:13px}
    label{font-size:13px;color:var(--muted);font-weight:600}
    input, select{
      background:rgba(0,0,0,.25); border:1px solid var(--line);
      color:var(--text); padding:10px; border-radius:8px; margin:6px 0;
      color-scheme:dark;
    }
    select option{ background-color:#0f1b33; color:var(--text); }
    input[type=file]{padding:8px}
    input[type=color]{ padding:0; width:42px; height:28px; border:1px solid var(--line); border-radius:6px; cursor:pointer; background:none; }
    .btn{
      display:inline-block; padding:10px 16px; border-radius:8px; border:1px solid var(--line);
      background:rgba(255,255,255,.06); cursor:pointer; font-weight:600;
      text-decoration:none; color:var(--text);
    }
    .btn:hover{background:rgba(255,255,255,.1)}
    .presets{display:flex; flex-wrap:wrap; gap:10px; margin:10px 0}
    .preset{
      display:inline-flex; align-items:center; gap:8px; border:1px solid var(--line);
      border-radius:999px; padding:8px 14px; cursor:pointer; font-size:13px; color:var(--text);
    }
    .preset:hover{background:rgba(255,255,255,.05)}
    .sw{width:16px;height:16px;border-radius:50%;display:inline-block;border:1px solid rgba(255,255,255,.25)}
    .sep{height:1px;background:var(--line);margin:14px 0}
    .hint{font-size:12px;color:var(--muted);margin-top:6px}
    .logo-prev{
      height:48px; max-width:220px; object-fit:contain; vertical-align:middle;
      background:rgba(255,255,255,.06); border:1px solid var(--line); border-radius:8px; padding:6px;
    }
    .success{background:rgba(52,211,153,.1); border:1px solid rgba(52,211,153,.3); padding:12px; border-radius:8px; margin:10px 0}
    .error{background:rgba(251,113,133,.1); border:1px solid rgba(251,113,133,.3); padding:12px; border-radius:8px; margin:10px 0}
  </style>
</head>
<body>
<div class="wrap">
""")

    if tenant:
        html.append(f"<a href='/admin/panel?tenant={quote(tenant)}' class='btn'>← Volver al panel</a>")
    else:
        html.append("<a href='/admin' class='btn'>← Volver al admin</a>")

    html.append("<div class='card'>")
    html.append("<h2>🎨 Personalización del portal</h2>")
    html.append("<div class='muted'>Logo y color de acento que ven los usuarios del portal de cada empresa.</div>")
    html.append("<form method='get' style='margin-top:12px'>")
    html.append("<label>Empresa</label><br>")
    html.append("<select name='tenant' onchange='this.form.submit()' style='min-width:280px'>")
    html.append("<option value=''>-- Seleccionar empresa --</option>")
    for t_ in tenants:
        s = " selected" if (t_.get("slug") == tenant) else ""
        html.append(f"<option value='{esc(t_.get('slug',''))}'{s}>{esc(t_.get('display_name',''))}</option>")
    html.append("</select>")
    html.append("</form>")
    html.append("</div>")

    if tenant:
        t_info = get_tenant(tenant)
        if not t_info:
            html.append("<div class='error'>❌ Empresa inválida.</div>")
        else:
            if msg in MSGS:
                css, texto = MSGS[msg]
                html.append(f"<div class='{css}'>{texto}</div>")

            html.append("<form method='post' enctype='multipart/form-data' class='card'>")
            html.append(f"<input type='hidden' name='tenant' value='{esc(tenant)}'>")
            html.append(f"<h3>{esc(t_info.get('display_name',''))}</h3>")

            html.append("<label>Color de acento</label>")
            html.append("<div class='presets'>")
            for key, label, hexv in BRANDING_PRESETS:
                ck = " checked" if preset_checked == key else ""
                html.append(
                    f"<label class='preset'><input type='radio' name='preset' value='{key}'{ck}>"
                    f"<span class='sw' style='background:{hexv}'></span>{esc(label)}</label>"
                )
            ck = " checked" if preset_checked == "custom" else ""
            html.append(
                f"""<label class='preset'><input type='radio' name='preset' value='custom'{ck}>"""
                f"""<input type='color' name='accent_custom' value='{esc(custom_value)}' """
                f"""onclick="this.parentNode.querySelector('input[type=radio]').checked=true">"""
                """Personalizado</label>"""
            )
            html.append("</div>")
            html.append("<div class='hint'>El acento se usa en resaltados y hovers del portal. El fondo es oscuro: evitá colores muy oscuros o quedan invisibles.</div>")

            html.append("<div class='sep'></div>")
            html.append("<label>Logo de la empresa</label><br>")
            if (b or {}).get("has_logo"):
                v = int((b or {}).get("updated_at") or 0)
                html.append(f"<img class='logo-prev' src='/branding/logo?t={quote(tenant)}&v={v}' alt='Logo actual'>")
                html.append("<label style='display:inline-flex;align-items:center;gap:6px;margin-left:12px;cursor:pointer'>"
                            "<input type='checkbox' name='clear_logo' value='1'> Quitar logo</label><br>")
            html.append("<input type='file' name='logo' accept='.png,.jpg,.jpeg,.webp'>")
            html.append("<div class='hint'>PNG, JPG o WebP · máximo 400 KB · ideal apaisado y con fondo transparente.</div>")

            html.append("<button class='btn' type='submit' style='margin-top:14px'>💾 Guardar</button>")
            html.append("</form>")

    html.append("</div></body></html>")
    return Response("".join(html), mimetype="text/html")


# =========================
# Routes
# =========================
@app.get("/")
def root():
    tok = request.args.get("token", "")
    if tok:
        return redirect("/admin")
    # El dominio publico (portalsia.com.ar) manda a la home del portal.
    # Cualquier otro host (URL interna de Render, localhost, etc.) sigue yendo a /admin.
    host = (request.host or "").split(":")[0].lower()
    if host == "portalsia.com.ar" or host.endswith(".portalsia.com.ar"):
        return redirect("/portal")
    return redirect("/admin")


# --- iconos pedidos por el browser en la raíz ---
from flask import send_from_directory

@app.get("/favicon.ico")
def favicon_root():
    return send_from_directory(
        app.static_folder, "favicon.ico", mimetype="image/vnd.microsoft.icon"
    )

@app.get("/apple-touch-icon.png")
@app.get("/apple-touch-icon-precomposed.png")
def apple_touch_icon_root():
    return send_from_directory(
        app.static_folder, "apple-touch-icon.png", mimetype="image/png"
    )

def save_template_sid(tenant: str, cuil: str, period: str, to_whatsapp: str, sid: str, nombre: str = ""):
    now = int(time.time())
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO message_status
        (message_sid, to_whatsapp, tenant, cuil, period, nombre, kind, created_at, last_status, last_status_at)
        VALUES (%s, %s, %s, %s, %s, %s, 'template', %s, 'sent', %s)
        ON CONFLICT (message_sid) DO NOTHING
    """, (sid, to_whatsapp, tenant, cuil, period, nombre, now, now))
    conn.commit()
    conn.close()

def already_sent_template(tenant: str, cuil: str, period: str, to_whatsapp: str) -> bool:
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT 1
            FROM message_status
            WHERE tenant=%s AND cuil=%s AND period=%s AND to_whatsapp=%s AND kind='template'
            LIMIT 1
        """, (tenant, cuil, period, to_whatsapp))

        return cur.fetchone() is not None
    finally:
        conn.close()



@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    # Si ya hay sesión activa, vamos directo al destino.
    nxt = request.args.get("next") or request.form.get("next") or "/admin"
    # Anti open-redirect: solo permitimos rutas internas.
    if not nxt.startswith("/") or nxt.startswith("//"):
        nxt = "/admin"

    if admin_session_active():
        return redirect(nxt)

    error = ""
    if request.method == "POST":
        pwd = (request.form.get("password") or "").strip()
        if not ADMIN_TOKEN:
            error = "El admin no está configurado (falta ADMIN_TOKEN en el servidor)."
        elif pwd and hmac.compare_digest(pwd, ADMIN_TOKEN):
            session["admin_authed"] = True
            session["admin_authed_at"] = int(time.time())
            session.permanent = False
            log.info("ADMIN LOGIN OK desde %s", request.remote_addr)
            return redirect(nxt)
        else:
            time.sleep(0.8)  # frena fuerza bruta contra el formulario
            error = "Clave incorrecta."
            log.warning("ADMIN LOGIN fallido desde %s", request.remote_addr)

    error_html = f"<div class='err'>{esc(error)}</div>" if error else ""
    page = f"""<!doctype html>
<html lang="es"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>Admin · Ingresar</title>
<style>
  :root {{ --bg:#0b1220; --text:#e8eefc; --line:rgba(255,255,255,.12); --accent:#5aa7ff; }}
  *{{box-sizing:border-box}}
  body{{margin:0;min-height:100vh;display:flex;align-items:center;justify-content:center;
    font-family:system-ui,-apple-system,Segoe UI,Roboto,sans-serif;
    background:radial-gradient(1000px 600px at 30% -10%, rgba(90,167,255,.20), transparent 60%), var(--bg);
    color:var(--text);padding:20px;}}
  .card{{width:100%;max-width:380px;border:1px solid var(--line);border-radius:16px;
    background:rgba(255,255,255,.04);padding:28px;box-shadow:0 20px 60px rgba(0,0,0,.4);}}
  h1{{font-size:20px;margin:0 0 6px}}
  p.sub{{margin:0 0 20px;color:#9fb2d6;font-size:14px}}
  label{{display:block;font-size:13px;margin-bottom:6px;color:#cdd9f0}}
  input[type=password]{{width:100%;padding:12px 14px;border-radius:10px;border:1px solid var(--line);
    background:rgba(0,0,0,.25);color:var(--text);font-size:15px;}}
  button{{margin-top:16px;width:100%;padding:12px;border:0;border-radius:10px;cursor:pointer;
    background:var(--accent);color:#04122b;font-weight:600;font-size:15px;}}
  .err{{margin-bottom:14px;padding:10px 12px;border-radius:10px;background:rgba(248,113,113,.15);
    border:1px solid rgba(248,113,113,.4);color:#fecaca;font-size:14px;}}
</style></head>
<body>
  <form class="card" method="post" action="/admin/login">
    <h1>🔐 Panel de administración</h1>
    <p class="sub">Ingresá la clave de acceso para continuar.</p>
    {error_html}
    <input type="hidden" name="next" value="{esc(nxt)}">
    <label for="password">Clave</label>
    <input id="password" name="password" type="password" autofocus autocomplete="current-password">
    <button type="submit">Ingresar</button>
  </form>
</body></html>"""
    return Response(page, mimetype="text/html")


@app.get("/admin/logout")
def admin_logout():
    session.pop("admin_authed", None)
    session.pop("admin_authed_at", None)
    return redirect("/admin/login")


@app.get("/admin/sentry_test")
def admin_sentry_test():
    # Dispara un error a propósito para verificar el circuito de alertas.
    # Solo accesible con sesión de admin (o header X-Admin-Token).
    auth = require_admin()
    if auth:
        return auth
    raise RuntimeError("Prueba de Sentry: si ves este error en sentry.io, las alertas funcionan")


@app.get("/admin")
def admin_home():
    auth = require_admin()
    if auth:
        return auth

    token = (request.args.get("token") or "").strip()
    tenants = load_tenants(force=True) or []

    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Admin</title>
  <style>
    :root{
      --bg:#0b1220;
      --muted:#9fb2d0;
      --text:#eaf0ff;
      --line:rgba(255,255,255,.08);
      --shadow: 0 10px 25px rgba(0,0,0,.25);
      --radius:14px;
      --mono: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", monospace;
      --sans: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial;
    }
    *{box-sizing:border-box}
    body{
      margin:0;
      font-family:var(--sans);
      background: radial-gradient(1200px 700px at 20% -20%, rgba(90,167,255,.25), transparent 60%),
                  radial-gradient(1200px 700px at 90% 0%, rgba(52,211,153,.18), transparent 55%),
                  var(--bg);
      color:var(--text);
    }
    .wrap{max-width:1100px;margin:0 auto;padding:22px}
    .topbar{
      display:flex;gap:14px;align-items:center;justify-content:space-between;
      padding:16px 18px;border:1px solid var(--line);border-radius:var(--radius);
      background:rgba(255,255,255,.03);box-shadow:var(--shadow);
      position:sticky;top:12px;backdrop-filter: blur(8px); z-index:10;
    }
    h2{margin:0;font-size:18px;letter-spacing:.2px}
    .muted{color:var(--muted);font-size:13px}
    code{font-family:var(--mono);font-size:12px;background:rgba(255,255,255,.06);padding:2px 6px;border-radius:8px;border:1px solid var(--line)}
    .card{
      margin-top:14px;
      border:1px solid var(--line);
      background:rgba(255,255,255,.03);
      border-radius:var(--radius);
      padding:16px;
      box-shadow:var(--shadow);
    }
    .row{display:flex;gap:10px;flex-wrap:wrap;align-items:center;justify-content:space-between}
    input[type="text"]{
      background:rgba(0,0,0,.25);
      border:1px solid var(--line);
      color:var(--text);
      padding:10px 10px;
      border-radius:12px;
      outline:none;
      min-width: 240px;
    }
    .grid{
      margin-top:12px;
      display:grid;
      grid-template-columns: repeat(3, 1fr);
      gap:12px;
    }
    @media(max-width:960px){ .grid{grid-template-columns:1fr} .topbar{flex-direction:column;align-items:flex-start}}
    .tile{
      border:1px solid var(--line);
      background:rgba(255,255,255,.02);
      border-radius:14px;
      padding:14px;
      transition:transform .05s ease, background .2s ease;
    }
    .tile:hover{transform:translateY(-1px);background:rgba(255,255,255,.04)}
    .tile h3{margin:0 0 6px 0;font-size:15px}
    .btn{
      display:inline-flex;align-items:center;justify-content:center;
      gap:8px;
      padding:10px 12px;
      border-radius:12px;
      border:1px solid var(--line);
      background:linear-gradient(180deg, rgba(255,255,255,.06), rgba(255,255,255,.02));
      cursor:pointer;
      font-weight:600;
      font-size:13px;
      text-decoration:none;
      color:var(--text);
    }
    .btn.secondary{background:rgba(255,255,255,.02)}
    .btn.small{padding:7px 10px;font-size:12px;border-radius:10px}
    .actions{display:flex;gap:10px;flex-wrap:wrap;margin-top:10px}
    .warn{
      border:1px solid rgba(251,191,36,.35);
      background: rgba(251,191,36,.10);
      padding:10px 12px;border-radius:12px;
      color: var(--text);
    }

    /* --- Blur empresas (selector /admin) --- */
    .secret-blur{
      filter: blur(6px);
      transition: .2s ease;
      user-select: none;
    }
    .secret-show{
      filter: blur(0);
    }
    .tile-head{
      display:flex;
      align-items:flex-start;
      justify-content:space-between;
      gap:10px;
    }
    .eye-btn{
      width: 36px;
      height: 36px;
      border-radius: 10px;
      border: 1px solid var(--line);
      background: rgba(255,255,255,.04);
      color: var(--text);
      cursor: pointer;
      display:inline-flex;
      align-items:center;
      justify-content:center;
      flex: 0 0 auto;
    }
    .eye-btn:hover{ background: rgba(255,255,255,.08); }
  </style>
</head>
<body>
<div class="wrap">
""")

    # Topbar
    html.append("<div class='topbar'>")
    html.append("<div>")
    html.append("<h2>Panel Admin</h2>")
    html.append("<div class='muted'>Elegí la empresa para abrir el panel.</div>")
    html.append("</div>")

    # Ojo general + token
    html.append("<div class='row' style='gap:10px;justify-content:flex-end'>")
    html.append("<button class='eye-btn' type='button' id='eyeAll' title='Mostrar/ocultar todas'>👁️</button>")
    html.append(f"<div class='muted'>Token: <code>{esc(token)}</code></div>")
    html.append("<a href='/admin/portal_users' class='btn'>👥 Usuarios del Portal</a>")
    html.append("<a href='/admin/portal_activity' class='btn'>🕑 Actividad del Portal</a>")
    html.append("<a href='/admin/logout' class='btn secondary' style='float:right'>Salir</a>")
    html.append("</div>")

    html.append("</div>")  # topbar

    # Warning si falta master
    if not EMPRESAS_FILE_ID:
        html.append("<div class='card'><div class='warn'>⚠️ Falta <code>EMPRESAS_FILE_ID</code> en ENV.</div></div>")

    if not tenants:
        html.append("<div class='card'>")
        html.append("<h3 style='margin:0 0 6px 0'>No hay empresas detectadas en el Excel maestro</h3>")
        html.append("<div class='muted'>Encabezados esperados: <code>Empresa</code> | <code>Envios_File_ID</code> | <code>Drive_Root_ID</code></div>")
        html.append("</div>")
    else:
        html.append("<div class='card'>")
        html.append("<div class='row'>")
        html.append("<div>")
        html.append("<h3 style='margin:0 0 6px 0'>Empresas</h3>")
        html.append(f"<div class='muted'>Detectadas: <b>{len(tenants)}</b></div>")
        html.append("</div>")
        html.append("<div>")
        html.append("<label class='muted' style='display:block;margin-bottom:6px'>Buscar</label>")
        html.append("<input id='q' type='text' placeholder='Escribí para filtrar...'>")
        html.append("</div>")
        html.append("</div>")

        html.append("<div class='grid' id='grid'>")
        for t in tenants:
            slug = t["slug"]
            name = t.get("display_name") or slug
            panel_url = f"/admin/panel?tenant={esc(slug)}"
            test_url = f"/admin/send_test?tenant={esc(slug)}"

            html.append(f"""
              <div class="tile" data-name="{esc(name).lower()} {esc(slug).lower()}" data-tenant="{esc(slug)}">
                <div class="tile-head">
                  <div>
                    <h3 class="company-name secret-blur" data-tenant="{esc(slug)}">{esc(name)}</h3>
                    <div class="muted company-tenant secret-blur" data-tenant="{esc(slug)}">tenant: <code>{esc(slug)}</code></div>
                  </div>
                  <button class="eye-btn eye-one" type="button" data-tenant="{esc(slug)}" title="Mostrar solo esta">👁️</button>
                </div>

                <div class="actions">
                  <a class="btn" href="{panel_url}">Abrir panel →</a>
                </div>
              </div>
            """)

        html.append("</div>")  # grid

        # JS filtro + blur/ojos
        html.append("""
        <script>
          (function(){
            const q = document.getElementById('q');
            const tiles = Array.from(document.querySelectorAll('#grid .tile'));

            const eyeAll = document.getElementById('eyeAll');
            const names  = Array.from(document.querySelectorAll('.company-name'));
            const tenants= Array.from(document.querySelectorAll('.company-tenant'));
            const eyeOne = Array.from(document.querySelectorAll('.eye-one'));

            function blurAll(){
              [...names, ...tenants].forEach(el => {
                el.classList.add('secret-blur');
                el.classList.remove('secret-show');
              });
              if(eyeAll) eyeAll.textContent = "👁️";
            }

            function showAll(){
              [...names, ...tenants].forEach(el => {
                el.classList.remove('secret-blur');
                el.classList.add('secret-show');
              });
              if(eyeAll) eyeAll.textContent = "🙈";
            }

            function showOnly(tenant){
              blurAll();
              [...names, ...tenants].forEach(el => {
                if(el.getAttribute('data-tenant') === tenant){
                  el.classList.remove('secret-blur');
                  el.classList.add('secret-show');
                }
              });
            }

            // Default: todo blureado
            blurAll();

            // Ojo general
            if(eyeAll){
              eyeAll.addEventListener('click', () => {
                const anyHidden = [...names, ...tenants].some(el => !el.classList.contains('secret-show'));
                if(anyHidden) showAll();
                else blurAll();
              });
            }

            // Ojo por empresa: muestra SOLO esa
            eyeOne.forEach(btn => {
              btn.addEventListener('click', (e) => {
                e.preventDefault();
                e.stopPropagation();
                const tenant = btn.getAttribute('data-tenant') || '';
                showOnly(tenant);
              });
            });

            // Filtro (tu lógica original)
            function apply(){
              const s = (q.value || '').trim().toLowerCase();
              for(const t of tiles){
                const hay = (t.getAttribute('data-name') || '');
                t.style.display = (!s || hay.includes(s)) ? '' : 'none';
              }
            }
            q.addEventListener('input', apply);

            // Opcional: al cambiar de pestaña, volver a ocultar
            document.addEventListener("visibilitychange", () => {
              if(document.hidden) blurAll();
            });
          })();
        </script>
        """)

        html.append("</div>")  # card

    html.append("</div></body></html>")
    return Response("".join(html), mimetype="text/html")

from flask import send_file
import pandas as pd
import io

from datetime import datetime, timezone


from flask import send_file
import io

@app.get("/admin/verifications_template.xlsx")
@admin_required
def admin_verifications_template_xlsx():
    output = io.BytesIO()
    import pandas as pd

    df = pd.DataFrame(columns=["cuil", "whatsapp", "dni", "nombre"])
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="verificaciones")

    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="verifications_template.xlsx",
    )

import time



def normalize_period_for_drive(period: str) -> str:
    """
    Convierte 'MM/AAAA' -> 'MM-AAAA'
    Acepta también 'MM-AAAA' y lo deja igual.
    """
    if not period:
        return ""
    p = period.strip()
    if "/" in p:
        return p.replace("/", "-")
    return p


def normalize_whatsapp(raw: str) -> str | None:
    """
    Recibe:
      - 'whatsapp:+54911...'
      - '+54911...'
      - '11 3622-2572'
      - '1136222572'
      - '2323 555360'

    Devuelve:
      - 'whatsapp:+549XXXXXXXXXX'
    """

    s = (raw or "").strip()
    if not s:
        return None

    if s.startswith("whatsapp:"):
        s = s.replace("whatsapp:", "").strip()

    d = _digits(s)
    if not d:
        return None

    # ya viene correcto
    if d.startswith("549"):
        return f"whatsapp:+{d}"

    # viene con 54 pero sin 9
    if d.startswith("54"):
        return f"whatsapp:+549{d[2:]}"

    # número local
    return f"whatsapp:+549{d}"


def ts_str(ts: int | None) -> str:
    if not ts:
        return ""
    try:
        return time.strftime("%d/%m/%Y %H:%M", time.gmtime(int(ts)))
    except Exception:
        return str(ts)


@app.post("/admin/verifications_import")
@admin_required
def admin_verifications_import():
    token = _get_admin_token_from_request()
    tenant = (request.form.get("tenant") or "").strip().lower()

    f = request.files.get("file")
    if not f:
        return Response("Falta archivo", status=400)

    import pandas as pd
    df = pd.read_excel(f)
    if df is None or df.empty:
        return redirect(f"/admin/panel?tenant={tenant}&msg=verif_import_empty")

    df.columns = [str(c).strip().lower() for c in df.columns]

    def pick(*names):
        for n in names:
            if n in df.columns:
                return n
        return None

    c_cuil = pick("cuil", "archivo", "archivo_norm")
    c_wpp  = pick("whatsapp", "telefono", "tel", "celular", "numero")
    c_nombre = pick("nombre", "name", "empleado", "persona")

    if not c_cuil or not c_wpp:
        return Response("El excel debe tener columnas cuil y whatsapp/telefono", status=400)

    rows = df.to_dict(orient="records")
    now = int(time.time())

    conn = get_db_connection()
    cur = conn.cursor()

    ok = 0
    skipped = 0

    for r in rows:
        cuil_raw = str(r.get(c_cuil, "")).strip()
        wpp_raw  = str(r.get(c_wpp, "")).strip()
        raw_nombre = r.get(c_nombre) if c_nombre else None
        nombre = ""
        if raw_nombre is not None and str(raw_nombre).strip().lower() != "nan":
            nombre = str(raw_nombre).strip()

        if not cuil_raw or not wpp_raw:
            skipped += 1
            continue

        cuil = strip_pdf(cuil_raw).strip()
        to_whatsapp = normalize_whatsapp(wpp_raw)
        if not to_whatsapp:
            skipped += 1
            continue

        # Import: marca como verificado "sin DNI" (solo vínculo número<->cuil)
        cur.execute("""
            INSERT INTO verifications (tenant, cuil, to_whatsapp, nombre, dni_hash, dni_last4, verified_at, updated_at)
            VALUES (%s, %s, %s, %s, NULL, NULL, %s, %s)
            ON CONFLICT(tenant, cuil, to_whatsapp)
            DO UPDATE SET
                nombre=excluded.nombre,
                updated_at=excluded.updated_at
        """, (tenant, cuil, to_whatsapp, nombre, now, now))

        ok += 1

    conn.commit()
    conn.close()

    return redirect(f"/admin/panel?tenant={tenant}&msg=verif_import_ok&n={ok}&skipped={skipped}")


@app.post("/admin/verifications_update")
@admin_required
def admin_verifications_update():
    token = _get_admin_token_from_request()
    tenant = (request.form.get("tenant") or "").strip().lower()
    return redirect(f"/admin/panel?tenant={tenant}&msg=verif_update_disabled")


@app.post("/admin/verifications_delete")
@admin_required
def admin_verifications_delete():
    token = _get_admin_token_from_request()
    tenant = (request.form.get("tenant") or "").strip().lower()
    cuil = (request.form.get("cuil") or "").strip()
    to_whatsapp = (request.form.get("to_whatsapp") or "").strip()

    if not (tenant and cuil and to_whatsapp):
        return Response("Faltan parámetros", status=400)

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
    DELETE FROM verifications
    WHERE tenant=%s AND cuil=%s AND to_whatsapp=%s
    """, (tenant, cuil, to_whatsapp))

    conn.commit()
    conn.close()

    return redirect(f"/admin/panel?tenant={tenant}&msg=verif_deleted")

from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from datetime import datetime, timezone
import datetime as dt

def ts_to_str(ts) -> str:
    if ts is None:
        return ""
    try:
        s = str(ts).strip()
        if s == "" or s.lower() == "nan":
            return ""

        ts_f = float(s)

        # ms -> s
        if ts_f > 10_000_000_000:
            ts_f = ts_f / 1000.0

        return dt.datetime.fromtimestamp(int(ts_f)).strftime("%d/%m/%Y %H:%M:%S")
    except Exception as e:
        log.exception("%s %s %s", "ts_to_str ERROR:", ts, repr(e))
        return ""






def generate_excel_report_v2(tenant: str, period_filter: str = "") -> BytesIO:
    tenant = (tenant or "").strip().lower()
    period_filter = norm_period_label(period_filter)

    conn = get_db_connection()
    cur = conn.cursor()

    # message_status
    cur.execute("""
        SELECT
            to_whatsapp, tenant, cuil, period, nombre, kind,
            created_at, delivered_at, read_at, failed_at
        FROM message_status
        WHERE tenant = %s
    """, (tenant,))
    msg_rows = cur.fetchall()

    # pending_views (último period por whatsapp)
    cur.execute("""
        SELECT DISTINCT ON (to_whatsapp)
            to_whatsapp, tenant, cuil, period, created_at AS last_ts
        FROM pending_views
        WHERE tenant = %s
        ORDER BY to_whatsapp, created_at DESC
    """, (tenant,))
    pv_rows = cur.fetchall()

    last_period_by_whatsapp = {}
    for r in pv_rows:
        w = (r["to_whatsapp"] or "").strip()
        p = norm_period_label(r["period"] or "")
        if w and p:
            last_period_by_whatsapp[w] = p

    # recibo_estado
    cur.execute("""
        SELECT tenant, cuil, period, estado, updated_at
        FROM recibo_estado
        WHERE tenant = %s
    """, (tenant,))
    estado_rows = cur.fetchall()

    # ✅ NUEVO: requests por período
    cur.execute("""
        SELECT tenant,cuil,period,to_whatsapp,request_count,last_requested_at
        FROM receipt_requests
        WHERE tenant = %s
    """, (tenant,))
    rr_rows = cur.fetchall()

    conn.close()

    estado_map = {}
    for r in estado_rows:
        c = norm_cuil(r["cuil"])
        p = norm_period_label(r["period"] or "")
        if c and p:
            estado_map[(c, p)] = {"estado": (r["estado"] or "").strip(), "ts": r["updated_at"]}

    rr_map = {}
    for r in rr_rows:
        c = norm_cuil(r["cuil"])
        p = norm_period_label(r["period"] or "")
        w = (r["to_whatsapp"] or "").strip()
        if c and p and w:
            rr_map[(w, c, p)] = {
                "count": int(r["request_count"] or 0),
                "last": r["last_requested_at"]
            }

    def key(whatsapp: str, period_norm: str):
        return ((whatsapp or "").strip(), (period_norm or "").strip())

    agg = {}

    for row in msg_rows:
        wpp = (row["to_whatsapp"] or "").strip()
        if not wpp:
            continue

        cuil = norm_cuil(row["cuil"])
        nombre = (row["nombre"] or "").strip()

        period_raw = (row["period"] or "").strip()
        period_norm = norm_period_label(period_raw)

        if not period_norm and wpp:
            period_norm = last_period_by_whatsapp.get(wpp, "")

        if period_filter and period_norm != period_filter:
            continue
        if not period_norm:
            continue

        k = key(wpp, period_norm)
        rec = agg.get(k)
        if not rec:
            rec = {
                "periodo": period_norm,
                "nombre": nombre,
                "cuil": cuil,
                "whatsapp": wpp,
                "plantilla_sent_at": None,
                "plantilla_delivered_at": None,
                "plantilla_read_at": None,
                "plantilla_failed_at": None,
                "pdf_sent_at": None,
                "pdf_delivered_at": None,
                "pdf_read_at": None,
                "pdf_failed_at": None,
                "respuesta_usuario": "",
                "respuesta_timestamp": None,
                "pedidos_recibo": 0,
                "ultimo_pedido": None,
            }
            agg[k] = rec

        if nombre and not rec["nombre"]:
            rec["nombre"] = nombre
        if cuil and not rec["cuil"]:
            rec["cuil"] = cuil

        kind = (row["kind"] or "").strip().lower()
        created_at = row["created_at"]
        delivered_at = row["delivered_at"]
        read_at = row["read_at"]
        failed_at = row["failed_at"]

        if kind == "template":
            if created_at and (not rec["plantilla_sent_at"] or created_at < rec["plantilla_sent_at"]):
                rec["plantilla_sent_at"] = created_at
            if delivered_at:
                rec["plantilla_delivered_at"] = delivered_at
            if read_at:
                rec["plantilla_read_at"] = read_at
            if failed_at:
                rec["plantilla_failed_at"] = failed_at

        elif kind in ("pdf", "media"):
            if created_at and (not rec["pdf_sent_at"] or created_at < rec["pdf_sent_at"]):
                rec["pdf_sent_at"] = created_at
            if delivered_at:
                rec["pdf_delivered_at"] = delivered_at
            if read_at:
                rec["pdf_read_at"] = read_at
            if failed_at:
                rec["pdf_failed_at"] = failed_at

    # mezclar recibo_estado
    for (wpp, per), rec in list(agg.items()):
        cuil = rec.get("cuil", "")
        st = estado_map.get((cuil, per))
        if st:
            rec["respuesta_usuario"] = st.get("estado", "") or ""
            rec["respuesta_timestamp"] = st.get("ts")

        rr = rr_map.get((wpp, cuil, per))
        if rr:
            rec["pedidos_recibo"] = rr["count"]
            rec["ultimo_pedido"] = rr["last"]

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Recibos"

    headers = [
        "Periodo","Nombre","CUIL","WhatsApp",
        "Plantilla_enviada","Plantilla_entregada","Plantilla_leida","Plantilla_fallida",
        "PDF_enviado","PDF_entregado","PDF_leido","PDF_fallido",
        "Respuesta_usuario","Respuesta_timestamp",
        "Pedidos_recibo","Ultimo_pedido"
    ]
    ws.append(headers)

    items = list(agg.values())
    items.sort(key=lambda r: (r.get("periodo") or "", r.get("nombre") or "", r.get("whatsapp") or ""))

    for rec in items:
        ws.append([
            rec.get("periodo",""),
            rec.get("nombre",""),
            rec.get("cuil",""),
            rec.get("whatsapp",""),
            ts_to_str(rec.get("plantilla_sent_at")),
            ts_to_str(rec.get("plantilla_delivered_at")),
            ts_to_str(rec.get("plantilla_read_at")),
            ts_to_str(rec.get("plantilla_failed_at")),
            ts_to_str(rec.get("pdf_sent_at")),
            ts_to_str(rec.get("pdf_delivered_at")),
            ts_to_str(rec.get("pdf_read_at")),
            ts_to_str(rec.get("pdf_failed_at")),
            rec.get("respuesta_usuario",""),
            ts_to_str(rec.get("respuesta_timestamp")),
            int(rec.get("pedidos_recibo") or 0),
            ts_to_str(rec.get("ultimo_pedido")),
        ])

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(10, max_len + 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

from googleapiclient.errors import HttpError
import time

def _is_retryable_drive_error(e: HttpError) -> bool:
    status = getattr(e.resp, "status", None)
    return status in (429, 500, 502, 503, 504)

def find_pdf_with_retry(tenant, cuil, period, tries=4):
    last = None
    for i in range(tries):
        try:
            return find_pdf_file_id_for_cuil_period(tenant, cuil, period)
        except HttpError as e:
            last = e
            if _is_retryable_drive_error(e):
                time.sleep(0.6 * (2 ** i))
                continue
            raise
    # si agotó reintentos, devolvemos None (no tumbamos la cola)
    log.info("%s %s %s %s %s", "DRIVE RETRY EXHAUSTED:", tenant, cuil, period, last)
    return None

def get_queue_stats(tenant: str, period: str) -> dict:
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
      SELECT status, COUNT(*) as n
      FROM template_send_queue
      WHERE tenant=%s AND period=%s
      GROUP BY status
    """, (tenant, period))
    d = {r['status']: r['n'] for r in cur.fetchall()}
    conn.close()
    # asegurar claves
    for k in ("PENDING","SENT","FAILED","SKIPPED"):
        d.setdefault(k, 0)
    return d

@app.get("/admin/seguimiento")
def admin_seguimiento():
    """
    Panel de seguimiento de pendientes (+7 días sin acción).
    """
    auth = require_admin()
    if auth:
        return auth

    token = (request.args.get("token") or "").strip()
    tenant = (request.args.get("tenant") or "").strip().lower()

    if not tenant:
        return Response("Falta tenant", status=400)

    t = get_tenant(tenant)
    if not t:
        return Response("Tenant inválido", status=400)

    # Períodos disponibles
    period_folders = list_tenant_period_folders(tenant)
    period_labels = []
    for p in period_folders:
        lbl = period_folder_to_label(p)
        if lbl:
            period_labels.append(lbl)

    # Período seleccionado
    panel_period = (request.args.get("period") or "").strip()
    if not panel_period and period_labels:
        panel_period = period_labels[0]

    # Cargar pendientes
    pending_views_7d = []
    pending_sigs_7d = []
    if panel_period:
        pending_views_7d = get_pending_views_over_7days(tenant, panel_period)
        pending_sigs_7d = get_pending_signatures_over_7days(tenant, panel_period)

    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Seguimiento de pendientes</title>
  <style>
    :root{
      --bg:#0b1220; --card:#0f1b33; --muted:#9fb2d0; --text:#eaf0ff;
      --line:rgba(255,255,255,.08); --accent:#5aa7ff; --ok:#34d399;
      --warn:#fbbf24; --bad:#fb7185; --shadow: 0 10px 25px rgba(0,0,0,.25);
      --radius:14px;
      --mono: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, monospace;
      --sans: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial;
    }
    *{box-sizing:border-box}
    body{
      margin:0; font-family:var(--sans);
      background: radial-gradient(1200px 700px at 20% -20%, rgba(90,167,255,.25), transparent 60%),
                  radial-gradient(1200px 700px at 90% 0%, rgba(52,211,153,.18), transparent 55%),
                  var(--bg);
      color:var(--text);
    }
    .wrap{max-width:1100px;margin:0 auto;padding:22px}
    .topbar{
      display:flex;gap:14px;align-items:center;justify-content:space-between;
      padding:16px 18px;border:1px solid var(--line);border-radius:var(--radius);
      background:rgba(255,255,255,.03);box-shadow:var(--shadow);
      position:sticky;top:12px;backdrop-filter:blur(8px);z-index:10;
    }
    .title h2{margin:0;font-size:18px}
    .subtitle{font-size:13px;color:var(--muted);margin-top:4px}
    .card{
      border:1px solid var(--line); background:rgba(255,255,255,.03);
      border-radius:var(--radius); padding:16px; box-shadow:var(--shadow);
      margin-top:16px;
    }
    .card h3{margin:0 0 10px 0;font-size:15px}
    .muted{color:var(--muted);font-size:13px}
    .row{display:flex;gap:10px;flex-wrap:wrap;align-items:center}
    .btn{
      display:inline-flex;align-items:center;gap:8px;
      padding:10px 12px;border-radius:12px;border:1px solid var(--line);
      background:linear-gradient(180deg, rgba(255,255,255,.06), rgba(255,255,255,.02));
      cursor:pointer;font-weight:600;font-size:13px;
      transition:transform .05s ease;
    }
    .btn:hover{transform:translateY(-1px)}
    .btn.secondary{background:rgba(255,255,255,.02)}
    .btn.small{padding:7px 10px;font-size:12px;border-radius:10px}
    select, input{
      background:rgba(0,0,0,.25);border:1px solid var(--line);
      color:var(--text);padding:10px;border-radius:12px;outline:none;
      color-scheme:dark;
    }
    select option{background-color:#0f1b33;color:var(--text)}
    .sep{height:1px;background:var(--line);margin:12px 0}
    table{width:100%;border-collapse:separate;border-spacing:0}
    th, td{padding:10px;border-bottom:1px solid var(--line);font-size:13px}
    th{font-size:12px;color:var(--muted);text-align:left}
    tr:hover td{background:rgba(255,255,255,.02)}
    .table-wrap{overflow:auto;border:1px solid var(--line);border-radius:14px;margin-top:10px}
    .mono{font-family:var(--mono)}
    .hint{font-size:12px;color:var(--muted);margin-top:6px}
    .badge{
      display:inline-block;padding:4px 10px;border-radius:999px;
      border:1px solid var(--line);background:rgba(255,255,255,.03);
      font-size:12px;font-weight:600;
    }
    .badge.red{color:#fb7185;border-color:rgba(251,113,133,.35)}
    .badge.yellow{color:#fbbf24;border-color:rgba(251,191,36,.35)}
  </style>
</head>
<body>
""")

    html.append("<div class='wrap'>")
    html.append("<div class='topbar'>")
    html.append("<div class='title'>")
    html.append("<h2>⚠️ Seguimiento de pendientes</h2>")
    html.append(f"<div class='subtitle'><b>Empresa:</b> {esc(t.get('display_name',''))}</div>")
    html.append("</div>")
    html.append("<div class='row'>")
    html.append(f"<a class='btn secondary' href='/admin/panel?tenant={esc(tenant)}'>← Volver al panel</a>")
    html.append("</div>")
    html.append("</div>")

    # Selector de período
    html.append("<div class='card'>")
    html.append("<form method='get'>")
    html.append(f"<input type='hidden' name='tenant' value='{esc(tenant)}'>")
    html.append("<div class='row'>")
    html.append("<div>")
    html.append("<label class='muted'>Período</label><br>")
    html.append("<select name='period'>")
    for lbl in period_labels:
        sel = "selected" if lbl == panel_period else ""
        html.append(f"<option value='{esc(lbl)}' {sel}>{esc(lbl)}</option>")
    html.append("</select>")
    html.append("</div>")
    html.append("<button class='btn' type='submit' style='margin-top:18px'>Aplicar</button>")
    html.append("</div>")
    html.append("</form>")
    html.append("</div>")

    if not panel_period:
        html.append("<div class='card'><div class='muted'>Elegí un período para ver los pendientes.</div></div>")
        html.append("</div></body></html>")
        return Response("".join(html), mimetype="text/html")

    # No vieron el recibo
    html.append("<div class='card'>")
    html.append(f"<h3>🔴 No vieron el recibo <span class='badge red'>{len(pending_views_7d)}</span></h3>")
    html.append("<div class='muted'>Template enviado hace más de 7 días, nunca hicieron click en VIEW_NOW</div>")
    if pending_views_7d:
        html.append(f"""
            <form method='post' action='/admin/resend_all_pending_views' style='margin-top:10px' 
                onsubmit='return confirm("¿Reenviar template a {len(pending_views_7d)} personas?");'>
                <input type='hidden' name='token' value='{esc(token)}'>
                <input type='hidden' name='tenant' value='{esc(tenant)}'>
                <input type='hidden' name='period' value='{esc(panel_period)}'>
                <button class='btn' type='submit' style='background:linear-gradient(180deg, rgba(251,113,133,.15), rgba(251,113,133,.08)); border-color:rgba(251,113,133,.4)'>
                    🔄 Reenviar a TODOS ({len(pending_views_7d)})
                </button>
            </form>
        """)
    if pending_views_7d:
        html.append("<div class='table-wrap'>")
        html.append("<table>")
        html.append("<thead><tr><th>Nombre</th><th>CUIL</th><th>WhatsApp</th><th>Días</th><th>Acción</th></tr></thead>")
        html.append("<tbody>")
        
        for p in pending_views_7d:
            html.append("<tr>")
            html.append(f"<td>{esc(p['nombre'])}</td>")
            html.append(f"<td class='mono'>{esc(p['cuil'])}</td>")
            html.append(f"<td class='mono'>{esc(p['whatsapp'])}</td>")
            html.append(f"<td>{p['days_ago']} días</td>")
            html.append("<td>")
            html.append(f"""
                <form method='post' action='/admin/resend_template' style='margin:0'>
                    <input type='hidden' name='token' value='{esc(token)}'>
                    <input type='hidden' name='tenant' value='{esc(tenant)}'>
                    <input type='hidden' name='period' value='{esc(panel_period)}'>
                    <input type='hidden' name='cuil' value='{esc(p["cuil"])}'>
                    <input type='hidden' name='whatsapp' value='{esc(p["whatsapp"])}'>
                    <button class='btn small' type='submit'>🔄 Reenviar</button>
                </form>
            """)
            html.append("</td>")
            html.append("</tr>")
        
        html.append("</tbody></table></div>")
    else:
        html.append("<div class='muted' style='margin-top:10px'>✅ No hay pendientes en esta categoría</div>")

    html.append("</div>")

    # No firmaron
    html.append("<div class='card'>")
    html.append(f"<h3>🟡 No firmaron <span class='badge yellow'>{len(pending_sigs_7d)}</span></h3>")
    html.append("<div class='muted'>PDF recibido hace más de 7 días, nunca firmaron (SIGN_OK/SIGN_OBS)</div>")
    if pending_sigs_7d:
        html.append(f"""
            <form method='post' action='/admin/remind_all_pending_signatures' style='margin-top:10px'
                onsubmit='return confirm("¿Enviar recordatorio de firma a {len(pending_sigs_7d)} personas?");'>
                <input type='hidden' name='token' value='{esc(token)}'>
                <input type='hidden' name='tenant' value='{esc(tenant)}'>
                <input type='hidden' name='period' value='{esc(panel_period)}'>
                <button class='btn' type='submit' style='background:linear-gradient(180deg, rgba(251,191,36,.15), rgba(251,191,36,.08)); border-color:rgba(251,191,36,.4)'>
                    📝 Recordar a TODOS ({len(pending_sigs_7d)})
                </button>
            </form>
        """)
    if pending_sigs_7d:
        html.append("<div class='table-wrap'>")
        html.append("<table>")
        html.append("<thead><tr><th>Nombre</th><th>CUIL</th><th>WhatsApp</th><th>Días</th><th>Acción</th></tr></thead>")
        html.append("<tbody>")
        
        for p in pending_sigs_7d:
            html.append("<tr>")
            html.append(f"<td>{esc(p['nombre'])}</td>")
            html.append(f"<td class='mono'>{esc(p['cuil'])}</td>")
            html.append(f"<td class='mono'>{esc(p['whatsapp'])}</td>")
            html.append(f"<td>{p['days_ago']} días</td>")
            html.append("<td>")
            html.append(f"""
                <form method='post' action='/admin/remind_signature' style='margin:0'>
                    <input type='hidden' name='token' value='{esc(token)}'>
                    <input type='hidden' name='tenant' value='{esc(tenant)}'>
                    <input type='hidden' name='period' value='{esc(panel_period)}'>
                    <input type='hidden' name='cuil' value='{esc(p["cuil"])}'>
                    <input type='hidden' name='whatsapp' value='{esc(p["whatsapp"])}'>
                    <button class='btn small' type='submit'>📝 Recordar</button>
                </form>
            """)
            html.append("</td>")
            html.append("</tr>")
        
        html.append("</tbody></table></div>")
    else:
        html.append("<div class='muted' style='margin-top:10px'>✅ No hay pendientes en esta categoría</div>")

    html.append("</div>")

    html.append("</div></body></html>")
    return Response("".join(html), mimetype="text/html")


@app.get("/admin/panel")
def admin_panel():
    auth = require_admin()
    if auth:
        return auth

    token = (request.args.get("token") or "").strip()
    tenant = (request.args.get("tenant") or "").strip().lower()

    if not tenant:
        return Response("Falta tenant. Volvé a /admin.", status=400)

    t = get_tenant(tenant)
    if not t:
        return Response("Tenant inválido. Volvé a /admin.", status=400)

    force = (request.args.get("refresh") in ("1", "true", "yes", "on"))
    envios_rows = load_envios_rows(tenant, force=force) or []

    panel_period = (request.args.get("period") or "").strip()

    # Reportes: periodos disponibles
    selected_period = panel_period
    period_folders = list_tenant_period_folders(tenant)  # ['01-2026','12-2025',...]
    period_labels = []
    for p in period_folders:
        lbl = period_folder_to_label(p)  # '01/2026'
        if lbl:
            period_labels.append(lbl)
    if not selected_period and period_labels:
        selected_period = period_labels[0]

    period_q = quote(selected_period or "", safe="")

    # Cola stats
    stats = None
    if panel_period:
        stats = get_queue_stats(tenant, panel_period)
    
    # Cargar pendientes de seguimiento
    pending_views_7d = []
    pending_sigs_7d = []
    if panel_period:
        pending_views_7d = get_pending_views_over_7days(tenant, panel_period)
        pending_sigs_7d = get_pending_signatures_over_7days(tenant, panel_period)

    # Verificaciones
    verifs = get_verifications_rows(tenant)

    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Panel empresa</title>
  <style>
    :root{
      --bg:#0b1220;
      --card:#0f1b33;
      --muted:#9fb2d0;
      --text:#eaf0ff;
      --line:rgba(255,255,255,.08);
      --accent:#5aa7ff;
      --ok:#34d399;
      --warn:#fbbf24;
      --bad:#fb7185;
      --btn:#14264a;
      --btn2:#1a2f5a;
      --shadow: 0 10px 25px rgba(0,0,0,.25);
      --radius:14px;
      --mono: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", monospace;
      --sans: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Helvetica, Arial, "Apple Color Emoji","Segoe UI Emoji";
    }
    *{box-sizing:border-box}
    body{
      margin:0;
      font-family:var(--sans);
      background: radial-gradient(1200px 700px at 20% -20%, rgba(90,167,255,.25), transparent 60%),
                  radial-gradient(1200px 700px at 90% 0%, rgba(52,211,153,.18), transparent 55%),
                  var(--bg);
      color:var(--text);
    }
    a{color:inherit;text-decoration:none}
    .wrap{max-width:1100px;margin:0 auto;padding:22px}
    .topbar{
      display:flex;gap:14px;align-items:center;justify-content:space-between;
      padding:16px 18px;border:1px solid var(--line);border-radius:var(--radius);
      background:rgba(255,255,255,.03);box-shadow:var(--shadow);
      position:sticky;top:12px;backdrop-filter: blur(8px); z-index:10;
    }
    .title{display:flex;flex-direction:column;gap:4px}
    .title h2{margin:0;font-size:18px;letter-spacing:.2px}
    .subtitle{font-size:13px;color:var(--muted)}
    .pill{
      display:inline-flex;align-items:center;gap:8px;
      font-size:12px;color:var(--muted)
    }
    code{font-family:var(--mono);font-size:12px;background:rgba(255,255,255,.06);padding:2px 6px;border-radius:8px;border:1px solid var(--line)}
    .grid{
      margin-top:16px;
      display:grid;
      grid-template-columns: 1.2fr .8fr;
      gap:14px;
    }
    @media(max-width:960px){ .grid{grid-template-columns:1fr} .topbar{flex-direction:column;align-items:flex-start}}
    .card{
      border:1px solid var(--line);
      background:rgba(255,255,255,.03);
      border-radius:var(--radius);
      padding:16px;
      box-shadow:var(--shadow);
    }
    .card h3{margin:0 0 10px 0;font-size:15px}
    .muted{color:var(--muted);font-size:13px}
    .row{display:flex;gap:10px;flex-wrap:wrap;align-items:center}
    .btn{
      display:inline-flex;align-items:center;justify-content:center;
      gap:8px;
      padding:10px 12px;
      border-radius:12px;
      border:1px solid var(--line);
      background:linear-gradient(180deg, rgba(255,255,255,.06), rgba(255,255,255,.02));
      cursor:pointer;
      font-weight:600;
      font-size:13px;
      transition:transform .05s ease, background .2s ease;
    }
    .btn:hover{transform:translateY(-1px);background:linear-gradient(180deg, rgba(255,255,255,.08), rgba(255,255,255,.03))}
    .btn.secondary{background:rgba(255,255,255,.02)}
    .btn.danger{border-color:rgba(251,113,133,.35); background:rgba(251,113,133,.10)}
    .btn.small{padding:7px 10px;font-size:12px;border-radius:10px}
    input[type="text"], input[type="number"], select{
      background:rgba(0,0,0,.25);
      border:1px solid var(--line);
      color:var(--text);
      padding:10px 10px;
      border-radius:12px;
      outline:none;
      min-width: 180px;
      color-scheme:dark;
    }
    select option{background-color:#0f1b33;color:var(--text)}
    label{font-size:12px;color:var(--muted)}
    .badge{
      display:inline-flex;align-items:center;gap:6px;
      padding:6px 10px;border-radius:999px;
      border:1px solid var(--line);
      background:rgba(255,255,255,.03);
      font-size:12px;color:var(--muted);
    }
    .badge b{color:var(--text)}
    .badge.ok{border-color:rgba(52,211,153,.35)}
    .badge.warn{border-color:rgba(251,191,36,.35)}ff
    .badge.bad{border-color:rgba(251,113,133,.35)}
    .sep{height:1px;background:var(--line);margin:12px 0}
    table{width:100%;border-collapse:separate;border-spacing:0}
    th, td{
      padding:10px 10px;
      border-bottom:1px solid var(--line);
      font-size:13px;
      vertical-align:middle;
    }
    th{font-size:12px;color:var(--muted);text-align:left}
    tr:hover td{background:rgba(255,255,255,.02)}
    .table-wrap{
    overflow:auto;
    max-height:360px;
    border:1px solid var(--line);
    border-radius:14px;
    }
    .table-wrap thead th{
    position:sticky;
    top:0;
    background:#0f1b33;
    z-index:1;
    }
    .right{display:flex;gap:10px;align-items:center;flex-wrap:wrap}
    .hint{font-size:12px;color:var(--muted);margin-top:6px}
    .kpi{display:flex;gap:8px;flex-wrap:wrap;margin-top:8px}
    .mono{font-family:var(--mono)}
  </style>
</head>
<body>
""")

    html.append("<div class='wrap'>")

    html.append("<div class='topbar'>")
    html.append("<div class='title'>")
    html.append("<h2>Panel empresa</h2>")
    html.append(f"<div class='subtitle'><b>Empresa:</b> {esc(t.get('display_name',''))} · <span class='pill'>slug <code>{esc(t.get('slug',''))}</code></span></div>")
    html.append("</div>")
    html.append("<div class='right'>")
    html.append(f"<a class='btn secondary' href='/admin'>← Volver</a>")
    html.append(f"<a class='btn' href='/admin/reenviar_template'>📤 Reenviar template</a>")
    html.append(f"<a class='btn' href='/admin/send_test?tenant={esc(tenant)}'>📤 Envío individual</a>")
    html.append(f"<a class='btn' href='/admin/reenviar_fallidos'>🔄 Reenviar fallidos</a>")
    html.append(f"<a class='btn' href='/admin/seguimiento?tenant={esc(tenant)}'>⚠️ Seguimiento</a>")
    html.append(f"<a class='btn' href='/admin/branding?tenant={esc(tenant)}'>🎨 Personalización</a>")
    html.append("</div>")
    html.append("</div>")

    html.append("<div class='grid'>")

    # Columna izquierda: acciones principales
    html.append("<div class='card'>")
    html.append("<h3>📩 Envío masivo</h3>")
    html.append("<div class='muted'>Enviá la plantilla a toda la empresa y gestioná la cola por período.</div>")
    html.append("<div class='sep'></div>")

    # Start queue
    html.append("<form method='post' action='/admin/send_template_queue_start'>")
    html.append(f"<input type='hidden' name='tenant' value='{esc(tenant)}'>")
    html.append("<div class='row'>")
    html.append("<div>")
    html.append("<label>Período (mm/aaaa)</label><br>")
    html.append("<select name='period' id='massPeriod' required>")
    for lbl in period_labels:
        sel = "selected" if lbl == (selected_period or "") else ""
        html.append(f"<option value='{esc(lbl)}' {sel}>{esc(lbl)}</option>")
    html.append("</select>")
    html.append("</div>")
    html.append("<div>")
    html.append("<label>Límite (0 = todos)</label><br>")
    html.append("<input type='number' name='limit' id='massLimit' min='0' value='0'>")
    html.append("</div>")
    html.append("</div>")
    html.append("<input type='hidden' name='require_pdf' value='true'>")
    html.append("<div style='margin-top:10px'>")
    html.append("<button class='btn' type='submit'>🚀 Encolar envío a toda la empresa</button>")
    html.append("</div>")
    html.append("</form>")
    # Queue section
    html.append("<div class='sep'></div>")
    html.append("<h3>⏳ Cola de envíos</h3>")

    if panel_period and stats:
        html.append(f"<div class='muted'>Período cola: <b>{esc(panel_period)}</b></div>")
        html.append("<div class='kpi'>")
        html.append(f"<span class='badge warn'>PENDING <b>{stats['PENDING']}</b></span>")
        html.append(f"<span class='badge ok'>SENT <b>{stats['SENT']}</b></span>")
        html.append(f"<span class='badge'>SKIPPED <b>{stats['SKIPPED']}</b></span>")
        html.append(f"<span class='badge bad'>FAILED <b>{stats['FAILED']}</b></span>")
        html.append("</div>")

        html.append("<form method='post' action='/admin/send_template_queue_tick' style='margin-top:10px;'>")
        html.append(f"<input type='hidden' name='tenant' value='{esc(tenant)}'>")
        html.append(f"<input type='hidden' name='period' value='{esc(panel_period)}'>")
        html.append("<div class='row'>")
        html.append("<div>")
        html.append("<label>Batch size</label><br>")
        html.append("<input type='number' name='batch_size' min='1' max='50' value='10'>")
        html.append("</div>")
        html.append("<div style='margin-top:18px'>")
        html.append("<button class='btn' type='submit'>▶️ Procesar ahora</button>")
        html.append("</div>")
        html.append("</div>")
        html.append("</form>")

        # NUEVO: Botón de envío automático
        html.append("<form method='post' action='/admin/send_auto' style='margin-top:10px;'>")
        html.append(f"<input type='hidden' name='tenant' value='{esc(tenant)}'>")
        html.append(f"<input type='hidden' name='period' value='{esc(panel_period)}'>")
        html.append("<input type='hidden' name='batch_size' value='10'>")
        html.append("<button class='btn' type='submit' style='background:linear-gradient(180deg, rgba(90,167,255,.15), rgba(90,167,255,.08)); border-color:rgba(90,167,255,.4)'>🚀 Envío automático (procesa todo)</button>")
        html.append("</form>")

        html.append("<div class='hint'>Cron URL (POST) sugerida:</div>")
        html.append(f"<div class='hint mono'><code>/admin/send_template_queue_tick?tenant={esc(tenant)}&period={esc(panel_period)}&batch_size=10&mode=json&token=TU_ADMIN_TOKEN</code> · para el cron: reemplazá TU_ADMIN_TOKEN por el valor real (o mandalo por header X-Admin-Token y ni siquiera queda en los access logs)</div>")
    else:
        html.append("<div class='muted'>Elegí un período en Reportes para ver la cola y poder procesarla.</div>")

    html.append("</div>")  # fin card izquierda

    # Columna derecha: reportes + herramientas rápidas
    html.append("<div class='card'>")
    html.append("<h3>📊 Reportes</h3>")
    html.append("<div class='muted'>Descargá reportes filtrando por período.</div>")
    html.append("<div class='sep'></div>")

    html.append("<form method='get' action='/admin/panel'>")
    html.append(f"<input type='hidden' name='tenant' value='{esc(tenant)}'>")
    html.append("<label>Período</label><br>")
    html.append("<div class='row' style='margin-top:6px'>")
    html.append("<select name='period'>")
    html.append("<option value=''>-- Todos / Sin filtro --</option>")
    for lbl in period_labels:
        sel = "selected" if lbl == selected_period else ""
        html.append(f"<option value='{esc(lbl)}' {sel}>{esc(lbl)}</option>")
    html.append("</select>")
    html.append("<button class='btn secondary' type='submit'>Aplicar</button>")
    html.append("</div>")
    html.append("</form>")

    html.append("<div class='sep'></div>")
    html.append("<div class='row'>")
    html.append(
        f"<a class='btn' href='/admin/report_recibos.xlsx?tenant={esc(tenant)}&period={period_q}'>📄 Reporte recibos (XLSX)</a>"
    )
    html.append(
        f"<a class='btn' href='/admin/report_recibos.pdf?tenant={esc(tenant)}&period={period_q}'>🧾 Informe (PDF)</a>"
    )
    html.append(
        f"<a class='btn secondary' href='/admin/report_envios.csv?tenant={esc(tenant)}'>📤 Envíos (CSV)</a>"
    )
    html.append("</div>")

    html.append("<div class='sep'></div>")
    html.append("<h3>🔎 Buscar períodos por CUIL</h3>")
    html.append(f"""
      <form method="get" action="/admin/periodos">
        <input type="hidden" name="tenant" value="{esc(tenant)}">
        <label>CUIL</label><br>
        <div class="row" style="margin-top:6px">
          <input type="text" name="cuil" placeholder="xx-xxxxxxxx-x" required>
          <button class="btn secondary" type="submit">Buscar</button>
        </div>
      </form>
    """)

    html.append("<div class='sep'></div>")
    html.append("<h3>🧹 Reset</h3>")
    html.append("<div class='muted'>Borra <code>pending_views</code> y <code>recibo_estado</code> para esta empresa (y período si lo completás).</div>")
    html.append(f"""
      <form method="post" action="/admin/reset" onsubmit="return confirm('¿Seguro? Esto borra pending y estados.');" style="margin-top:10px">
        <input type="hidden" name="tenant" value="{esc(tenant)}">
        <label>Período (opcional, mm/aaaa)</label><br>
        <div class="row" style="margin-top:6px">
          <input type="text" name="period" placeholder="01/2026">
          <button class="btn danger" type="submit">Resetear</button>
        </div>
      </form>
    """)

    html.append("</div>")  # fin card derecha

    html.append("</div>")  # fin grid

    # Verificaciones
    html.append("<div class='card' style='margin-top:14px'>")
    html.append("<div class='row' style='justify-content:space-between'>")
    html.append("<div>")
    html.append("<h3>✅ Verificaciones</h3>")
    html.append(f"<div class='muted'>Registros: <b>{len(verifs)}</b></div>")
    html.append("</div>")
    html.append("</div>")
    html.append("<div class='sep'></div>")

    if verifs:
        html.append(f"""
          <form id="bulkForm" method="post" action="/admin/verifications_delete_bulk"
                onsubmit="return confirm('¿Borrar verificaciones seleccionadas?');">
                <input type="hidden" name="tenant" value="{esc(tenant)}">
          </form>
        """)

        html.append("<div class='table-wrap'>")
        html.append("<table>")
        html.append("""
          <thead>
            <tr>
              <th></th>
              <th>CUIL</th>
              <th>Nombre</th>
              <th>WhatsApp</th>
              <th>Verificado</th>
              <th>Acciones</th>
            </tr>
          </thead>
          <tbody>
        """)

        for r in verifs[:500]:
            key = f"{r['cuil']}|{r['to_whatsapp']}"
            html.append("<tr>")
            # checkbox asociado al bulkForm sin anidarlo
            html.append(f"<td><input form='bulkForm' type='checkbox' name='keys' value='{esc(key)}'></td>")
            html.append(f"<td>{esc(r['cuil'])}</td>")
            html.append(f"<td>{esc(r.get('nombre','') or '')}</td>")
            html.append(f"<td>{esc(r['to_whatsapp'])}</td>")
            html.append(f"<td>{esc(ts_str(r.get('verified_at')))}</td>")
            html.append("<td class='row'>")

            # form independiente por fila
            html.append(f"""
              <form method="post" action="/admin/verifications_delete"
                    onsubmit="return confirm('¿Borrar verificación?');" style="margin:0">
                        <input type="hidden" name="tenant" value="{esc(tenant)}">
                <input type="hidden" name="cuil" value="{esc(r['cuil'])}">
                <input type="hidden" name="to_whatsapp" value="{esc(r['to_whatsapp'])}">
                <button class="btn small danger" type="submit">Borrar</button>
              </form>
            """)
            html.append("</td>")
            html.append("</tr>")

        html.append("</tbody></table></div>")
        html.append("<div style='margin-top:10px'>")
        html.append("<button class='btn danger' form='bulkForm' type='submit'>🗑️ Borrar seleccionados</button>")
        html.append("</div>")
    else:
        html.append("<div class='muted'>No hay verificaciones cargadas.</div>")

    html.append("<div class='sep'></div>")

    # Importar verificaciones
    html.append(f"""
      <h3>📥 Importar verificaciones</h3>
      <div class="muted">Subí un Excel con columnas: <code>cuil</code>, <code>whatsapp</code> (o teléfono). Opcional: <code>dni</code>.</div>
      <form method="post" action="/admin/verifications_import" enctype="multipart/form-data" style="margin-top:10px">
        <input type="hidden" name="tenant" value="{esc(tenant)}">
        <div class="row">
          <input type="file" name="file" accept=".xlsx" required>
          <button class="btn secondary" type="submit">Importar</button>
        </div>
      </form>
    """)

    html.append("</div>")  # fin card verifs

    # Preview envíos
    html.append("<div class='card' style='margin-top:14px'>")
    html.append("<div class='row' style='justify-content:space-between'>")
    html.append("<div>")
    html.append("<h3>👀 Preview Excel de envíos</h3>")
    html.append(f"<div class='muted'>Filas: <b>{len(envios_rows)}</b></div>")
    html.append("</div>")
    html.append(f"<a class='btn secondary' href='/admin/panel?tenant={esc(tenant)}&refresh=1&period={esc(selected_period or '')}'>🔄 Refrescar</a>")
    html.append("</div>")
    html.append("<div class='sep'></div>")

    sample = envios_rows
    if sample:
        cols = list(sample[0].keys())
        html.append("<div class='table-wrap'>")
        html.append("<table>")
        html.append("<thead><tr>" + "".join(f"<th>{esc(c)}</th>" for c in cols) + "</tr></thead>")
        html.append("<tbody>")
        for r in sample:
            html.append("<tr>" + "".join(f"<td>{esc(str(r.get(c,'')))}</td>" for c in cols) + "</tr>")
        html.append("</tbody></table></div>")
    else:
        html.append("<div class='muted'>No se pudo leer el Excel de envíos o está vacío.</div>")

    html.append("</div>")  # fin preview card

    html.append("</div></body></html>")
    return Response("".join(html), mimetype="text/html")

def get_all_tenants():
    """
    Obtiene todas las empresas (tenants) del master file.
    """
    tenants = load_tenants()
    return tenants if tenants else []

@app.route("/admin/portal_users", methods=["GET", "POST"])
def admin_portal_users():
    """
    Gestión de usuarios del portal de clientes.
    """
    auth = require_admin()
    if auth:
        return auth
    
    token = request.args.get("token") or request.form.get("token") or ""
    
    # Crear nuevo usuario
    if request.method == "POST":
        action = request.form.get("action", "")
        
        if action == "create":
            tenant = request.form.get("tenant", "").strip().lower()
            username = request.form.get("username", "").strip().lower()
            email = request.form.get("email", "").strip()
            full_name = request.form.get("full_name", "").strip()
            
            result = create_client_user(tenant, username, email, full_name, created_by="admin")
            
            if result['ok']:
                msg = f"success&temp_pwd={result.get('temp_password', '')}"
            else:
                msg = f"error&details={result['message']}"
            
            return redirect(f"/admin/portal_users?msg={msg}")
        
        elif action == "toggle":
            user_id = int(request.form.get("user_id", 0))
            toggle_client_user_active(user_id)
            return redirect(f"/admin/portal_users?msg=toggled")
        
        elif action == "delete":
            user_id = int(request.form.get("user_id", 0))
            delete_client_user(user_id)
            return redirect(f"/admin/portal_users?msg=deleted")
    
    # Listar usuarios
    users = get_all_client_users()
    tenants = get_all_tenants()
    
    msg = request.args.get("msg", "")
    temp_pwd = request.args.get("temp_pwd", "")
    
    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Usuarios del Portal</title>
  <style>
    :root{
      --bg:#0b1220; --card:#0f1b33; --muted:#9fb2d0; --text:#eaf0ff;
      --line:rgba(255,255,255,.08); --ok:#34d399; --bad:#fb7185;
      --radius:14px; --mono: monospace;
    }
    *{box-sizing:border-box}
    body{
      margin:0; font-family:system-ui;
      background: radial-gradient(1200px 700px at 20% -20%, rgba(90,167,255,.25), transparent 60%),
                  var(--bg);
      color:var(--text); padding:20px;
    }
    .wrap{max-width:1100px;margin:0 auto}
    .card{
      border:1px solid var(--line); background:rgba(255,255,255,.03);
      border-radius:var(--radius); padding:20px; margin-bottom:20px;
    }
    h2{margin:0 0 10px 0}
    .muted{color:var(--muted);font-size:13px}
    input, select{
      background:rgba(0,0,0,.25); border:1px solid var(--line);
      color:var(--text); padding:10px; border-radius:8px; margin:5px 0;
      color-scheme:dark;
    }
    select option{background-color:#0f1b33;color:var(--text)}
    .btn{
      display:inline-block; padding:10px 16px; border-radius:8px;
      border:1px solid var(--line); background:rgba(255,255,255,.06);
      cursor:pointer; font-weight:600; text-decoration:none; color:var(--text);
      margin:5px;
    }
    .btn:hover{background:rgba(255,255,255,.1)}
    .btn.danger{border-color:rgba(251,113,133,.35); background:rgba(251,113,133,.10)}
    table{width:100%; border-collapse:collapse; margin-top:15px}
    th, td{padding:12px; text-align:left; border-bottom:1px solid var(--line)}
    th{color:var(--muted); font-size:12px}
    .success{background:rgba(52,211,153,.1); border:1px solid rgba(52,211,153,.3); padding:12px; border-radius:8px; margin:10px 0}
    .error{background:rgba(251,113,133,.1); border:1px solid rgba(251,113,133,.3); padding:12px; border-radius:8px; margin:10px 0}
    .mono{font-family:var(--mono); background:rgba(0,0,0,.3); padding:2px 6px; border-radius:4px}
  </style>
</head>
<body>
<div class="wrap">
""")
    
    html.append(f"<a href='/admin' class='btn'>← Volver al admin</a>")
    
    html.append("<div class='card'>")
    html.append("<h2>👥 Usuarios del Portal</h2>")
    html.append("<div class='muted'>Gestión de accesos para clientes</div>")
    
    # Mensajes
    if msg == "success":
        html.append(f"<div class='success'>✅ Usuario creado exitosamente")
        if temp_pwd:
            html.append(f"<br>Contraseña temporal: <span class='mono'>{esc(temp_pwd)}</span>")
            html.append("<br><small>(El usuario recibió un email con estos datos)</small>")
        html.append("</div>")
    elif msg.startswith("error"):
        details = request.args.get("details", "")
        html.append(f"<div class='error'>❌ Error: {esc(details)}</div>")
    elif msg == "toggled":
        html.append("<div class='success'>✅ Estado actualizado</div>")
    elif msg == "deleted":
        html.append("<div class='success'>✅ Usuario eliminado</div>")
    
    html.append("</div>")
    
    # Formulario crear usuario
    html.append("<div class='card'>")
    html.append("<h3>➕ Crear nuevo usuario</h3>")
    html.append(f"<form method='post'>")
    html.append(f"<input type='hidden' name='action' value='create'>")
    
    html.append("<label>Empresa:</label><br>")
    html.append("<select name='tenant' required style='width:100%;max-width:400px'>")
    html.append("<option value=''>-- Seleccionar empresa --</option>")
    for t in tenants:
        html.append(f"<option value='{esc(t['slug'])}'>{esc(t['display_name'])}</option>")
    html.append("</select><br>")
    
    html.append("<label>Email (será el usuario de acceso):</label><br>")
    html.append("<input type='email' name='email' required placeholder='rrhh@empresa.com' style='width:100%;max-width:400px'><br>")
    
    html.append("<label>Nombre completo:</label><br>")
    html.append("<input type='text' name='full_name' placeholder='María González' style='width:100%;max-width:400px'><br>")
    
    html.append("<button type='submit' class='btn'>Crear usuario</button>")
    html.append("</form>")
    html.append("</div>")
    # Lista de usuarios
    html.append("<div class='card'>")
    html.append(f"<h3>📋 Usuarios existentes ({len(users)})</h3>")
    
    if users:
        html.append("<table>")
        html.append("<thead><tr>")
        html.append("<th>Empresa</th><th>Usuario</th><th>Email</th><th>Nombre</th><th>Estado</th><th>Último login</th><th>Acciones</th>")
        html.append("</tr></thead><tbody>")
        
        for u in users:
            status = "✅ Activo" if u['active'] else "❌ Inactivo"
            last_login = ts_str(u['last_login']) if u['last_login'] else "Nunca"
            must_change = " (debe cambiar pwd)" if u['must_change_password'] else ""
            
            html.append("<tr>")
            html.append(f"<td>{esc(u['tenant'])}</td>")
            html.append(f"<td class='mono'>{esc(u['username'])}</td>")
            html.append(f"<td>{esc(u['email'] or '')}</td>")
            html.append(f"<td>{esc(u['full_name'] or '')}</td>")
            html.append(f"<td>{status}{must_change}</td>")
            html.append(f"<td>{last_login}</td>")
            html.append("<td>")
            
            # Toggle activo/inactivo
            html.append(f"<form method='post' style='display:inline'>")
            html.append(f"<input type='hidden' name='action' value='toggle'>")
            html.append(f"<input type='hidden' name='user_id' value='{u['id']}'>")
            toggle_text = "Desactivar" if u['active'] else "Activar"
            html.append(f"<button type='submit' class='btn'>{toggle_text}</button>")
            html.append("</form>")
            
            # Eliminar
            html.append(f"<form method='post' style='display:inline' onsubmit='return confirm(\"¿Eliminar usuario?\")'>")
            html.append(f"<input type='hidden' name='action' value='delete'>")
            html.append(f"<input type='hidden' name='user_id' value='{u['id']}'>")
            html.append(f"<button type='submit' class='btn danger'>Eliminar</button>")
            html.append("</form>")
            
            html.append("</td>")
            html.append("</tr>")
        
        html.append("</tbody></table>")
    else:
        html.append("<div class='muted'>No hay usuarios creados todavía</div>")
    
    html.append("</div>")
    
    html.append("</div></body></html>")
    return Response("".join(html), mimetype="text/html")
@app.route("/admin/portal_activity")
def admin_portal_activity():
    """Historial de actividad de los usuarios del portal (vista admin)."""
    auth = require_admin()
    if auth:
        return auth

    filtro_tenant = (request.args.get("tenant") or "").strip().lower()
    filtro_action = (request.args.get("action") or "").strip()

    eventos = get_admin_audit_log(tenant=filtro_tenant, action=filtro_action, limit=300)
    tenants = get_all_tenants()
    acciones = ['login', 'logout', 'change_password', 'password_reset', 'delete_certificado']

    html = []
    html.append("""<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Actividad del Portal</title>
  <style>
    :root{
      --bg:#0b1220; --card:#0f1b33; --muted:#9fb2d0; --text:#eaf0ff;
      --line:rgba(255,255,255,.08); --ok:#34d399; --bad:#fb7185;
      --radius:14px; --mono: monospace;
    }
    *{box-sizing:border-box}
    body{
      margin:0; font-family:system-ui;
      background: radial-gradient(1200px 700px at 20% -20%, rgba(90,167,255,.25), transparent 60%), var(--bg);
      color:var(--text); padding:20px;
    }
    .wrap{max-width:1100px;margin:0 auto}
    .card{ border:1px solid var(--line); background:rgba(255,255,255,.03); border-radius:var(--radius); padding:20px; margin-bottom:20px; }
    h2{margin:0 0 10px 0}
    .muted{color:var(--muted);font-size:13px}
    select{ background:rgba(0,0,0,.25); border:1px solid var(--line); color:var(--text); padding:10px; border-radius:8px; margin:5px 0; color-scheme:dark; }
    select option{ background-color:#0f1b33; color:var(--text); }
    .btn{ display:inline-block; padding:10px 16px; border-radius:8px; border:1px solid var(--line); background:rgba(255,255,255,.06); cursor:pointer; font-weight:600; text-decoration:none; color:var(--text); margin:5px; }
    .btn:hover{background:rgba(255,255,255,.1)}
    table{width:100%; border-collapse:collapse; margin-top:15px}
    th, td{padding:10px 12px; text-align:left; border-bottom:1px solid var(--line); font-size:13px}
    th{color:var(--muted); font-size:12px}
    .mono{font-family:var(--mono); font-size:12px; color:var(--muted)}
  </style>
</head>
<body>
<div class="wrap">
""")
    html.append("<a href='/admin' class='btn'>← Volver al admin</a>")

    html.append("<div class='card'>")
    html.append("<h2>🕑 Actividad del Portal</h2>")
    html.append("<div class='muted'>Accesos y acciones de los usuarios de los clientes</div>")

    # Filtros
    html.append("<form method='get' style='margin-top:14px'>")
    html.append("<select name='tenant'>")
    html.append("<option value=''>-- Todas las empresas --</option>")
    for t in tenants:
        sel = " selected" if (t['slug'] == filtro_tenant) else ""
        html.append(f"<option value='{esc(t['slug'])}'{sel}>{esc(t['display_name'])}</option>")
    html.append("</select> ")
    html.append("<select name='action'>")
    html.append("<option value=''>-- Todas las acciones --</option>")
    for a in acciones:
        sel = " selected" if (a == filtro_action) else ""
        html.append(f"<option value='{esc(a)}'{sel}>{esc(_accion_label(a))}</option>")
    html.append("</select> ")
    html.append("<button type='submit' class='btn'>Filtrar</button>")
    html.append("</form>")
    html.append("</div>")

    # Tabla
    html.append("<div class='card'>")
    html.append(f"<h3>Últimos {len(eventos)} movimientos</h3>")
    if eventos:
        html.append("<table>")
        html.append("<thead><tr><th>Fecha</th><th>Empresa</th><th>Usuario</th><th>Acción</th><th>Detalle</th><th>IP</th></tr></thead>")
        html.append("<tbody>")
        for e in eventos:
            quien = e.get('full_name') or e.get('email') or e.get('username') or f"(user #{e.get('user_id','?')})"
            html.append("<tr>")
            html.append(f"<td>{esc(ts_str(e.get('created_at')))}</td>")
            html.append(f"<td>{esc(e.get('tenant','') or '')}</td>")
            html.append(f"<td>{esc(quien)}</td>")
            html.append(f"<td>{esc(_accion_label(e.get('action','')))}</td>")
            html.append(f"<td>{esc(e.get('details','') or '')}</td>")
            html.append(f"<td class='mono'>{esc(e.get('ip_address','') or '')}</td>")
            html.append("</tr>")
        html.append("</tbody></table>")
    else:
        html.append("<div class='muted'>No hay actividad registrada con esos filtros.</div>")
    html.append("</div>")

    html.append("</div></body></html>")
    return Response("".join(html), mimetype="text/html")

@app.get("/admin/periodos")
def admin_periodos():
    auth = require_admin()
    if auth:
        return auth

    tenant = (request.args.get("tenant") or "").strip().lower()
    cuil = (request.args.get("cuil") or "").strip()
    if not get_tenant(tenant):
        return Response("Tenant inválido", status=400)
    if not cuil:
        return Response("Falta CUIL", status=400)

    periods = list_periods_for_cuil(tenant, cuil)
    return jsonify({"tenant": tenant, "cuil": cuil, "periodos": periods})

import pandas as pd

def get_pending_views_over_7days(tenant: str, period: str) -> list:
    """
    Devuelve personas que recibieron el template hace +7 días pero nunca pidieron el PDF.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    
    seven_days_ago = int(time.time()) - (7 * 24 * 60 * 60)
    
    # Buscar templates enviados hace más de 7 días SIN PDF enviado
    cur.execute("""
        SELECT 
            ms.cuil,
            ms.to_whatsapp,
            ms.nombre,
            ms.created_at as sent_at,
            ms.message_sid as sent_sid
        FROM message_status ms
        WHERE ms.tenant = %s
          AND ms.period = %s
          AND ms.kind = 'template'
          AND ms.created_at < %s
          AND ms.created_at IS NOT NULL
          AND NOT EXISTS (
            SELECT 1 FROM sent_pdfs sp
            WHERE sp.tenant = ms.tenant
              AND sp.cuil = ms.cuil
              AND sp.period = ms.period
              AND sp.to_whatsapp = ms.to_whatsapp
          )
        ORDER BY ms.created_at ASC
    """, (tenant, period, seven_days_ago))
    
    rows = cur.fetchall()
    conn.close()
    
    result = []
    for r in rows:
        days_ago = int((time.time() - r['sent_at']) / 86400) if r['sent_at'] else 0
        result.append({
            'cuil': r['cuil'],
            'whatsapp': r['to_whatsapp'],
            'nombre': r['nombre'] or '',
            'sent_at': r['sent_at'],
            'days_ago': days_ago,
            'sid': r['sent_sid']
        })
    
    return result

def get_pending_signatures_over_7days(tenant: str, period: str) -> list:
    """
    Devuelve personas que recibieron el PDF hace +7 días pero nunca firmaron ni observaron.
    """
    conn = get_db_connection()
    cur = conn.cursor()
    
    seven_days_ago = int(time.time()) - (7 * 24 * 60 * 60)
    
    # Buscar PDFs enviados hace más de 7 días sin firma ni observación
    cur.execute("""
        SELECT 
            sp.cuil,
            sp.to_whatsapp,
            sp.created_at,
            sp.message_sid
        FROM sent_pdfs sp
        WHERE sp.tenant = %s
          AND sp.period = %s
          AND sp.created_at < %s
          AND sp.created_at IS NOT NULL
          AND NOT EXISTS (
            SELECT 1 FROM recibo_estado re
            WHERE re.tenant = sp.tenant
              AND re.cuil = sp.cuil
              AND re.period = sp.period
              AND re.estado IN ('FIRMADO', 'OBSERVADO')
          )
        ORDER BY sp.created_at ASC
    """, (tenant, period, seven_days_ago))
    
    rows = cur.fetchall()
    conn.close()
    
    # Enriquecer con datos de envios
    envios = load_envios_rows(tenant)
    
    result = []
    for r in rows:
        person = find_person_by_cuil(envios, r['cuil'])
        days_ago = int((time.time() - r['created_at']) / 86400) if r['created_at'] else 0
        
        result.append({
            'cuil': r['cuil'],
            'whatsapp': r['to_whatsapp'],
            'nombre': person.get('nombre', '') if person else '',
            'sent_at': r['created_at'],
            'days_ago': days_ago,
            'sid': r['message_sid']
        })
    
    return result

@app.post("/admin/resend_all_pending_views")
def admin_resend_all_pending_views():
    """
    Reenvía el template a TODOS los que no vieron (>7 días).
    """
    auth = require_admin()
    if auth:
        return auth
    
    token = _get_admin_token_from_request()
    tenant = (request.form.get("tenant") or "").strip().lower()
    period = (request.form.get("period") or "").strip()
    
    if not tenant or not period:
        return Response("Faltan parámetros", status=400)
    
    # Obtener lista de pendientes
    pending = get_pending_views_over_7days(tenant, period)
    
    if not pending:
        return redirect(f"/admin/seguimiento?tenant={tenant}&period={period}&msg=no_pending")
    
    # Encolar todos para envío automático
    base_url = request.host_url.rstrip('/')
    
    def process_in_background():
        import requests
        sent = 0
        failed = 0
        
        for p in pending:
            try:
                # Reenviar template
                envios = load_envios_rows(tenant)
                person = find_person_by_cuil(envios, p['cuil'])
                nombre = person.get("nombre", "") if person else ""
                
                sid = send_whatsapp_template(
                    p['whatsapp'],
                    content_vars={
                        "1": nombre or "Hola",
                        "2": period,
                    },
                    template_sid=TWILIO_TEMPLATE_SID or None,
                    status_callback=STATUS_CALLBACK_URL,
                )
                
                # Registrar
                save_template_sid(tenant, p['cuil'], period, p['whatsapp'], sid, nombre=nombre)
                
                # Actualizar pending_view
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute("""
                    UPDATE pending_views
                    SET origin = 'RESEND_MASS', created_at = %s
                    WHERE tenant = %s AND cuil = %s AND period = %s AND to_whatsapp = %s
                """, (int(time.time()), tenant, p['cuil'], period, p['whatsapp']))
                
                if cur.rowcount == 0:
                    add_pending_view(p['whatsapp'], tenant, p['cuil'], period, origin="RESEND_MASS")
                
                conn.commit()
                conn.close()
                
                sent += 1
                log.info(f"[RESEND_MASS] Enviado a {p['cuil']}: {sid}")
                
                # Pausa entre envíos
                time.sleep(1)
                
            except Exception as e:
                failed += 1
                log.exception(f"[RESEND_MASS] Error enviando a {p['cuil']}: {e}")
        
        log.info(f"[RESEND_MASS] Completado. Enviados: {sent}, Fallidos: {failed}")
    
    # Disparar en background
    thread = threading.Thread(target=process_in_background, daemon=True)
    thread.start()
    
    return redirect(f"/admin/seguimiento?tenant={tenant}&period={period}&msg=resend_started&total={len(pending)}")


@app.post("/admin/remind_all_pending_signatures")
def admin_remind_all_pending_signatures():
    """
    Envía recordatorio de firma a TODOS los que no firmaron (>7 días).
    """
    auth = require_admin()
    if auth:
        return auth
    
    token = _get_admin_token_from_request()
    tenant = (request.form.get("tenant") or "").strip().lower()
    period = (request.form.get("period") or "").strip()
    
    if not tenant or not period:
        return Response("Faltan parámetros", status=400)
    
    # Obtener lista de pendientes
    pending = get_pending_signatures_over_7days(tenant, period)
    
    if not pending:
        return redirect(f"/admin/seguimiento?tenant={tenant}&period={period}&msg=no_pending")
    
    # Verificar que exista el template de firma
    if not TWILIO_SIGN_TEMPLATE_SID:
        return Response("TWILIO_SIGN_TEMPLATE_SID no configurado", status=500)
    
    def process_in_background():
        sent = 0
        failed = 0
        
        for p in pending:
            try:
                # Enviar template de firma
                sid = send_whatsapp_template(
                    p['whatsapp'],
                    content_vars={
                        "1": p['nombre'] or "Hola",
                        "2": period,
                    },
                    template_sid=TWILIO_SIGN_TEMPLATE_SID,
                    status_callback=STATUS_CALLBACK_URL,
                )
                
                # Registrar
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute("""
                    INSERT INTO message_status (tenant, cuil, period, kind, message_sid, status, to_whatsapp, created_at)
                    VALUES (%s, %s, %s, 'REMIND_SIGN_MASS', %s, 'sent', %s, %s)
                """, (tenant, p['cuil'], period, sid, p['whatsapp'], int(time.time())))
                conn.commit()
                conn.close()
                
                sent += 1
                log.info(f"[REMIND_MASS] Enviado a {p['cuil']}: {sid}")
                
                # Pausa entre envíos
                time.sleep(1)
                
            except Exception as e:
                failed += 1
                log.exception(f"[REMIND_MASS] Error enviando a {p['cuil']}: {e}")
        
        log.info(f"[REMIND_MASS] Completado. Enviados: {sent}, Fallidos: {failed}")
    
    # Disparar en background
    thread = threading.Thread(target=process_in_background, daemon=True)
    thread.start()
    
    return redirect(f"/admin/seguimiento?tenant={tenant}&period={period}&msg=remind_started&total={len(pending)}")

def get_envios_df_for_tenant(tenant_slug: str, force: bool = False) -> pd.DataFrame:
    """
    Devuelve DataFrame del Excel de envíos de la empresa (tenant).
    Usa tu función existente load_envios_rows(tenant,...).
    """
    rows = load_envios_rows(tenant_slug, force=force) or []
    return pd.DataFrame(rows)

def get_estado_report(tenant: str, period: str | None = None):
    conn = get_db_connection()
    cur = conn.cursor()

    if period:
        cur.execute("""
            SELECT cuil, period, estado, updated_at
            FROM recibo_estado
            WHERE tenant = %s AND period = %s
            ORDER BY cuil
        """, (tenant, period))
    else:
        cur.execute("""
            SELECT cuil, period, estado, updated_at
            FROM recibo_estado
            WHERE tenant = %s
            ORDER BY period DESC, cuil
        """, (tenant,))

    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from flask import send_file
import io
import datetime


def _fmt_ts(ts: int | None) -> str:
    if not ts:
        return ""
    # tu server está en UTC, si querés BA: ajustá acá (UTC-3) o dejalo así
    return datetime.datetime.fromtimestamp(int(ts), datetime.timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

@app.route("/admin/reenviar_template", methods=["GET", "POST"])
def admin_reenviar_template():
    """
    Reenviar template INITIAL a personas específicas (override).
    """
    auth = require_admin()
    if auth:
        return auth
    
    if request.method == "POST":
        tenant = request.form.get("tenant", "").strip()
        period = request.form.get("period", "").strip()
        cuils_text = request.form.get("cuils", "").strip()
        
        # Parsear CUILs
        cuils = []
        for line in cuils_text.replace(",", "\n").split("\n"):
            cuil = norm_cuil(line.strip())
            if cuil and len(cuil) == 11:
                cuils.append(cuil)
        
        if not tenant or not period or not cuils:
            return Response("Falta tenant, period o CUILs", status=400)
        
        # Cargar envios
        envios = load_envios_rows(tenant)
        
        resultados = []
        for cuil in cuils:
            person = find_person_by_cuil(envios, cuil)
            if not person:
                resultados.append(f"❌ {cuil}: No encontrado en envíos")
                continue
            
            nombre = person.get('nombre', '')
            
            # Buscar WhatsApp
            whatsapp = person.get('to_whatsapp', '')
            if not whatsapp:
                whatsapp = person.get('telefono_raw', '')
            
            if not whatsapp:
                # Fallback a BD
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute("""
                    SELECT to_whatsapp FROM message_status
                    WHERE tenant = %s AND cuil = %s
                    ORDER BY created_at DESC LIMIT 1
                """, (tenant, cuil))
                row = cur.fetchone()
                conn.close()

                if row and row['to_whatsapp']:
                    whatsapp = row['to_whatsapp']
                else:
                    resultados.append(f"❌ {cuil} ({nombre}): Sin WhatsApp")
                    continue
            
            # Normalizar
            if not whatsapp.startswith('whatsapp:'):
                whatsapp = norm_whatsapp(whatsapp)
            
            # ✅ ENVIAR TEMPLATE (OVERRIDE - no verificar si ya existe)
            try:
                sid = send_whatsapp_template(
                    whatsapp,
                    content_vars={"1": (nombre or "Hola")},
                    template_sid=TWILIO_TEMPLATE_SID,
                    status_callback=STATUS_CALLBACK_URL,
                )
                
                # Guardar en BD
                save_template_sid(tenant, cuil, period, whatsapp, sid, nombre=nombre)
                add_pending_view(whatsapp, tenant, cuil, period, origin="INITIAL")
                
                resultados.append(f"✅ {cuil} ({nombre}): Template reenviado - SID: {sid}")
                
            except Exception as e:
                resultados.append(f"❌ {cuil} ({nombre}): Error - {str(e)}")
        
        # Mostrar resultados
        html = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Resultados reenvío template</title>
  <style>
    body {{ font-family: monospace; padding: 20px; background: #0f1629; color: #eaf0ff; }}
    .success {{ color: #10b981; }}
    .error {{ color: #ef4444; }}
    a {{ color: #5aa7ff; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    .result {{ padding: 8px; border-bottom: 1px solid #333; }}
  </style>
</head>
<body>
  <h1>📊 Resultados del reenvío de templates</h1>
  <p><strong>Tenant:</strong> {esc(tenant)}</p>
  <p><strong>Período:</strong> {esc(period)}</p>
  <p><strong>Total procesados:</strong> {len(resultados)}</p>
  <hr>
"""
        
        for r in resultados:
            css_class = "success" if "✅" in r else "error"
            html += f"<div class='result {css_class}'>{esc(r)}</div>"
        
        html += f"""
  <hr>
  <p><a href="/admin/reenviar_template">← Volver</a></p>
</body>
</html>
"""
        return Response(html, mimetype="text/html")
    
    # GET: Formulario
    html = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Reenviar Template</title>
  <style>
    body {{
      font-family: system-ui, sans-serif;
      background: #0f1629;
      color: #eaf0ff;
      padding: 20px;
      max-width: 800px;
      margin: 0 auto;
    }}
    h1 {{ color: #5aa7ff; }}
    .card {{
      background: rgba(255,255,255,0.05);
      border: 1px solid rgba(255,255,255,0.1);
      border-radius: 12px;
      padding: 24px;
      margin: 20px 0;
    }}
    label {{
      display: block;
      margin: 16px 0 8px 0;
      font-weight: 600;
      color: #9fb2d0;
    }}
    input, textarea {{
      width: 100%;
      padding: 12px;
      border: 1px solid rgba(255,255,255,0.2);
      border-radius: 8px;
      background: rgba(0,0,0,0.3);
      color: #eaf0ff;
      font-family: monospace;
      font-size: 14px;
    }}
    textarea {{
      min-height: 200px;
      resize: vertical;
    }}
    button {{
      background: linear-gradient(135deg, #F4C430, #d4a514);
      color: #1f2766;
      border: none;
      padding: 14px 32px;
      border-radius: 8px;
      font-weight: 700;
      font-size: 16px;
      cursor: pointer;
      margin-top: 20px;
    }}
    button:hover {{
      background: linear-gradient(135deg, #d4a514, #F4C430);
    }}
    .hint {{
      font-size: 13px;
      color: #9fb2d0;
      margin-top: 6px;
    }}
    .example {{
      background: rgba(0,0,0,0.3);
      padding: 12px;
      border-radius: 6px;
      font-family: monospace;
      font-size: 13px;
      margin-top: 8px;
    }}
    a {{
      color: #5aa7ff;
      text-decoration: none;
    }}
    a:hover {{
      text-decoration: underline;
    }}
    .warning {{
      background: rgba(251, 191, 36, 0.1);
      border: 1px solid rgba(251, 191, 36, 0.3);
      padding: 16px;
      border-radius: 8px;
      margin: 16px 0;
    }}
  </style>
</head>
<body>
  <h1>📤 Reenviar Template INITIAL</h1>
  
  <div class="card warning">
    <p><strong>⚠️ IMPORTANTE - Usar solo en emergencias:</strong></p>
    <ul>
      <li>Esto reenviará el template INITIAL a personas que YA lo recibieron</li>
      <li>Úsalo solo si borraron el chat y necesitan volver a iniciar</li>
      <li>NO verifica si ya tienen template activo (OVERRIDE)</li>
    </ul>
  </div>
  
  <form method="post">
    
    <div class="card">
      <label>Tenant</label>
      <input type="text" name="tenant" placeholder="san-patricio" required>
      
      <label>Período</label>
      <input type="text" name="period" placeholder="04/2026" required>
      <div class="hint">Formato: MM/YYYY</div>
      
      <label>CUILs (uno por línea o separados por coma)</label>
      <textarea name="cuils" placeholder="27357233831
27131464563
20-12345678-9" required></textarea>
      <div class="hint">Personas que borraron el chat y necesitan el template de nuevo</div>
      
      <div class="example">
        <strong>Ejemplo:</strong><br>
        27357233831<br>
        27131464563<br>
        20-12345678-9
      </div>
      
      <button type="submit">🚀 Reenviar Templates</button>
    </div>
  </form>
  
  <p><a href="/admin">← Volver al admin</a></p>
</body>
</html>
"""
    
    return Response(html, mimetype="text/html")


@app.get("/admin/verifications.xlsx")
@admin_required
def admin_verifications_xlsx():
    token = _get_admin_token_from_request()
    tenant = (request.args.get("tenant") or "").strip().lower()
    if not tenant:
        return Response("Falta tenant", status=400)

    rows = get_verifications_rows(tenant)  # debe leer de verifications
    import pandas as pd
    df = pd.DataFrame(rows or [])

    # orden y nombres
    if not df.empty:
        cols = [c for c in ["tenant","cuil","to_whatsapp","nombre","verified_at","updated_at"] if c in df.columns]
        df = df[cols]

    import io
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="verifications")
    out.seek(0)

    resp = Response(out.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp.headers["Content-Disposition"] = f'attachment; filename="verifications_{tenant}.xlsx"'
    return resp

@app.post("/admin/verifications_delete_bulk")
@admin_required
def admin_verifications_delete_bulk():
    token = _get_admin_token_from_request()
    tenant = (request.form.get("tenant") or "").strip().lower()
    keys = request.form.getlist("keys")

    if not tenant or not keys:
        return redirect(f"/admin/panel?tenant={tenant}&msg=verif_bulk_empty")

    conn = get_db_connection()
    cur = conn.cursor()

    n = 0
    for k in keys:
        try:
            cuil, to_whatsapp = k.split("|", 1)
        except Exception:
            continue
        cur.execute(
            "DELETE FROM verifications WHERE tenant=%s AND cuil=%s AND to_whatsapp=%s",
            (tenant, cuil, to_whatsapp)
        )
        n += cur.rowcount

    conn.commit()
    conn.close()

    return redirect(f"/admin/panel?tenant={tenant}&msg=verif_bulk_deleted&n={n}")



@app.get("/admin/report_estado.csv")
def admin_report_estado():
    auth = require_admin()
    if auth:
        return auth

    tenant = request.args.get("tenant", "").strip().lower()
    period = request.args.get("period")

    rows = get_estado_report(tenant, period)

    def generate():
        yield "cuil,period,estado,updated_at\n"
        for r in rows:
            yield f"{r['cuil']},{r['period']},{r['estado']},{r['updated_at']}\n"

    return Response(
        generate(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=estado_{tenant}.csv"
        }
    )

def get_sent_pdfs_report(tenant: str):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT cuil, period, to_whatsapp, message_sid, created_at, sign_sent_at
        FROM sent_pdfs
        WHERE tenant = %s
        ORDER BY created_at DESC
    """, (tenant,))
    rows = cur.fetchall()
    conn.close()
    return [dict(r) for r in rows]



@app.get("/admin/report_envios.csv")
def admin_report_envios():
    auth = require_admin()
    if auth:
        return auth

    tenant = request.args.get("tenant", "").strip().lower()
    rows = get_sent_pdfs_report(tenant)

    def generate():
        yield "cuil,period,whatsapp,message_sid,created_at,sign_sent_at\n"
        for r in rows:
            yield f"{r['cuil']},{r['period']},{r['to_whatsapp']},{r['message_sid']},{r['created_at']},{r['sign_sent_at'] or ''}\n"

    return Response(
        generate(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=envios_{tenant}.csv"
        }
    )

def _period_variants(period: str) -> list[str]:
    # period esperado: "01/2026" (o "01-2026")
    p = (period or "").strip()
    p = p.replace("-", "/").replace("_", "/").replace(".", "/").replace(" ", "/")
    # si viene "012026" intentar normalizar
    if len(p) == 6 and p.isdigit():
        p = f"{p[:2]}/{p[2:]}"
    if "/" not in p:
        return [p]

    mm, yyyy = p.split("/", 1)
    mm = mm.zfill(2)
    yyyy = yyyy.strip()

    return [
        f"{mm}/{yyyy}",
        f"{mm}-{yyyy}",
        f"{mm}_{yyyy}",
        f"{mm} {yyyy}",
        f"{mm}.{yyyy}",
        f"{mm}{yyyy}",
    ]


import re
from urllib.parse import quote

FOLDER_MIME = "application/vnd.google-apps.folder"

def period_folder_to_label(folder_name: str) -> str:
    # '01-2026' -> '01/2026'
    m = re.match(r"^(\d{2})-(\d{4})$", (folder_name or "").strip())
    if not m:
        return ""
    return f"{m.group(1)}/{m.group(2)}"

def normalize_period_label(s: str) -> str:
    # '1/2026' -> '01/2026' ; '01/2026' queda igual
    m = re.match(r"^\s*(\d{1,2})\s*/\s*(\d{4})\s*$", (s or "").strip())
    if not m:
        return (s or "").strip()
    mm = int(m.group(1))
    yyyy = int(m.group(2))
    return f"{mm:02d}/{yyyy:04d}"

import re

def label_to_period_folder(label: str) -> str:
    # "01/2026" -> "01-2026"
    label = (label or "").strip()
    m = re.match(r"^(\d{2})/(\d{4})$", label)
    if not m:
        return ""
    return f"{m.group(1)}-{m.group(2)}"


def _drive_find_child_folder_id(service, parent_id: str, folder_name: str) -> str:
    """
    Busca una carpeta hija por nombre exacto y devuelve su id (o "").
    Usa tu helper _drive_list_children.
    """
    children = _drive_list_children(
        service,
        parent_id=parent_id,
        mime_type=FOLDER_MIME,
        page_size=500
    )
    for c in children:
        if (c.get("name") or "").strip() == folder_name:
            return (c.get("id") or "").strip()
    return ""

##############################################################
def _drive_ensure_child_folder_shared(service, parent_id: str, folder_name: str) -> str:
    """
    Igual que _drive_ensure_child_folder pero con soporte de Shared Drives.
    Busca la carpeta hija por nombre; si no existe, la crea. Devuelve su id.
    """
    # Buscar (con flags de shared drive)
    safe_name = folder_name.replace("'", "\\'")
    q = (
        f"'{parent_id}' in parents "
        f"and name = '{safe_name}' "
        f"and mimeType = 'application/vnd.google-apps.folder' "
        f"and trashed = false"
    )
    res = service.files().list(
        q=q,
        fields="files(id,name)",
        pageSize=5,
        supportsAllDrives=True,
        includeItemsFromAllDrives=True,
    ).execute()
    files = res.get("files", [])
    if files:
        return (files[0].get("id") or "").strip()
 
    # Crear (con soporte shared drive)
    metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id],
    }
    folder = service.files().create(
        body=metadata,
        fields="id",
        supportsAllDrives=True,
    ).execute()
    return (folder.get("id") or "").strip()
 
CERTIFICADOS_DRIVE_ID = os.environ.get("CERTIFICADOS_SHARED_DRIVE_ID", "").strip()
 
 
# Nombre de la subcarpeta de Drive según el tipo de certificado.
CERT_SUBCARPETA = {"medico": "Medicos", "familia": "Familia"}

def get_certificados_folder_id(tenant_slug: str, tipo: str = "medico") -> str:
    """
    Devuelve (creando si hace falta) el id de la subcarpeta del tipo de
    certificado, DENTRO de la carpeta del tenant, DENTRO del Shared Drive:
        <SharedDrive>/<tenant>/<Medicos|Familia>/
    """
    if not CERTIFICADOS_DRIVE_ID:
        raise RuntimeError("Falta CERTIFICADOS_SHARED_DRIVE_ID en ENV")

    sub = CERT_SUBCARPETA.get(tipo, "Medicos")
    service = drive_service()
    # En un Shared Drive, el ID del drive funciona como id de la carpeta raíz.
    tenant_folder = _drive_ensure_child_folder_shared(service, CERTIFICADOS_DRIVE_ID, tenant_slug)
    return _drive_ensure_child_folder_shared(service, tenant_folder, sub)
 
 
def upload_certificado_to_drive(tenant: str, cuil: str, data: bytes, content_type: str, tipo: str = "medico") -> tuple[str, str]:
    """
    Sube el certificado a <SharedDrive>/<tenant>/<Medicos|Familia>/<cuil>_<fecha>_<hora>.<ext>
    Devuelve (file_id, file_name).
    """
    folder_id = get_certificados_folder_id(tenant, tipo)
    if not folder_id:
        raise RuntimeError(f"No se pudo obtener/crear carpeta certificados para tenant={tenant} tipo={tipo}")
 
    # extensión según content_type
    if "pdf" in content_type:
        ext = "pdf"
    elif "jpeg" in content_type or "jpg" in content_type:
        ext = "jpg"
    elif "png" in content_type:
        ext = "png"
    else:
        ext = "bin"
 
    now = time.localtime()
    fecha = time.strftime("%Y-%m-%d", now)
    hora = time.strftime("%H%M%S", now)
    file_name = f"{cuil}_{fecha}_{hora}.{ext}"
 
    service = drive_service()
    media = MediaIoBaseUpload(io.BytesIO(data), mimetype=content_type, resumable=False)
    metadata = {"name": file_name, "parents": [folder_id]}
    f = service.files().create(
        body=metadata,
        media_body=media,
        fields="id",
        supportsAllDrives=True,   # 👈 clave para Shared Drive
    ).execute()
    file_id = (f.get("id") or "").strip()
    return file_id, file_name
 
 
 
def download_twilio_media(media_url: str) -> tuple[bytes, str]:
    """
    Descarga un archivo de media de Twilio (requiere auth con las credenciales).
    Devuelve (bytes, content_type).
    """
    r = requests.get(
        media_url,
        auth=(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN),
        timeout=30,
    )
    r.raise_for_status()
    content_type = r.headers.get("Content-Type", "application/octet-stream")
    return r.content, content_type
 
def save_certificado(tenant: str, cuil: str, nombre: str, to_whatsapp: str,
                     file_id: str, file_name: str, mime_type: str, tipo: str = "medico"):
    """Registra el certificado en la BD."""
    conn = get_db_connection()
    cur = conn.cursor()
    now = int(time.time())
    cur.execute("""
        INSERT INTO certificados (tenant, cuil, nombre, to_whatsapp, file_id, file_name, mime_type, created_at, tipo)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, (tenant, cuil, nombre, to_whatsapp, file_id, file_name, mime_type, now, tipo))
    conn.commit()
    conn.close()
    log.info(f"✅ Certificado guardado: {tenant}/{cuil} tipo={tipo} -> {file_name} (file_id={file_id})")
 
 
def get_certificados(tenant: str, limit: int = 500, tipo: str = "medico") -> list[dict]:
    """Lista los certificados de un tenant y tipo, más recientes primero."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, tenant, cuil, nombre, to_whatsapp, file_id, file_name, mime_type, created_at, tipo
        FROM certificados
        WHERE tenant = %s AND COALESCE(tipo, 'medico') = %s
        ORDER BY created_at DESC
        LIMIT %s
    """, (tenant, tipo, limit))
    rows = cur.fetchall()
    conn.close()
    result = []
    for r in rows:
        try:
            result.append(dict(r))
        except Exception:
            keys = ["id","tenant","cuil","nombre","to_whatsapp","file_id","file_name","mime_type","created_at"]
            result.append(dict(zip(keys, r)))
    return result
 
##############################################################
def _drive_child_file_exists(service, parent_id: str, filename: str) -> bool:
    """
    True si existe un archivo con nombre exacto dentro de parent_id.
    Para no listar 500 siempre, podés implementar query directa,
    pero con tus helpers lo hacemos por listado.
    """
    children = _drive_list_children(
        service,
        parent_id=parent_id,
        mime_type=None,   # archivos (no filtramos por mime)
        page_size=500
    )
    for c in children:
        if (c.get("name") or "").strip() == filename:
            return True
    return False



def get_tenant_period_folder_id(tenant_slug: str, period_label: str) -> str:
    """
    Devuelve el folder_id del período dentro del root del tenant.
    period_label esperado: "MM/YYYY"
    """
    t = get_tenant(tenant_slug)
    if not t:
        return ""

    root_id = (t.get("drive_root_id") or t.get("recibos_root_id") or "").strip()
    if not root_id:
        return ""

    period_folder_name = label_to_period_folder(period_label)  # "MM-YYYY"
    if not period_folder_name:
        return ""

    service = drive_service()
    return _drive_find_child_folder_id(service, root_id, period_folder_name)


import re


def list_tenant_period_folders(tenant_slug: str) -> list[str]:
    """
    Devuelve nombres de carpetas período existentes en Drive para el tenant.
    Formato carpeta esperado: 'MM-AAAA'
    """
    t = get_tenant(tenant_slug)
    if not t:
        return []

    root_id = (t.get("drive_root_id") or t.get("recibos_root_id") or "").strip()
    if not root_id:
        return []

    service = drive_service()

    children = _drive_list_children(
        service,
        parent_id=root_id,
        mime_type=FOLDER_MIME,
        page_size=500
    )

    names = [
        c.get("name", "")
        for c in children
        if re.match(r"^\d{2}-\d{4}$", (c.get("name", "") or "").strip())
    ]

    # ordenar desc por (yyyy, mm)
    def key(name: str):
        mm, yyyy = name.split("-")
        return (int(yyyy), int(mm))

    names.sort(key=key, reverse=True)
    return names


def list_tenant_period_labels(service, tenant: str) -> list[str]:
    # ['01-2026', '12-2025'] -> ['01/2026','12/2025']
    folders = list_tenant_period_folders(tenant)
    labels = []
    for f in folders:
        lbl = period_folder_to_label(f)
        if lbl:
            labels.append(lbl)
    return labels



def find_pdf_file_id(tenant: str, cuil: str, period: str) -> str | None:
    """
    Busca el PDF {cuil}.pdf en el root del tenant.
    (Misma lógica que /media/pdf)
    """
    t = get_tenant(tenant)
    if not t:
        log.error("%s %s", "❌ tenant inválido:", tenant)
        return None

    root_id = (t.get("drive_root_id") or t.get("recibos_root_id") or "").strip()
    log.info("%s %s", "ROOT_ID:", root_id)

    if not root_id:
        log.error("%s %s %s", "❌ tenant sin drive_root_id/recibos_root_id:", tenant, t)
        return None

    cuil = strip_pdf(cuil).strip()
    filename = f"{cuil}.pdf"

    service = drive_service()

    q = f"'{root_id}' in parents and trashed=false and name='{filename}'"
    res = service.files().list(q=q, fields="files(id,name)", pageSize=5).execute()
    files = res.get("files", [])

    # debug útil
    if not files:
        log.info("%s %s %s %s", "🔎 NO ENCONTRÉ:", filename, "en root:", root_id)

    return files[0]["id"] if files else None



@app.post("/twilio/webhook")
def twilio_webhook():
    return twilio_inbound()

@app.post("/admin/reset_tenant")
@admin_required
def admin_reset_tenant():
    token = _get_admin_token_from_request()
    tenant = (request.form.get("tenant") or "").strip().lower()
    period_raw = (request.form.get("period") or "").strip()

    if not tenant:
        return Response("Falta tenant", status=400)

    # Siempre limpiar pendings del tenant
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM pending_views WHERE tenant=%s;", (tenant,))
    deleted_pending = cur.rowcount
    log.info(f"RESET: DELETE FROM pending_views WHERE tenant=%s; args= ({tenant},) deleted= {deleted_pending}")

    if period_raw:
        p_norm = norm_period_label(period_raw)  # MM/AAAA
        # si no se pudo normalizar, usamos el raw igual
        candidates = set([period_raw])

        if p_norm:
            candidates.add(p_norm)
            candidates.add(p_norm.replace("/", "-"))  # MM-AAAA (legacy)

        # También cubrimos el caso inverso: si mandan MM-AAAA, generamos MM/AAAA
        if "-" in period_raw and not p_norm:
            # por las dudas
            candidates.add(period_raw.replace("-", "/"))

        # borramos en todas las tablas por cada candidato
        for p in candidates:
            cur.execute("DELETE FROM recibo_estado WHERE tenant=%s AND period=%s;", (tenant, p))
            log.info(f"RESET: DELETE FROM recibo_estado WHERE tenant=%s AND period=%s; args= {(tenant, p)} deleted= {cur.rowcount}")
            
            cur.execute("DELETE FROM message_status WHERE tenant=%s AND period=%s;", (tenant, p))
            log.info(f"RESET: DELETE FROM message_status WHERE tenant=%s AND period=%s; args= {(tenant, p)} deleted= {cur.rowcount}")
            
            cur.execute("DELETE FROM sent_pdfs WHERE tenant=%s AND period=%s;", (tenant, p))
            log.info(f"RESET: DELETE FROM sent_pdfs WHERE tenant=%s AND period=%s; args= {(tenant, p)} deleted= {cur.rowcount}")
            
            cur.execute("DELETE FROM receipt_request_events WHERE tenant=%s AND period=%s;", (tenant, p))
            log.info(f"RESET: DELETE FROM receipt_request_events WHERE tenant=%s AND period=%s; args= {(tenant, p)} deleted= {cur.rowcount}")
            
            cur.execute("DELETE FROM receipt_requests WHERE tenant=%s AND period=%s;", (tenant, p))
            log.info(f"RESET: DELETE FROM receipt_requests WHERE tenant=%s AND period=%s; args= {(tenant, p)} deleted= {cur.rowcount}")
            
            cur.execute("DELETE FROM template_send_queue WHERE tenant=%s AND period=%s;", (tenant, p))
            log.info(f"RESET: DELETE FROM template_send_queue WHERE tenant=%s AND period=%s; args= {(tenant, p)} deleted= {cur.rowcount}")
            
            # ✅ NUEVO: Borrar pending_terms de este tenant/período
            cur.execute("DELETE FROM pending_terms WHERE tenant=%s AND period=%s;", (tenant, p))
            log.info(f"RESET: DELETE FROM pending_terms WHERE tenant=%s AND period=%s; args= {(tenant, p)} deleted= {cur.rowcount}")
        
        # ✅ NUEVO: Borrar terms_accepted de usuarios de este tenant/período
        cur.execute("""
            DELETE FROM terms_accepted 
            WHERE whatsapp IN (
                SELECT DISTINCT to_whatsapp 
                FROM message_status 
                WHERE tenant = %s AND period IN ({})
            )
        """.format(','.join(['%s'] * len(candidates))), (tenant, *candidates))
        log.info(f"RESET: DELETE FROM terms_accepted (tenant={tenant}, periods={candidates}); deleted= {cur.rowcount}")
        
    else:
        # Sin período: borrar todo del tenant
        cur.execute("DELETE FROM recibo_estado WHERE tenant=%s;", (tenant,))
        log.info(f"RESET: DELETE FROM recibo_estado WHERE tenant=%s; args= ({tenant},) deleted= {cur.rowcount}")
        
        cur.execute("DELETE FROM message_status WHERE tenant=%s;", (tenant,))
        log.info(f"RESET: DELETE FROM message_status WHERE tenant=%s; args= ({tenant},) deleted= {cur.rowcount}")
        
        cur.execute("DELETE FROM sent_pdfs WHERE tenant=%s;", (tenant,))
        log.info(f"RESET: DELETE FROM sent_pdfs WHERE tenant=%s; args= ({tenant},) deleted= {cur.rowcount}")
        
        cur.execute("DELETE FROM receipt_request_events WHERE tenant=%s;", (tenant,))
        log.info(f"RESET: DELETE FROM receipt_request_events WHERE tenant=%s; args= ({tenant},) deleted= {cur.rowcount}")
        
        cur.execute("DELETE FROM receipt_requests WHERE tenant=%s;", (tenant,))
        log.info(f"RESET: DELETE FROM receipt_requests WHERE tenant=%s; args= ({tenant},) deleted= {cur.rowcount}")
        
        cur.execute("DELETE FROM template_send_queue WHERE tenant=%s;", (tenant,))
        log.info(f"RESET: DELETE FROM template_send_queue WHERE tenant=%s; args= ({tenant},) deleted= {cur.rowcount}")
        
        # ✅ NUEVO: Borrar pending_terms de este tenant (sin período)
        cur.execute("DELETE FROM pending_terms WHERE tenant=%s;", (tenant,))
        log.info(f"RESET: DELETE FROM pending_terms WHERE tenant=%s; args= ({tenant},) deleted= {cur.rowcount}")
        
        # ✅ NUEVO: Borrar terms_accepted de usuarios de este tenant (sin período)
        cur.execute("""
            DELETE FROM terms_accepted 
            WHERE whatsapp IN (
                SELECT DISTINCT to_whatsapp 
                FROM message_status 
                WHERE tenant = %s
            )
        """, (tenant,))
        log.info(f"RESET: DELETE FROM terms_accepted (tenant={tenant}, all periods); deleted= {cur.rowcount}")

    conn.commit()
    conn.close()

    # devolvemos SIEMPRE el normalizado para que el panel quede prolijo
    p_show = norm_period_label(period_raw) if period_raw else ""
    return redirect(f"/admin/panel?tenant={tenant}&msg=reset_ok&period={p_show or period_raw}")

from flask import redirect

from flask import redirect

def _period_candidates(period_raw: str) -> list[str]:
    period_raw = (period_raw or "").strip()
    if not period_raw:
        return []
    p_norm = norm_period_label(period_raw)  # MM/AAAA o ""
    s = {period_raw}
    if p_norm:
        s.add(p_norm)
        s.add(p_norm.replace("/", "-"))
    else:
        # si no normaliza, probamos el cambio simple
        s.add(period_raw.replace("-", "/"))
        s.add(period_raw.replace("/", "-"))
    return sorted(x for x in s if x)

@app.post("/admin/reset")
@admin_required
def admin_reset():
    token = _get_admin_token_from_request()
    tenant = (request.form.get("tenant") or "").strip().lower()
    period_raw = (request.form.get("period") or "").strip()

    if not tenant:
        return Response("Falta tenant", status=400)

    periods = _period_candidates(period_raw)

    conn = get_db_connection()
    cur = conn.cursor()

    def _del(sql, args):
        cur.execute(sql, args)
        log.info("%s %s %s %s %s %s", "RESET:", sql.split("\n")[0][:80], "args=", args, "deleted=", cur.rowcount)

    _del("DELETE FROM pending_views WHERE tenant=%s;", (tenant,))

    if periods:
        for p in periods:
            _del("DELETE FROM recibo_estado WHERE tenant=%s AND period=%s;", (tenant, p))
            _del("DELETE FROM message_status WHERE tenant=%s AND period=%s;", (tenant, p))
            _del("DELETE FROM sent_pdfs WHERE tenant=%s AND period=%s;", (tenant, p))
            _del("DELETE FROM receipt_request_events WHERE tenant=%s AND period=%s;", (tenant, p))
            _del("DELETE FROM receipt_requests WHERE tenant=%s AND period=%s;", (tenant, p))
            _del("DELETE FROM template_send_queue WHERE tenant=%s AND period=%s;", (tenant, p))
            _del("DELETE FROM pending_terms WHERE tenant=%s AND period=%s;", (tenant, p))  # ✅ NUEVO
        
        # ✅ NUEVO: Borrar terms_accepted de usuarios de este tenant/período
        placeholders = ','.join('%s' for _ in periods)
        cur.execute(f"""
            DELETE FROM terms_accepted 
            WHERE whatsapp IN (
                SELECT DISTINCT to_whatsapp 
                FROM message_status 
                WHERE tenant = %s AND period IN ({placeholders})
            )
        """, (tenant, *periods))
        log.info(f"RESET: DELETE FROM terms_accepted (tenant={tenant}, periods={periods}); deleted= {cur.rowcount}")
        
    else:
        _del("DELETE FROM recibo_estado WHERE tenant=%s;", (tenant,))
        _del("DELETE FROM message_status WHERE tenant=%s;", (tenant,))
        _del("DELETE FROM sent_pdfs WHERE tenant=%s;", (tenant,))
        _del("DELETE FROM receipt_request_events WHERE tenant=%s;", (tenant,))
        _del("DELETE FROM receipt_requests WHERE tenant=%s;", (tenant,))
        _del("DELETE FROM template_send_queue WHERE tenant=%s;", (tenant,))
        _del("DELETE FROM pending_terms WHERE tenant=%s;", (tenant,))  # ✅ NUEVO
        
        # ✅ NUEVO: Borrar terms_accepted de usuarios de este tenant
        cur.execute("""
            DELETE FROM terms_accepted 
            WHERE whatsapp IN (
                SELECT DISTINCT to_whatsapp 
                FROM message_status 
                WHERE tenant = %s
            )
        """, (tenant,))
        log.info(f"RESET: DELETE FROM terms_accepted (tenant={tenant}, all periods); deleted= {cur.rowcount}")

    conn.commit()
    conn.close()

    p_show = norm_period_label(period_raw) if period_raw else ""
    url = f"/admin/panel?tenant={tenant}&msg=reset_ok"
    if p_show:
        url += f"&period={p_show}"
    return redirect(url)

def queue_template_send(tenant: str, period: str, to_whatsapp: str, cuil: str,
                        nombre: str = "", require_pdf: bool = True) -> str:
    """
    Devuelve: 'inserted', 'requeued', 'noop'
    - inserted: no existía
    - requeued: existía SKIPPED/FAILED y la volvimos a PENDING
    - noop: existía y estaba PENDING/SENT (no tocamos)
    """
    now = int(time.time())
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
      INSERT INTO template_send_queue
        (tenant, period, to_whatsapp, cuil, nombre, require_pdf, status, created_at, updated_at, error)
      VALUES
        (%s, %s, %s, %s, %s, %s, 'PENDING', %s, %s, '')
      ON CONFLICT(tenant, period, to_whatsapp, cuil) DO UPDATE SET
        nombre=excluded.nombre,
        require_pdf=excluded.require_pdf,
        updated_at=excluded.updated_at,
        status=CASE
          WHEN template_send_queue.status IN ('SKIPPED','FAILED') THEN 'PENDING'
          ELSE template_send_queue.status
        END,
        error=CASE
          WHEN template_send_queue.status IN ('SKIPPED','FAILED') THEN ''
          ELSE template_send_queue.error
        END
    """, (tenant, period, to_whatsapp, cuil, (nombre or ""), bool(require_pdf), now, now))

    # Determinar resultado
    cur.execute("""
      SELECT status FROM template_send_queue
      WHERE tenant=%s AND period=%s AND to_whatsapp=%s AND cuil=%s
      LIMIT 1
    """, (tenant, period, to_whatsapp, cuil))
    _row = cur.fetchone()
    status = _row['status'] if _row else ""

    conn.commit()
    conn.close()

    # heurística simple
    if status == "PENDING":
        # pudo ser inserted o requeued; si querés exactitud, guardá rowcount antes/después
        return "requeued_or_inserted"
    return "noop"


def count_queue_status(tenant: str, period: str) -> dict:
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
      SELECT status, COUNT(*) as n
      FROM template_send_queue
      WHERE tenant=%s AND period=%s
      GROUP BY status
    """, (tenant, period))
    d = {r['status']: r['n'] for r in cur.fetchall()}
    conn.close()
    return d


@app.post("/admin/send_template_queue_start")
@admin_required
def admin_send_template_queue_start():
    token = _get_admin_token_from_request()

    tenant = (request.form.get("tenant") or "").strip().lower()
    period = (request.form.get("period") or "").strip()
    require_pdf = (request.form.get("require_pdf") or "true").lower() in ("1", "true", "yes", "on")

    if not tenant:
        return Response("Falta tenant", status=400)
    if not period:
        return Response("Falta period", status=400)

    df = get_envios_df_for_tenant(tenant)
    if df is None or df.empty:
        return Response("Excel de envíos vacío o no encontrado", status=400)

    df.columns = [str(c).strip().lower() for c in df.columns]

    def pick(*names):
        for n in names:
            if n in df.columns:
                return n
        return None

    c_nombre = pick("nombre", "name", "empleado", "persona")
    c_tel = pick("telefono", "tel", "celular", "whatsapp", "numero")
    c_arch = pick("archivo", "cuil", "archivo_norm")

    if not c_tel or not c_arch:
        return Response("El Excel debe tener columnas telefono y archivo (cuil).", status=400)

    rows = df.to_dict(orient="records")

    try:
        limit = int(request.form.get("limit") or "0")
    except ValueError:
        limit = 0
    if limit > 0:
        rows = rows[:limit]

    enqueued = 0
    skipped_bad = 0
    skipped_dup = 0

    for r in rows:
        nombre = str(r.get(c_nombre, "")).strip() if c_nombre else ""
        tel_raw = str(r.get(c_tel, "")).strip()
        arch_raw = str(r.get(c_arch, "")).strip()

        if not tel_raw or not arch_raw:
            skipped_bad += 1
            continue

        to_whatsapp = normalize_whatsapp(tel_raw)
        if not to_whatsapp:
            skipped_bad += 1
            continue

        # CUIL desde archivo
        try:
            cuil = strip_pdf(arch_raw)
        except Exception:
            skipped_bad += 1
            continue

        cuil_digits = norm_digits(cuil)
        if len(cuil_digits) != 11:
            skipped_bad += 1
            continue
        cuil = cuil_digits

        ok = queue_template_send(
            tenant=tenant,
            period=period,
            to_whatsapp=to_whatsapp,
            cuil=cuil,
            nombre=nombre,
            require_pdf=require_pdf,
        )
        if ok:
            enqueued += 1
        else:
            skipped_dup += 1

    stats = count_queue_status(tenant, period)
    return redirect(
        f"/admin/panel?tenant={tenant}&msg=queue_enqueued"
        f"&period={period}&enqueued={enqueued}&dup={skipped_dup}&bad={skipped_bad}"
        f"&pending={stats.get('PENDING',0)}&sent={stats.get('SENT',0)}"
        f"&failed={stats.get('FAILED',0)}&skipped={stats.get('SKIPPED',0)}"
    )

def _fetch_queue_batch(tenant: str, period: str, batch_size: int = 10) -> list[dict]:
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
      SELECT id, tenant, period, to_whatsapp, cuil, nombre, require_pdf
      FROM template_send_queue
      WHERE tenant=%s AND period=%s AND status='PENDING'
      ORDER BY id ASC
      LIMIT %s
    """, (tenant, period, batch_size))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows


def _mark_queue_row(row_id: int, status: str, error: str = "", sent_sid: str | None = None):
    now = int(time.time())
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
      UPDATE template_send_queue
      SET status=%s, error=%s, updated_at=%s, sent_sid=COALESCE(%s, sent_sid),
          sent_at=CASE WHEN %s IS NOT NULL THEN %s ELSE sent_at END
      WHERE id=%s
    """, (status, (error or ""), now, sent_sid, sent_sid, now, row_id))
    conn.commit()
    conn.close()

def notify_billing_batch_done(tenant: str, period: str, enviados: int) -> None:
    """
    Avisa por mail (1 sola vez por tanda) que terminó el envío INITIAL de un período.
    El candado en template_batch_notice evita duplicar el aviso si el tick se vuelve
    a ejecutar con la cola ya vacía.
    """
    if not BILLING_NOTIFY_EMAIL:
        return

    now = int(time.time())

    conn = get_db_connection()
    cur = conn.cursor()

    # Self-healing: asegura la tabla candado si init_db no la creó todavía.
    cur.execute("""
        CREATE TABLE IF NOT EXISTS template_batch_notice (
            tenant TEXT NOT NULL,
            period TEXT NOT NULL,
            enviados INTEGER,
            notified_at BIGINT NOT NULL,
            PRIMARY KEY (tenant, period)
        )
    """)

    # Claim atómico: si la fila ya existe, no reenviamos.
    cur.execute("""
        INSERT INTO template_batch_notice (tenant, period, enviados, notified_at)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (tenant, period) DO NOTHING
    """, (tenant, period, enviados, now))
    claimed = (cur.rowcount == 1)
    conn.commit()
    conn.close()

    if not claimed:
        return  # ya se había avisado para esta tanda

    t = get_tenant(tenant)
    empresa = (t.get("display_name") if t else tenant) or tenant
    momento = ts_str(now)

    subject = f"[Facturación] Envío completado · {empresa} · {period}"
    html_body = f"""
    <div style="font-family:system-ui,Arial,sans-serif;font-size:15px;color:#1a1a1a">
      <h2 style="margin:0 0 12px">Envío de recibos completado</h2>
      <table cellpadding="6" style="border-collapse:collapse">
        <tr><td><strong>Empresa</strong></td><td>{esc(empresa)}</td></tr>
        <tr><td><strong>Período</strong></td><td>{esc(period)}</td></tr>
        <tr><td><strong>Momento</strong></td><td>{esc(momento)}</td></tr>
        <tr><td><strong>Mensajes enviados</strong></td><td>{enviados}</td></tr>
      </table>
    </div>
    """
    try:
        send_email(BILLING_NOTIFY_EMAIL, subject, html_body)
        log.info(f"📧 Aviso de facturación enviado: {empresa} {period} ({enviados} msgs)")
    except Exception as e:
        log.exception(f"No se pudo enviar el aviso de facturación: {e}")
        
@app.post("/admin/send_template_queue_tick")
def admin_send_template_queue_tick():
    # Único endpoint del panel que acepta token por URL/form: lo usan el cron
    # (query string) y el self-call del envío automático (form). También valen
    # la sesión (humano) y el header X-Admin-Token (otras automatizaciones).
    if not (admin_authenticated() or admin_token_valid()):
        return Response("Unauthorized", status=401)

    token = _get_admin_token_from_request()

    tenant = (request.form.get("tenant") or request.args.get("tenant") or "").strip().lower()
    period = (request.form.get("period") or request.args.get("period") or "").strip()

    try:
        batch_size = int(request.form.get("batch_size") or request.args.get("batch_size") or "10")
    except ValueError:
        batch_size = 10
    batch_size = max(1, min(batch_size, 50))

    if not tenant:
        return Response("Falta tenant", status=400)
    if not period:
        return Response("Falta period", status=400)

    # ✅ NUEVO: Recibir cache si send_auto lo pasó
    use_cache = (request.form.get("use_cache") or request.args.get("use_cache") or "").lower() == "true"
    pdf_cache = {}
    if use_cache:
        pdf_cache = precache_pdfs_for_period(tenant, period)

    rows = _fetch_queue_batch(tenant, period, batch_size=batch_size)

    processed = 0
    sent = 0
    skipped = 0
    failed = 0

    for r in rows:
        processed += 1
        row_id = r["id"]
        to_whatsapp = (r["to_whatsapp"] or "").strip()
        cuil = (r["cuil"] or "").strip()
        nombre = (r.get("nombre") or "").strip()
        require_pdf = bool(r.get("require_pdf", True))

        try:
            # 1) Si require_pdf, validamos que exista PDF
            if require_pdf:
                # ✅ NUEVO: Usar cache si existe
                if use_cache and cuil in pdf_cache:
                    pdf_file_id = pdf_cache[cuil]
                else:
                    pdf_file_id = find_pdf_with_retry(tenant, cuil, period)
                
                if not pdf_file_id:
                    _mark_queue_row(row_id, "SKIPPED", error="NO_PDF")
                    skipped += 1
                    continue

            # 2) Si ya lo mandamos, skip
            if already_sent_template(tenant, cuil, period, to_whatsapp):
                _mark_queue_row(row_id, "SKIPPED", error="ALREADY_SENT")
                skipped += 1
                continue

            # 3) Enviar template
            sid = send_whatsapp_template(
                to_whatsapp,
                content_vars={"1": (nombre or "Hola")},
                template_sid=TWILIO_TEMPLATE_SID,
                status_callback=STATUS_CALLBACK_URL,
            )

            # 4) Persistir
            save_template_sid(tenant, cuil, period, to_whatsapp, sid, nombre=nombre)
            add_pending_view(to_whatsapp, tenant, cuil, period, origin="INITIAL")

            _mark_queue_row(row_id, "SENT", sent_sid=sid)
            sent += 1

        except Exception as e:
            _mark_queue_row(row_id, "FAILED", error=str(e)[:250])
            failed += 1

    stats = count_queue_status(tenant, period)

    # 🧾 Facturación: si la cola quedó vacía, avisar (1 sola vez) que terminó la tanda
    if stats.get("PENDING", 0) == 0 and stats.get("SENT", 0) > 0:
        notify_billing_batch_done(tenant, period, stats.get("SENT", 0))

    if (request.form.get("mode") or request.args.get("mode") or "").lower() == "json":
        return {
            "tenant": tenant,
            "period": period,
            "processed": processed,
            "sent": sent,
            "skipped": skipped,
            "failed": failed,
            "stats": stats,
        }

    return redirect(
        f"/admin/panel?tenant={tenant}&msg=queue_tick"
        f"&period={period}&processed={processed}&sent={sent}&skipped={skipped}&failed={failed}"
        f"&pending={stats.get('PENDING',0)}"
    )

@app.post("/admin/send_auto")
@admin_required
def admin_send_auto():
    """
    Procesa la cola automáticamente en background usando threading.
    Versión simple sin Celery - gratis pero con límite de ~30 minutos.
    """
    # Token para el self-call de background a /send_template_queue_tick.
    # Si el admin disparó esto por sesión (sin token en el form), usamos el ADMIN_TOKEN real:
    # el thread no tiene sesión, así que necesita autenticarse por token.
    token = _get_admin_token_from_request() or ADMIN_TOKEN
    
    tenant = (request.form.get("tenant") or "").strip().lower()
    period = (request.form.get("period") or "").strip()
    
    try:
        batch_size = int(request.form.get("batch_size") or "10")
    except ValueError:
        batch_size = 10
    batch_size = max(1, min(batch_size, 50))
    
    if not tenant:
        return Response("Falta tenant", status=400)
    if not period:
        return Response("Falta period", status=400)
    
    # ✅ CAPTURAR LA URL ANTES DEL THREAD
    base_url = request.host_url.rstrip('/')
    
    # Función que corre en el thread
    def process_in_background():
        import requests
        
        # ✅ PRE-CACHE AL INICIO
        pdf_cache = precache_pdfs_for_period(tenant, period)
        
        processed_total = 0
        sent_total = 0
        iterations = 0
        max_iterations = 200
        
        while iterations < max_iterations:
            iterations += 1
            
            try:
                response = requests.post(
                    f"{base_url}/admin/send_template_queue_tick",
                    data={
                        "tenant": tenant,
                        "period": period,
                        "batch_size": batch_size,
                        "mode": "json",
                        "token": token,
                        # ✅ PASAR EL CACHE
                        "use_cache": "true",
                    },
                    timeout=60
                )
                
                if response.status_code != 200:
                    log.error(f"[AUTO] Error HTTP {response.status_code}")
                    break
                
                data = response.json()
                processed = data.get("processed", 0)
                sent = data.get("sent", 0)
                
                processed_total += processed
                sent_total += sent
                
                log.info(f"[AUTO] Iteración {iterations}: procesados={processed}, enviados={sent}, total={sent_total}")
                
                # Si no procesó nada, terminamos
                if processed == 0:
                    log.info(f"[AUTO] Completado. Total enviados: {sent_total}")
                    break
                
                # Pausa de 2 segundos entre lotes
                time.sleep(5)
                
            except Exception as e:
                log.exception(f"[AUTO] Error: {e}")
                break
        
        log.info(f"[AUTO] Finalizó. Iteraciones: {iterations}, Total enviado: {sent_total}")
    
    # Disparar el thread en background
    thread = threading.Thread(target=process_in_background, daemon=True)
    thread.start()
    
    # Respuesta inmediata al usuario
    return redirect(
        f"/admin/panel?tenant={tenant}&msg=auto_started"
        f"&period={period}"
    )



def debug_list_root_pdfs(tenant: str, limit=20):
    t = get_tenant(tenant)
    root_id = t.get("drive_root_id") or t.get("recibos_root_id")
    service = drive_service()

    q = f"'{root_id}' in parents and trashed=false"
    res = service.files().list(
        q=q,
        fields="files(id,name,mimeType)",
        pageSize=limit
    ).execute()

    log.info("\n=== ROOT FILES ===")
    for f in res.get("files", []):
        log.info("%s %s", f["name"], f["mimeType"])

from io import BytesIO
from flask import send_file

import pandas as pd
import time

def _db_row_to_dict(r):
    return dict(r) if r is not None else None

def _get_last_msg_status(tenant: str, cuil: str, period: str, kind: str):
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT *
        FROM message_status
        WHERE tenant=%s AND cuil=%s AND period=%s AND kind=%s
        ORDER BY COALESCE(created_at, 0) DESC, id DESC
        LIMIT 1
    """, (tenant, cuil, period, kind))
    row = cur.fetchone()
    conn.close()
    return _db_row_to_dict(row)



@app.route("/admin/send_test", methods=["GET", "POST"])
def admin_send_test():
    auth = require_admin()
    if auth:
        return auth

    # request.values lee de la query (GET) y del form (POST) por igual.
    token = request.values.get("token", "")
    tenant = (request.values.get("tenant") or "").strip().lower()
    cuil = (request.values.get("cuil") or "").strip()
    period = (request.values.get("period") or "").strip()
    whatsapp_override = (request.values.get("whatsapp_override") or "").strip()

    t = get_tenant(tenant)
    if not t:
        return Response("Tenant inválido", status=400)

    html = []
    html.append("<h2>Envío individual a persona específica</h2>")
    html.append(f"<p><b>Empresa:</b> {esc(t['display_name'])}</p>")
    html.append(f"<p><a href='/admin/panel?tenant={esc(tenant)}'>← volver al panel</a></p>")

    html.append(f"""
    <style>
      body {{ font-family: system-ui; max-width: 700px; margin: 40px auto; padding: 20px; background: #f5f5f5; }}
      .card {{ background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); margin-bottom: 20px; }}
      label {{ display: block; margin-bottom: 5px; font-weight: 600; color: #333; }}
      input {{ width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 4px; margin-bottom: 15px; }}
      button {{ background: #0066cc; color: white; padding: 12px 24px; border: none; border-radius: 4px; cursor: pointer; font-size: 14px; font-weight: 600; }}
      button:hover {{ background: #0052a3; }}
      .success {{ color: #059669; background: #d1fae5; padding: 12px; border-radius: 4px; border-left: 4px solid #059669; }}
      .error {{ color: #dc2626; background: #fee2e2; padding: 12px; border-radius: 4px; border-left: 4px solid #dc2626; }}
      .info {{ color: #0284c7; background: #e0f2fe; padding: 12px; border-radius: 4px; border-left: 4px solid #0284c7; margin: 10px 0; }}
      .mono {{ font-family: monospace; background: #f3f4f6; padding: 2px 6px; border-radius: 3px; }}
    </style>
    """)

    html.append("<div class='card'>")
    html.append(f"""
    <form method="post">
      <input type="hidden" name="tenant" value="{esc(tenant)}">
      <input type="hidden" name="token" value="{esc(token)}">

      <label>CUIL</label>
      <input type="text" name="cuil" value="{esc(cuil)}" placeholder="20-12345678-9" required>

      <label>Período (mm/aaaa)</label>
      <input type="text" name="period" value="{esc(period)}" placeholder="04/2025" required>

      <label>WhatsApp (opcional — solo si cambió de número)</label>
      <input type="text" name="whatsapp_override" value="{esc(whatsapp_override)}" placeholder="+54 9 11 1234-5678 (dejar vacío = usar el del Excel)">

      <button type="submit">🔍 Buscar y enviar</button>
    </form>
    """)
    html.append("</div>")

    # El envío solo se dispara por POST (botón del form) o con ADMIN_TOKEN por
    # header X-Admin-Token (automatización). Un GET con la sesión activa NO envía:
    # evita que un link malicioso dispare mensajes usando la cookie del admin (CSRF).
    _envio_ok = (request.method == "POST") or admin_token_valid_header()
    if cuil and period and not _envio_ok:
        html.append("<div class='card'><div class='info'>ℹ️ Por seguridad, el envío se confirma desde el botón del formulario.</div></div>")

    if cuil and period and _envio_ok:
        html.append("<div class='card'>")
        html.append("<h3>Resultado del envío</h3>")

        # Normalizar CUIL
        cuil_digits = norm_cuil(cuil)
        if len(cuil_digits) != 11:
            html.append("<div class='error'>❌ CUIL inválido. Debe tener 11 dígitos.</div>")
            return Response("".join(html) + "</div></body></html>", mimetype="text/html")

        # Buscar persona en Excel de envíos
        envios = load_envios_rows(tenant)
        person = find_person_by_cuil(envios, cuil_digits)

        if not person:
            html.append("<div class='error'>❌ No se encontró esa persona en el Excel de envíos.</div>")
            return Response("".join(html) + "</div></body></html>", mimetype="text/html")

        to_whatsapp = person.get("to_whatsapp", "")
        if whatsapp_override:
            ow = normalize_whatsapp(whatsapp_override)
            if not ow:
                html.append("<div class='error'>❌ El WhatsApp ingresado no es válido.</div>")
                return Response("".join(html) + "</div></body></html>", mimetype="text/html")
            to_whatsapp = ow
            html.append(f"<div class='info'>📱 Enviando al número ingresado manualmente: <span class='mono'>{esc(to_whatsapp)}</span> (no al del Excel).</div>")
        if not to_whatsapp:
            html.append("<div class='error'>❌ Esa persona no tiene WhatsApp configurado en el Excel.</div>")
            return Response("".join(html) + "</div></body></html>", mimetype="text/html")

        nombre = person.get("nombre", "")

        html.append(f"<div class='info'>✅ Persona encontrada<br>")
        html.append(f"<b>Nombre:</b> {esc(nombre)}<br>")
        html.append(f"<b>WhatsApp:</b> <span class='mono'>{esc(to_whatsapp)}</span><br>")
        html.append(f"<b>CUIL:</b> <span class='mono'>{esc(cuil_digits)}</span></div>")

        # Verificar que exista el PDF
        pdf_file_id = find_pdf_file_id_for_cuil_period(tenant, cuil_digits, period)
        if not pdf_file_id:
            html.append(f"<div class='error'>❌ No se encontró el PDF para el período {esc(period)}.</div>")
            periods = list_periods_for_cuil(tenant, cuil_digits)
            if periods:
                html.append(f"<div class='info'>Períodos disponibles: {esc(', '.join(periods))}</div>")
            return Response("".join(html) + "</div></body></html>", mimetype="text/html")

        html.append(f"<div class='info'>✅ PDF encontrado para {esc(period)}</div>")

        # Verificar si ya fue enviado antes
        already = already_sent_template(tenant, cuil_digits, period, to_whatsapp)
        if already:
            html.append(f"<div class='info'>ℹ️ Esta persona ya recibió el template anteriormente.</div>")

        # Enviar template
        try:
            sid_tpl = send_whatsapp_template(
                to_whatsapp,
                content_vars={
                    "1": nombre or "Hola",
                    "2": period,
                },
                template_sid=TWILIO_TEMPLATE_SID or None,
                status_callback=STATUS_CALLBACK_URL,
            )
            
            # Registrar en las tablas correspondientes
            save_template_sid(tenant, cuil_digits, period, to_whatsapp, sid_tpl, nombre=nombre)
            add_pending_view(to_whatsapp, tenant, cuil_digits, period, origin="MANUAL")
            
            # Marcar en la cola como enviado (si existía)
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("""
                UPDATE template_send_queue
                SET status = 'SENT', sent_at = %s, sent_sid = %s
                WHERE tenant = %s AND cuil = %s AND period = %s AND to_whatsapp = %s
            """, (int(time.time()), sid_tpl, tenant, cuil_digits, period, to_whatsapp))
            conn.commit()
            conn.close()

            html.append(f"<div class='success'>✅ Template enviado exitosamente<br>")
            html.append(f"<b>Message SID:</b> <span class='mono'>{esc(sid_tpl)}</span></div>")

        except Exception as e:
            html.append(f"<div class='error'>❌ Error enviando template:<br>{esc(str(e))}</div>")

        html.append("</div>")

    return Response("".join(html) + "</body></html>", mimetype="text/html")

@app.post("/admin/resend_template")
def admin_resend_template():
    """
    Reenvía el template a una persona que no lo vio después de 7+ días.
    """
    auth = require_admin()
    if auth:
        return auth
    
    token = _get_admin_token_from_request()
    tenant = (request.form.get("tenant") or "").strip().lower()
    period = (request.form.get("period") or "").strip()
    cuil = (request.form.get("cuil") or "").strip()
    whatsapp = (request.form.get("whatsapp") or "").strip()
    
    if not all([tenant, period, cuil, whatsapp]):
        return Response("Faltan parámetros", status=400)
    
    # Buscar nombre en envios
    envios = load_envios_rows(tenant)
    person = find_person_by_cuil(envios, cuil)
    nombre = person.get("nombre", "") if person else ""
    
    try:
        # Enviar template
        sid = send_whatsapp_template(
            whatsapp,
            content_vars={
                "1": nombre or "Hola",
                "2": period,
            },
            template_sid=TWILIO_TEMPLATE_SID or None,
            status_callback=STATUS_CALLBACK_URL,
        )
        
        # Registrar (sobreescribe el anterior)
        save_template_sid(tenant, cuil, period, whatsapp, sid, nombre=nombre)
        
        # Actualizar pending_view origin a RESEND
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            UPDATE pending_views
            SET origin = 'RESEND', created_at = %s
            WHERE tenant = %s AND cuil = %s AND period = %s AND to_whatsapp = %s
        """, (int(time.time()), tenant, cuil, period, whatsapp))
        
        # Si no existía, crear uno
        if cur.rowcount == 0:
            add_pending_view(whatsapp, tenant, cuil, period, origin="RESEND")
        
        conn.commit()
        conn.close()
        
        msg = f"resend_ok&cuil={cuil}"
        
    except Exception as e:
        msg = f"resend_error&error={str(e)[:100]}"
    
    return redirect(f"/admin/panel?tenant={tenant}&period={period}&msg={msg}")


@app.post("/admin/remind_signature")
def admin_remind_signature():
    """
    Envía recordatorio para firmar usando el template aprobado de firma.
    """
    auth = require_admin()
    if auth:
        return auth
    
    token = _get_admin_token_from_request()
    tenant = (request.form.get("tenant") or "").strip().lower()
    period = (request.form.get("period") or "").strip()
    cuil = (request.form.get("cuil") or "").strip()
    whatsapp = (request.form.get("whatsapp") or "").strip()
    
    if not all([tenant, period, cuil, whatsapp]):
        return Response("Faltan parámetros", status=400)
    
    # Buscar nombre
    envios = load_envios_rows(tenant)
    person = find_person_by_cuil(envios, cuil)
    nombre = person.get("nombre", "") if person else ""
    
    try:
        # Enviar template de firma (con botones SIGN_OK/SIGN_OBS)
        if not TWILIO_SIGN_TEMPLATE_SID:
            return Response("TWILIO_SIGN_TEMPLATE_SID no configurado", status=500)
        
        sid = send_whatsapp_template(
            whatsapp,
            content_vars={
                "1": nombre or "Hola",
                "2": period,
            },
            template_sid=TWILIO_SIGN_TEMPLATE_SID,
            status_callback=STATUS_CALLBACK_URL,
        )
        
        # Registrar el recordatorio
        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO message_status (tenant, cuil, period, kind, message_sid, status, to_whatsapp, created_at)
            VALUES (%s, %s, %s, 'REMIND_SIGN', %s, 'sent', %s, %s)
        """, (tenant, cuil, period, sid, whatsapp, int(time.time())))
        conn.commit()
        conn.close()
        
        msg = f"remind_ok&cuil={cuil}"
        
    except Exception as e:
        msg = f"remind_error&error={str(e)[:100]}"
    
    return redirect(f"/admin/panel?tenant={tenant}&period={period}&msg={msg}")

# =========================
# Twilio inbound: VIEW_NOW + firma/observa
# =========================
import time
from flask import Response, request

def twiml(msg: str):
    return Response(
        f"<Response><Message>{msg}</Message></Response>",
        mimetype="application/xml",
        status=200
    )

import os, json
from twilio.rest import Client

WHATSAPP_MENU_CONTENT_SID = os.getenv("WHATSAPP_MENU_CONTENT_SID", "")

# ============================================================================

TERMS_TEMPLATE_SID = os.getenv("TWILIO_TERMS_TEMPLATE_SID", "")
TERMS_PDF_FILE_ID = os.getenv("TERMS_PDF_FILE_ID", "")

# =========================
# TÉRMINOS Y CONDICIONES
# =========================

def has_accepted_terms(whatsapp: str) -> bool:
    """Verifica si el usuario ya aceptó los T&C."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT 1 FROM terms_accepted WHERE whatsapp = %s LIMIT 1", (whatsapp,))
    exists = cur.fetchone() is not None
    conn.close()
    return exists


def save_terms_acceptance(whatsapp: str):
    """Guarda que el usuario aceptó los T&C."""
    conn = get_db_connection()
    cur = conn.cursor()
    now = int(time.time())
    cur.execute("""
        INSERT INTO terms_accepted (whatsapp, accepted_at)
        VALUES (%s, %s)
        ON CONFLICT (whatsapp) DO UPDATE SET
            accepted_at = EXCLUDED.accepted_at
    """, (whatsapp, now))
    conn.commit()
    conn.close()
    log.info(f"✅ T&C aceptados: {whatsapp}")


def set_pending_terms_acceptance(whatsapp: str, tenant: str, cuil: str, period: str, origin: str = "INITIAL"):
    """Marca que el usuario está esperando aceptar T&C."""
    conn = get_db_connection()
    cur = conn.cursor()
    now = int(time.time())
    cur.execute("""
        INSERT INTO pending_terms (whatsapp, tenant, cuil, period, origin, created_at)
        VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT (whatsapp) DO UPDATE SET
            tenant = EXCLUDED.tenant,
            cuil = EXCLUDED.cuil,
            period = EXCLUDED.period,
            origin = EXCLUDED.origin,
            created_at = EXCLUDED.created_at
    """, (whatsapp, tenant, cuil, period, origin, now))
    conn.commit()
    conn.close()
    log.info(f"✅ Pending T&C: {whatsapp} -> {tenant}/{cuil}/{period}")


def get_pending_terms(whatsapp: str):
    """Obtiene los datos del usuario que está esperando aceptar T&C."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT tenant, cuil, period, origin FROM pending_terms WHERE whatsapp = %s LIMIT 1
    """, (whatsapp,))
    row = cur.fetchone()
    conn.close()
    
    if row:
        return {"tenant": row['tenant'], "cuil": row['cuil'], "period": row['period'], "origin": row['origin']}
    return None


def clear_pending_terms(whatsapp: str):
    """Limpia el estado de espera después de aceptar."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM pending_terms WHERE whatsapp = %s", (whatsapp,))
    conn.commit()
    conn.close()
    log.info(f"✅ Cleared pending T&C: {whatsapp}")

@app.get("/media/terms")
def media_terms():
    """Sirve el PDF de Términos y Condiciones."""
    token = request.args.get("token", "").strip()
    if not check_media_token(token, "terms"):
        return Response("Unauthorized", status=401)
    
    if not TERMS_PDF_FILE_ID:
        return Response("T&C PDF no configurado", status=404)
    
    try:
        service = drive_service()
        
        # Obtener metadata
        file_metadata = service.files().get(fileId=TERMS_PDF_FILE_ID, fields='size,name').execute()
        file_size = int(file_metadata.get('size', 0))
        
        # Descargar con chunks
        req = service.files().get_media(fileId=TERMS_PDF_FILE_ID)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, req, chunksize=5*1024*1024)
        
        done = False
        while not done:
            status, done = downloader.next_chunk()
        
        fh.seek(0)
        data = fh.read()
        
        resp = Response(data, mimetype="application/pdf")
        resp.headers["Content-Disposition"] = 'inline; filename="terminos-y-condiciones.pdf"'
        resp.headers["Content-Length"] = str(len(data))
        resp.headers["Cache-Control"] = "public, max-age=86400"  # 24h cache
        
        return resp
        
    except Exception as e:
        log.exception(f"❌ Error descargando T&C: {e}")
        return Response("Error descargando PDF de T&C", status=500)


def send_terms_and_conditions(to_whatsapp: str, tenant: str, cuil: str, period: str):
    """Envía el PDF de T&C con template de aceptación."""
    try:
        # Validar que estén configuradas las variables
        if not TERMS_TEMPLATE_SID:
            log.error(f"❌ ERROR: TERMS_TEMPLATE_SID no está configurado")
            return None
        
        if not TERMS_PDF_FILE_ID:
            log.error(f"❌ ERROR: TERMS_PDF_FILE_ID no está configurado")
            return None
        
        client = _twilio_client()  # ✅ Usar la función correcta
        
        # Preparar payload del PDF
        payload_pdf = {
            "to": to_whatsapp,
            "body": "📄 Antes de continuar, por favor leé nuestros Términos y Condiciones adjuntos.",
            "media_url": [f"{request.host_url.rstrip('/')}/media/terms?token={make_media_token('terms')}"],
        }
        
        # Preparar payload del botón
        payload_button = {
            "to": to_whatsapp,
            "content_sid": TERMS_TEMPLATE_SID,
        }
        
        # Agregar from_ o messaging_service_sid (mismo patrón que send_whatsapp_pdf)
        if TWILIO_MESSAGING_SERVICE_SID:
            payload_pdf["messaging_service_sid"] = TWILIO_MESSAGING_SERVICE_SID
            payload_button["messaging_service_sid"] = TWILIO_MESSAGING_SERVICE_SID
        elif TWILIO_WHATSAPP_FROM:
            payload_pdf["from_"] = TWILIO_WHATSAPP_FROM
            payload_button["from_"] = TWILIO_WHATSAPP_FROM
        else:
            log.error(f"❌ ERROR: No hay TWILIO_WHATSAPP_FROM ni TWILIO_MESSAGING_SERVICE_SID configurado")
            return None
        
        # Agregar status callback si existe
        if STATUS_CALLBACK_URL:
            payload_pdf["status_callback"] = STATUS_CALLBACK_URL
            payload_button["status_callback"] = STATUS_CALLBACK_URL
        
        # 1. Enviar PDF de T&C
        msg_pdf = client.messages.create(**payload_pdf)
        
        # 2. Enviar template con botón "Acepto"
        msg_button = client.messages.create(**payload_button)
        
        log.info(f"✅ T&C enviados a {to_whatsapp}: PDF={msg_pdf.sid}, Button={msg_button.sid}")
        return msg_button.sid
        
    except Exception as e:
        log.exception(f"❌ Error enviando T&C: {type(e).__name__}: {e}")
        return None
# ============================================================================

def send_whatsapp_menu_template(to_whatsapp: str, nombre: str = "") -> str | None:
    """
    Envía la plantilla del menú (Quick Reply) vía Content API.
    Devuelve Message SID o None.
    """
    if not WHATSAPP_MENU_CONTENT_SID:
        log.error("ERROR: falta WHATSAPP_MENU_CONTENT_SID")
        return None

    client = Client(TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN)

    nombre = (nombre or "").strip()

    # 🔒 Twilio NO acepta variables vacías -> usamos espacio como fallback
    vars_ = {"1": nombre if nombre else " "}

    try:
        msg = client.messages.create(
            to=to_whatsapp,
            from_=TWILIO_WHATSAPP_FROM,
            content_sid=WHATSAPP_MENU_CONTENT_SID,
            content_variables=json.dumps(vars_)
        )
        return msg.sid
    except Exception as e:
        log.exception("%s %s %s %s", "ERROR send_whatsapp_menu_template:", e, "vars_=", vars_)
        return None

def list_previous_periods_excluding_current(tenant: str, cuil: str, limit: int = 3) -> list[str]:
    import datetime as _dt

    now = _dt.datetime.now()
    current = f"{now.month:02d}/{now.year:04d}"   # ej "03/2026"

    periods = list_periods_for_cuil2(tenant, cuil) or []  # ya viene ordenado desc
    # sacamos el mes corriente si aparece
    periods = [p for p in periods if (p or "").replace("-", "/") != current]

    # ahora periods[0] es el último real anterior al mes actual
    return periods[:limit]


# FUNCIONES MULTI-TENANT
# ============================================================================

def find_all_tenants_for_whatsapp(whatsapp: str) -> List[dict]:
    """
    Busca en qué tenants (empresas) aparece este número de WhatsApp.
    
    Retorna lista de dicts con:
    [
        {"tenant": "san-patricio", "display_name": "San Patricio", "cuil": "27123456789"},
        ...
    ]
    """
    whatsapp_normalized = norm_whatsapp(whatsapp) if whatsapp else ""
    if not whatsapp_normalized:
        return []
    
    result = []
    tenants = load_tenants()
    
    for tenant_info in tenants:
        tenant_slug = tenant_info.get("slug", "")
        display_name = tenant_info.get("display_name", tenant_slug)
        
        try:
            envios = load_envios_rows(tenant_slug)
        except Exception as e:
            log.exception(f"Error loading envios for {tenant_slug}: {e}")
            continue
        
        for row in envios:
            # ✅ FIX: Convertir a string antes de strip()
            tel = str(
                row.get("telefono") or 
                row.get("teléfono") or 
                row.get("Telefono") or 
                row.get("whatsapp") or 
                row.get("celular") or 
                row.get("phone") or 
                ""
            ).strip()
            
            if not tel:
                continue
            
            tel_normalized = norm_whatsapp(tel)
            
            if tel_normalized == whatsapp_normalized:
                archivo = strip_pdf(str(row.get("archivo") or row.get("Archivo") or ""))
                cuil = norm_cuil(archivo)
                
                if cuil:
                    # ✅ FIX: Convertir nombre a string también
                    nombre = str(row.get("nombre") or row.get("Nombre") or "").strip()
                    
                    result.append({
                        "tenant": tenant_slug,
                        "display_name": display_name,
                        "cuil": cuil,
                        "nombre": nombre
                    })
                    break
    
    return result


def save_multi_tenant_selection_state(whatsapp: str, tenants: List[dict] = None, action: str = "RESEND", period_offset: int = 0):
    """
    Guarda las opciones de tenant disponibles para que el usuario elija.
    
    Args:
        whatsapp: número de WhatsApp normalizado
        tenants: lista de dicts con tenant, display_name, cuil (puede ser None si solo guardamos offset)
        action: "RESEND" para pedir recibo, "SEE_PREVIOUS" para ver períodos
        period_offset: offset de períodos (para paginación)
    """
    now = int(time.time())
    expires_at = now + (10 * 60)  # 10 minutos
    
    # Si no pasamos tenants, obtener el estado actual para preservarlo
    if tenants is None:
        current = get_multi_tenant_selection_state(whatsapp)
        if current:
            tenants = current.get("tenants", [])
        else:
            tenants = []
    
    context = {
        "tenants": tenants,
        "action": action,
        "period_offset": period_offset
    }
    context_json = json.dumps(context)
    
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO multi_tenant_selection (whatsapp, tenants_json, created_at, expires_at)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT(whatsapp) DO UPDATE SET
            tenants_json = excluded.tenants_json,
            created_at = excluded.created_at,
            expires_at = excluded.expires_at
    """, (whatsapp, context_json, now, expires_at))
    conn.commit()
    conn.close()

def get_multi_tenant_selection_state(whatsapp: str) -> Optional[dict]:
    """Recupera el estado de selección multi-tenant."""
    conn = get_db_connection()
    cur = conn.cursor()
    now = int(time.time())
    
    cur.execute("""
        SELECT tenants_json, created_at
        FROM multi_tenant_selection
        WHERE whatsapp = %s AND expires_at > %s
        LIMIT 1
    """, (whatsapp, now))
    
    row = cur.fetchone()
    conn.close()
    
    if not row:
        return None
    
    try:
        data = json.loads(row['tenants_json'])
        # Compatibilidad con formato antiguo
        if isinstance(data, list):
            return {
                "tenants": data,
                "action": "RESEND",
                "period_offset": 0,
                "created_at": row['created_at']
            }
        # Formato nuevo
        return {
            "tenants": data.get("tenants", []),
            "action": data.get("action", "RESEND"),
            "period_offset": data.get("period_offset", 0),
            "created_at": row['created_at']
        }
    except Exception as e:
        log.exception(f"Error parsing multi_tenant_selection: {e}")
        return None
    
def show_previous_periods_paginated(from_whatsapp: str, tenant: str, cuil: str, offset: int = 0):
    """
    Muestra períodos anteriores con paginación.
    """
    # Obtener más períodos de los que vamos a mostrar para saber si hay más
    all_prev = list_previous_periods_excluding_current(tenant, cuil, limit=20)
    
    # Aplicar offset
    available = all_prev[offset:]
    
    if not available:
        return ("ℹ️ No hay más períodos anteriores disponibles.", False)
    
    # Mostrar hasta 3 períodos
    to_show = available[:3]
    has_more = len(available) > 3
    
    # Guardar el primer período en pending para manejar la selección
    add_pending_view(from_whatsapp, tenant, cuil, to_show[0], origin="SEE_PREVIOUS")
    pending = get_latest_pending_view(from_whatsapp)
    set_pending_step(pending["id"], "CHOOSE_PREVIOUS")
    
    # ✅ NUEVO: Guardar también el offset en el pending
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE pending_views 
        SET period_offset = %s 
        WHERE id = %s
    """, (offset, pending["id"]))
    conn.commit()
    conn.close()
    
    log.info(f"✅ SAVED OFFSET IN PENDING: {offset}")  # Debug
    
    # Construir mensaje
    msg = "🗂️ Períodos anteriores:\n\n"
    for i, p in enumerate(to_show, start=1):
        msg += f"{i}. {p}\n"
    
    if has_more:
        msg += f"\n4️⃣ Ver más períodos anteriores"
    
    msg += "\nRespondé con el número para elegir."
    
    return (msg, has_more)


def clear_multi_tenant_selection_state(whatsapp: str):
    """Limpia el estado de selección después de que el usuario eligió."""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM multi_tenant_selection WHERE whatsapp = %s", (whatsapp,))
    conn.commit()
    conn.close()


TWILIO_SIGN_TEMPLATE_SID = os.environ.get("TWILIO_SIGN_TEMPLATE_SID", "").strip()

@app.post("/twilio/inbound")
def twilio_inbound():
    if not _twilio_signature_ok():
        log.warning("⚠️ /twilio/inbound: firma Twilio inválida, request rechazado")
        return Response("Forbidden", status=403)

    from_whatsapp = (request.form.get("From") or "").strip()
    button = (request.form.get("ButtonPayload") or "").strip()
    body = (request.form.get("Body") or "").strip()
    in_sid = (request.form.get("MessageSid") or "").strip()

    # ✅ List Picker manda el Item ID en Body (no en ButtonPayload como el Quick Reply).
    # Si el body es un ID de opción conocido y no vino button, lo tratamos como button.
    _MENU_IDS = {
        "RESEND_LAST", "SEE_PREVIOUS", "MORE_OPTIONS",
        "VIEW_NOW", "NO_NEED", "SIGN_OK", "SIGN_OBS", "ACCEPT_TERMS",
        "CERT_MEDICO", "CERT_FAMILIA",
    }
    if not button and body in _MENU_IDS:
        button = body
        body = ""

    log.info("%s %s %s %s %s %s %s %s", "INBOUND:", from_whatsapp, "MessageSid:", in_sid, "ButtonPayload:", button, "Body:", body)

    # ✅ DEDUP global: si Twilio reintenta el mismo inbound, no hacemos nada
    # ✅ DEDUP global
    if inbound_seen(in_sid):
        log.info("%s %s", "DEDUP inbound:", in_sid)
        return Response("OK", status=200)

    # ============================================================================
    # ✅ DETECTAR ACEPTACIÓN DE T&C
    # ============================================================================
    if button == "ACCEPT_TERMS" or "acepto" in body.lower():
        pending_terms = get_pending_terms(from_whatsapp)
        
        if pending_terms:
            save_terms_acceptance(from_whatsapp)
            
            tenant = pending_terms['tenant']
            cuil = pending_terms['cuil']
            period = pending_terms['period']
            origin = pending_terms.get('origin', 'INITIAL')
            
            sid_pdf = _send_pdf_flow(from_whatsapp, tenant, cuil, period, origin=origin)
            
            clear_pending_terms(from_whatsapp)
            
            if not sid_pdf:
                return twiml("❌ Hubo un error enviando el PDF. Intentá de nuevo o contactá a RRHH.")
            
            return Response("OK", status=200)

    def _is_receipt_request_text(t: str) -> bool:
        t = (t or "").strip().lower()
        if not t:
            return False
        if t.startswith("recibo"):
            return True
        return t in ("pdf", "reenviar", "reenviar recibo", "reenvio", "reenvío", "pedir recibo", "quiero mi recibo")

    pending = get_latest_pending_view(from_whatsapp)
    log.info("%s %s", "PENDING:", pending)
    
    # =========================
    # SEE_PREVIOUS debe funcionar incluso sin pending
    # y SIEMPRE excluir el mes corriente
    # =========================
    if button == "SEE_PREVIOUS":
        # SIEMPRE verificar si el usuario está en múltiples tenants
        tenants_found = find_all_tenants_for_whatsapp(from_whatsapp)
        
        if len(tenants_found) == 0:
            return twiml("👋 Para ver períodos anteriores, necesitás el mensaje inicial de RRHH. Si no lo tenés, avisá a RRHH.")
        
        elif len(tenants_found) > 1:
            # Multi-tenant: preguntar primero
            save_multi_tenant_selection_state(from_whatsapp, tenants_found, action="SEE_PREVIOUS", period_offset=0)
            
            msg = "Trabajás en varias empresas. ¿De cuál querés ver los períodos?\n\n"
            for i, t in enumerate(tenants_found, 1):
                msg += f"{i}️⃣ {t['display_name']}\n"
            msg += "\nRespondé con el número."
            
            return twiml(msg)
        
        # Si llegamos acá, es porque hay UN SOLO tenant
        tenant0 = tenants_found[0]["tenant"]
        cuil0 = tenants_found[0]["cuil"]

        msg, has_more = show_previous_periods_paginated(from_whatsapp, tenant0, cuil0, offset=0)
        
        # Guardar estado para manejar "4" (ver más)
        if has_more:
            save_multi_tenant_selection_state(
                from_whatsapp, 
                tenants=[{"tenant": tenant0, "cuil": cuil0}],
                action="SEE_PREVIOUS_PAGINATION",
                period_offset=3
            )
        
        return twiml(msg)

    # =========================================================================
    # 🩺 CERT — botón "Certificado médico" (early, igual patrón que SEE_PREVIOUS)
    # Funciona aunque no haya pending. Si está en varias empresas, pregunta a cuál.
    # =========================================================================
    if button in ("CERT_MEDICO", "CERT_FAMILIA"):
        # El tipo viaja en origin para recordarlo hasta que llegue el archivo.
        cert_tipo = "familia" if button == "CERT_FAMILIA" else "medico"
        cert_origin = "CERT_FAMILIA" if cert_tipo == "familia" else "CERT"
        cert_label = "certificado de asignaciones familiares" if cert_tipo == "familia" else "certificado médico"
        cert_emoji = "🧑\u200d🧑\u200d🧒" if cert_tipo == "familia" else "🩺"

        tenants_found = find_all_tenants_for_whatsapp(from_whatsapp)

        if len(tenants_found) == 0:
            return twiml("👋 Para enviar tu certificado, primero necesitás el mensaje inicial de RRHH. Si no lo tenés, avisá a RRHH.")

        elif len(tenants_found) > 1:
            # Multi-empresa: preguntar a cuál mandar el certificado (recordando el tipo)
            save_multi_tenant_selection_state(from_whatsapp, tenants_found, action=cert_origin)

            msg = "Trabajás en varias empresas. ¿A cuál querés enviar el certificado?\n\n"
            for i, t in enumerate(tenants_found, 1):
                msg += f"{i}️⃣ {t['display_name']}\n"
            msg += "\nRespondé con el número."

            return twiml(msg)

        # Un solo tenant: armar pending y pasar a AWAIT_CERT
        tenant0 = tenants_found[0]["tenant"]
        cuil0 = tenants_found[0]["cuil"]
        add_pending_view(from_whatsapp, tenant0, cuil0, "", origin=cert_origin)
        p = get_latest_pending_view(from_whatsapp)
        set_pending_step(p["id"], "AWAIT_CERT")
        return twiml(
            f"{cert_emoji} Mandame una *foto* o *PDF* de tu {cert_label}.\n\n"
            "Si querés salir, escribí *CANCELAR*."
        )

    # ============================================================================
    # MANEJO DE SELECCIÓN MULTI-TENANT (1, 2, 3, 4)
    # ============================================================================
    if not button and body.strip() in ("1", "2", "3", "4"):
        step_now = (pending.get("step") or "READY").upper() if isinstance(pending, dict) else "READY"
        
        # ============================================================
        # CASO ESPECIAL: "4" cuando está en modo CHOOSE_PREVIOUS
        # ============================================================
        if body.strip() == "4" and step_now == "CHOOSE_PREVIOUS":
            # Verificar si hay paginación activa
            multi_state = get_multi_tenant_selection_state(from_whatsapp)
            
            if multi_state and multi_state.get("action") == "SEE_PREVIOUS_PAGINATION":
                tenant_info = multi_state.get("tenants", [{}])[0]
                tenant = tenant_info.get("tenant")
                cuil = tenant_info.get("cuil")
                current_offset = multi_state.get("period_offset", 0)
                
                log.info(f"🔍 PAGING FORWARD: current_offset={current_offset}")
                
                msg, has_more = show_previous_periods_paginated(from_whatsapp, tenant, cuil, offset=current_offset)
                
                # Actualizar offset para próxima paginación
                if has_more:
                    new_offset = current_offset + 3
                    log.info(f"✅ SAVING NEXT OFFSET: {new_offset}")
                    save_multi_tenant_selection_state(
                        from_whatsapp,
                        tenants=[{"tenant": tenant, "cuil": cuil}],
                        action="SEE_PREVIOUS_PAGINATION",
                        period_offset=new_offset
                    )
                else:
                    # No hay más, limpiar
                    clear_multi_tenant_selection_state(from_whatsapp)
                
                return twiml(msg)
            else:
                # Si no hay estado de paginación, buscar tenant del pending
                log.warning(f"⚠️ NO PAGINATION STATE FOUND, using pending tenant/cuil")
                
                if not pending:
                    return twiml("❌ Hubo un error. Por favor, pedí tus períodos anteriores nuevamente.")
                
                tenant = (pending.get("tenant") or "").strip().lower()
                cuil = (pending.get("cuil") or "").strip()
                
                # Empezar desde offset 3 (segunda página)
                msg, has_more = show_previous_periods_paginated(from_whatsapp, tenant, cuil, offset=3)
                
                if has_more:
                    save_multi_tenant_selection_state(
                        from_whatsapp,
                        tenants=[{"tenant": tenant, "cuil": cuil}],
                        action="SEE_PREVIOUS_PAGINATION",
                        period_offset=6
                    )
                
                return twiml(msg)
        
        # Si NO está en modo CHOOSE_PREVIOUS, entonces puede ser multi-tenant
        if step_now != "CHOOSE_PREVIOUS":
            multi_state = get_multi_tenant_selection_state(from_whatsapp)
            
            if multi_state:
                try:
                    idx = int(body.strip()) - 1
                    tenants_list = multi_state.get("tenants", [])
                    action = multi_state.get("action", "RESEND")
                    
                    if idx < 0 or idx >= len(tenants_list):
                        return twiml("❌ Opción inválida. Respondé con el número correcto.")
                    
                    selected = tenants_list[idx]
                    tenant = selected["tenant"]
                    cuil = selected["cuil"]
                    
                    log.info(f"🔍 MULTI-TENANT SELECT: tenant={tenant}, action={action}")
                    
                    # ============================================================
                    # ACCIÓN: SEE_PREVIOUS → Mostrar períodos anteriores
                    # ============================================================
                    if action == "SEE_PREVIOUS":
                        # ⚠️ NO limpiar aquí para SEE_PREVIOUS
                        msg, has_more = show_previous_periods_paginated(from_whatsapp, tenant, cuil, offset=0)
                        
                        log.info(f"🔍 AFTER SELECT COMPANY: has_more={has_more}")
                        
                        # Guardar estado para manejar "4" (ver más)
                        if has_more:
                            log.info(f"✅ SAVING PAGINATION STATE: offset=3")
                            save_multi_tenant_selection_state(
                                from_whatsapp,
                                tenants=[{"tenant": tenant, "cuil": cuil}],
                                action="SEE_PREVIOUS_PAGINATION",
                                period_offset=3
                            )
                        else:
                            # Si no hay más páginas, limpiar
                            clear_multi_tenant_selection_state(from_whatsapp)
                        
                        return twiml(msg)

                    # ========================================================
                    # 🩺 CERT — ACCIÓN: eligió empresa para mandar certificado
                    # ========================================================
                    if action in ("CERT", "CERT_FAMILIA"):
                        _tipo = "familia" if action == "CERT_FAMILIA" else "medico"
                        _label = "certificado de asignaciones familiares" if _tipo == "familia" else "certificado médico"
                        _emoji = "🧑\u200d🧑\u200d🧒" if _tipo == "familia" else "🩺"
                        clear_multi_tenant_selection_state(from_whatsapp)
                        add_pending_view(from_whatsapp, tenant, cuil, "", origin=action)
                        p = get_latest_pending_view(from_whatsapp)
                        set_pending_step(p["id"], "AWAIT_CERT")
                        return twiml(
                            f"{_emoji} Mandame una *foto* o *PDF* de tu {_label}.\n\n"
                            "Si querés salir, escribí *CANCELAR*."
                        )

                    # ============================================================
                    # ACCIÓN: RESEND → Enviar último recibo
                    # ============================================================
                    # Para RESEND sí limpiamos porque ya no necesitamos el estado
                    clear_multi_tenant_selection_state(from_whatsapp)
                    
                    log.info(f"🔍 RESEND: tenant={tenant}, cuil={cuil}")
                    
                    period = resolve_best_period_with_pdf(tenant, cuil)
                    log.info(f"🔍 RESEND: period={period}")
                    
                    if not period:
                        _log_receipt_request_event(tenant, cuil, "", from_whatsapp, "MULTI_SELECT", "NO_PDF")
                        return twiml("⚠️ No encontré recibos disponibles. Avisá a RRHH.")
                    
                    add_pending_view(from_whatsapp, tenant, cuil, period, origin="RESEND_LAST")
                    pending = get_latest_pending_view(from_whatsapp)
                    
                    cnt = get_receipt_request_count(tenant, cuil, period, from_whatsapp)
                    log.info(f"🔍 RESEND: cnt={cnt}")
                    
                    if cnt >= 3:
                        _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "MULTI_SELECT", "BLOCKED_LIMIT")
                        return twiml(f"⚠️ Ya pediste este recibo {cnt}/3 veces para {period}. Si necesitás más, avisá a RRHH.")
                    
                    is_verified = is_verified_contact(tenant, cuil, from_whatsapp)
                    log.info(f"🔍 RESEND: is_verified={is_verified}")
                    
                    if not is_verified:
                        set_pending_step(pending["id"], "AWAIT_DNI")
                        _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "MULTI_SELECT", "ASK_DNI")
                        return twiml("🔐 Para reenviar tu recibo, enviá tu DNI (solo números, sin puntos).")
                    
                    log.info(f"🔍 RESEND: Calling _send_pdf_flow...")
                    
                    sid_pdf = _send_pdf_flow(from_whatsapp, tenant, cuil, period, origin="RESEND_LAST")
                    log.info(f"🔍 RESEND: sid_pdf={sid_pdf}")
                    
                    if not sid_pdf:
                        _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "MULTI_SELECT", "ERROR")
                        return twiml("❌ No pude enviar el PDF en este momento. Probá de nuevo o avisá a RRHH.")
                    
                    n = inc_receipt_request_count(tenant, cuil, period, from_whatsapp)
                    _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "MULTI_SELECT", "SENT", message_sid=sid_pdf)
                    
                    log.info(f"✅ RESEND: PDF sent successfully, sid={sid_pdf}")
                    
                    return Response("OK", status=200)
                    
                except Exception as e:
                    log.exception(f"Error en selección multi-tenant: {e}")
                    clear_multi_tenant_selection_state(from_whatsapp)
                    return twiml("❌ Hubo un error. Por favor, pedí tu recibo nuevamente.")
            
    # =========================
    # REGLA: cualquier texto (sin botón) dispara menú,
    # EXCEPTO cuando estamos esperando DNI o selección de períodos
    # =========================
    if not button:
        step_now = (pending.get("step") or "READY").upper() if isinstance(pending, dict) else "READY"
        body_norm = (body or "").strip()

        # AWAIT_DNI: dejamos pasar para que lo procese el bloque AWAIT_DNI
        if step_now == "AWAIT_DNI":
            pass

        # 🩺 CERT — AWAIT_CERT: dejamos pasar para que lo procese el bloque AWAIT_CERT
        # (importante: si no, la FOTO dispararía el menú)
        elif step_now == "AWAIT_CERT":
            pass

        # CHOOSE_PREVIOUS: si no es 1/2/3/4, devolvemos ayuda (no menú)
        elif step_now == "CHOOSE_PREVIOUS":
            if body_norm in ("1", "2", "3", "4"):
                pass
            else:
                return twiml("🗂️ Respondé con 1, 2, 3 o 4 para elegir un período anterior.")

        else:
            # ✅ Enviar menú y terminar (SIN TwiML al usuario)
            nombre = ""
            if isinstance(pending, dict):
                nombre = get_nombre_for_cuil(pending["tenant"], pending["cuil"])
            sid = send_whatsapp_menu_template(from_whatsapp, nombre=nombre)
            return Response("OK", status=200)


    # =========================
    # SIN PENDING: o guía o reconstrucción para "RECIBO"
    # =========================
    if not pending:
        if not _is_receipt_request_text(body) and not button:
            in_sid = (request.form.get("MessageSid") or "").strip()

            # 🔥 En vez de contestar texto, mandamos la plantilla menú
            sid = send_whatsapp_menu_template(from_whatsapp, nombre="")
            return Response("OK", status=200)


        # si es pedido de recibo -> buscar en qué tenants está
        tenants_found = find_all_tenants_for_whatsapp(from_whatsapp)
        
        if len(tenants_found) == 0:
            _log_receipt_request_event("", "", "", from_whatsapp, "USER_TEXT", "NO_CONTEXT")
            return twiml("👋 Para enviarte tu recibo, primero necesitás el mensaje inicial de RRHH. Si no lo tenés, avisá a RRHH.")
        
        elif len(tenants_found) == 1:
            # ✅ Un solo tenant → flow normal
            tenant = tenants_found[0]["tenant"]
            cuil = tenants_found[0]["cuil"]
            
        elif len(tenants_found) > 1:
            # ✅ Multi-tenant: guardar estado y preguntar
            save_multi_tenant_selection_state(from_whatsapp, tenants_found, action="RESEND")
            
            msg = "Trabajás en varias empresas. ¿De cuál querés el recibo?\n\n"
            for i, t in enumerate(tenants_found, 1):
                msg += f"{i}️⃣ {t['display_name']}\n"
            msg += "\nRespondé con el número."
            
            return twiml(msg)

        period = resolve_best_period_with_pdf(tenant, cuil)
        if not period:
            _log_receipt_request_event(tenant, cuil, "", from_whatsapp, "USER_TEXT", "NO_PDF")
            return twiml("⚠️ No encontré ningún recibo disponible para tu CUIL. Avisá a RRHH.")

        add_pending_view(from_whatsapp, tenant, cuil, period, origin="RESEND_LAST")

        pending = get_latest_pending_view(from_whatsapp)

    if not pending:
        return Response("OK", status=200)

    tenant = (pending.get("tenant") or "").strip().lower()
    cuil = (pending.get("cuil") or "").strip()
    period = (pending.get("period") or "").strip()
    step = (pending.get("step") or "READY").upper()

    # 🔒 Si ya cerró
    # 🔒 Bloquear SOLO si intenta firmar/observar
    if button in ("SIGN_OK", "SIGN_OBS") or body in ("SIGN_OK", "SIGN_OBS"):
        estado = get_recibo_estado(tenant, cuil, period)
        if estado in ("FIRMADO", "OBSERVADO"):
            msg = "✅ Este recibo ya fue firmado." if estado == "FIRMADO" else "📝 Este recibo quedó como observado."
            return twiml(msg)


    # =========================
    # BOTONES PLANTILLA: RESEND_LAST / SEE_PREVIOUS / MORE_OPTIONS
    # =========================

    # MORE_OPTIONS (respuesta simple por ahora)
    if button == "MORE_OPTIONS":
        return twiml("ℹ️ Esta opción estará disponible próximamente.")

    # RESEND_LAST -> envía el último recibo disponible (mes actual si existe, sino el último real)
    if button == "RESEND_LAST":
        # ✅ NUEVO: Verificar si está en múltiples empresas
        tenants_found = find_all_tenants_for_whatsapp(from_whatsapp)
        
        if len(tenants_found) == 0:
            _log_receipt_request_event("", "", "", from_whatsapp, "RESEND_LAST", "NO_CONTEXT")
            return twiml("👋 Para enviarte tu recibo, primero necesitás el mensaje inicial de RRHH. Si no lo tenés, avisá a RRHH.")
        
        elif len(tenants_found) > 1:
            # Multi-tenant: preguntar de cuál empresa quiere el recibo
            save_multi_tenant_selection_state(from_whatsapp, tenants_found, action="RESEND")
            
            msg = "Trabajás en varias empresas. ¿De cuál querés el último recibo?\n\n"
            for i, t in enumerate(tenants_found, 1):
                msg += f"{i}️⃣ {t['display_name']}\n"
            msg += "\nRespondé con el número."
            
            return twiml(msg)
        
        # Si llegamos acá, solo está en 1 empresa
        tenant = tenants_found[0]["tenant"]
        cuil = tenants_found[0]["cuil"]
        
        best_period = resolve_best_period_with_pdf(tenant, cuil)
        if not best_period:
            _log_receipt_request_event(tenant, cuil, "", from_whatsapp, "RESEND_LAST", "NO_PDF")
            return twiml("⚠️ No encontré recibos disponibles para tu CUIL. Avisá a RRHH.")

        cnt = get_receipt_request_count(tenant, cuil, best_period, from_whatsapp)
        if cnt >= 3:
            _log_receipt_request_event(tenant, cuil, best_period, from_whatsapp, "RESEND_LAST", "BLOCKED_LIMIT")
            return twiml(f"⚠️ Ya pediste este recibo {cnt}/3 veces para {best_period}. Si necesitás más, avisá a RRHH.")

        if not is_verified_contact(tenant, cuil, from_whatsapp):
            # guardamos el período que vamos a reenviar
            add_pending_view(from_whatsapp, tenant, cuil, best_period, origin="RESEND_LAST")
            pending = get_latest_pending_view(from_whatsapp)
            set_pending_step(pending["id"], "AWAIT_DNI")

            _log_receipt_request_event(tenant, cuil, best_period, from_whatsapp, "RESEND_LAST", "ASK_DNI")
            return twiml("🔐 Para reenviar tu recibo, enviá tu DNI (solo números, sin puntos).")


        sid_pdf = _send_pdf_flow(from_whatsapp, tenant, cuil, best_period, origin="RESEND_LAST")
        if not sid_pdf:
            _log_receipt_request_event(tenant, cuil, best_period, from_whatsapp, "RESEND_LAST", "ERROR")
            return twiml("❌ No pude enviar el PDF en este momento. Probá de nuevo o avisá a RRHH.")
        # ✅ sumar el pedido (esto es lo que faltaba)
        n = inc_receipt_request_count(tenant, cuil, best_period, from_whatsapp)
        # ✅ no mandamos texto antes del PDF
        _log_receipt_request_event(tenant, cuil, best_period, from_whatsapp, "RESEND_LAST", "SENT", message_sid=sid_pdf, origin="RESEND_LAST")
        return Response("OK", status=200)



    # =========================
    # SELECCIÓN de período anterior (1/2/3) cuando step=CHOOSE_PREVIOUS
    # =========================
    if step == "CHOOSE_PREVIOUS" and (not button) and (body or "").strip() in ("1", "2", "3"):
        idx = int((body or "").strip()) - 1

        # ✅ Leer offset directamente del pending
        current_offset = pending.get("period_offset", 0)
        
        log.info(f"🔍 PAGINATION DEBUG: current_offset={current_offset} (from pending)")
        
        # Obtener períodos desde el offset correcto
        all_prev = list_previous_periods_excluding_current(tenant, cuil, limit=20)
        prev = all_prev[current_offset:current_offset + 3]
        
        log.info(f"🔍 PERIODS: all={all_prev[:10]}, showing={prev}")
        
        if idx >= len(prev):
            return twiml("❌ Opción inválida. Respondé con 1, 2 o 3.")

        chosen_period = prev[idx]
        log.info(f"✅ SELECTED: idx={idx}, period={chosen_period}")

        # Limpiar estado de paginación
        clear_multi_tenant_selection_state(from_whatsapp)
        
        # volvemos a READY
        set_pending_step(pending["id"], "READY")

        cnt = get_receipt_request_count(tenant, cuil, chosen_period, from_whatsapp)
        if cnt >= 3:
            _log_receipt_request_event(tenant, cuil, chosen_period, from_whatsapp, "CHOOSE_PREVIOUS", "BLOCKED_LIMIT")
            return twiml(f"⚠️ Ya pediste este recibo {cnt}/3 veces para {chosen_period}. Si necesitás más, avisá a RRHH.")

        if not is_verified_contact(tenant, cuil, from_whatsapp):
            add_pending_view(from_whatsapp, tenant, cuil, chosen_period)
            pending = get_latest_pending_view(from_whatsapp)
            set_pending_step(pending["id"], "AWAIT_DNI")
            _log_receipt_request_event(tenant, cuil, chosen_period, from_whatsapp, "CHOOSE_PREVIOUS", "ASK_DNI", origin="CHOOSE_PREVIOUS")
            return twiml("🔐 Para reenviar tu recibo, enviá tu DNI (solo números, sin puntos).")

        sid_pdf = _send_pdf_flow(from_whatsapp, tenant, cuil, chosen_period)
        if not sid_pdf:
            _log_receipt_request_event(tenant, cuil, chosen_period, from_whatsapp, "CHOOSE_PREVIOUS", "ERROR")
            return twiml("❌ No pude enviar el PDF en este momento. Probá de nuevo o avisá a RRHH.")

        n = inc_receipt_request_count(tenant, cuil, chosen_period, from_whatsapp)
        _log_receipt_request_event(tenant, cuil, chosen_period, from_whatsapp, "CHOOSE_PREVIOUS", "SENT", message_sid=sid_pdf)
        return twiml(f"📄 Listo. Te reenvié el recibo {chosen_period}. (Pedido {n}/3)")

    # =========================================================================
    # 🩺 CERT — AWAIT_CERT: esperando el archivo del certificado médico
    # (va ANTES de AWAIT_DNI)
    # =========================================================================
    if step == "AWAIT_CERT":
        # ¿quiere cancelar?
        if body.strip().lower() in ("cancelar", "cancel", "salir"):
            set_pending_step(pending["id"], "READY")
            return twiml("✅ Cancelado. Cuando quieras, volvé a abrir el menú.")

        # ¿mandó un archivo? Twilio lo indica con NumMedia > 0
        try:
            num_media = int(request.form.get("NumMedia") or "0")
        except ValueError:
            num_media = 0

        if num_media < 1:
            # mandó texto en vez de archivo
            return twiml(
                "📎 Necesito una *foto* o *PDF* del certificado.\n\n"
                "Si querés salir, escribí *CANCELAR*."
            )

        media_url = (request.form.get("MediaUrl0") or "").strip()
        media_ct = (request.form.get("MediaContentType0") or "").strip().lower()

        # validar tipo: solo imagen o pdf
        if not (media_ct.startswith("image/") or "pdf" in media_ct):
            return twiml(
                "⚠️ Solo acepto *imágenes* o *PDF*. Probá de nuevo.\n\n"
                "Si querés salir, escribí *CANCELAR*."
            )

        try:
            # tipo según el origin guardado al iniciar el flujo (CERT_FAMILIA -> familia)
            cert_tipo = "familia" if (pending.get("origin") or "") == "CERT_FAMILIA" else "medico"

            # 1) descargar de Twilio
            data, content_type = download_twilio_media(media_url)
            if content_type == "application/octet-stream" and media_ct:
                content_type = media_ct

            # 2) subir a Drive (a la subcarpeta del tipo)
            file_id, file_name = upload_certificado_to_drive(tenant, cuil, data, content_type, cert_tipo)

            # 3) registrar en BD
            nombre = pending.get("nombre", "") or get_nombre_for_cuil(tenant, cuil)
            save_certificado(tenant, cuil, nombre, from_whatsapp, file_id, file_name, content_type, cert_tipo)

            # 3.b) notificar por email a RRHH (en segundo plano para no demorar la respuesta a WhatsApp)
            try:
                threading.Thread(
                    target=send_certificado_notification,
                    args=(tenant, nombre, cuil, cert_tipo),
                    daemon=True
                ).start()
            except Exception as e:
                log.warning(f"No se pudo lanzar la notificación de certificado: {e}")
            # 4) volver a READY y confirmar
            set_pending_step(pending["id"], "READY")
            return twiml("✅ Recibí tu certificado. RRHH lo va a revisar. ¡Gracias!")

        except Exception as e:
            log.exception(f"❌ Error guardando certificado: {type(e).__name__}: {e}")
            return twiml(
                "❌ Hubo un error guardando tu certificado. Probá de nuevo en un momento, o avisá a RRHH."
            )

    # =========================
    # AWAIT_DNI
    # =========================
    if step == "AWAIT_DNI":
        if button:
            return twiml("🔐 Para continuar, enviá tu DNI (solo números).")

        dni_user = _digits(body)
        dni_expected = cuil_to_dni(cuil)

        if not dni_expected or len(dni_user) < 7:
            inc_pending_dni_attempts(pending["id"])
            return twiml("🔐 Enviá tu DNI (solo números, sin puntos). Ej: 28169249")

        if dni_user != dni_expected:
            tries = inc_pending_dni_attempts(pending["id"])
            if tries >= 3:
                consume_pending_view(pending["id"])
                return twiml("❌ DNI incorrecto (3 intentos). Volvé a solicitar el recibo desde el mensaje inicial.")
            return twiml(f"❌ DNI incorrecto. Intento {tries}/3. Probá de nuevo (solo números).")

        # DNI OK
        set_verified_contact(tenant, cuil, from_whatsapp, dni_user, nombre=pending.get("nombre",""))
        set_pending_step(pending["id"], "READY")

        cnt = get_receipt_request_count(tenant, cuil, period, from_whatsapp)
        if cnt >= 3:
            _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "DNI_OK", "BLOCKED_LIMIT")
            consume_pending_view(pending["id"])
            return twiml(f"⚠️ Ya pediste este recibo {cnt}/3 veces para {period}. Si necesitás más, avisá a RRHH.")

        # ✅ VERIFICAR T&C antes de enviar PDF
        already_accepted = has_accepted_terms(from_whatsapp)
        log.info(f"🔍 T&C CHECK (AWAIT_DNI): whatsapp={from_whatsapp}, already_accepted={already_accepted}")
        log.info(f"🔍 T&C CONFIG: TERMS_TEMPLATE_SID={TERMS_TEMPLATE_SID[:10] if TERMS_TEMPLATE_SID else 'EMPTY'}, TERMS_PDF_FILE_ID={TERMS_PDF_FILE_ID[:10] if TERMS_PDF_FILE_ID else 'EMPTY'}")
        
        if not already_accepted:
            origin = (pending.get("origin") or "INITIAL")
            log.info(f"🔍 T&C: Enviando T&C a {from_whatsapp}")
            send_terms_and_conditions(from_whatsapp, tenant, cuil, period)
            set_pending_terms_acceptance(from_whatsapp, tenant, cuil, period, origin)
            _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "DNI_OK", "SENT_TERMS")
            return twiml("✅ DNI verificado.")
        else:
            log.info(f"🔍 T&C: Ya aceptó, enviando PDF directo")

        # ✅ usar origin del pending (INITIAL si vino del admin)
        origin = (pending.get("origin") or "INITIAL")

        sid_pdf = _send_pdf_flow(from_whatsapp, tenant, cuil, period, origin=origin)
        if not sid_pdf:
            _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "DNI_OK", "ERROR", origin=origin)
            return twiml("✅ DNI verificado, pero hubo un error enviando el recibo. Avisá a RRHH.")

        n = inc_receipt_request_count(tenant, cuil, period, from_whatsapp)
        _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "DNI_OK", "SENT", message_sid=sid_pdf, origin=origin)
        return twiml(f"✅ DNI verificado.")

    # =========================
    # VIEW_NOW
    # =========================
    if button == "VIEW_NOW" or body == "VIEW_NOW":
        if not is_verified_contact(tenant, cuil, from_whatsapp):
            set_pending_step(pending["id"], "AWAIT_DNI")
            return twiml("🔐 Para confirmar tu identidad, enviá tu DNI (solo números, sin puntos).")

        cnt = get_receipt_request_count(tenant, cuil, period, from_whatsapp)
        if cnt >= 3:
            _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "VIEW_NOW", "BLOCKED_LIMIT")
            return twiml(f"⚠️ Ya pediste este recibo {cnt}/3 veces para {period}. Si necesitás más, avisá a RRHH.")

        # ✅ VERIFICAR T&C antes de enviar PDF
        if not has_accepted_terms(from_whatsapp):
            origin = (pending.get("origin") or "INITIAL")
            send_terms_and_conditions(from_whatsapp, tenant, cuil, period)
            set_pending_terms_acceptance(from_whatsapp, tenant, cuil, period, origin)
            _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "VIEW_NOW", "SENT_TERMS")
            return Response("OK", status=200)

        sid_pdf = _send_pdf_flow(from_whatsapp, tenant, cuil, period)
        if not sid_pdf:
            _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "VIEW_NOW", "ERROR")
            return twiml("❌ No pude enviar el PDF en este momento. Avisá a RRHH.")

        n = inc_receipt_request_count(tenant, cuil, period, from_whatsapp)
        _log_receipt_request_event(tenant, cuil, period, from_whatsapp, "VIEW_NOW", "SENT", message_sid=sid_pdf)
        return Response("OK", status=200)

    # NO_NEED
    if button == "NO_NEED" or body == "NO_NEED":
        set_recibo_estado(tenant, cuil, period, "NO_NEED")
        consume_pending_view(pending["id"])
        return twiml("✅ Perfecto, no hay problema.")

    # SIGN_OK / SIGN_OBS
    if button in ("SIGN_OK", "SIGN_OBS") or body in ("SIGN_OK", "SIGN_OBS"):
        if button == "SIGN_OK" or body == "SIGN_OK":
            set_recibo_estado(tenant, cuil, period, "FIRMADO")
            consume_pending_view(pending["id"])
            return twiml("✅ Recibo firmado. ¡Gracias!")
        else:
            set_recibo_estado(tenant, cuil, period, "OBSERVADO")
            consume_pending_view(pending["id"])
            return twiml("📝 Recibo observado. Contacte RRHH para más información.")

    return Response("OK", status=200)

@app.route("/admin/reenviar_fallidos", methods=["GET", "POST"])
def admin_reenviar_fallidos():
    """
    Reenviar PDFs a empleados específicos.
    """
    auth = require_admin()
    if auth:
        return auth

    if request.method == "POST":
        tenant = request.form.get("tenant", "").strip()
        period = request.form.get("period", "").strip()
        cuils_text = request.form.get("cuils", "").strip()
        
        # Parsear CUILs (uno por línea o separados por coma)
        cuils = []
        for line in cuils_text.replace(",", "\n").split("\n"):
            cuil = norm_cuil(line.strip())
            if cuil and len(cuil) == 11:
                cuils.append(cuil)
        
        if not tenant or not period or not cuils:
            return Response("Falta tenant, period o CUILs", status=400)
        
        # Cargar envios para obtener WhatsApp
        envios = load_envios_rows(tenant)
        
        resultados = []
        for cuil in cuils:
            # Buscar datos del empleado
            person = find_person_by_cuil(envios, cuil)
            if not person:
                resultados.append(f"❌ {cuil}: No encontrado en envíos")
                continue
            
            nombre = person.get('nombre', '')

            # Buscar WhatsApp en múltiples columnas posibles del Excel
            whatsapp = (
                person.get('whatsapp') or 
                person.get('telefono') or 
                person.get('celular') or 
                person.get('phone') or 
                person.get('numero') or
                ''
            ).strip()

            # Si no hay WhatsApp en el Excel, buscar en la BD
            if not whatsapp:
                # Intentar obtener desde message_status
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute("""
                    SELECT to_whatsapp FROM message_status
                    WHERE tenant = %s AND cuil = %s
                    ORDER BY created_at DESC LIMIT 1
                """, (tenant, cuil))
                row = cur.fetchone()
                conn.close()

                if row and row['to_whatsapp']:
                    whatsapp = row['to_whatsapp']
                else:
                    resultados.append(f"❌ {cuil} ({nombre}): Sin WhatsApp en Excel ni BD")
                    continue

            # Normalizar formato WhatsApp
            if not whatsapp.startswith('whatsapp:'):
                whatsapp = norm_whatsapp(whatsapp)
            
            # Enviar PDF
            try:
                sid = _send_pdf_flow(whatsapp, tenant, cuil, period, origin="INITIAL")
                if sid:
                    resultados.append(f"✅ {cuil} ({nombre}): PDF reenviado - SID: {sid}")
                else:
                    resultados.append(f"❌ {cuil} ({nombre}): PDF no encontrado en Drive")
            except Exception as e:
                resultados.append(f"❌ {cuil} ({nombre}): Error - {str(e)}")
        
        # Mostrar resultados
        html = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Resultados del reenvío</title>
  <style>
    body {{ font-family: monospace; padding: 20px; background: #0f1629; color: #eaf0ff; }}
    .success {{ color: #10b981; }}
    .error {{ color: #ef4444; }}
    a {{ color: #5aa7ff; text-decoration: none; }}
    a:hover {{ text-decoration: underline; }}
    .result {{ padding: 8px; border-bottom: 1px solid #333; }}
  </style>
</head>
<body>
  <h1>📊 Resultados del reenvío</h1>
  <p><strong>Tenant:</strong> {esc(tenant)}</p>
  <p><strong>Período:</strong> {esc(period)}</p>
  <p><strong>Total procesados:</strong> {len(resultados)}</p>
  <hr>
"""
        
        for r in resultados:
            css_class = "success" if "✅" in r else "error"
            html += f"<div class='result {css_class}'>{esc(r)}</div>"
        
        html += f"""
  <hr>
  <p><a href="/admin/reenviar_fallidos">← Volver</a></p>
</body>
</html>
"""
        return Response(html, mimetype="text/html")
    
    # GET: Mostrar formulario
    html = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Reenviar PDFs fallidos</title>
  <style>
    body {{
      font-family: system-ui, sans-serif;
      background: #0f1629;
      color: #eaf0ff;
      padding: 20px;
      max-width: 800px;
      margin: 0 auto;
    }}
    h1 {{ color: #5aa7ff; }}
    .card {{
      background: rgba(255,255,255,0.05);
      border: 1px solid rgba(255,255,255,0.1);
      border-radius: 12px;
      padding: 24px;
      margin: 20px 0;
    }}
    label {{
      display: block;
      margin: 16px 0 8px 0;
      font-weight: 600;
      color: #9fb2d0;
    }}
    input, textarea {{
      width: 100%;
      padding: 12px;
      border: 1px solid rgba(255,255,255,0.2);
      border-radius: 8px;
      background: rgba(0,0,0,0.3);
      color: #eaf0ff;
      font-family: monospace;
      font-size: 14px;
    }}
    textarea {{
      min-height: 200px;
      resize: vertical;
    }}
    button {{
      background: linear-gradient(135deg, #F4C430, #d4a514);
      color: #1f2766;
      border: none;
      padding: 14px 32px;
      border-radius: 8px;
      font-weight: 700;
      font-size: 16px;
      cursor: pointer;
      margin-top: 20px;
    }}
    button:hover {{
      background: linear-gradient(135deg, #d4a514, #F4C430);
    }}
    .hint {{
      font-size: 13px;
      color: #9fb2d0;
      margin-top: 6px;
    }}
    .example {{
      background: rgba(0,0,0,0.3);
      padding: 12px;
      border-radius: 6px;
      font-family: monospace;
      font-size: 13px;
      margin-top: 8px;
    }}
    a {{
      color: #5aa7ff;
      text-decoration: none;
    }}
    a:hover {{
      text-decoration: underline;
    }}
  </style>
</head>
<body>
  <h1>🔄 Reenviar PDFs fallidos</h1>
  
  <div class="card">
    <p><strong>⚠️ Importante:</strong></p>
    <ul>
      <li>Esto reenviará el PDF a los CUILs especificados</li>
      <li>Cuando el PDF sea entregado, se enviará automáticamente el botón de firma</li>
      <li>Solo funciona si el PDF existe en Google Drive</li>
    </ul>
  </div>
  
  <form method="post">
    
    <div class="card">
      <label>Tenant</label>
      <input type="text" name="tenant" placeholder="san-patricio" required>
      
      <label>Período</label>
      <input type="text" name="period" placeholder="04/2026" required>
      <div class="hint">Formato: MM/YYYY</div>
      
      <label>CUILs (uno por línea o separados por coma)</label>
      <textarea name="cuils" placeholder="27357233831
27131464563
20-12345678-9" required></textarea>
      <div class="hint">Podés pegar desde Excel, uno por línea o separados por comas</div>
      
      <div class="example">
        <strong>Ejemplo:</strong><br>
        27357233831<br>
        27131464563<br>
        20-12345678-9
      </div>
      
      <button type="submit">🚀 Reenviar PDFs</button>
    </div>
  </form>
  
  <p><a href="/admin">← Volver al admin</a></p>
</body>
</html>
"""
    
    return Response(html, mimetype="text/html")

@app.get("/media/certificado")
def media_certificado():
    #"\"\"Sirve un certificado médico desde Drive (para el panel admin).\"\"\"
    token = request.args.get("token", "").strip()
    if not ADMIN_TOKEN or token != ADMIN_TOKEN:
        # Fail-closed: sin ADMIN_TOKEN configurado este endpoint queda bloqueado.
        return Response("Unauthorized", status=401)
 
    file_id = (request.args.get("file_id") or "").strip()
    if not file_id:
        return Response("Falta file_id", status=400)
 
    try:
        service = drive_service()
        meta = service.files().get(fileId=file_id, fields='name,mimeType').execute()
        file_name = meta.get('name', 'certificado')
        mime = meta.get('mimeType', 'application/octet-stream')
 
        req = service.files().get_media(fileId=file_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, req, chunksize=5*1024*1024)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        data = fh.read()
 
        resp = Response(data, mimetype=mime)
        resp.headers["Content-Disposition"] = f'inline; filename="{file_name}"'
        resp.headers["Content-Length"] = str(len(data))
        resp.headers["Cache-Control"] = "private, max-age=60"
        return resp
    except Exception as e:
        log.exception(f"❌ Error descargando certificado: {e}")
        return Response("Error descargando certificado", status=500)


def _send_pdf_flow(from_whatsapp: str, tenant: str, cuil: str, period: str, origin: str = "INITIAL") -> str | None:
    """
    Envía el PDF del recibo.
    - origin="INITIAL": envío inicial (RRHH) -> puede disparar firma al entregar.
    - origin="RESEND": reenvío pedido por el usuario -> NO debe disparar firma.
    """
    file_id = find_pdf_file_id_for_cuil_period(tenant, cuil, period)
    if not file_id:
        return None


    # ✅ URL firmada y con vencimiento (ya no viaja ADMIN_TOKEN a Twilio)
    t_q = (tenant or "").strip().lower()
    c_q = (cuil or "").strip()
    p_q = (period or "").strip()
    qs = urlencode({
        "tenant": t_q,
        "cuil": c_q,
        "period": p_q,
        "token": make_media_token("pdf", tenant=t_q, cuil=c_q, period=p_q),
    })
    pdf_url = f"{request.host_url.rstrip('/')}/media/pdf?{qs}"
    try:
        # ✅ Para controlar el orden:
        # - En INITIAL podemos incluir body (si querés).
        # - En RESEND lo mandamos vacío y notificamos después del delivered desde /twilio/status.
        body_text = f"Acá tenés tu recibo {period}." if origin == "INITIAL" else ""

        sid_pdf = send_whatsapp_pdf(
            from_whatsapp,
            pdf_url,
            body=body_text,
            status_callback=STATUS_CALLBACK_URL,
        )


        # ✅ Guardamos SID + ORIGIN (para decidir si enviar firma o no al delivered)
        save_pdf_sid(tenant, cuil, period, from_whatsapp, sid_pdf, origin=origin)

        return sid_pdf

    except Exception as e:
        log.exception("%s %s", "ERROR sending PDF:", e)
        return None


@app.get("/health")
def health():
    return jsonify({"ok": True, "ts": int(time.time()), "db_path": DB_PATH})
